def build_admin_dashboard_data(user, branch_id=None):
    """Collect metrics, listings, and charts for the admin/HR dashboard using the current schema.
    
    Args:
        user: The current user object
        branch_id: Optional branch_id for filtering. If None, uses branch_scope for HR users with assigned branch.
                   For admin users, None means all branches.
    """
    role = user.get('role') if user else None
    role_label = 'System Administrator' if role == 'admin' else 'HR Manager'
    # Initialize dashboard containers to safe defaults to avoid UnboundLocalError
    stats = {}
    metrics = {}
    chart_data = {
        'applications_over_time': {'labels': [], 'data': []},
        'status_distribution': {'labels': [], 'data': []},
        'top_jobs': {'labels': [], 'data': []},
        'branch_performance': {'labels': [], 'data': []},
    }
    formatted_recent_applications = []
    formatted_recent_jobs = []
    trends = {}
    alerts = []
    branch_info = {}
    
    # STRICT BRANCH SCOPING: If branch_id not provided, get user's assigned branch
    # Admin users: branch_id remains None (access all branches)
    # HR users with assigned branch: use that branch_id (ENFORCE strict scoping)
    # HR users without assigned branch: branch_id remains None (access all branches)
    if branch_id is None:
        if role == 'hr':
            # Get HR user's assigned branch (if any)
            branch_id = get_branch_scope(user)
    # Prepare pending applications query and params (ensure variables are defined)
    pending_query = "SELECT COUNT(*) AS count FROM applications a JOIN jobs j ON j.job_id = a.job_id WHERE a.status = 'pending'"
    pending_params = []
    if branch_id:
        pending_query += " AND j.branch_id = %s"
        pending_params.append(branch_id)
    stats['pending_this_week'] = fetch_count(
        pending_query + " AND a.applied_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)",
        tuple(pending_params),
    )

    def dated_application_count(condition):
        query = f"""
            SELECT COUNT(*) AS count
            FROM applications a
            JOIN jobs j ON j.job_id = a.job_id
            WHERE {condition}
        """
        params = []
        if branch_id:
            query += " AND j.branch_id = %s"
            params.append(branch_id)
        return fetch_count(query, tuple(params))

    stats['new_applications_today'] = dated_application_count("DATE(a.applied_at) = CURDATE()")
    stats['new_applications_week'] = dated_application_count("a.applied_at >= DATE_SUB(NOW(), INTERVAL 7 DAY)")

    def interviews_count(condition):
        query = f"""
            SELECT COUNT(*) AS count
            FROM interviews i
            JOIN applications a ON a.application_id = i.application_id
            JOIN jobs j ON j.job_id = a.job_id
            WHERE {condition}
        """
        params = []
        if branch_id:
            query += " AND j.branch_id = %s"
            params.append(branch_id)
        return fetch_count(query, tuple(params))

    stats['total_interviews'] = interviews_count('1=1')
    stats['interviews_today'] = interviews_count('DATE(i.scheduled_date) = CURDATE()')

    # Compute additional dashboard totals (open jobs, total applicants, hired apps, total apps)
    try:
        # Open jobs: count jobs with status indicating visible/open
        if branch_id:
            stats['open_jobs'] = fetch_count("SELECT COUNT(*) AS count FROM jobs j WHERE j.branch_id = %s AND (COALESCE(j.status,'open') IN ('open','active'))", (branch_id,))
            stats['total_applications'] = fetch_count("SELECT COUNT(*) AS count FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = %s", (branch_id,))
            stats['hired_applications'] = fetch_count("SELECT COUNT(*) AS count FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE LOWER(COALESCE(a.status,'')) = 'hired' AND j.branch_id = %s", (branch_id,))
        else:
            stats['open_jobs'] = fetch_count("SELECT COUNT(*) AS count FROM jobs WHERE COALESCE(status,'open') IN ('open','active')")
            stats['total_applications'] = fetch_count("SELECT COUNT(*) AS count FROM applications")
            stats['hired_applications'] = fetch_count("SELECT COUNT(*) AS count FROM applications WHERE LOWER(COALESCE(status,'')) = 'hired'")

        # Total applicants (global)
        stats['total_applicants'] = fetch_count('SELECT COUNT(*) AS count FROM applicants')

        # Success rate: percent hired / total_applications
        try:
            total_apps = int(stats.get('total_applications') or 0)
            hired = int(stats.get('hired_applications') or 0)
            stats['success_rate'] = int((hired / total_apps * 100) if total_apps > 0 else 0)
        except Exception:
            stats['success_rate'] = 0
    except Exception as totals_err:
        print(f'⚠️ Error computing dashboard totals: {totals_err}')
        stats.setdefault('open_jobs', 0)
        stats.setdefault('total_applicants', 0)
        stats.setdefault('total_applications', 0)
        stats.setdefault('hired_applications', 0)
        stats.setdefault('success_rate', 0)

    interview_filter = "AND j.branch_id = %s" if branch_id else ""
    interview_params = [branch_id] if branch_id else []
    # Determine if `location` column exists on interviews to avoid SQL errors
    interview_location_expr = "i.location"
    try:
        db = get_db()
        if db:
            cur_check = db.cursor()
            try:
                cur_check.execute("SHOW COLUMNS FROM interviews LIKE 'location'")
                if not cur_check.fetchone():
                    interview_location_expr = "'' AS location"
            finally:
                cur_check.close()
    except Exception:
        interview_location_expr = "'' AS location"

    upcoming_interviews = fetch_rows(
        f"""
        SELECT i.interview_id,
               ap.full_name AS applicant_name,
               j.job_title AS job_title,
               i.scheduled_date,
               COALESCE(i.interview_mode, 'in-person') AS interview_mode,
               i.interview_mode,
               a.status AS application_status,
               {interview_location_expr}
        FROM interviews i
        JOIN applications a ON a.application_id = i.application_id
        JOIN applicants ap ON ap.applicant_id = a.applicant_id
        JOIN jobs j ON j.job_id = a.job_id
        WHERE i.scheduled_date >= NOW() {interview_filter}
        ORDER BY i.scheduled_date ASC
        LIMIT 5
        """,
        tuple(interview_params),
    )

    formatted_upcoming_interviews = [
        {
            'interview_id': row.get('interview_id'),
            'applicant_name': row.get('applicant_name'),
            'position_applied': row.get('job_title'),
            'date_time': format_human_datetime(row.get('scheduled_date')),
            'type': (row.get('interview_mode') or 'in-person').replace('-', ' ').title(),
            'location': row.get('location'),
            'status_key': (row.get('application_status') or '').lower(),
            'status_label': (row.get('application_status') or '').replace('_', ' ').title(),
        }
        for row in upcoming_interviews
    ]

    recent_activity = []
    for item in formatted_recent_applications[:5]:
        recent_activity.append({
            'type': 'application',
            'description': f"{item['applicant_name']} applied for {item['position_applied']}",
            'timestamp': format_human_datetime(item['applied_at']),
            'status': item['status_key'],
        })
    for item in formatted_upcoming_interviews[:5]:
        recent_activity.append({
            'type': 'interview',
            'description': f"Interview scheduled for {item['applicant_name']} ({item['position_applied']})",
            'timestamp': item['date_time'],
            'status': item['status_key'],
        })
    for item in formatted_recent_jobs[:3]:
        recent_activity.append({
            'type': 'job',
            'description': f"Job update: {item['title']}",
            'timestamp': item['posted_at'],
            'status': item['status_key'],
        })

    # Chart data - applications over last 30 days
    chart_params = []
    applications_over_time_query = """
        SELECT DATE(a.applied_at) AS submitted_date,
               COUNT(*) AS total
        FROM applications a
        JOIN jobs j ON j.job_id = a.job_id
        WHERE a.applied_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)
    """
    if branch_id:
        applications_over_time_query += " AND j.branch_id = %s"
        chart_params.append(branch_id)
    applications_over_time_query += " GROUP BY DATE(a.applied_at) ORDER BY submitted_date ASC"
    applications_over_time = fetch_rows(applications_over_time_query, tuple(chart_params))
    chart_data['applications_over_time']['labels'] = [
        row['submitted_date'].strftime('%b %d') if row.get('submitted_date') else ''
        for row in applications_over_time
    ]
    chart_data['applications_over_time']['data'] = [row['total'] for row in applications_over_time]

    status_params = []
    status_distribution_query = """
        SELECT a.status, COUNT(*) AS total
        FROM applications a
        JOIN jobs j ON j.job_id = a.job_id
        WHERE 1=1
    """
    if branch_id:
        status_distribution_query += " AND j.branch_id = %s"
        status_params.append(branch_id)
    status_distribution_query += " GROUP BY a.status"
    status_distribution = fetch_rows(status_distribution_query, tuple(status_params))
    
    # Map statuses: normalize legacy statuses to canonical ones
    # Note: 'withdrawn' status has been removed - applications with withdrawn status should be manually reviewed
    status_map = {}
    status_normalization = {
        'pending': 'pending',
        'scheduled': 'scheduled',
        'interviewed': 'interviewed',
        'hired': 'hired',
        'rejected': 'rejected',
        'reviewed': 'pending',
        'shortlisted': 'pending',
        'accepted': 'hired',
        'applied': 'pending',
        'under_review': 'pending',
        'interview': 'interviewed'
    }
    
    for row in status_distribution:
        if row.get('status') and row.get('total', 0) > 0:  # Only include statuses with count > 0
            status = row['status'].lower()
            # Skip withdrawn status completely - don't include it in chart
            if status == 'withdrawn':
                continue
            # Normalize status to canonical one
            normalized_status = status_normalization.get(status, status)
            # Skip if normalized to withdrawn (shouldn't happen, but safety check)
            if normalized_status == 'withdrawn':
                continue
            # Combine counts for same normalized status
            if normalized_status in status_map:
                status_map[normalized_status] += row['total']
            else:
                status_map[normalized_status] = row['total']
    
    # Only include statuses with count > 0, in a consistent order
    # Define the order: scheduled, interviewed, hired, rejected (4 statuses only)
    # Note: withdrawn and pending are explicitly excluded
    status_order = ['scheduled', 'interviewed', 'hired', 'rejected']
    ordered_statuses = []
    for status in status_order:
        if status in status_map and status_map[status] > 0:
            ordered_statuses.append((status, status_map[status]))
    
    # Add any other statuses not in the standard order (but exclude withdrawn)
    for status, count in status_map.items():
        if status not in status_order and status != 'withdrawn' and count > 0:
            ordered_statuses.append((status, count))
    
    chart_data['status_distribution']['labels'] = [
        status.replace('_', ' ').title() for status, count in ordered_statuses
    ]
    chart_data['status_distribution']['data'] = [
        count for status, count in ordered_statuses
    ]

    top_job_params = []
    # Determine job title expression based on current jobs schema
    job_title_expr = job_column_expr('job_title', alias='j', alternatives=['title'], default="'Untitled Job'")
    top_job_params = []
    top_jobs_query = f"""
        SELECT {job_title_expr} AS job_title, COUNT(a.application_id) AS total
        FROM jobs j
        LEFT JOIN applications a ON a.job_id = j.job_id
        WHERE 1=1
    """
    if branch_id:
        top_jobs_query += " AND j.branch_id = %s"
        top_job_params.append(branch_id)
    top_jobs_query += f" GROUP BY j.job_id, {job_title_expr} ORDER BY total DESC, {job_title_expr} ASC LIMIT 5"
    top_jobs = fetch_rows(top_jobs_query, tuple(top_job_params))
    chart_data['top_jobs']['labels'] = [row['job_title'] for row in top_jobs]
    chart_data['top_jobs']['data'] = [row['total'] for row in top_jobs]

    if branch_id is None:
        branch_performance = fetch_rows(
            """
            SELECT b.branch_name, COUNT(a.application_id) AS total
            FROM branches b
            LEFT JOIN jobs j ON j.branch_id = b.branch_id
            LEFT JOIN applications a ON a.job_id = j.job_id
            GROUP BY b.branch_id, b.branch_name
            ORDER BY total DESC
            LIMIT 5
            """
        )
        chart_data['branch_performance']['labels'] = [row['branch_name'] for row in branch_performance]
        chart_data['branch_performance']['data'] = [row['total'] for row in branch_performance]

    # Active sessions snapshot
    active_sessions = fetch_rows(
        """
        SELECT s.session_id,
               s.created_at,
               u.email,
               u.user_type
        FROM auth_sessions s
        JOIN users u ON u.user_id = s.user_id
        WHERE s.is_active = 1
        ORDER BY s.created_at DESC
        LIMIT 5
        """
    )

    dashboard = {
        'user': {
            'full_name': user.get('name') if user else '',
            'email': user.get('email') if user else '',
            'user_type': role_label,
            'role': role,
        },
        'branch_info': branch_info,
        'stats': stats,
        'metrics': metrics,
        'chart_data': chart_data,
        'trends': trends,
        'system_alerts': alerts,
        'recent_jobs': formatted_recent_jobs,
        'recent_applications': formatted_recent_applications,
        'upcoming_interviews': formatted_upcoming_interviews,
        'recent_activity': recent_activity,
        'notifications': [],
        'notifications_count': 0,
        'active_sessions': [
            {
                'session_id': row.get('session_id'),
                'login_time': format_human_datetime(row.get('login_time')),
                'email': row.get('email'),
                'user_type': row.get('user_type'),
            }
            for row in active_sessions
        ],
        'branches': [] if branch_id else fetch_branches(),
        'hr_accounts': fetch_hr_accounts() if role == 'admin' else (
            fetch_rows(
                """
                SELECT a.admin_id, a.full_name, u.email
                FROM admins a
                JOIN users u ON u.user_id = a.user_id
                WHERE u.user_type = 'hr'
                ORDER BY a.full_name ASC
                """,
                None,
            )
        ),
    }

    return dashboard

import os
import mimetypes
import re
import html as _html
import traceback
import logging

log = logging.getLogger(__name__)
from dotenv import load_dotenv

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g, send_file, send_from_directory, Response
from flask_wtf.csrf import CSRFProtect, CSRFError, generate_csrf
from functools import wraps
from datetime import datetime, date, timedelta, timezone
from uuid import uuid4
from decimal import Decimal, InvalidOperation
from threading import Lock

load_dotenv()
from config import Config
from utils.database import get_db, close_db, execute_query
from utils.auth import (
    hash_password,
    check_password,
    login_user,
    logout_user,
    get_current_user,
    is_logged_in,
)
from utils.helpers import save_uploaded_file
from utils.mailer import send_email
from utils.rate_limit import rate_limit
from utils.theme import get_branch_theme_css, get_branch_logo_html, get_branch_banner_style

def _track_failed_login():
    """Helper function to track failed login attempts for rate limiting."""
    from utils.rate_limit import _rate_limit_store, _rate_limit_lock
    from datetime import datetime
    from flask import request
    
    identifier = request.remote_addr or 'unknown'
    key = f"login:{identifier}"
    with _rate_limit_lock:
        if key not in _rate_limit_store:
            _rate_limit_store[key] = []
        _rate_limit_store[key].append(datetime.now())


# ============================================================================
# AUTOMATION SYSTEM - Automatic Notifications and Emails
# ============================================================================

def auto_notify_and_email(cursor, application_id, message, email_subject=None, email_body=None, recipient_email=None, recipient_name=None):
    """
    Automatically create notification and send email for any system event.
    This ensures all events are automatically communicated to users.
    Prevents duplicate notifications by checking if the same message already exists.
    Also prevents JSON responses from being saved as notifications.
    """
    try:
        # Prevent JSON responses from being saved as notifications
        message_str = str(message).strip()
        if message_str.startswith('{') and ('"success"' in message_str or '"message"' in message_str or '"error"' in message_str):
            print(f'⚠️ Blocked JSON response from being saved as notification in auto_notify_and_email: {message_str[:100]}')
            return
        
        ensure_schema_compatibility()
        notification_columns = set()
        try:
            cursor.execute('SHOW COLUMNS FROM notifications')
            notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        except Exception as col_error:
            print(f"⚠️ Error checking notification columns: {col_error}")
        
        # Check if notification already exists to prevent duplicates
            cursor.execute(
                '''
                SELECT notification_id FROM notifications
                WHERE application_id = %s AND message = %s
                LIMIT 1
                ''',
                (application_id, message),
            )
        existing_notification = cursor.fetchone()
        
        if existing_notification:
            print(f'⚠️ Duplicate notification detected for application {application_id}. Skipping notification creation and email to prevent duplicates.')
            # Don't send email if notification already exists (prevents duplicate emails)
            return
        
        # Automatically create notification in system (only if not duplicate)
        if 'sent_at' in notification_columns:
            cursor.execute(
                '''
                INSERT INTO notifications (application_id, message, sent_at, is_read)
                VALUES (%s, %s, NOW(), 0)
                ''',
                (application_id, message),
            )
        else:
            cursor.execute(
                '''
                INSERT INTO notifications (application_id, message, is_read)
                VALUES (%s, %s, 0)
                ''',
                (application_id, message),
            )
        print(f'✅ Notification created in system for application {application_id}: {message[:50]}...')
        
        # Email notifications are always enabled (system settings removed)
        email_enabled = True
        # Automatically send email if enabled and recipient info provided
        # Only send email if notification was created (not duplicate)
        if email_enabled and recipient_email and email_subject and email_body:
            try:
                # Validate email parameters before sending
                if recipient_email and email_subject and email_body:
                    send_email(recipient_email, email_subject, email_body)
                    print(f'✅ Email sent to {recipient_email} for application {application_id}')
            except Exception as email_error:
                print(f"⚠️ Auto-email error (non-blocking): {email_error}")
                # Continue even if email fails - notification is created in system
        else:
            if not email_enabled:
                print(f'⚠️ Email not sent (disabled) for application {application_id}')
            elif not recipient_email:
                print(f'⚠️ Email not sent (missing recipient email) for application {application_id}')
            elif not email_subject or not email_body:
                print(f'⚠️ Email not sent (missing subject or body) for application {application_id}')
    except Exception as notify_error:
        log.exception(f"⚠️ Auto-notification error (non-blocking): {notify_error}")
        # Continue even if notification fails


def auto_update_application_status(cursor, application_id, new_status, reason=''):
    """
    Automatically update application status and notify applicant via system notification and email.
    """
    try:
        # Ensure schema compatibility to get correct job column names
        ensure_schema_compatibility()
        db_temp = get_db()
        if db_temp:
            cursor_temp = db_temp.cursor()
            try:
                _update_job_columns(cursor_temp)
            finally:
                cursor_temp.close()
        
        # Get job title expression - ensure it uses 'j' alias for jobs table
        job_title_expr = job_column_expr('job_title', alias='j', alternatives=['title'], default="'Untitled Job'")
        
        # Get applicant info before updating - ensure we get the correct job for THIS specific application
        cursor.execute(
            f'''
            SELECT ap.applicant_id, ap.email, ap.full_name, a.status AS old_status, 
                   a.job_id,
                   COALESCE({job_title_expr}, 'Untitled Job') AS job_title
            FROM applicants ap
            JOIN applications a ON ap.applicant_id = a.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            WHERE a.application_id = %s
            LIMIT 1
            ''',
            (application_id,)
        )
        applicant_info = cursor.fetchone()
        
        if applicant_info:
            old_status = applicant_info.get('old_status')
            
            # Update status in database - guard against missing `updated_at` column
            cursor.execute("SHOW COLUMNS FROM applications LIKE 'updated_at'")
            _col = cursor.fetchone()
            if _col:
                cursor.execute(
                    'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s',
                    (new_status, application_id),
                )
            else:
                cursor.execute(
                    'UPDATE applications SET status = %s WHERE application_id = %s',
                    (new_status, application_id),
                )
            print(f'✅ Status updated: Application {application_id} -> {new_status}')
            # If status is hired, ensure interviews are marked completed or created
            try:
                if str(new_status).lower() == 'hired':
                    _sync_interviews_on_application_hired(cursor, application_id)
            except Exception as _sync_err:
                print(f'⚠️ Error syncing interviews after hire: {_sync_err}')
            
            # Auto-notify and email applicant
            status_display = new_status.replace('_', ' ').title()
            job_title = applicant_info.get('job_title') or 'Your Application'
            applicant_name = applicant_info.get('full_name') or 'Applicant'
            applicant_email = applicant_info.get('email')
            
            # Special handling for "hired" status
            if new_status.lower() == 'hired':
                message = f'Congratulations! You have been hired for the position: "{job_title}". Welcome to the team!'
                email_subject = f'Congratulations! You\'ve Been Hired - {job_title}'
                email_body = f"""Dear {applicant_name},

Congratulations! We are pleased to inform you that you have been selected for the position of {job_title}.

We are excited to welcome you to our team and look forward to working with you.

Please expect a call within 24hrs.

{f"Additional Notes: {reason}" if reason else ""}

Please log in to your account to view more details and next steps.

J&T Express Recruitment Team
                """.strip()
            elif new_status.lower() == 'rejected':
                message = f'Your application status for "{job_title}" has been updated to: {status_display}'
                if reason:
                    message += f' - {reason}'
                email_subject = f'Application Status Update - {job_title}'
                email_body = f"""Dear {applicant_name},

Thank you for your interest in the position "{job_title}".

After careful consideration, we regret to inform you that we have decided to move forward with other candidates at this time.

{f"Reason: {reason}" if reason else ""}

We appreciate your time and interest in our company. We encourage you to apply for future positions that match your qualifications.

Best regards,
J&T Express Recruitment Team
                """.strip()
            else:
                message = f'Your application status for "{job_title}" has been updated to: {status_display}'
                if reason:
                    message += f' - {reason}'
                email_subject = f'Application Status Update - {job_title}'
                email_body = f"""Dear {applicant_name},

Your application status for the job position "{job_title}" has been updated.

New Status: {status_display}
{f"Reason: {reason}" if reason else ""}

Please log in to your account to view more details.

Best regards,
J&T Express Recruitment Team
                """.strip()
            
            # Create notification and send email to applicant
            if applicant_email:
                auto_notify_and_email(
                    cursor, application_id, message,
                    email_subject, email_body,
                    applicant_email,
                    applicant_name
                )
                print(f'✅ Notification created and email sent to applicant {applicant_info.get("applicant_id")} ({applicant_email}) for application {application_id}')
            else:
                # Still create notification even if no email
                auto_notify_and_email(
                    cursor, application_id, message,
                    None, None,
                    None, None
                )
                print(f'✅ Notification created for application {application_id} (no email - applicant email missing)')
            
            return True
        else:
            print(f'⚠️ No applicant info found for application {application_id}')
    except Exception as auto_error:
        log.exception(f"⚠️ Auto-status update error: {auto_error}")
    return False


def auto_handle_job_status(cursor, job_id, new_status, old_status=None):
    """
    Automatically handle job status changes (posted_at, notifications, etc.)
    """
    try:
        # Check if posted_at column exists
        posted_at_col = job_column_name('posted_at')
        
        # Automatically set posted_at when status becomes 'active' or 'open'
        # Always update to current server time (NOW()) to ensure accurate posting time
        if new_status in ('active', 'open'):
            if posted_at_col:
                cursor.execute(
                    f'UPDATE jobs SET {posted_at_col} = NOW() WHERE job_id = %s',
                    (job_id,)
                )
            else:
                # If posted_at doesn't exist, try created_at
                if 'created_at' in JOB_COLUMNS:
                    cursor.execute(
                        'UPDATE jobs SET created_at = NOW() WHERE job_id = %s',
                        (job_id,)
                    )
        # Automatically clear posted_at when status becomes 'closed'
        elif new_status == 'closed':
            # Keep posted_at for historical record, just update status
            pass
        
        return True
    except Exception as auto_error:
        print(f"⚠️ Auto-job status error: {auto_error}")
    return False


VALID_JOB_STATUSES = ('active', 'closed')
PUBLISHABLE_JOB_STATUSES = ('open',)  # Only 'open' status is visible to applicants (database enum: 'open', 'closed')
VALID_EMPLOYMENT_TYPES = ('full_time', 'part_time', 'internship')
VALID_WORK_ARRANGEMENTS = ('onsite', 'remote', 'hybrid', 'field', 'flexible')
VALID_EXPERIENCE_LEVELS = ('entry', 'mid', 'senior', 'lead', 'manager')

# Database enum values: ENUM('pending', 'scheduled', 'interviewed', 'hired', 'rejected')
APPLICATION_STATUSES = ('pending', 'scheduled', 'interviewed', 'hired', 'rejected')
APPLICATION_PIPELINE_STATUSES = ('pending', 'scheduled', 'interviewed')
APPLICATION_REVIEW_STATUSES = ('pending',)
APPLICATION_SUCCESS_STATUSES = ('hired',)
APPLICATION_FAILED_STATUSES = ('rejected',)
APPLICATION_ACTIVE_STATUSES = ('pending', 'scheduled', 'interviewed')
APPLICATION_TERMINAL_STATUSES = ('hired', 'rejected')
APPLICATION_STATUS_LABELS = {
    'pending': 'Pending',
    'scheduled': 'Scheduled',
    'interviewed': 'Interviewed',
    'hired': 'Hired',
    'rejected': 'Rejected',
    # Legacy status mappings for backward compatibility
    'reviewed': 'Pending',
    'accepted': 'Hired',
    'applied': 'Pending',
    'under_review': 'Pending',
    'interview': 'Interviewed',
}
APPLICATION_STATUS_FLOW = ('pending', 'scheduled', 'interviewed', 'hired', 'rejected')


app = Flask(__name__)
app.config.from_object(Config)
csrf = CSRFProtect(app)
# Configure CSRF settings - Disable in development to avoid token mismatch issues with rate limiting
# CSRF configuration: enable by default, allow override via env var
app.config['WTF_CSRF_ENABLED'] = os.environ.get('WTF_CSRF_ENABLED', 'true').lower() in ('1', 'true', 'yes')
app.config['WTF_CSRF_TIME_LIMIT'] = None  # No time limit for CSRF tokens
# Disable redirect pages - use HTTP 302 redirects directly
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Configure secure session cookies
# Note: SECURE should be True in production (HTTPS); default to false for local dev
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', 'false').lower() in ('1', 'true', 'yes')
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript access
app.config['SESSION_COOKIE_SAMESITE'] = os.environ.get('SESSION_COOKIE_SAMESITE', 'Lax')
# Use PERMANENT_SESSION_LIFETIME from Config (seconds) when available
try:
    app.config['PERMANENT_SESSION_LIFETIME'] = int(os.environ.get('PERMANENT_SESSION_LIFETIME_SECONDS', app.config.get('PERMANENT_SESSION_LIFETIME', 3600)))
except Exception:
    app.config['PERMANENT_SESSION_LIFETIME'] = 3600
app.config['SESSION_REFRESH_EACH_REQUEST'] = True  # Refresh session on each request

# Ensure uploads folder exists at startup and respect absolute paths
_upload_folder_cfg = app.config.get('UPLOAD_FOLDER', 'uploads/resumes')
if os.path.isabs(_upload_folder_cfg):
    _upload_path = _upload_folder_cfg
else:
    _upload_path = os.path.join(app.instance_path, _upload_folder_cfg)
os.makedirs(_upload_path, exist_ok=True)
app.config['UPLOAD_FOLDER'] = _upload_path

_schema_lock = Lock()
_schema_checked = False
JOB_COLUMNS = set()


def immediate_redirect(location, code=302):
    """Create an immediate HTTP redirect without showing redirect page."""
    from flask import Response
    response = Response(status=code)
    response.headers['Location'] = location
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    # Ensure minimal response body for immediate redirect
    response.data = b''
    return response


def _update_job_columns(cursor):
    """Refresh the cached set of columns available on the jobs table."""
    global JOB_COLUMNS
    try:
        cursor.execute('SHOW COLUMNS FROM jobs')
        rows = cursor.fetchall() or []
        JOB_COLUMNS = {
            row.get('Field') if isinstance(row, dict) else row[0]
            for row in rows
        }
    except Exception as exc:
        print(f'⚠️ Failed to inspect jobs table columns: {exc}')
        JOB_COLUMNS = set()
    return JOB_COLUMNS


def job_column(preferred, *alternatives):
    """Return the present column name on jobs table, preferring the modern schema."""
    for candidate in (preferred,) + alternatives:
        if candidate in JOB_COLUMNS:
            return candidate
    return None


def job_column_expr(preferred, alias='j', default='NULL', alternatives=None):
    """Return a SQL expression pointing at an existing jobs column or a safe fallback."""
    alternatives = alternatives or []
    column_name = job_column(preferred, *alternatives)
    if column_name:
        return f'{alias}.{column_name}'
    return default


def job_column_name(preferred, alternatives=None, default=None):
    """Return the raw column name for updates/inserts, or a provided default."""
    alternatives = alternatives or []
    column_name = job_column(preferred, *alternatives)
    if column_name:
        return column_name
    return default


def ensure_schema_compatibility():
    """Best-effort guard to align dynamic queries with the current MySQL schema."""
    global _schema_checked
    if _schema_checked:
        return

    # Try to acquire lock non-blocking - if another thread is checking, just return
    lock_acquired = False
    try:
        lock_acquired = _schema_lock.acquire(blocking=False)
        if not lock_acquired:
            # Another thread is already checking schema, skip this call
            return
        
        if _schema_checked:
            if lock_acquired:
                _schema_lock.release()
            return

        db = get_db()
        if not db:
            if lock_acquired:
                _schema_lock.release()
            return

        cursor = None
        updates_applied = False
        success = False

        def ensure_column(cur, table_name, column_name, column_definition, post_add=None):
            # Whitelist allowed table names to prevent SQL injection
            ALLOWED_TABLES = {'users', 'admins', 'applicants', 'jobs', 'interviews', 'notifications', 
                             'applications', 'auth_sessions', 'resumes', 'application_attachments'}
            if table_name not in ALLOWED_TABLES:
                print(f'⚠️ Invalid table name for column addition: {table_name}')
                return False
            # Validate column name contains only safe characters
            if not column_name.replace('_', '').isalnum():
                print(f'⚠️ Invalid column name: {column_name}')
                return False
            cur.execute(f"SHOW COLUMNS FROM `{table_name}` LIKE %s", (column_name,))
            if cur.fetchone():
                return False
            dangerous_pattern = re.compile(r'\b(DROP|DELETE|TRUNCATE|EXEC(?:UTE)?)\b', re.IGNORECASE)
            if dangerous_pattern.search(column_definition or ''):
                print(f'⚠️ Dangerous keyword in column definition: {column_definition}')
                return False
            cur.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{column_name}` {column_definition}")
            if callable(post_add):
                try:
                    post_add()
                except Exception as copy_exc:
                    print(f'⚠️ Post-add hook for {table_name}.{column_name} failed: {copy_exc}')
            return True

        def ensure_table(cur, table_name, create_sql):
            """Ensure a table exists, create it if it doesn't."""
            # Whitelist allowed table names
            ALLOWED_TABLES = {'positions', 'application_attachments', 'activity_log_deletions'}
            if table_name not in ALLOWED_TABLES:
                print(f'⚠️ Invalid table name: {table_name}')
                return False
            try:
                cur.execute(f"SHOW TABLES LIKE %s", (table_name,))
                if cur.fetchone():
                    return False
                dangerous_pattern = re.compile(r'\b(DROP|DELETE|TRUNCATE|EXEC(?:UTE)?)\b', re.IGNORECASE)
                if dangerous_pattern.search(create_sql or ''):
                    print(f'⚠️ Dangerous keyword in table creation SQL')
                    return False
                cur.execute(create_sql)
                return True
            except Exception as e:
                print(f'⚠️ Failed to ensure table {table_name}: {e}')
                return False

        try:
            cursor = db.cursor()
            _update_job_columns(cursor)

            # Only ensure logout_time exists (last_activity is not in actual schema)
            updates_applied |= ensure_column(cursor, 'auth_sessions', 'logout_time', 'DATETIME NULL DEFAULT NULL')
            try:
                cursor.execute("ALTER TABLE auth_sessions MODIFY logout_time DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP")
            except Exception:
                pass
            try:
                # Some schemas used login_time historically; use created_at as the canonical timestamp
                cursor.execute("UPDATE auth_sessions SET logout_time = created_at WHERE logout_time IS NULL")
                if cursor.rowcount:
                    updates_applied = True
            except Exception:
                pass
            updates_applied |= ensure_column(cursor, 'interviews', 'interview_mode', "VARCHAR(50) NULL DEFAULT NULL")
            
            # Ensure interview status column exists
            # Ensure status column exists with all needed values including 'confirmed' and 'rescheduled'
            try:
                cursor.execute("SHOW COLUMNS FROM interviews LIKE 'status'")
                status_col = cursor.fetchone()
                if status_col:
                    # Check if 'confirmed' is in the enum
                    col_type = status_col.get('Type', '') if isinstance(status_col, dict) else str(status_col[1]) if len(status_col) > 1 else ''
                    if 'confirmed' not in col_type.lower():
                        # Modify enum to include 'confirmed' and 'rescheduled'
                        try:
                            cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                            updates_applied = True
                            print('✅ Added "confirmed" and "rescheduled" to interviews.status enum')
                        except Exception as enum_err:
                            print(f'⚠️ Could not modify status enum: {enum_err}')
                else:
                    # Column doesn't exist, create it
                    updates_applied |= ensure_column(cursor, 'interviews', 'status', "ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
            except Exception as status_check_err:
                print(f'⚠️ Error checking status column: {status_check_err}')
                # Fallback: try to ensure column exists
                updates_applied |= ensure_column(cursor, 'interviews', 'status', "ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
            
            # Ensure applications table status enum includes 'scheduled'
            try:
                cursor.execute("""
                    SELECT COLUMN_TYPE 
                    FROM INFORMATION_SCHEMA.COLUMNS 
                    WHERE TABLE_SCHEMA = DATABASE() 
                    AND TABLE_NAME = 'applications' 
                    AND COLUMN_NAME = 'status'
                """)
                result = cursor.fetchone()
                if result:
                    current_enum = result[0].lower()
                    if 'scheduled' not in current_enum:
                        # Update enum to include 'scheduled'
                        cursor.execute("""
                            ALTER TABLE applications 
                            MODIFY COLUMN status ENUM('pending', 'scheduled', 'interviewed', 'hired', 'rejected', 'withdrawn') 
                            NOT NULL DEFAULT 'pending'
                        """)
                        updates_applied = True
                        print('✅ Updated applications.status enum to include "scheduled"')
            except Exception as enum_error:
                print(f'⚠️ Could not update applications.status enum: {enum_error}')
                # Continue - enum might already be correct or table might not exist yet
            
            # Ensure positions table exists
            positions_sql = """
                CREATE TABLE IF NOT EXISTS positions (
                    position_id INT AUTO_INCREMENT PRIMARY KEY,
                    title VARCHAR(200) NOT NULL,
                    department VARCHAR(200) NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME NULL DEFAULT NULL ON UPDATE CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
            updates_applied |= ensure_table(cursor, 'positions', positions_sql)
            
            job_columns = _update_job_columns(cursor)
            # Ensure jobs table can store per-job file rules
            updates_applied |= ensure_column(cursor, 'jobs', 'allowed_extensions', "VARCHAR(255) NULL DEFAULT NULL")
            updates_applied |= ensure_column(cursor, 'jobs', 'max_file_size_mb', "INT NULL DEFAULT NULL")
            # Store required file types as JSON array (e.g., ["resume", "letter", "license", "certificate"])
            updates_applied |= ensure_column(cursor, 'jobs', 'required_file_types', "TEXT NULL DEFAULT NULL")
            # Ensure application_attachments table exists for storing multiple attachments per application
            attachments_sql = '''
                CREATE TABLE IF NOT EXISTS application_attachments (
                    attachment_id INT AUTO_INCREMENT PRIMARY KEY,
                    application_id INT NOT NULL,
                    resume_id INT NOT NULL,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    INDEX (application_id),
                    INDEX (resume_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            '''
            updates_applied |= ensure_table(cursor, 'application_attachments', attachments_sql)

            # Ensure activity_log_deletions table exists to record bulk-deletes
            activity_deletions_sql = '''
                CREATE TABLE IF NOT EXISTS activity_log_deletions (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    branch_id INT NULL,
                    deleted_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    INDEX (branch_id),
                    INDEX (deleted_at)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            '''
            updates_applied |= ensure_table(cursor, 'activity_log_deletions', activity_deletions_sql)

            updates_applied |= ensure_column(
                cursor,
                'notifications',
                'sent_at',
                'DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP'
            )
            updates_applied |= ensure_column(
                cursor,
                'notifications',
                'is_read',
                'TINYINT(1) NOT NULL DEFAULT 0'
            )
            # Ensure last login/logout columns exist
            updates_applied |= ensure_column(
                cursor,
                'users',
                'last_login',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'users',
                'last_logout',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'admins',
                'last_login',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'admins',
                'last_logout',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'last_login',
                'DATETIME NULL DEFAULT NULL'
            )
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'last_logout',
                'DATETIME NULL DEFAULT NULL'
            )
            
            # Ensure applications.viewed_at column exists
            updates_applied |= ensure_column(
                cursor,
                'applications',
                'viewed_at',
                'DATETIME NULL'
            )
            
            # Ensure users.email_verified column exists for email verification
            updates_applied |= ensure_column(
                cursor,
                'users',
                'email_verified',
                'TINYINT(1) DEFAULT 0'
            )
            
            # Ensure applicants.verification_token column exists
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'verification_token',
                'VARCHAR(255) NULL DEFAULT NULL'
            )
            
            # Ensure applicants.verification_token_expires column exists for 60-second expiration
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'verification_token_expires',
                'DATETIME NULL DEFAULT NULL'
            )
            
            # Ensure applicants.last_profile_update column exists for tracking profile updates
            updates_applied |= ensure_column(
                cursor,
                'applicants',
                'last_profile_update',
                'DATETIME NULL DEFAULT NULL'
            )
            

            # Ensure resumes.file_type column exists (resume, letter, license)
            updates_applied |= ensure_column(
                cursor,
                'resumes',
                'file_type',
                "VARCHAR(50) NOT NULL DEFAULT 'resume'"
            )

            if updates_applied:
                try:
                    db.commit()
                except Exception:
                    pass

            success = True
        except Exception as schema_err:
            log.warning('⚠️ Schema compatibility check error: %s', schema_err)
        finally:
            if cursor:
                try:
                    cursor.close()
                except Exception:
                    # Ignore errors closing cursor
                    pass
            # Always release lock
            if lock_acquired:
                try:
                    _schema_lock.release()
                except Exception:
                    # Ignore lock release errors
                    pass

        if success:
            _schema_checked = True
    except Exception as outer_err:
        # Handle any errors in lock acquisition or early returns
        if lock_acquired:
            try:
                _schema_lock.release()
            except Exception:
                # Ignore lock release errors
                pass
        log.exception(f'⚠️ Schema compatibility outer error: {outer_err}')


# Register template filters
@app.template_filter('format_human_datetime')
def format_human_datetime_filter(value):
    """Produce a human-readable timestamp in 12-hour format (AM/PM)."""
    if isinstance(value, (datetime, date)):
        dt_value = value if isinstance(value, datetime) else datetime.combine(value, datetime.min.time())
        return dt_value.strftime('%b %d, %Y %I:%M %p')
    elif isinstance(value, str):
        try:
            # Try parsing common datetime formats
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d', '%Y-%m-%d %H:%M']:
                try:
                    dt_value = datetime.strptime(value, fmt)
                    return dt_value.strftime('%b %d, %Y %I:%M %p')
                except ValueError:
                    continue
        except Exception:
            # Ignore parse errors
            pass
    return value or ''


@app.teardown_appcontext
def teardown_database(exception=None):
    """Ensure database connections are closed after each request."""
    close_db(exception)


def fetch_notifications_for(scope=None, limit=5):
    """Fetch notifications with optional scoping.
    
    scope: dict with any of:
      - applicant_id: int → fetch applicant's notifications (application-related)
      - branch_id: int → fetch branch-scoped application notifications
      - system_only: bool → fetch system (application_id IS NULL)
    Returns (formatted_list, unread_count)
    """
    scope = scope or {}
    db = get_db()
    if not db:
        return ([], 0)
    cursor = db.cursor(dictionary=True)
    try:
        ensure_schema_compatibility()
        # Introspect columns
        cursor.execute('SHOW COLUMNS FROM notifications')
        notif_cols = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notif_cols else 'COALESCE(n.created_at, NOW())'
        is_read_expr = 'COALESCE(n.is_read, 0)' if 'is_read' in notif_cols else '0'
        has_application_fk = 'application_id' in notif_cols
        params = []
        joins = ''
        where_parts = []
        # Scope by system vs application related
        if scope.get('system_only'):
            where_parts.append('n.application_id IS NULL')
            # Ensure system-level notifications do not contain applicant-facing messages
            applicant_only_filters_sys = [
                "n.message NOT LIKE 'You applied for%'",
                "n.message NOT LIKE 'Congratulations! You have been hired%'",
                "n.message NOT LIKE 'Your application status%'",
                "n.message NOT LIKE 'Congratulations! You%'",
                "n.message NOT LIKE '%application status%'",
                "n.message NOT LIKE '%status has been updated%'",
            ]
            if where_parts:
                where_parts.append(' AND '.join(applicant_only_filters_sys))
        elif has_application_fk:
            # Application-related joins if needed
            joins = 'JOIN applications a ON n.application_id = a.application_id'
            if scope.get('applicant_id'):
                where_parts.append('a.applicant_id = %s')
                params.append(scope['applicant_id'])
            if scope.get('branch_id'):
                joins += ' JOIN jobs j ON a.job_id = j.job_id'
                where_parts.append('j.branch_id = %s')
                params.append(scope['branch_id'])
        # Build WHERE
        where_sql = ''
        if where_parts:
            where_sql = 'WHERE ' + ' AND '.join(where_parts)
        
            # If fetching for HR (branch_id scope but not applicant_id), exclude applicant-only notifications
            # CRITICAL: HR should NEVER see applicant-facing notifications - these are for applicants only
            if scope.get('branch_id') and not scope.get('applicant_id') and not scope.get('system_only'):
                applicant_only_filters = [
                    "n.message NOT LIKE 'You applied for%'",
                    "n.message NOT LIKE 'Congratulations! You have been hired%'",
                    "n.message NOT LIKE 'Your application status%'",
                    "n.message NOT LIKE 'Congratulations! You%'",  # Catch any variation of congratulations messages to applicants
                    "n.message NOT LIKE '%application status%'",  # Exclude all status update notifications
                    "n.message NOT LIKE '%status has been updated%'",  # Exclude status update messages
                ]
                if where_sql:
                    where_sql += ' AND ' + ' AND '.join(applicant_only_filters)
                else:
                    where_sql = 'WHERE ' + ' AND '.join(applicant_only_filters)
        
        # Fetch list
        cursor.execute(
            f'''
            SELECT n.notification_id,
                   n.message,
                   {sent_at_expr} AS sent_at,
                   {is_read_expr} AS is_read
            FROM notifications n
            {joins}
            {where_sql}
            ORDER BY {sent_at_expr} DESC
            LIMIT %s
            ''',
            tuple(params + [limit]),
        )
        rows = cursor.fetchall() or []
        formatted = []
        for r in rows:
            message = r.get('message', '')
            # Filter out JSON responses that might have been incorrectly stored
            if message and message.strip().startswith('{') and '"success"' in message:
                continue  # Skip JSON responses
            formatted.append({
                'title': 'Update' if not scope.get('system_only') else 'System',
                'message': message,
                'time': format_human_datetime(r.get('sent_at')),
                'is_read': r.get('is_read', False),
            })
        # Unread count
        unread_where = where_sql
        if unread_where:
            unread_where = unread_where + f' AND {is_read_expr} = 0'
        else:
            unread_where = f'WHERE {is_read_expr} = 0'
        cursor.execute(
            f'''
            SELECT COUNT(*) AS unread
            FROM notifications n
            {joins}
            {unread_where}
            ''',
            tuple(params),
        )
        unread_row = cursor.fetchone() or {}
        unread_count = unread_row.get('unread', 0) or 0
        return (formatted, unread_count)
    except Exception:
        return ([], 0)
    finally:
        try:
            cursor.close()
        except Exception:
            pass

def login_required(*roles):
    """Decorator enforcing authentication and optional role-based access control."""

    def decorator(view_func):
        @wraps(view_func)
        def wrapped_view(*args, **kwargs):
            # Prefer redirect for normal browser navigation (GET). Only treat as JSON/AJAX
            # when the request explicitly is JSON, is an XHR, or is a non-GET request that
            # prefers JSON. This prevents direct link access from returning JSON errors.
            wants_json = (
                request.is_json
                or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                or (request.method != 'GET' and (request.accept_mimetypes.accept_json or 'application/json' in (request.headers.get('Accept') or '').lower()))
                or (request.path or '').startswith('/api/')
                or (request.accept_mimetypes.best == 'application/json')
                or ('application/json' in (request.headers.get('Accept') or '').lower())
            )
            if not is_logged_in():
                if wants_json:
                    return jsonify({'success': False, 'error': 'Authentication required'}), 401
                flash('Please login to access this page.', 'error')
                return immediate_redirect(url_for('login', _external=False))

            if roles:
                # Resolve current role robustly: prefer session, fallback to get_current_user()
                current_role = (session.get('user_role') or '')
                if not current_role:
                    user = get_current_user()
                    if user:
                        # `user` may contain 'role' or 'user_type'
                        current_role = (user.get('role') or user.get('user_type') or '')

                current_role = (current_role or '').lower()
                # Normalize common variants (e.g., super_admin -> admin)
                if current_role in ('super_admin', 'superadmin'):
                    current_role = 'admin'
                # Normalize broader variants: treat roles containing 'hr' as 'hr' and containing 'admin' as 'admin'
                try:
                    if 'hr' in current_role:
                        current_role = 'hr'
                    elif 'admin' in current_role:
                        current_role = 'admin'
                except Exception:
                    pass

                allowed = {r.lower() for r in roles}
                if current_role not in allowed:
                    # Do not flash a permission message here to avoid showing the
                    # generic "You do not have permission to access this page." banner
                    # when users navigate via the sidebar. Redirect silently.
                    if wants_json:
                        return jsonify({'success': False, 'error': 'Not authorized'}), 403
                    return immediate_redirect(url_for('index', _external=False))

            return view_func(*args, **kwargs)

        return wrapped_view

    return decorator


@app.before_request
def load_logged_in_user():
    """Attach the current user to the Flask global context for easy template access."""
    # Skip for logout, login, resend_verification, and index routes to prevent blocking
    if request.endpoint in ('logout', 'login', 'resend_verification', 'index', 'favicon', 'chrome_devtools'):
        # Still set current_user for index if logged in, but skip schema check
        if request.endpoint == 'index':
            g.current_user = get_current_user()
        return
    
    # Only run schema compatibility check once, not on every request
    # It's already checked during startup, so skip it here to avoid blocking
    # ensure_schema_compatibility()  # Commented out to prevent blocking on every request
    g.current_user = get_current_user()

def determine_user_friendly_action(path, form_action, target_table):
    """Determine a user-friendly action name from path and form action."""
    path_lower = path.lower()
    form_action_lower = (form_action or '').lower()
    
    # Check form action first (most reliable)
    if form_action_lower in ['add', 'create', 'post']:
        return 'Add'
    elif form_action_lower in ['update', 'edit', 'modify']:
        return 'Update'
    elif form_action_lower == 'delete':
        return 'Delete'
    elif form_action_lower == 'schedule':
        return 'Schedule'
    elif form_action_lower == 'reschedule':
        return 'Reschedule'
    elif form_action_lower == 'cancel':
        return 'Cancel'
    elif form_action_lower == 'update_status':
        return 'Update Status'
    elif form_action_lower == 'bulk_update_status':
        return 'Bulk Update Status'
    
    # Check path patterns
    if '/add' in path_lower or '/create' in path_lower or '/post' in path_lower:
        return 'Add'
    elif '/update' in path_lower or '/edit' in path_lower:
        return 'Update'
    elif '/delete' in path_lower:
        return 'Delete'
    elif '/schedule' in path_lower:
        return 'Schedule'
    elif '/reschedule' in path_lower:
        return 'Reschedule'
    elif '/cancel' in path_lower:
        return 'Cancel'
    elif '/status' in path_lower:
        return 'Update Status'
    
    # Check target table for context
    if target_table == 'jobs' or 'job' in path_lower:
        if 'add' in path_lower or 'create' in path_lower or 'post' in path_lower:
            return 'Add Job'
        elif 'update' in path_lower or 'edit' in path_lower:
            return 'Update Job'
        elif 'delete' in path_lower:
            return 'Delete Job'
    elif target_table == 'applications' or 'application' in path_lower:
        if 'status' in path_lower or 'update_status' in form_action_lower:
            return 'Update Application Status'
        elif 'bulk' in path_lower or 'bulk_update' in form_action_lower:
            return 'Bulk Update Applications'
    elif target_table == 'interviews' or 'interview' in path_lower:
        if 'schedule' in path_lower or form_action_lower == 'schedule':
            return 'Schedule Interview'
        elif 'reschedule' in path_lower or form_action_lower == 'reschedule':
            return 'Reschedule Interview'
        elif 'cancel' in path_lower or form_action_lower == 'cancel':
            return 'Cancel Interview'
        elif 'update' in path_lower or 'edit' in path_lower:
            return 'Update Interview'
        elif 'delete' in path_lower:
            return 'Delete Interview'
    
    # Default fallback
    if form_action:
        return form_action.replace('_', ' ').title()
    elif path:
        path_parts = [p for p in path.split('/') if p and p not in ['admin', 'hr', 'applicant']]
        if path_parts:
            last_part = path_parts[-1]
            if last_part.isdigit() and len(path_parts) > 1:
                return path_parts[-2].replace('-', ' ').replace('_', ' ').title()
            else:
                return last_part.replace('-', ' ').replace('_', ' ').title()
    
    return 'Action'

def log_hr_activity(admin_id, action, target_table, target_id, details=None, skip_notification=False):
    """Log HR activity for admin monitoring."""
    try:
        db = get_db()
        if not db:
            return False
        cursor = db.cursor()
        try:
            # Ensure activity_logs table exists
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS activity_logs (
                    log_id INT AUTO_INCREMENT PRIMARY KEY,
                    admin_id INT,
                    action VARCHAR(255) NOT NULL,
                    target_table VARCHAR(255) NOT NULL,
                    target_id INT,
                    details TEXT,
                    created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    INDEX idx_admin_id (admin_id),
                    INDEX idx_created_at (created_at),
                    INDEX idx_target (target_table, target_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            
            # Check if target_table column exists, add it if not
            try:
                cursor.execute("SHOW COLUMNS FROM activity_logs LIKE 'target_table'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_table VARCHAR(255) NOT NULL DEFAULT '' AFTER action")
                    print('✅ Added target_table column to activity_logs table')
            except Exception as col_err:
                # Column might already exist or table structure is different
                print(f'⚠️ Could not check/add target_table column: {col_err}')
            
            # Check if target_id column exists, add it if not
            try:
                cursor.execute("SHOW COLUMNS FROM activity_logs LIKE 'target_id'")
                if not cursor.fetchone():
                    # Determine position - after target_table if it exists, otherwise after action
                    try:
                        cursor.execute("SHOW COLUMNS FROM activity_logs LIKE 'target_table'")
                        if cursor.fetchone():
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL AFTER target_table")
                        else:
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL AFTER action")
                    except Exception as pos_err:
                        # Fallback: just add after action
                        try:
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL AFTER action")
                        except Exception:
                            # Last resort: add at the end
                            cursor.execute("ALTER TABLE activity_logs ADD COLUMN target_id INT NULL")
                    print('✅ Added target_id column to activity_logs table')
            except Exception as col_err:
                # Column might already exist or table structure is different
                print(f'⚠️ Could not check/add target_id column: {col_err}')
            
            # Insert activity log - determine actual columns present and build INSERT accordingly
            try:
                cursor.execute("SHOW COLUMNS FROM activity_logs")
                col_rows = cursor.fetchall() or []
                col_names = []
                for r in col_rows:
                    # row format may be tuple like (Field, Type, ...)
                    if isinstance(r, dict):
                        col_names.append(r.get('Field'))
                    elif isinstance(r, (list, tuple)) and len(r) > 0:
                        col_names.append(r[0])

                insert_cols = []
                insert_vals = []

                # Prefer user_id if present, otherwise admin_id
                if 'user_id' in col_names:
                    insert_cols.append('user_id')
                    insert_vals.append(admin_id)
                elif 'admin_id' in col_names:
                    insert_cols.append('admin_id')
                    insert_vals.append(admin_id)

                # action column expected
                if 'action' in col_names:
                    insert_cols.append('action')
                    insert_vals.append(action)

                # target_table / target_id best-effort
                if 'target_table' in col_names:
                    insert_cols.append('target_table')
                    insert_vals.append(target_table or '')
                if 'target_id' in col_names:
                    insert_cols.append('target_id')
                    insert_vals.append(target_id)

                # description/details mapping
                if 'description' in col_names:
                    insert_cols.append('description')
                    insert_vals.append(details)
                elif 'details' in col_names:
                    insert_cols.append('details')
                    insert_vals.append(details)

                # Build and execute dynamic INSERT only if we found at least action and one id column
                if insert_cols and 'action' in insert_cols:
                    cols_sql = ', '.join(insert_cols)
                    vals_sql = ', '.join(['%s'] * len(insert_vals))
                    sql = f'INSERT INTO activity_logs ({cols_sql}) VALUES ({vals_sql})'
                    try:
                        cursor.execute(sql, tuple(insert_vals))
                    except Exception as e:
                        err_text = str(e).lower()
                        # Handle foreign key failure for user_id by retrying with admin_id when available
                        if 'foreign key' in err_text and 'user_id' in insert_cols:
                            if 'admin_id' in col_names:
                                try:
                                    alt_cols = [c if c != 'user_id' else 'admin_id' for c in insert_cols]
                                    alt_vals = [v for (c, v) in zip(insert_cols, insert_vals) if c != 'user_id']
                                    # insert the admin_id value in place of user_id
                                    # find index where user_id was and insert admin_id value
                                    user_idx = insert_cols.index('user_id')
                                    alt_vals.insert(user_idx, admin_id)
                                    alt_cols_sql = ', '.join(alt_cols)
                                    alt_vals_sql = ', '.join(['%s'] * len(alt_vals))
                                    alt_sql = f'INSERT INTO activity_logs ({alt_cols_sql}) VALUES ({alt_vals_sql})'
                                    cursor.execute(alt_sql, tuple(alt_vals))
                                except Exception as e2:
                                    print(f'⚠️ Retry insert with admin_id failed: {e2}')
                            else:
                                # As fallback, remove user_id and include admin id in details to avoid FK constraint
                                try:
                                    reduced_cols = [c for c in insert_cols if c != 'user_id']
                                    reduced_vals = [v for (c, v) in zip(insert_cols, insert_vals) if c != 'user_id']
                                    # append admin id to details
                                    new_details = (details or '') + f' [actor_admin_id={admin_id}]'
                                    if 'description' in reduced_cols:
                                        desc_idx = reduced_cols.index('description')
                                        reduced_vals[desc_idx] = new_details
                                    elif 'details' in reduced_cols:
                                        det_idx = reduced_cols.index('details')
                                        reduced_vals[det_idx] = new_details
                                    else:
                                        reduced_cols.append('details')
                                        reduced_vals.append(new_details)
                                    reduced_cols_sql = ', '.join(reduced_cols)
                                    reduced_vals_sql = ', '.join(['%s'] * len(reduced_vals))
                                    reduced_sql = f'INSERT INTO activity_logs ({reduced_cols_sql}) VALUES ({reduced_vals_sql})'
                                    cursor.execute(reduced_sql, tuple(reduced_vals))
                                except Exception as e3:
                                    print(f'⚠️ Fallback insert after FK failure also failed: {e3}')
                        else:
                            raise
                else:
                    # Last resort: try minimal known schema variations
                    try:
                        cursor.execute('INSERT INTO activity_logs (user_id, action, description) VALUES (%s, %s, %s)', (admin_id, action, details))
                    except Exception as fallback_err:
                        try:
                            cursor.execute('INSERT INTO activity_logs (admin_id, action, details) VALUES (%s, %s, %s)', (admin_id, action, details))
                        except Exception as final_err:
                            print(f'⚠️ Failed to log HR activity (all attempts failed): {final_err}')
            except Exception as insert_err:
                print(f'⚠️ Failed to prepare/insert activity log: {insert_err}')
            
            # Only create notification if not skipped (to avoid duplicates)
            # notify_admin_on_hr_actions already creates proper user-friendly notifications
            if not skip_notification:
                msg = f'HR Action: {action} on {target_table}'
                if target_id:
                    msg += f' (ID: {target_id})'
                if details:
                    msg += f' - {details}'
                
                # Prevent JSON responses from being saved as notifications
                msg_str = str(msg).strip()
                if msg_str.startswith('{') and ('"success"' in msg_str or '"message"' in msg_str or '"error"' in msg_str):
                    print(f'⚠️ Blocked JSON response from being saved as notification in log_hr_activity: {msg_str[:100]}')
                    return True
                
                cursor.execute(
                    '''
                    INSERT INTO notifications (application_id, message, sent_at, is_read)
                    VALUES (NULL, %s, NOW(), 0)
                    ''',
                    (msg,)
                )
                db.commit()
            return True
        except Exception as e:
            db.rollback()
            print(f'⚠️ Failed to log HR activity: {e}')
            return False
        finally:
            cursor.close()
    except Exception:
        return False


@app.before_request
def notify_admin_on_hr_actions():
    """Mirror all HR branch actions (POST) to Admin as system notifications and activity logs."""
    try:
        user = getattr(g, 'current_user', None)
        if not user or user.get('role') != 'hr':
            return
        if request.method != 'POST':
            return
        
        # Skip notification creation for delete notifications routes to prevent JSON responses from being saved
        if '/notifications/delete' in request.path or '/notifications/delete-all' in request.path:
            return
        
        hr_id = user.get('id')
        hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
        if not hr_id:
            return
            
        db = get_db()
        if not db:
            return
        
        cursor = db.cursor(dictionary=True)
        try:
            # Extract action details from request
            action = request.form.get('action', '')
            path = request.path
            branch_id = session.get('branch_id')
            
            # Determine target table and ID from path and form data
            target_table = 'unknown'
            target_id = None
            details = None  # Will be set based on action type
            admin_message = None
            
            # Parse common actions and create detailed admin notifications
            if 'applications' in path:
                target_table = 'applications'
                target_id = request.form.get('application_id', type=int)
                if action == 'update_status':
                    new_status = request.form.get('status', '')
                    details = f'Status changed to: {new_status}'
                    # Fetch applicant name and job title for better notification message
                    applicant_name = 'applicant'
                    job_title = 'position'
                    if target_id:
                        try:
                            cursor.execute('''
                                SELECT ap.full_name, COALESCE(j.job_title, 'N/A') AS job_title
                                FROM applications a
                                JOIN applicants ap ON a.applicant_id = ap.applicant_id
                                LEFT JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s
                                LIMIT 1
                            ''', (target_id,))
                            app_info = cursor.fetchone()
                            if app_info:
                                applicant_name = app_info.get('full_name') or 'applicant'
                                job_title = app_info.get('job_title') or 'position'
                        except Exception as fetch_err:
                            print(f'⚠️ Error fetching applicant info for admin notification: {fetch_err}')
                    status_display = new_status.replace('_', ' ').title()
                    admin_message = f'HR {hr_name} updated application status to {status_display} for {applicant_name} ({job_title}).'
            elif 'applicants' in path:
                target_table = 'applicants'
                target_id = request.form.get('applicant_id', type=int)
                if action == 'bulk_update_status':
                    bulk_status = request.form.get('bulk_status', '')
                    details = f'Bulk status update: {bulk_status}'
                    admin_message = f'HR {hr_name} performed bulk status update to {bulk_status.replace("_", " ").title()}'
            elif 'job-postings' in path or 'job_postings' in path:
                target_table = 'jobs'
                target_id = request.form.get('job_id', type=int) or (path.split('/')[-1] if path.split('/')[-1].isdigit() else None)
                
                # Determine if this is an add or update based on path and action
                is_update = 'update' in path or action in ['update', 'edit']
                is_add = action in ['add', 'create'] and not is_update
                
                # Get job title if available for better notification
                job_title = None
                if target_id:
                    try:
                        cursor.execute('SELECT job_title FROM jobs WHERE job_id = %s LIMIT 1', (target_id,))
                        job_row = cursor.fetchone()
                        if job_row:
                            job_title = job_row.get('title') if isinstance(job_row, dict) else job_row[0]
                    except Exception:
                        # Ignore non-critical DB errors while fetching job title
                        pass
                
                if is_add:
                    details = f'Job: {job_title}' if job_title else 'New job posting added'
                    if job_title:
                        admin_message = f'HR {hr_name} posted a new job: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} posted a new job posting'
                elif is_update:
                    details = f'Job Updated: {job_title}' if job_title else 'Job posting updated'
                    if job_title:
                        admin_message = f'HR {hr_name} updated job posting: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} updated a job posting'
                elif action == 'delete':
                    details = f'Job Deleted: {job_title}' if job_title else 'Job posting deleted'
                    if job_title:
                        admin_message = f'HR {hr_name} deleted job posting: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} deleted a job posting'
                else:
                    # Fallback for other actions
                    details = f'Job Modified: {job_title}' if job_title else 'Job posting modified'
                    if job_title:
                        admin_message = f'HR {hr_name} modified job posting: "{job_title}"'
                    else:
                        admin_message = f'HR {hr_name} modified a job posting'
            elif 'interviews' in path:
                target_table = 'interviews'
                target_id = request.form.get('interview_id', type=int)
                if action in ['schedule', 'update', 'reschedule', 'cancel']:
                    details = f'Interview {action.title()}ed'
                    admin_message = f'HR {hr_name} {action}d interview'
                    if target_id:
                        admin_message += f' (Interview ID: {target_id})'
            elif 'communications' in path:
                target_table = 'communications'
                # HR communications route has been removed - no longer tracking
                admin_message = None
                details = None
            elif 'reports' in path:
                target_table = 'reports'
                details = 'Accessed reports'
                admin_message = f'HR {hr_name} accessed reports/analytics'
            
            # Determine user-friendly action name for logging early
            user_friendly_action = determine_user_friendly_action(path, action, target_table)

            # Default details if not set by any action handler — use friendly action as fallback
            if not details:
                details = f'Action: {user_friendly_action}' if user_friendly_action else 'System action'

            # Create admin notification if we have a message
            if admin_message:
                # Don't add branch ID to notification message - it's not user-friendly
                try:
                    create_admin_notification(cursor, admin_message)
                    db.commit()
                    print(f'✅ Admin notification created: {admin_message}')
                except Exception as notify_err:
                    print(f'⚠️ Error creating admin notification: {notify_err}')

            # Log the activity (for activity logs only, not notifications)
            # Don't create duplicate notification in log_hr_activity
            # Only append branch info when we have a string in details
            try:
                if branch_id and details:
                    details = f'{details} | Branch: {branch_id}'
            except Exception:
                # Safety: ensure details is a string
                details = str(details or user_friendly_action or 'System action')
                if branch_id:
                    details = f'{details} | Branch: {branch_id}'

            # Pass skip_notification=True to prevent duplicate notification creation
            log_hr_activity(hr_id, user_friendly_action, target_table, target_id, details, skip_notification=True)
        finally:
            cursor.close()
        
    except Exception as e:
        # Non-blocking; never break the request due to admin-notify mirror
        print(f'⚠️ Error in notify_admin_on_hr_actions: {e}')
        pass

@app.context_processor
def inject_user():
    """Provide current user information to all Jinja templates.

    Use session-scoped `branch_id` if present (branch selector), otherwise fall back
    to the HR user's assigned `branch_id`. This ensures the notifications dropdown
    matches the HR notifications page when HR staff switch branch context.
    """
    user = getattr(g, 'current_user', None)
    # Provide HR notifications (branch-scoped for specific branch, all notifications if managing all branches)
    notifications = []
    notif_count = 0
    branches = []
    if user and user.get('role') == 'hr':
        # Respect session branch (HR branch selector) when present; fallback to user's assignment
        branch_id = session.get('branch_id') if session.get('branch_id') is not None else user.get('branch_id')
        if branch_id:
            # HR with specific branch: show branch-scoped notifications
            notifications, notif_count = fetch_notifications_for({'branch_id': branch_id}, limit=5)
        else:
            # HR managing all branches: show all notifications (no scope)
            notifications, notif_count = fetch_notifications_for(None, limit=5)
        # Fetch all branches for HR branch selector dropdown
        try:
            branches = fetch_branches()
        except Exception:
            branches = []
    return {
        'current_user': user,
        'hr_notifications': notifications,
        'hr_notif_count': notif_count,
        'hr_branches': branches,
        'is_logged_in': is_logged_in,
    }

@app.context_processor
def inject_csrf_token():
    """Make CSRF token function available in all templates."""
    def csrf_token():
        """Generate CSRF token for forms."""
        return generate_csrf()
    return dict(csrf_token=csrf_token)


@app.template_global()
def count_uploaded_documents(applicant_id=None, application_id=None):
    """Return number of uploaded documents for an applicant.

    Accepts either `applicant_id` or `application_id`. If only
    `application_id` is provided the function will resolve the
    corresponding `applicant_id` from `applications`.
    """
    try:
        db = get_db()
        if not db:
            return 0
        cur = db.cursor(dictionary=True)
        try:
            # If application_id is provided, count attachments linked to that application
            if application_id:
                try:
                    cur.execute('SELECT COUNT(*) AS cnt FROM application_attachments WHERE application_id = %s', (application_id,))
                    row = cur.fetchone() or {}
                    return int(row.get('cnt', 0) or 0)
                except Exception:
                    # Fallback to counting resumes by applicant if attachments table/query fails
                    pass

            if not applicant_id and application_id:
                cur.execute('SELECT applicant_id FROM applications WHERE application_id = %s LIMIT 1', (application_id,))
                row = cur.fetchone() or {}
                applicant_id = row.get('applicant_id')

            if not applicant_id:
                return 0

            cur.execute('SELECT COUNT(*) AS cnt FROM resumes WHERE applicant_id = %s', (applicant_id,))
            row = cur.fetchone() or {}
            return int(row.get('cnt', 0) or 0)
        finally:
            try:
                cur.close()
            except Exception:
                pass
    except Exception:
        return 0

@app.context_processor
def inject_applicant_notifications():
    """Provide notification data to applicant templates."""
    if is_logged_in() and session.get('user_role') == 'applicant':
        applicant_id = session.get('user_id')
        if applicant_id:
            db = get_db()
            if db:
                cursor = db.cursor(dictionary=True)
                try:
                    ensure_schema_compatibility()
                    # Check notifications table columns
                    cursor.execute('SHOW COLUMNS FROM notifications')
                    notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                    
                    has_application_fk = 'application_id' in notification_columns
                    if not has_application_fk:
                        return {'unread_count': 0, 'recent_notifications': []}
                    
                    # Build sent_at expression
                    if 'sent_at' in notification_columns:
                        sent_at_expr = 'n.sent_at'
                    elif 'created_at' in notification_columns:
                        sent_at_expr = 'n.created_at'
                    else:
                        sent_at_expr = 'NOW()'
                    
                    # Build is_read expression
                    if 'is_read' in notification_columns:
                        is_read_expr = 'COALESCE(n.is_read, 0)'
                    else:
                        is_read_expr = '0'
                    
                    # Get dynamic job column expressions
                    _update_job_columns(cursor)
                    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'System Notification'")
                    
                    select_fields = [
                        'n.notification_id',
                        'n.message',
                        f'{sent_at_expr} AS sent_at',
                        f'{is_read_expr} AS is_read',
                        'n.application_id',
                        'COALESCE(a.status, \'\') AS application_status',
                        f'COALESCE({job_title_expr}, \'System Notification\') AS job_title',
                    ]
                    query = f'''
                        SELECT DISTINCT {', '.join(select_fields)}
                        FROM notifications n
                        JOIN applications a ON n.application_id = a.application_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.applicant_id = %s
                        AND (
                            n.message LIKE 'You applied for%'
                            OR n.message LIKE 'Your application status%'
                            OR n.message LIKE 'Congratulations! You have been hired%'
                        )
                        ORDER BY {sent_at_expr} DESC
                        LIMIT 5
                    '''
                    cursor.execute(query, (applicant_id,))
                    notifications = cursor.fetchall() or []
                    
                    # Deduplicate by notification_id to prevent duplicates
                    seen_ids = set()
                    unique_notifications = []
                    for notif in notifications:
                        notif_id = notif.get('notification_id')
                        if notif_id and notif_id not in seen_ids:
                            seen_ids.add(notif_id)
                            unique_notifications.append(notif)
                    notifications = unique_notifications
                    
                    # Count unread
                    unread_count = len([n for n in notifications if not n.get('is_read')])
                    
                    # Format notifications same as communications page
                    formatted_notifications = []
                    for notif in notifications:
                        formatted_notifications.append({
                            'notification_id': notif.get('notification_id'),
                            'message': notif.get('message'),
                            'sent_at': format_human_datetime(notif.get('sent_at')),
                            'is_read': notif.get('is_read', False),
                            'application_id': notif.get('application_id'),
                            'job_title': notif.get('job_title'),
                            'application_status': notif.get('application_status'),
                        })
                    
                    return {'unread_count': unread_count, 'recent_notifications': formatted_notifications}
                except Exception as e:
                    print(f'⚠️ Error fetching applicant notifications: {e}')
                    return {'unread_count': 0, 'recent_notifications': []}
                finally:
                    if cursor:
                        cursor.close()
    return {'unread_count': 0, 'recent_notifications': []}

@app.context_processor
def inject_admin_notifications():
    """Provide admin notification data to all admin templates."""
    user = getattr(g, 'current_user', None)
    if user and user.get('role') in ('admin', 'hr'):
        try:
            # Fetch system-level notifications (application_id IS NULL)
            formatted, unread = fetch_notifications_for({'system_only': True}, limit=5)
            return {
                'admin_notifs': formatted,
                'admin_notif_count': unread,
                'admin_notif_display': '99+' if unread > 99 else unread
            }
        except Exception as e:
            print(f'⚠️ Error fetching admin notifications: {e}')
            return {
                'admin_notifs': [],
                'admin_notif_count': 0,
                'admin_notif_display': 0
            }
    return {
        'admin_notifs': [],
        'admin_notif_count': 0,
        'admin_notif_display': 0
    }

# ------------------------------


# ---------------------------------------------------------------------------
# HR Notifications endpoint (alias for HR communications/logs)
# Fix for templates expecting 'hr_notifications' endpoint.
# ---------------------------------------------------------------------------
@app.route('/hr/notifications')
@login_required('hr')
def hr_notifications():
    """View all HR notifications scoped to the HR user's branch."""
    user = get_current_user() or {}
    if not isinstance(user, dict):
        user = {}
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('hr/notifications.html', notifications=[], unread_count=0)
    ensure_schema_compatibility()
    cursor = db.cursor(dictionary=True)
    try:
        # Clean up any JSON response notifications first (AGGRESSIVE CLEANUP)
        branch_id = session.get('branch_id')
        try:
            # First, delete JSON notifications that match the pattern
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                """)
            json_cleaned = cursor.rowcount
            if json_cleaned > 0:
                print(f'✅ Cleaned up {json_cleaned} JSON response notification(s) from HR notifications page')
                db.commit()
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_err}')
            db.rollback()
        
        # Ensure notifications table and columns
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notification_columns else 'COALESCE(n.created_at, NOW())'
        is_read_expr = 'COALESCE(n.is_read, 0)' if 'is_read' in notification_columns else '0'
        # Scope to HR branch via jobs
        params = []
        where_sql = ''
        if branch_id:
            where_sql = 'WHERE j.branch_id = %s'
            params.append(branch_id)
        # HR focuses on application-related notifications
        # CRITICAL: Exclude ALL applicant-only notifications - these are for applicants only, NOT HR
        # HR should only see notifications about branch activities, not applicant-facing messages
        hr_where_sql = where_sql
        applicant_only_filters = [
            'n.message NOT LIKE \'You applied for%\'',
            'n.message NOT LIKE \'Congratulations! You have been hired%\'',
            'n.message NOT LIKE \'Your application status%\'',
            'n.message NOT LIKE \'Congratulations! You%\'',  # Catch any variation of congratulations messages to applicants
        ]
        if hr_where_sql:
            hr_where_sql += ' AND ' + ' AND '.join(applicant_only_filters)
        else:
            hr_where_sql = 'WHERE ' + ' AND '.join(applicant_only_filters)
        
        # Build joins and where clause to scope notifications to branch and exclude applicant-only messages
        joins = 'JOIN applications a ON n.application_id = a.application_id LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id'
        where_clauses = ["n.message NOT LIKE '{%'"]
        if branch_id:
            joins += ' JOIN jobs j ON a.job_id = j.job_id'
            where_clauses.append('j.branch_id = %s')
            # params already contains branch_id

        # Exclude applicant-only messages for HR view
        applicant_only_filters = [
            "n.message NOT LIKE 'You applied for%'",
            "n.message NOT LIKE 'Congratulations! You have been hired%'",
            "n.message NOT LIKE 'Your application status%'",
            "n.message NOT LIKE 'Congratulations! You%'",
        ]
        where_clauses.extend(applicant_only_filters)

        where_sql = 'WHERE ' + ' AND '.join(where_clauses) if where_clauses else ''

        cursor.execute(
            f'''
            SELECT n.notification_id,
                   n.message,
                   {sent_at_expr} AS sent_at,
                   {is_read_expr} AS is_read,
                   a.application_id,
                   COALESCE(ap.full_name, 'Unknown') AS applicant_name
            FROM notifications n
            {joins}
            {where_sql}
            ORDER BY {sent_at_expr} DESC
            LIMIT 200
            ''',
            tuple(params)
        )
        rows = cursor.fetchall() or []
        # Filter out any JSON response notifications that might have slipped through
        filtered_rows = []
        for r in rows:
            message = r.get('message', '')
            # Skip JSON responses - check for exact JSON string and patterns
            if message:
                message_str = str(message).strip()
                # Skip if it's a JSON response
                if (message_str.startswith('{') and ('"success"' in message_str or '"message"' in message_str or '"error"' in message_str)):
                    continue
                # Skip exact JSON response string
                if message_str == '{"message":"Notifications deleted.","success":true}':
                    continue
                # Skip if message contains the JSON response pattern
                if '{"message":"Notifications deleted.' in message_str:
                    continue
            filtered_rows.append(r)
        
        unread_count = len([r for r in filtered_rows if not r.get('is_read')])
        notifications = [
            {
                'notification_id': r.get('notification_id'),
                'message': r.get('message', ''),
                'sent_at': format_human_datetime(r.get('sent_at')),
                'is_read': r.get('is_read', False),
                'applicant_name': r.get('applicant_name'),
            }
            for r in filtered_rows
        ]
        return render_template('hr/notifications.html', notifications=notifications, unread_count=unread_count)
    except Exception as exc:
        log.exception('❌ HR notifications error: %s', exc)
        flash('Unable to load HR notifications. Please try again later.', 'error')
        return render_template('hr/notifications.html', notifications=[], unread_count=0)
    finally:
        cursor.close()


def fetch_archived_applicants_data():
    """Helper: return (archived_list, branches_list) for archived applicants pages."""
    db = get_db()
    if not db:
        return ([], [])

    cursor = db.cursor(dictionary=True)
    try:
        user = get_current_user()
        branch_id = get_branch_scope(user)
        # Determine archiving strategy (robust): consider status values and archived_at column
        try:
            cursor.execute('SHOW COLUMNS FROM applications')
            cols = {row.get('Field') for row in cursor.fetchall() or []}
            has_archived_at = 'archived_at' in cols
        except Exception:
            has_archived_at = False

        # Build a resilient WHERE clause that handles legacy values and schema differences:
        # - status values (case-insensitive) like 'archived', 'removed', 'deleted'
        # - any status that contains the word 'archive'
        # - OR archived_at IS NOT NULL when the column exists
        archived_status_clause = "(LOWER(COALESCE(a.status, '')) IN ('archived','removed','deleted') OR LOWER(COALESCE(a.status, '')) LIKE '%archive%')"
        if has_archived_at:
            archived_clause = f"({archived_status_clause} OR a.archived_at IS NOT NULL)"
        else:
            archived_clause = archived_status_clause

        job_title_expr = job_column_expr('job_title', alias='j', alternatives=['title'], default="'Untitled Job'")
        archived_at_select = 'COALESCE(a.archived_at, a.applied_at) AS archived_at' if has_archived_at else 'a.applied_at AS archived_at'
        order_by_expr = 'COALESCE(a.archived_at, a.applied_at) DESC' if has_archived_at else 'a.applied_at DESC'

        query = f"""
            SELECT a.application_id, a.applicant_id, a.resume_id, a.status, a.applied_at AS submitted_at,
                   {archived_at_select},
                   ap.full_name AS applicant_name, ap.email AS applicant_email, ap.phone_number AS applicant_phone,
                   {job_title_expr} AS job_title, j.branch_id, COALESCE(b.branch_name, 'All Branches') AS branch_name,
                   CASE WHEN a.resume_id IS NOT NULL THEN 1 ELSE 0 END AS has_resume
            FROM applications a
            LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {archived_clause}""" + ((" AND j.branch_id = %s" ) if branch_id else "") + f"""
            ORDER BY {order_by_expr}
        """

        if branch_id:
            cursor.execute(query, (branch_id,))
        else:
            cursor.execute(query)
        archived = cursor.fetchall() or []

        # Format dates for display
        for row in archived:
            row['submitted_at'] = format_human_datetime(row.get('submitted_at'))
            row['archived_at'] = format_human_datetime(row.get('archived_at')) if row.get('archived_at') else None

        branches = fetch_branches()
        return (archived, branches)
    finally:
        try:
            cursor.close()
        except Exception:
            pass


@app.route('/admin/archived-applicants')
@login_required('admin')
def admin_archived_applicants():
    """Admin view: show archived applicants using admin template."""
    try:
        archived, branches = fetch_archived_applicants_data()
        return render_template('admin/archived_applicants.html', archived=archived, branches=branches)
    
    except Exception as exc:
        log.exception(f'❌ Admin archived applicants error: {exc}')
        flash('Unable to load archived applicants. Please try again later.', 'error')
        return render_template('admin/archived_applicants.html', archived=[], branches=[])


@app.route('/admin/activity-logs/<int:log_id>/delete', methods=['POST'])
@login_required('admin', 'hr')
def admin_delete_activity_log(log_id):
    """Delete an activity_log entry (admin/hr)."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_security'))

    cursor = db.cursor()
    try:
        # Verify exists
        cursor.execute('SELECT log_id FROM activity_logs WHERE log_id = %s LIMIT 1', (log_id,))
        if not cursor.fetchone():
            flash('Activity log not found.', 'error')
            return redirect(url_for('admin_security'))

        cursor.execute('DELETE FROM activity_logs WHERE log_id = %s', (log_id,))
        db.commit()
        flash('Activity log deleted.', 'success')
    except Exception as exc:
        db.rollback()
        print(f'⚠️ Failed to delete activity log {log_id}: {exc}')
        flash('Failed to delete activity log.', 'error')
    finally:
        try:
            cursor.close()
        except Exception:
            pass

    return redirect(url_for('admin_security', just_deleted=1))


@app.route('/admin/activity-logs/delete-all', methods=['POST'])
@login_required('admin', 'hr')
def admin_delete_all_activity_logs():
    """Delete all activity_log entries. Admin can delete all; HR deletes only their branch's logs."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_security'))

    user = get_current_user()
    branch_scope = session.get('branch_id')

    cursor = db.cursor()
    try:
        if user and user.get('role') == 'admin':
            cursor.execute('DELETE FROM activity_logs')
            msg = 'All activity logs deleted.'
        elif branch_scope:
            cursor.execute('DELETE FROM activity_logs WHERE branch_id = %s', (branch_scope,))
            msg = 'Activity logs for your branch deleted.'
        else:
            # Fallback: delete nothing and warn
            flash('Unable to determine scope for deletion.', 'error')
            return redirect(url_for('admin_security'))

        db.commit()
        # Record deletion marker to prevent immediately re-showing old logs
        try:
            try:
                marker_cursor = db.cursor()
                marker_cursor.execute("INSERT INTO activity_log_deletions (branch_id, deleted_at) VALUES (%s, NOW())", (branch_scope,))
                db.commit()
            finally:
                try:
                    marker_cursor.close()
                except Exception:
                    pass
        except Exception as marker_exc:
            print(f'⚠️ Failed to record activity log deletion marker: {marker_exc}')
        flash(msg, 'success')
    except Exception as exc:
        db.rollback()
        print(f'⚠️ Failed to delete all activity logs: {exc}')
        flash('Failed to delete activity logs.', 'error')
    finally:
        try:
            cursor.close()
        except Exception:
            pass

    return redirect(url_for('admin_security', just_deleted=1))


@app.route('/hr/notifications/read-all', methods=['POST'])
@login_required('hr')
def mark_all_hr_notifications_read():
    """Mark all HR notifications as read, scoped to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor()
    try:
        # Enforce branch scoping: HR users with branch scope can only delete applications from their branch
        try:
            branch_scope = session.get('branch_id')
        except Exception:
            branch_scope = None
        try:
            cursor.execute('''
                SELECT j.branch_id
                FROM applications a
                JOIN jobs j ON a.job_id = j.job_id
                WHERE a.application_id = %s
                LIMIT 1
            ''', (application_id,))
            row = cursor.fetchone() or {}
            app_branch = row.get('branch_id') if isinstance(row, dict) else (row[0] if row else None)
            if branch_scope and str(app_branch) != str(branch_scope):
                flash('Access denied: cannot delete application outside your branch.', 'error')
                return redirect(url_for('archived_applicants'))
        except Exception:
            # If scope check fails, continue cautiously
            pass
        ensure_schema_compatibility()
        branch_id = session.get('branch_id')
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if not has_is_read:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification read status not available.'}), 400
            flash('Notification read status not available.', 'error')
            return redirect(url_for('hr_notifications'))
        
        # Clean up any JSON responses that might have been stored as notifications
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE j.branch_id = %s
                    AND n.message LIKE '{%' 
                    AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%')
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE message LIKE '{%' 
                    AND (message LIKE '%"success"%' OR message LIKE '%"message"%')
                """)
        except Exception as cleanup_error:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_error}')
        
        # Mark all notifications as read, scoped to branch
        if branch_id:
            cursor.execute("""
                UPDATE notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                SET n.is_read = 1
                WHERE j.branch_id = %s AND n.is_read = 0
            """, (branch_id,))
        else:
            cursor.execute('UPDATE notifications SET is_read = 1 WHERE is_read = 0')
        
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'All notifications marked as read'})
        flash('All notifications marked as read.', 'success')
    except Exception as exc:
        db.rollback()
        log.exception(f'❌ Mark all HR notifications read error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark all notifications as read.'}), 500
        flash('Failed to mark all notifications as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


@app.route('/hr/notifications/<int:notification_id>/read', methods=['POST'])
@login_required('hr')
def mark_hr_notification_read(notification_id):
    """Mark a notification as read, verifying it belongs to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor(dictionary=True)
    try:
        ensure_schema_compatibility()
        branch_id = session.get('branch_id')
        
        # Verify notification belongs to HR branch
        if branch_id:
            cursor.execute("""
                SELECT n.notification_id
                FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE n.notification_id = %s AND j.branch_id = %s
                LIMIT 1
            """, (notification_id, branch_id))
        else:
            cursor.execute("""
                SELECT notification_id FROM notifications WHERE notification_id = %s LIMIT 1
            """, (notification_id,))
        
        notif_record = cursor.fetchone()
        if not notif_record:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification not found or access denied.'}), 404
            flash('Notification not found or access denied.', 'error')
            return redirect(url_for('hr_notifications'))
        
        # Check if this notification is a JSON response and delete it if so
        cursor.execute('SELECT message FROM notifications WHERE notification_id = %s', (notification_id,))
        notif_msg = cursor.fetchone()
        if notif_msg:
            message = notif_msg.get('message', '') if isinstance(notif_msg, dict) else (notif_msg[0] if notif_msg else '')
            if message and message.strip().startswith('{') and '"success"' in message:
                cursor.execute('DELETE FROM notifications WHERE notification_id = %s', (notification_id,))
                db.commit()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': True, 'message': 'Invalid notification removed', 'notification_id': notification_id})
                flash('Invalid notification removed.', 'success')
                return redirect(url_for('hr_notifications'))
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if has_is_read:
            cursor.execute(
                'UPDATE notifications SET is_read = 1 WHERE notification_id = %s',
                (notification_id,)
            )
            db.commit()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': True, 'message': 'Notification marked as read', 'notification_id': notification_id})
            flash('Notification marked as read.', 'success')
        else:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification read status not available.'}), 400
            flash('Notification read status not available.', 'error')
    except Exception as exc:
        db.rollback()
        log.exception(f'❌ Mark HR notification read error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark notification as read.'}), 500
        flash('Failed to mark notification as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


@app.route('/hr/notifications/<int:notification_id>/delete', methods=['POST'])
@login_required('hr')
def delete_hr_notification(notification_id):
    """Delete a notification, verifying it belongs to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor(dictionary=True)
    try:
        branch_id = session.get('branch_id')
        
        # Verify notification belongs to HR branch
        if branch_id:
            cursor.execute("""
                SELECT n.notification_id
                FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE n.notification_id = %s AND j.branch_id = %s
                LIMIT 1
            """, (notification_id, branch_id))
        else:
            cursor.execute("""
                SELECT notification_id FROM notifications WHERE notification_id = %s LIMIT 1
            """, (notification_id,))
        
        notif_record = cursor.fetchone()
        if not notif_record:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification not found or access denied.'}), 404
            flash('Notification not found or access denied.', 'error')
            return redirect(url_for('hr_notifications'))
        
        # Clean up any JSON response notifications before deleting (AGGRESSIVE)
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                """)
            json_cleaned = cursor.rowcount
            if json_cleaned > 0:
                print(f'✅ Cleaned up {json_cleaned} JSON response notification(s) before delete')
                db.commit()
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_err}')
            db.rollback()
        
        # Delete only the specific notification
        cursor.execute('DELETE FROM notifications WHERE notification_id = %s', (notification_id,))
        deleted_count = cursor.rowcount
        
        if deleted_count == 0:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification not found or already deleted.'}), 404
            flash('Notification not found or already deleted.', 'error')
            return redirect(url_for('hr_notifications'))
        
        db.commit()
        
        # Final cleanup AFTER commit to catch any JSON notifications created by other processes
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                """)
            final_cleaned = cursor.rowcount
            if final_cleaned > 0:
                print(f'✅ Final cleanup after delete: Removed {final_cleaned} JSON response notification(s)')
                db.commit()
        except Exception as final_cleanup_err:
            print(f'⚠️ Error in final cleanup: {final_cleanup_err}')
            db.rollback()
        
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'Notification deleted successfully', 'notification_id': notification_id})
        flash('Notification deleted successfully.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        log.exception(f'❌ Delete HR notification error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to delete notification.'}), 500
        flash('Failed to delete notification.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


@app.route('/hr/notifications/delete-all', methods=['POST'])
@login_required('hr')
def delete_all_hr_notifications():
    """Delete all HR notifications, scoped to the HR user's branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_notifications'))
    
    cursor = db.cursor()
    try:
        # Enforce branch scoping for immediate deletes as well
        try:
            branch_scope = session.get('branch_id')
        except Exception:
            branch_scope = None
        try:
            cursor.execute('''
                SELECT j.branch_id
                FROM applications a
                JOIN jobs j ON a.job_id = j.job_id
                WHERE a.application_id = %s
                LIMIT 1
            ''', (application_id,))
            row = cursor.fetchone() or {}
            app_branch = row.get('branch_id') if isinstance(row, dict) else (row[0] if row else None)
            if branch_scope and str(app_branch) != str(branch_scope):
                flash('Access denied: cannot delete application outside your branch.', 'error')
                return redirect(url_for('applicants'))
        except Exception:
            pass
        branch_id = session.get('branch_id')
        
        # First, clean up any JSON response notifications
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE j.branch_id = %s
                    AND n.message LIKE '{%' 
                    AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%')
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE message LIKE '{%' 
                    AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%')
                """)
            json_cleaned = cursor.rowcount
            if json_cleaned > 0:
                print(f'✅ Cleaned up {json_cleaned} JSON response notification(s) before delete-all')
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_err}')
        
        # Delete notifications scoped to branch (excluding JSON responses)
        if branch_id:
            cursor.execute("""
                DELETE n FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE j.branch_id = %s
            """, (branch_id,))
        else:
            cursor.execute('DELETE FROM notifications')
        
        deleted_count = cursor.rowcount
        db.commit()
        
        # CRITICAL: Final cleanup BEFORE returning response to prevent JSON from being saved
        # This must run AFTER commit but BEFORE any response is returned
        try:
            if branch_id:
                cursor.execute("""
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE (j.branch_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                        OR n.message = '{"message":"Notifications deleted.","success":true}'
                        OR n.message LIKE '%{"message":"Notifications deleted.%'
                    )
                """, (branch_id,))
            else:
                cursor.execute("""
                    DELETE FROM notifications 
                    WHERE (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                        OR message = '{"message":"Notifications deleted.","success":true}'
                        OR message LIKE '%{"message":"Notifications deleted.%'
                    )
                """)
            final_cleaned = cursor.rowcount
            if final_cleaned > 0:
                print(f'✅ Final cleanup: Removed {final_cleaned} JSON response notification(s) after delete-all')
                db.commit()
        except Exception as final_cleanup_err:
            print(f'⚠️ Error in final cleanup: {final_cleanup_err}')
            db.rollback()
        
        # Return response - JSON response should NEVER be saved as notification
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': f'All notifications deleted successfully ({deleted_count} notification(s) removed)'})
        flash(f'All notifications deleted successfully ({deleted_count} notification(s) removed).', 'success')
    except Exception as exc:
        db.rollback()
        log.exception(f'❌ Delete all HR notifications error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to delete all notifications.'}), 500
        flash('Failed to delete all notifications.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('hr_notifications'))


def generate_token():
    """Generate a random token suitable for email verification and password resets."""
    return uuid4().hex


def send_verification_email(email, token, applicant_name=None):
    """Send a verification email containing the confirmation link with HTML template."""
    from flask import render_template_string
    import os
    
    verification_url = url_for('verify_email', token=token, _external=True)
    # Generate absolute URL for logo image (needed for email clients)
    logo_url = url_for('static', filename='images/whitehat_logo.jpg', _external=True)
    subject = 'Verify your J&T Express applicant account'
    
    # Plain text version
    body = f"""
Hi{(' ' + applicant_name) if applicant_name else ''},

Thank you for registering with J&T Express.

Please confirm your email address by clicking the link below:
{verification_url}

If you did not create this account, you can safely ignore this email.

This verification link will expire after 60 seconds.

Regards,
J&T Express Recruitment Team
"""
    
    # HTML version
    html_template_path = os.path.join('templates', 'emails', 'verification_email.html')
    html_body = None
    
    try:
        if os.path.exists(html_template_path):
            with open(html_template_path, 'r', encoding='utf-8') as f:
                html_template = f.read()
            html_body = render_template_string(html_template, verification_url=verification_url, logo_url=logo_url)
    except Exception as e:
        print(f'⚠️ Could not load HTML email template: {e}')
    
    send_email(email, subject, body, html_body)


def send_password_reset_email(email, token):
    """Send password reset instructions to the applicant."""
    reset_url = url_for('reset_password', token=token, _external=True)
    subject = 'Reset your J&T Express password'
    body = f"""
    Hi,

    A password reset was requested for your account. To set a new password, please click the link below:
    {reset_url}

    This link will expire in 30 minutes. If you did not request a reset, you can ignore this email.

    Regards,
    J&T Express Recruitment Team
    """
    send_email(email, subject, body)


def generate_2fa_code():
    """Generate a 6-digit 2FA verification code."""
    import random
    return ''.join([str(random.randint(0, 9)) for _ in range(6)])


def send_2fa_email(email, verification_code, user_name=None, role='admin'):
    """Send 2FA verification code via email."""
    role_labels = {
        'admin': 'Admin',
        'hr': 'HR Manager',
        'applicant': 'User'
    }
    role_label = role_labels.get(role, 'User')
    
    subject = f'Your J&T Express {role_label} Verification Code'
    
    body = f"""
Hi{(' ' + user_name) if user_name else ''},

Your {role_label} login verification code is: {verification_code}

This code will expire in 10 minutes.

If you did not attempt to log in, please ignore this email.

Regards,
J&T Express Security Team
"""
    
    try:
        send_email(email, subject, body)
        return True
    except Exception as exc:
        print(f'❌ Failed to send 2FA email: {exc}')
        return False


def create_2fa_verification(user_id, email, user_name=None, role='admin'):
    """Create a 2FA verification request for a user (admin, hr, or applicant)."""
    # Applicant 2FA has been disabled application-wide — don't create DB records for applicants.
    if role == 'applicant':
        print('ℹ️ Applicant 2FA disabled: skipping creation of verification record')
        return None, None

    db = get_db()
    if not db:
        print('❌ Database connection failed for 2FA')
        return None, None
    
    cursor = db.cursor(dictionary=True)
    try:
        # Map role to table name
        table_map = {
            'admin': 'admin_2fa_verification',
            'hr': 'hr_2fa_verification',
            'applicant': 'applicant_2fa_verification'
        }
        table_name = table_map.get(role, 'admin_2fa_verification')
        
        # Generate verification code and temp token
        verification_code = generate_2fa_code()
        temp_token = generate_token()
        
        print(f'🔐 Creating 2FA verification for user_id={user_id}, email={email}, role={role}')
        
        # Delete any existing verification codes for this user
        cursor.execute(
            f'DELETE FROM {table_name} WHERE user_id = %s AND verified = 0',
            (user_id,)
        )
        
        # Insert new verification record (valid for 10 minutes)
        cursor.execute(
            f'''
            INSERT INTO {table_name}
            (user_id, email, verification_code, temp_token, attempts, verified, expired_at)
            VALUES (%s, %s, %s, %s, 0, 0, DATE_ADD(NOW(), INTERVAL 10 MINUTE))
            ''',
            (user_id, email, verification_code, temp_token)
        )
        db.commit()
        
        print(f'✓ 2FA record created: code={verification_code}, token={temp_token[:10]}...')
        
        # Send verification code via email
        email_sent = send_2fa_email(email, verification_code, user_name, role)
        if not email_sent:
            print(f'⚠️ Warning: Email not sent for 2FA, but record created')
        
        return temp_token, verification_code
    except Exception as exc:
        db.rollback()
        log.exception(f'❌ Failed to create 2FA verification: {exc}')
        return None, None
    finally:
        cursor.close()


def verify_2fa_code(temp_token, verification_code, role='admin'):
    """Verify the 2FA code entered by the user."""
    # Applicant 2FA has been disabled — always fail verification for applicant role.
    if role == 'applicant':
        print('ℹ️ Applicant 2FA disabled: skipping verification')
        return False, None

    db = get_db()
    if not db:
        return False, None
    
    cursor = db.cursor(dictionary=True)
    try:
        # Map role to table name
        table_map = {
            'admin': 'admin_2fa_verification',
            'hr': 'hr_2fa_verification',
            'applicant': 'applicant_2fa_verification'
        }
        table_name = table_map.get(role, 'admin_2fa_verification')
        
        # Get the verification record
        cursor.execute(
            f'''
            SELECT user_id, email, verification_code, attempts, max_attempts, expired_at
            FROM {table_name}
            WHERE temp_token = %s AND verified = 0
            LIMIT 1
            ''',
            (temp_token,)
        )
        record = cursor.fetchone()
        
        if not record:
            return False, None
        
        # Check if expired
        if record['expired_at'] < datetime.now():
            cursor.execute(f'DELETE FROM {table_name} WHERE temp_token = %s', (temp_token,))
            db.commit()
            return False, None
        
        # Check if max attempts exceeded
        if record['attempts'] >= record['max_attempts']:
            cursor.execute(f'DELETE FROM {table_name} WHERE temp_token = %s', (temp_token,))
            db.commit()
            return False, None
        
        # Verify code
        if record['verification_code'] == verification_code:
            # Mark as verified
            cursor.execute(
                f'UPDATE {table_name} SET verified = 1 WHERE temp_token = %s',
                (temp_token,)
            )
            db.commit()
            return True, record['user_id']
        else:
            # Increment attempts
            cursor.execute(
                f'UPDATE {table_name} SET attempts = attempts + 1 WHERE temp_token = %s',
                (temp_token,)
            )
            db.commit()
            remaining = record['max_attempts'] - record['attempts'] - 1
            return False, remaining
    except Exception as exc:
        db.rollback()
        print(f'❌ Failed to verify 2FA code: {exc}')
        return False, None
    finally:
        cursor.close()


def get_valid_admin_id(admin_id):
    """Validate that admin_id exists in admins table. Returns admin_id if valid, None otherwise."""
    if not admin_id:
        return None
    
    db = get_db()
    if not db:
        return None
    
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            'SELECT admin_id FROM admins WHERE admin_id = %s LIMIT 1',
            (admin_id,)
        )
        result = cursor.fetchone()
        return result['admin_id'] if result else None
    except Exception:
        return None
    finally:
        cursor.close()


def log_profile_change(user_id, role, field, old_value, new_value):
    """Profile change logging is disabled - changes are not persisted to database."""
    # Profile change logging disabled - no database entries are created
    pass


def ensure_default_accounts():
    """Create baseline System Administrator account if it does not exist. Only System Administrator can create HR accounts."""
    try:
        db = get_db()
        if not db:
            print('❌ Unable to verify default accounts because the database connection failed.')
            print('   Please ensure MySQL is running and the database credentials are correct.')
            return False
    except Exception as db_error:
        print(f'❌ Database connection error: {db_error}')
        print('   Please ensure MySQL is running and the database credentials are correct.')
        return False
    
    cursor = None
    try:
        cursor = db.cursor(dictionary=True)
        
        # Only create System Administrator - no HR accounts
        # Use environment variables for default admin creation on initial setup
        admin_email = os.environ.get('DEFAULT_ADMIN_EMAIL')
        admin_password = os.environ.get('DEFAULT_ADMIN_PASSWORD')

        account = {
            'full_name': 'System Administrator',
            'email': admin_email,
            'role': 'admin',
            'password': admin_password,
        }

        # Validate environment-provided credentials before proceeding
        if not admin_email or not admin_password:
            print('⚠️ Default admin account not created: please set DEFAULT_ADMIN_EMAIL and DEFAULT_ADMIN_PASSWORD environment variables before starting the app.')
            print("   Alternatively, create one explicitly using the management CLI: `python manage.py create_admin --email admin@example.com` (password will be prompted).")
            return False

        created_any = False

        # Check if System Administrator already exists
        cursor.execute(
            """
            SELECT u.user_id, u.user_type, a.admin_id
            FROM users u
            LEFT JOIN admins a ON a.user_id = u.user_id
            WHERE u.email = %s
            LIMIT 1
            """,
            (account['email'],),
        )

        exists = cursor.fetchone()

        # If exists and has admin record with correct settings, skip
        if exists and exists.get('admin_id'):
            # Verify it's set up correctly
            if exists.get('user_type') == 'super_admin' and exists.get('branch_id') is None:
                return True  # Already correctly configured
            else:
                # Fix existing account
                user_id = exists['user_id']
                # Update user_type to super_admin and ensure email is verified
                cursor.execute(
                    "UPDATE users SET user_type = 'super_admin', email_verified = 1 WHERE user_id = %s",
                    (user_id,)
                )
                # Update admin record
                cursor.execute(
                    "UPDATE admins SET full_name = 'System Administrator' WHERE admin_id = %s",
                    (exists['admin_id'],)
                )
                # Update password_hash if provided
                try:
                    password_hash = hash_password(account['password'])
                except Exception as e:
                    print(f"⚠️ Skipping password update for existing admin because password hashing failed: {e}")
                    password_hash = None

                if password_hash:
                    cursor.execute(
                        "UPDATE users SET password_hash = %s WHERE user_id = %s",
                        (password_hash, user_id)
                    )
                    db.commit()
                    print(f"✅ Fixed System Administrator account: {account['email']}")
                else:
                    # Commit the other fixes but don't force a password change
                    try:
                        db.commit()
                    except Exception:
                        pass
                    print(f"⚠️ Fixed System Administrator account metadata but left password unchanged: {account['email']}")
                return True

        user_type = 'super_admin'  # Always super_admin for System Administrator
            
        if exists and exists.get('user_id') and not exists.get('admin_id'):
            # User exists but no admin record, create admin record
            user_id = exists['user_id']
            user_email = account['email']
            # Update user_type to super_admin if needed and ensure email is verified
            cursor.execute(
                "UPDATE users SET user_type = 'super_admin', email_verified = 1 WHERE user_id = %s",
                (user_id,)
            )
        else:
            # Create new user account
            # Admin/HR accounts are automatically verified (no email verification required)
            try:
                password_hash = hash_password(account['password'])
            except Exception as e:
                print(f"❌ Error creating default accounts: failed to hash password: {e}")
                return False

            cursor.execute(
                """
                INSERT INTO users (email, password_hash, user_type, is_active, email_verified)
                VALUES (%s, %s, %s, %s, 1)
                """,
                (account['email'], password_hash, user_type, True),
            )
            user_id = cursor.lastrowid
            user_email = account['email']

        # Create or update admin record
        cursor.execute(
            """
            SELECT admin_id FROM admins WHERE user_id = %s LIMIT 1
            """,
            (user_id,),
        )
        admin_exists = cursor.fetchone()
        
        if not admin_exists:
            # Get password_hash from users table to copy to admins table
            cursor.execute(
                "SELECT password_hash FROM users WHERE user_id = %s LIMIT 1",
                (user_id,)
            )
            user_record = cursor.fetchone()
            password_hash = user_record.get('password_hash') if user_record else None
            
            # If password_hash is still None, create it
            if not password_hash:
                try:
                    password_hash = hash_password(account['password'])
                except Exception as e:
                    print(f"❌ Error creating admin record: failed to hash password: {e}")
                    return False

                cursor.execute(
                    "UPDATE users SET password_hash = %s WHERE user_id = %s",
                    (password_hash, user_id)
                )
            
            # Insert admin record with password_hash
            cursor.execute(
                """
                INSERT INTO admins (user_id, full_name, email, password_hash)
                VALUES (%s, %s, %s, %s)
                """,
                (user_id, account['full_name'], user_email, password_hash),
            )
            created_any = True
            print(f"✅ Created System Administrator account: {account['email']}")
        else:
            # Update existing admin record
            cursor.execute(
                "UPDATE admins SET full_name = 'System Administrator' WHERE admin_id = %s",
                (admin_exists['admin_id'],)
            )

        if created_any:
            try:
                db.commit()
            except Exception as commit_error:
                log.warning('⚠️ Error committing default accounts: %s', commit_error)
                try:
                    db.rollback()
                except Exception:
                    # Ignore rollback errors
                    pass

        return True
        
    except Exception as exc:
        try:
            if db:
                db.rollback()
        except Exception:
            # Ignore rollback errors
            pass
        log.exception(f'❌ Error creating default accounts: {exc}')
        return False
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                # Ignore errors while closing cursor
                pass


def _first_value(row, default=0):
    """Return the first value from a dictionary result or a default."""
    if not row:
        return default

    for value in row.values():
        if value is None:
            return default
        if isinstance(value, Decimal):
            return float(value)
        return value

    return default


def fetch_count(query, params=None, default=0):
    """Execute a COUNT-style query and safely extract the resulting value."""
    row = execute_query(query, params, fetch_one=True)
    return _first_value(row, default)


def fetch_rows(query, params=None):
    """Execute a query returning multiple rows, ensuring a list result."""
    rows = execute_query(query, params, fetch_all=True)
    return rows or []


def fetch_branches():
    """Return all branches ordered alphabetically."""
    return fetch_rows(
        '''
        SELECT
            b.branch_id,
            b.branch_name,
            b.address,
            b.operating_hours,
            b.is_active,
            b.created_at
        FROM branches b
        ORDER BY b.branch_name ASC
        '''
    )


@app.context_processor
def inject_activity_log_deletion_marker():
    """Provide the latest activity_log_deletions.deleted_at timestamp for the current branch (if any).
    Templates can use `activity_logs_deleted_at` to hide logs older than the marker.
    """
    db = get_db()
    if not db:
        return {'activity_logs_deleted_at': None}
    branch_scope = session.get('branch_id')
    cursor = db.cursor()
    try:
        if branch_scope:
            cursor.execute('''
                SELECT deleted_at FROM activity_log_deletions
                WHERE branch_id = %s
                ORDER BY deleted_at DESC LIMIT 1
            ''', (branch_scope,))
            row = cursor.fetchone()
        else:
            cursor.execute('''
                SELECT deleted_at FROM activity_log_deletions
                WHERE branch_id IS NULL
                ORDER BY deleted_at DESC LIMIT 1
            ''')
            row = cursor.fetchone()
        if not row:
            return {'activity_logs_deleted_at': None}
        # row may be dict or tuple
        deleted_at = row.get('deleted_at') if isinstance(row, dict) else (row[0] if row else None)
        return {'activity_logs_deleted_at': deleted_at}
    except Exception:
        return {'activity_logs_deleted_at': None}
    finally:
        try:
            cursor.close()
        except Exception:
            pass


def fetch_positions():
    """Return all job positions ordered alphabetically."""
    # Positions table has been removed, return empty list
    return []


def get_branch_scope(user):
    """Return the branch_id associated with an HR user, otherwise None.
    For HR users with a specific branch assignment, return that branch_id.
    For HR users with all branches access (branch_id=None), return None."""
    if not user:
        return None
    # Only HR users may have a branch scope
    if user.get('role') != 'hr':
        return None

    # Return the branch_id from the user object (already fetched in get_current_user)
    return user.get('branch_id')


def fetch_hr_accounts():
    """Fetch HR administrator accounts. HR accounts may be scoped to a branch (branch_id NULL => All Branches)."""
    db = get_db()
    if not db:
        print('⚠️ No database connection in fetch_hr_accounts')
        return []
    
    cursor = db.cursor(dictionary=True)
    try:
        # First, check if there are any HR users in the users table
        cursor.execute("SELECT COUNT(*) as count FROM users WHERE user_type = 'hr'")
        hr_count = cursor.fetchone()
        print(f'🔍 Total HR users in users table: {hr_count.get("count", 0) if hr_count else 0}')
        
        # Check if there are any admins linked to HR users
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM admins a
            JOIN users u ON u.user_id = a.user_id
            WHERE u.user_type = 'hr'
        """)
        admin_hr_count = cursor.fetchone()
        print(f'🔍 Total HR admins (joined): {admin_hr_count.get("count", 0) if admin_hr_count else 0}')
        
        # Now fetch the actual HR accounts
        cursor.execute(
        """
        SELECT
            a.admin_id,
            a.full_name,
            u.email,
            a.branch_id AS branch_id,
            COALESCE(b.branch_name, 'All Branches') AS branch_name,
            u.is_active,
            CASE WHEN a.branch_id IS NULL THEN 0 ELSE 1 END AS assigned_branch_count
        FROM admins a
        JOIN users u ON u.user_id = a.user_id
        LEFT JOIN branches b ON a.branch_id = b.branch_id
        WHERE u.user_type = 'hr'
        ORDER BY a.full_name ASC
        """
    )
        rows = cursor.fetchall()
        print(f'✅ fetch_hr_accounts: Found {len(rows) if rows else 0} HR accounts')
        if rows:
            print(f'🔍 Sample HR account: {rows[0]}')
        return rows or []
    except Exception as e:
        log.exception(f'❌ Error fetching HR accounts: {e}')
        return []
    finally:
        cursor.close()


def fetch_all_applications(user=None):
    """Fetch applications, automatically scoping to HR user's branch."""
    branch_id = get_branch_scope(user)

    if branch_id:
        # HR users: Only see applications for jobs in their branch
        query = """
            SELECT a.application_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   j.job_title AS job_title,
                   COALESCE(b.branch_name,'Unassigned') AS branch_name,
                   a.status,
                   a.applied_at
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE j.branch_id = %s
            ORDER BY a.applied_at DESC
        """
        params = (branch_id,)
        print(f'🔍 fetch_all_applications: HR user - filtering by branch_id={branch_id}')
    else:
        # Admin users: See all applications from all branches
        query = """
            SELECT a.application_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   j.job_title AS job_title,
                   COALESCE(b.branch_name,'Unassigned') AS branch_name,
                   a.status,
                   a.applied_at
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            ORDER BY a.applied_at DESC
        """
        params = None
        print(f'🔍 fetch_all_applications: Admin user - showing all applications from all branches')

    return fetch_rows(query, params)


def build_report_stats(user=None, date_filter='', date_params=None):
    """Build report statistics with optional date filtering."""
    branch_id = get_branch_scope(user)
    date_params = date_params or []

    # Status normalization: map legacy statuses to canonical ones
    pending_statuses = ['pending', 'reviewed', 'applied', 'under_review', 'shortlisted']
    interviewed_statuses = ['interviewed', 'interview']
    hired_statuses = ['hired', 'accepted']
    rejected_statuses = ['rejected']

    if branch_id:
        base = "FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = %s"
        params = (branch_id,)
        date_clause = f" {date_filter}" if date_filter else ""
        all_params = list(params) + date_params

        # Build IN clauses for status groups
        pending_placeholders = ','.join(['%s'] * len(pending_statuses))
        interviewed_placeholders = ','.join(['%s'] * len(interviewed_statuses))
        hired_placeholders = ','.join(['%s'] * len(hired_statuses))
        rejected_placeholders = ','.join(['%s'] * len(rejected_statuses))

        stats = {
            'total_applications': fetch_count(f'SELECT COUNT(*) AS count {base}{date_clause}', tuple(all_params)),
            'pending': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'interviewed': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({interviewed_placeholders})", tuple(all_params + interviewed_statuses)),
            'hired': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
            'rejected': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({rejected_placeholders})", tuple(all_params + rejected_statuses)),
            'total_interviews': fetch_count(f"SELECT COUNT(DISTINCT i.interview_id) AS count FROM applications a JOIN jobs j ON a.job_id = j.job_id LEFT JOIN interviews i ON i.application_id = a.application_id WHERE j.branch_id = %s{date_clause.replace('a.applied_at', 'i.scheduled_date') if date_filter else ''}", tuple(all_params)),
            # Legacy mappings for backward compatibility
            'applied': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'reviewed': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'accepted': fetch_count(f"SELECT COUNT(*) AS count {base}{date_clause} AND a.status IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
        }
    else:
        # Build IN clauses for status groups
        pending_placeholders = ','.join(['%s'] * len(pending_statuses))
        interviewed_placeholders = ','.join(['%s'] * len(interviewed_statuses))
        hired_placeholders = ','.join(['%s'] * len(hired_statuses))
        rejected_placeholders = ','.join(['%s'] * len(rejected_statuses))

        if date_filter:
            # Remove leading 'AND ' only, not all occurrences
            date_filter_clean = date_filter.strip()
            if date_filter_clean.startswith('AND '):
                date_filter_clean = date_filter_clean[4:]  # Remove 'AND ' from start
            # Replace 'a.applied_at' with 'applied_at' since queries without branch_id don't use alias 'a'
            date_filter_clean = date_filter_clean.replace('a.applied_at', 'applied_at')
            date_clause = f" WHERE {date_filter_clean}"
            base_query = f'SELECT COUNT(*) AS count FROM applications{date_clause}'
            status_query_base = f'SELECT COUNT(*) AS count FROM applications{date_clause} AND status'
            all_params = date_params
            # For interview query, we need to use alias 'a', so create separate clause
            interview_date_filter = date_filter.strip()
            if interview_date_filter.startswith('AND '):
                interview_date_filter = interview_date_filter[4:]
            interview_date_clause = f" WHERE {interview_date_filter.replace('a.applied_at', 'i.scheduled_date')}"
        else:
            base_query = 'SELECT COUNT(*) AS count FROM applications'
            status_query_base = 'SELECT COUNT(*) AS count FROM applications WHERE status'
            all_params = []
            date_clause = ''
            interview_date_clause = ''

        stats = {
            'total_applications': fetch_count(base_query, tuple(all_params) if all_params else None),
            'pending': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'interviewed': fetch_count(f"{status_query_base} IN ({interviewed_placeholders})", tuple(all_params + interviewed_statuses)),
            'hired': fetch_count(f"{status_query_base} IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
            'rejected': fetch_count(f"{status_query_base} IN ({rejected_placeholders})", tuple(all_params + rejected_statuses)),
            'total_interviews': fetch_count(f"SELECT COUNT(DISTINCT i.interview_id) AS count FROM applications a LEFT JOIN interviews i ON i.application_id = a.application_id{interview_date_clause}", tuple(all_params) if all_params else None),
            # Legacy mappings for backward compatibility
            'applied': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'reviewed': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'under_review': fetch_count(f"{status_query_base} IN ({pending_placeholders})", tuple(all_params + pending_statuses)),
            'interview': fetch_count(f"{status_query_base} IN ({interviewed_placeholders})", tuple(all_params + interviewed_statuses)),
            'accepted': fetch_count(f"{status_query_base} IN ({hired_placeholders})", tuple(all_params + hired_statuses)),
        }

    return stats


def to_iso(value):
    """Convert date/datetime objects to ISO strings for JavaScript consumption."""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    return value


def format_human_datetime(value):
    """Produce a human-readable timestamp in 12-hour format (AM/PM)."""
    if isinstance(value, (datetime, date)):
        dt_value = value if isinstance(value, datetime) else datetime.combine(value, datetime.min.time())
        return dt_value.strftime('%b %d, %Y %I:%M %p')
    elif isinstance(value, str):
        try:
            # Try parsing common datetime formats
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d', '%Y-%m-%d %H:%M']:
                try:
                    dt_value = datetime.strptime(value, fmt)
                    return dt_value.strftime('%b %d, %Y %I:%M %p')
                except ValueError:
                    continue
        except Exception:
            # Ignore parse errors
            pass
    return value or ''


def create_admin_notification(cursor, message, application_id=None):
    """Insert a general notification entry for administrators/admin feed.
    Prevents duplicates by checking if notification with same message and application_id already exists.
    Also prevents JSON responses from being saved as notifications."""
    if not cursor or not message:
        return
    
    # Prevent JSON responses from being saved as notifications
    message_str = str(message).strip()
    if message_str.startswith('{') and ('"success"' in message_str or '"message"' in message_str or '"error"' in message_str):
        print(f'⚠️ Blocked JSON response from being saved as notification: {message_str[:100]}')
        return

    # Block or transform obvious applicant-facing messages from being saved to admin/HR feeds.
    # If message is applicant-facing but `application_id` is provided, attempt to rewrite it
    # into a third-person admin-friendly message (e.g., prefix with applicant name and job),
    # otherwise block and log.
    try:
        import re
        applicant_patterns = re.compile(r"^(\s)*(you|your)\b|\bcongratulations\b|you have been|you applied for|we are pleased to inform you",
                                        flags=re.IGNORECASE)
        if applicant_patterns.search(message_str):
            # Try to transform into admin-friendly message if we have an application_id to lookup
            if application_id is not None:
                try:
                    # Attempt to get applicant full name and job title for a better admin message
                    cursor.execute(
                        '''
                        SELECT a.application_id, ap.full_name AS applicant_name, COALESCE(j.job_title, '') AS job_title
                        FROM applications a
                        LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.application_id = %s
                        LIMIT 1
                        ''',
                        (application_id,)
                    )
                    app_row = cursor.fetchone()
                    applicant_name = None
                    job_title = None
                    if app_row:
                        applicant_name = app_row.get('applicant_name') if isinstance(app_row, dict) else (app_row[1] if len(app_row) > 1 else None)
                        job_title = app_row.get('job_title') if isinstance(app_row, dict) else (app_row[2] if len(app_row) > 2 else None)

                    # Remove leading applicant-directed fragments for a cleaner admin message
                    transformed = re.sub(r'^(congratulations[!]*\s*)', '', message_str, flags=re.IGNORECASE)
                    transformed = re.sub(r'^\s*(you|your)\b[:,]?\s*', '', transformed, flags=re.IGNORECASE)
                    transformed = transformed.strip().rstrip('.')

                    if applicant_name:
                        admin_message = f'Applicant {applicant_name}'
                        if job_title:
                            admin_message += f' ({job_title})'
                        admin_message += f': {transformed}'
                    else:
                        admin_message = transformed

                    print(f'ℹ️ Transformed applicant-facing message into admin message: {admin_message[:200]}')
                    # Replace original message with transformed admin_message for insertion below
                    message = admin_message
                    message_str = str(message).strip()
                except Exception as e:
                    print(f'⚠️ Could not rewrite applicant-facing notification to admin message: {e}')
                    # If we cannot safely rewrite, block to avoid leaking applicant-facing text
                    print(f'⚠️ Blocked applicant-facing notification from being saved to admin feed: {message_str[:200]}')
                    return
            else:
                print(f'⚠️ Blocked applicant-facing notification from being saved to admin feed (no application_id): {message_str[:200]}')
                return
    except Exception:
        # If regex check fails for any reason, continue — prefer not to block valid admin messages inadvertently
        pass
    
    try:
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        if not cursor.fetchone():
            return
        
        cursor.execute('SHOW COLUMNS FROM notifications')
        columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in (cursor.fetchall() or []) if row}
        if 'message' not in columns:
            return
        
        # Check if notification already exists to prevent duplicates
        if application_id is not None and 'application_id' in columns:
            cursor.execute(
                '''
                SELECT notification_id FROM notifications
                WHERE application_id = %s AND message = %s
                LIMIT 1
                ''',
                (application_id, message)
            )
            existing_notification = cursor.fetchone()
            if existing_notification:
                # Notification already exists, skip creation
                return
        
        fields = []
        values = []
        params = []
        
        if application_id is not None and 'application_id' in columns:
            fields.append('application_id')
            values.append('%s')
            params.append(application_id)
        
        fields.append('message')
        values.append('%s')
        params.append(message)
        
        if 'sent_at' in columns:
            fields.append('sent_at')
            values.append('NOW()')
        if 'is_read' in columns:
            fields.append('is_read')
            values.append('0')
        
        sql = f"INSERT INTO notifications ({', '.join(fields)}) VALUES ({', '.join(values)})"
        cursor.execute(sql, tuple(params))
    except Exception as notify_err:
        print(f'⚠️ Notification insert error: {notify_err}')


def _sync_interviews_on_application_hired(cursor, application_id):
    """Ensure interviews are completed or created when an application is marked as 'hired'.

    Behavior:
    - If interviews exist for the application, mark any non-final interviews as 'completed'.
    - If no interview exists, create a completed interview record (best-effort; uses available columns).
    The function uses the provided cursor but DOES NOT commit; the caller should commit the transaction.
    """
    if not cursor or not application_id:
        return
    try:
        # Inspect interviews columns
        cursor.execute("SHOW COLUMNS FROM interviews")
        columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in (cursor.fetchall() or [])}

        # Check if any interview exists for this application
        cursor.execute('SELECT interview_id FROM interviews WHERE application_id = %s', (application_id,))
        existing = cursor.fetchall() or []

        final_statuses = ('completed', 'cancelled', 'no_show')

        if existing:
            # Update any non-final interviews to completed
            if 'updated_at' in columns:
                cursor.execute(
                    'UPDATE interviews SET status = %s, updated_at = NOW() WHERE application_id = %s AND status NOT IN (%s, %s, %s)',
                    ('completed', application_id, *final_statuses),
                )
            else:
                cursor.execute(
                    'UPDATE interviews SET status = %s WHERE application_id = %s AND status NOT IN (%s, %s, %s)',
                    ('completed', application_id, *final_statuses),
                )
        else:
            # Create a completed interview record (best-effort)
            fields = []
            placeholders = []
            params = []

            # application_id (required)
            if 'application_id' in columns:
                fields.append('application_id')
                placeholders.append('%s')
                params.append(application_id)

            # scheduled_date -> use NOW() if available
            use_now_for_scheduled = 'scheduled_date' in columns
            if use_now_for_scheduled:
                fields.append('scheduled_date')
                placeholders.append('NOW()')

            # interview_mode
            if 'interview_mode' in columns:
                fields.append('interview_mode')
                placeholders.append('%s')
                params.append('in-person')

            # status
            if 'status' in columns:
                fields.append('status')
                placeholders.append('%s')
                params.append('completed')

            # notes
            if 'notes' in columns:
                fields.append('notes')
                placeholders.append('%s')
                params.append('Auto-created completed interview because application was marked hired')

            if fields:
                sql = f"INSERT INTO interviews ({', '.join(fields)}) VALUES ({', '.join(placeholders)})"
                cursor.execute(sql, tuple(params))
    except Exception as e:
        print(f'⚠️ Failed to sync/create interview for application {application_id}: {e}')


def format_file_size(num_bytes):
    """Convert a byte value into a human-readable string."""
    if not isinstance(num_bytes, (int, float)) or num_bytes < 0:
        return 'Unknown'

    step = 1024.0
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    size = float(num_bytes)

    for unit in units:
        if size < step or unit == units[-1]:
            return f"{size:.1f} {unit}" if unit != 'B' else f"{int(size)} {unit}"
        size /= step


def parse_decimal_value(raw_value):
    """Convert an arbitrary numeric string into a Decimal or None."""
    if raw_value is None:
        return None

    cleaned = str(raw_value).strip().replace(',', '')
    if not cleaned:
        return None

    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def normalize_choice(value, valid_choices, default):
    """Return a sanitized enum choice using the provided defaults."""
    choice = (value or '').strip().lower()
    return choice if choice in valid_choices else default


def format_salary_range(currency_code, minimum, maximum):
    """Return a human-readable salary range string."""
    if minimum is None and maximum is None:
        return '—'

    def _format_amount(amount):
        if amount is None:
            return ''
        try:
            quantized = Decimal(amount)
        except (InvalidOperation, ValueError):
            return ''
        if quantized == quantized.to_integral():
            return f"{int(quantized):,}"
        return f"{quantized:,.2f}"

    currency = (currency_code or 'PHP').upper()
    min_str = _format_amount(minimum)
    max_str = _format_amount(maximum)

    if min_str and max_str:
        if min_str == max_str:
            return f"{currency} {min_str}"
        return f"{currency} {min_str} - {max_str}"
    if min_str:
        return f"{currency} {min_str}+"
    if max_str:
        return f"Up to {currency} {max_str}"
    return '—'


def get_application_status_label(value):
    """Return a user-friendly label for an application status."""
    status_key = (value or '').strip().lower()
    if not status_key:
        return 'Applied'
    return APPLICATION_STATUS_LABELS.get(status_key, status_key.replace('_', ' ').title())


def fetch_jobs_for_user(user):
    """Fetch job postings, automatically scoping to an HR user's branch."""
    params = []
    where_clause = ''

    branch_scope = get_branch_scope(user)
    if branch_scope is not None:
        where_clause = 'WHERE (j.branch_id = %s OR j.branch_id IS NULL)'
        params.append(branch_scope)

    # Ensure schema compatibility and update job columns cache
    ensure_schema_compatibility()
    db = get_db()
    if db:
        cursor = db.cursor()
        try:
            _update_job_columns(cursor)
        finally:
            cursor.close()

    job_title_col = job_column('job_title', 'title')
    job_description_col = job_column('job_description', 'description')
    job_requirements_col = job_column('job_requirements', 'requirements')
    posted_at_col = job_column('posted_at', 'created_at') or 'created_at'
    created_at_col = job_column('created_at', 'posted_at') or 'created_at'

    job_title_expr = f'j.{job_title_col}' if job_title_col else "'Untitled Job'"
    job_description_expr = f'j.{job_description_col}' if job_description_col else 'NULL'
    job_requirements_expr = f'j.{job_requirements_col}' if job_requirements_col else 'NULL'
    
    # Build admin join for posted_by (actual schema)
    if 'posted_by' in JOB_COLUMNS:
        admin_join = 'LEFT JOIN admins a_posted ON j.posted_by = a_posted.admin_id'
    else:
        admin_join = ''

    query = f'''
        SELECT
            j.job_id,
            {job_title_expr} AS job_title,
            {job_description_expr} AS job_description,
            {job_requirements_expr} AS job_requirements,
            j.status,
            j.branch_id,
            j.created_at,
            j.created_at AS posted_at,
            COALESCE(b.branch_name, 'Unassigned') AS branch_name,
            {job_title_expr} AS position_name,
            'General' AS department,
            COALESCE(a_posted.full_name, 'System') AS posted_by_name,
            (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
        FROM jobs j
        LEFT JOIN branches b ON j.branch_id = b.branch_id
        {admin_join}
        {where_clause}
        ORDER BY COALESCE(j.created_at, j.created_at) DESC
    '''

    return fetch_rows(query, tuple(params) if params else None)


def build_applicant_dashboard_data(applicant_id):
    """Collect metrics and listings tailored to a specific applicant."""
    dashboard = {
        'stats': {
            'total_applications': 0,
            'pending': 0,  # Maps from 'pending' status in database
            'interviewed': 0,  # Maps from 'interviewed' status in database
            'hired': 0,  # Maps from 'hired' status in database
            'rejected': 0,  # Maps from 'rejected' status in database
            # Legacy keys for backward compatibility
            'applied': 0,
            'under_review': 0,
            'interview': 0,
        },
        'applications': [],
        'upcoming_interviews': [],
        'notifications': [],
    }

    if not applicant_id:
        print(f'⚠️ build_applicant_dashboard_data: No applicant_id provided')
        return dashboard

    try:
        db = get_db()
        if not db:
            print(f'⚠️ build_applicant_dashboard_data: No database connection')
            return dashboard
        
        cursor = db.cursor(dictionary=True)
        job_title_expr = "'Untitled Job'"
        job_location_expr = "'Unassigned'"
        
        try:
            # Ensure schema compatibility and get dynamic job column expressions
            ensure_schema_compatibility()
            _update_job_columns(cursor)
            job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
            job_location_expr = job_column_expr('job_location', alternatives=['location'], default="'Unassigned'")

            try:
                dashboard['stats']['total_applications'] = fetch_count(
                    'SELECT COUNT(*) AS count FROM applications WHERE applicant_id = %s',
                    (applicant_id,),
                )
            except Exception as count_error:
                print(f'⚠️ Error fetching total_applications: {count_error}')
                dashboard['stats']['total_applications'] = 0

            try:
                status_rows = fetch_rows(
                    """
                    SELECT status, COUNT(*) AS count
                    FROM applications
                    WHERE applicant_id = %s
                    GROUP BY status
                    """,
                    (applicant_id,),
                )
                for row in (status_rows or []):
                    if row and row.get('status'):
                        status = row['status']
                        count = row.get('count', 0)
                        # Map database statuses to stats keys
                        if status == 'pending':
                            dashboard['stats']['pending'] = count
                            dashboard['stats']['under_review'] = count  # Legacy compatibility
                            dashboard['stats']['applied'] = count  # Legacy compatibility
                        elif status == 'interviewed':
                            dashboard['stats']['interviewed'] = count
                            dashboard['stats']['interview'] = count  # Legacy compatibility
                        elif status == 'hired':
                            dashboard['stats']['hired'] = count
                        elif status == 'rejected':
                            dashboard['stats']['rejected'] = count
                        # Also set directly if key exists (for backward compatibility)
                        if status in dashboard['stats']:
                            dashboard['stats'][status] = count
            except Exception as status_error:
                print(f'⚠️ Error fetching status rows: {status_error}')

            try:
                applications_rows = fetch_rows(
                    f"""
                    SELECT a.application_id,
                           {job_title_expr} AS job_title,
                           COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                           a.status,
                           a.applied_at,
                           COALESCE({job_location_expr}, b.branch_name) AS job_location,
                           CASE 
                               WHEN EXISTS (
                                   SELECT 1 
                                   FROM interviews i 
                                   WHERE i.application_id = a.application_id 
                                   LIMIT 1
                               ) THEN 1 
                               ELSE 0 
                           END AS has_interview
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    WHERE a.applicant_id = %s
                    ORDER BY a.applied_at DESC
                    LIMIT 10
                    """,
                    (applicant_id,),
                )
                dashboard['applications'] = []
                for row in (applications_rows or []):
                    try:
                        submitted_at = row.get('applied_at')
                        dashboard['applications'].append({
                            'application_id': row.get('application_id'),
                            'job_title': row.get('job_title') or 'Untitled Job',
                            'branch_name': row.get('branch_name') or 'Unassigned',
                            'company_name': row.get('branch_name') or 'Unassigned',
                            'job_location': row.get('job_location'),
                            'location': row.get('job_location'),
                            'status': row.get('status'),
                            'applied_at': format_human_datetime(submitted_at) if submitted_at else '',
                            'submitted_at': format_human_datetime(submitted_at) if submitted_at else '',
                            'applied_date': format_human_datetime(submitted_at) if submitted_at else '',
                            'updated_at': format_human_datetime(submitted_at) if submitted_at else '',
                            'has_interview': bool(row.get('has_interview')),
                        })
                    except Exception as row_error:
                        print(f'⚠️ Error processing application row: {row_error}')
                        continue
            except Exception as apps_error:
                print(f'⚠️ Error fetching applications: {apps_error}')
                dashboard['applications'] = []

            try:
                interviews_rows = fetch_rows(
                    f"""
                    SELECT i.scheduled_date,
                           COALESCE(i.interview_mode, 'in-person') AS interview_mode,
                           i.interview_mode,
                           {job_title_expr} AS job_title
                    FROM interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND i.scheduled_date >= NOW()
                    ORDER BY i.scheduled_date ASC
                    LIMIT 5
                    """,
                    (applicant_id,),
                )
                dashboard['upcoming_interviews'] = []
                for row in (interviews_rows or []):
                    try:
                        scheduled_date = row.get('scheduled_date')
                        dashboard['upcoming_interviews'].append({
                            'job_title': row.get('job_title') or 'Interview',
                            'date_time': format_human_datetime(scheduled_date) if scheduled_date else '',
                            'type': (row.get('interview_mode') or 'in-person').replace('-', ' ').title(),
                            'location': row.get('location') or 'To be confirmed',
                        })
                    except Exception as row_error:
                        print(f'⚠️ Error processing interview row: {row_error}')
                        continue
            except Exception as interviews_error:
                print(f'⚠️ Error fetching interviews: {interviews_error}')
                dashboard['upcoming_interviews'] = []

            # Check if sent_at column exists in notifications table
            try:
                cursor.execute('SHOW COLUMNS FROM notifications')
                notification_columns_raw = cursor.fetchall()
                notification_columns = [col.get('Field') if isinstance(col, dict) else col[0] for col in (notification_columns_raw or [])]
                sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notification_columns else 'COALESCE(n.created_at, NOW())'
            except Exception:
                sent_at_expr = 'COALESCE(n.created_at, NOW())'
            
            try:
                notification_rows = fetch_rows(
                    f"""
                    SELECT n.message, {sent_at_expr} AS sent_at
                    FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    WHERE a.applicant_id = %s
                    ORDER BY {sent_at_expr} DESC
                    LIMIT 5
                    """,
                    (applicant_id,),
                )
                dashboard['notifications'] = []
                for row in (notification_rows or []):
                    try:
                        sent_at = row.get('sent_at')
                        dashboard['notifications'].append({
                            'message': row.get('message') or '',
                            'time': format_human_datetime(sent_at) if sent_at else '',
                        })
                    except Exception as row_error:
                        log.warning('⚠️ Error processing notification row: %s', row_error)
                        continue
            except Exception as notif_error:
                log.warning('⚠️ Error fetching notifications: %s', notif_error)
                dashboard['notifications'] = []
        finally:
            cursor.close()

    except Exception as exc:
        log.exception('❌ Failed to build applicant dashboard data: %s', exc)
        # Return empty dashboard on error to prevent 500 error

    return dashboard


def fetch_open_jobs(filters=None, applicant_id=None):
    """Retrieve open job postings with optional filters and smart matching.
    
    IMPORTANT: Shows jobs with status 'open' that are either:
    1. Assigned to ALL branches (branch_id IS NULL) - visible to all applicants
    2. Assigned to specific branch - visible when explicitly filtered or to applicants of that branch
    
    Applicants can see:
    - Jobs with branch_id IS NULL (assigned to all branches) - always visible
    - Jobs from ANY branch if no branch filter is applied
    - Jobs from specific branch if branch_id filter is applied
    """
    # Ensure schema compatibility before querying
    ensure_schema_compatibility()
    db = get_db()
    has_position_name = False
    if db:
        cursor = db.cursor()
        try:
            _update_job_columns(cursor)
            # Check if position_name column exists
            cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
            has_position_name = cursor.fetchone() is not None
        finally:
            cursor.close()
    
    # Use 'status' column directly - we know the schema uses 'status'
    # Filter by publishable statuses - only jobs with status 'open' are visible to applicants
    # Database enum supports: 'open', 'closed' - only 'open' is visible
    status_placeholders = ','.join(['%s'] * len(PUBLISHABLE_JOB_STATUSES))
    where_clauses = ['j.status IN ({})'.format(status_placeholders)]
    params = list(PUBLISHABLE_JOB_STATUSES)

    # Apply filters only if provided and not empty
    if filters and isinstance(filters, dict) and len(filters) > 0:
        if filters.get('keyword'):
            keyword = f"%{filters['keyword']}%"
            keyword_fields = [
                col for col in [
                    job_column('job_title', 'title'),
                    job_column('job_description', 'description'),
                ]
                if col
            ]
            # Build keyword search clauses for job fields
            keyword_clauses = []
            if keyword_fields:
                job_like_clauses = [f"j.{column} LIKE %s" for column in keyword_fields]
                keyword_clauses.extend(job_like_clauses)
                params.extend([keyword] * len(keyword_fields))
            
            # Add branch name to keyword search
            keyword_clauses.append("b.branch_name LIKE %s")
            params.append(keyword)
            
            if keyword_clauses:
                where_clauses.append(f"({' OR '.join(keyword_clauses)})")

        # Filter by branch if explicitly provided
        if filters.get('branch_id'):
            # Show only jobs assigned to that specific branch
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])
        else:
            # No branch filter - show jobs assigned to ALL branches (branch_id IS NULL) or any branch
            # This allows applicants to see jobs meant for everyone
            pass  # By default, all non-filtered jobs are shown

        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])
    
    # Check if saved_only filter is set - need to handle this before building where_sql
    saved_only_filter = filters and filters.get('saved_only') == '1' if filters else False
    
    # Add saved_jobs join if applicant is logged in AND table exists
    saved_join = ''
    is_saved_select = '0 AS is_saved'  # Default to 0 if not logged in
    saved_jobs_table_exists = False
    if applicant_id:
        # Check if saved_jobs table exists before trying to join
        db_check = get_db()
        if db_check:
            cursor_check = db_check.cursor()
            try:
                cursor_check.execute("SHOW TABLES LIKE 'saved_jobs'")
                if cursor_check.fetchone():
                    # Table exists, use the join
                    saved_jobs_table_exists = True
                    # Filter by saved jobs only if saved_only filter is set
                    if saved_only_filter:
                        # Use INNER JOIN to only show saved jobs
                        saved_join = f'INNER JOIN saved_jobs sj ON sj.job_id = j.job_id AND sj.applicant_id = {applicant_id}'
                        is_saved_select = '1 AS is_saved'
                        where_clauses.append('sj.job_id IS NOT NULL')  # ✅ FIX: Use job_id instead of saved_job_id
                    else:
                        # Use LEFT JOIN to show all jobs with saved status
                        saved_join = f'LEFT JOIN saved_jobs sj ON sj.job_id = j.job_id AND sj.applicant_id = {applicant_id}'
                        is_saved_select = 'CASE WHEN sj.job_id IS NOT NULL THEN 1 ELSE 0 END AS is_saved'  # ✅ FIX: Use job_id instead of saved_job_id
                # If table doesn't exist, just use default 0 AS is_saved
            except Exception:
                # If check fails, just use default
                pass
            finally:
                cursor_check.close()

    where_sql = ' AND '.join(where_clauses)

    # Use direct column names since we know the schema: title, description, requirements, status, branch_id, posted_at
    # IMPORTANT: Show ALL jobs with status 'open' from ALL branches - no restrictions
    # This ensures applicants can see ALL jobs posted by HR or Admin with status 'open'
    # Build position_name expression conditionally based on whether column exists
    if has_position_name:
        position_name_expr = 'COALESCE(j.position_name, j.job_title)'
    else:
        position_name_expr = 'j.job_title'
    
    query = f"""
        SELECT
            j.job_id,
            j.job_title AS job_title,
            j.job_description AS job_description,
            '' AS job_requirements,
            j.branch_id,
            j.status,
            COALESCE(j.created_at, j.created_at) AS posted_at,
            COALESCE(b.branch_name, 'Unassigned') AS branch_name,
            {position_name_expr} AS position_title,
            'General' AS department,
            (SELECT COUNT(*) FROM applications a WHERE a.job_id = j.job_id) AS application_count,
            {is_saved_select}
        FROM jobs j
        LEFT JOIN branches b ON j.branch_id = b.branch_id
        {saved_join}
        WHERE {where_sql}
        ORDER BY COALESCE(j.created_at, j.created_at) DESC
    """
    
    try:
        rows = fetch_rows(query, tuple(params))
    except Exception as query_error:
        log.exception('❌ fetch_open_jobs query error: %s', query_error)
        rows = []
    

    jobs = []
    for row in rows:
        salary_display = format_salary_range(
            row.get('salary_currency'),
            row.get('salary_min'),
            row.get('salary_max'),
        )

        jobs.append(
            {
                'job_id': row.get('job_id'),
                'title': row.get('job_title'),
                'summary': (row.get('job_description') or '')[:200] if row.get('job_description') else '',
                'description': row.get('job_description'),
                'requirements': row.get('job_requirements'),
                'employment_type': row.get('employment_type'),
                'work_arrangement': row.get('work_arrangement'),
                'experience_level': row.get('experience_level'),
                'location': row.get('job_location'),
                'salary_currency': row.get('salary_currency'),
                'salary_min': row.get('salary_min'),
                'salary_max': row.get('salary_max'),
                'salary_display': salary_display,
                'branch_id': row.get('branch_id'),
                'branch_name': row.get('branch_name'),
                'position_id': row.get('position_id'),
                'position': row.get('position_title'),
                'position_title': row.get('position_title'),
                'position_name': row.get('position_title'),
                'department': row.get('department'),
                'application_deadline': row.get('application_deadline'),
                'status': row.get('status'),
                'posted_at': format_human_datetime(row.get('posted_at') or row.get('created_at')),
                'application_count': row.get('application_count', 0),
                'is_saved': bool(row.get('is_saved', 0)),
            }
        )

    # Smart matching: if applicant_id provided, calculate match scores
    if applicant_id:
        jobs = add_smart_matching(jobs, applicant_id)

    return jobs


def add_smart_matching(jobs, applicant_id):
    """Add smart matching scores based on applicant profile and application history."""
    db = get_db()
    if not db:
        return jobs
    
    cursor = db.cursor(dictionary=True)
    try:
        # Get applicant's application history
        cursor.execute(
            '''
            SELECT application_id, job_id, status, viewed_at
            FROM applications
            WHERE applicant_id = %s
            ''',
            (applicant_id,),
        )
        fetched = cursor.fetchall()
        # Build a mapping of job_id -> application info
        app_history = {row['job_id']: {'status': row['status'], 'application_id': row['application_id'], 'viewed_at': row.get('viewed_at')} for row in fetched}
        
        # Get applicant's resume keywords (simplified - could be enhanced with NLP)
        # Calculate match scores and annotate each job
        for job in jobs:
            score = 0
            info = app_history.get(job['job_id'])
            if not info:
                score += 10
            else:
                if info.get('status') in {'rejected', 'hired'}:
                    score -= 5
            
            job['match_score'] = max(0, min(100, score))
            job['already_applied'] = job['job_id'] in app_history
            job['application_status'] = info.get('status') if info else None
            job['application_id'] = info.get('application_id') if info else None
            job['is_viewed'] = bool(info.get('viewed_at')) if info else False
    except Exception as exc:
        print(f'⚠️ Smart matching error: {exc}')
    finally:
        cursor.close()
    
    # Sort by match score (highest first)
    jobs.sort(key=lambda x: x.get('match_score', 0), reverse=True)
    return jobs




def fetch_applicants_summary(user=None):
    """Return applicants with aggregated application info, scoped for HR users."""
    branch_id = get_branch_scope(user)

    if branch_id:
        query = """
            SELECT ap.applicant_id,
                   ap.full_name,
                   ap.email,
                   ap.phone_number,
                   ap.created_at,
                   COUNT(a.application_id) AS total_applications,
                   SUM(CASE WHEN a.status = 'applied' THEN 1 ELSE 0 END) AS applied,
                   SUM(CASE WHEN a.status = 'under_review' THEN 1 ELSE 0 END) AS under_review,
                   SUM(CASE WHEN a.status = 'interview' THEN 1 ELSE 0 END) AS interview,
                   SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS hired,
                   SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected
            FROM applicants ap
            JOIN applications a ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE j.branch_id = %s
            GROUP BY ap.applicant_id, ap.full_name, ap.email, ap.phone_number, ap.created_at
            ORDER BY ap.created_at DESC
        """
        params = (branch_id,)
    else:
        query = """
            SELECT ap.applicant_id,
                   ap.full_name,
                   ap.email,
                   ap.phone_number,
                   ap.created_at,
                   COUNT(a.application_id) AS total_applications,
                   SUM(CASE WHEN a.status = 'applied' THEN 1 ELSE 0 END) AS applied,
                   SUM(CASE WHEN a.status = 'under_review' THEN 1 ELSE 0 END) AS under_review,
                   SUM(CASE WHEN a.status = 'interview' THEN 1 ELSE 0 END) AS interview,
                   SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS hired,
                   SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected
            FROM applicants ap
            LEFT JOIN applications a ON a.applicant_id = ap.applicant_id
            GROUP BY ap.applicant_id, ap.full_name, ap.email, ap.phone_number, ap.created_at
            ORDER BY ap.created_at DESC
        """
        params = None

    rows = fetch_rows(query, params)

    for r in rows:
        r['created_at'] = format_human_datetime(r.get('created_at'))
    return rows

@app.route('/')
def index():
    """Landing page."""
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle login for applicants, HR managers, and administrators."""
    # Apply rate limiting only to POST requests (actual login attempts), not GET requests (viewing page)
    if request.method == 'POST':
        from utils.rate_limit import _rate_limit_store, _rate_limit_lock
        from datetime import datetime
        
        identifier = request.remote_addr or 'unknown'
        key = f"login:{identifier}"
        now = datetime.now()
        max_requests = 10  # Allow 10 login attempts
        window_seconds = 900  # Within 15 minutes (was 5 attempts per 5 minutes)
        
        with _rate_limit_lock:
            # Clean old entries (older than window_seconds)
            if key in _rate_limit_store:
                _rate_limit_store[key] = [
                    timestamp for timestamp in _rate_limit_store[key]
                    if (now - timestamp).total_seconds() < window_seconds
                ]
            else:
                _rate_limit_store[key] = []
            
            # Check if limit exceeded
            if len(_rate_limit_store[key]) >= max_requests:
                flash('Too many login attempts. Please wait 15 minutes before trying again.', 'error')
                return immediate_redirect(url_for('login', _external=True))
    
    db = None
    cursor = None
    try:
        if request.method == 'POST':
            # Track this login attempt (only for failed attempts)
            from utils.rate_limit import _rate_limit_store, _rate_limit_lock
            from datetime import datetime
            
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '').strip()
            user_type = request.form.get('user_type', 'applicant').strip().lower()
                
            if not email or not password:
                flash('Please fill in all required fields.', 'error')
                _track_failed_login()
                return render_template('login.html')
                
            # Get database connection with timeout protection
            try:
                db = get_db()
                if not db:
                    flash('Database connection error. Please ensure MySQL is running in XAMPP.', 'error')
                    return render_template('login.html')
            except Exception as db_err:
                flash(f'Database connection failed: {str(db_err)}. Please ensure MySQL is running.', 'error')
                return render_template('login.html')

            cursor = db.cursor(dictionary=True)

            try:
                type_map = {
                    'admin': 'super_admin',
                    'hr': 'hr',
                    'applicant': 'applicant'
                }
                target_type = type_map.get(user_type, 'applicant')

                if target_type in {'super_admin', 'hr'}:
                    cursor.execute(
                    """
                    SELECT
                        u.user_id,
                        u.password_hash,
                        u.user_type,
                        u.is_active,
                        a.admin_id,
                        a.full_name
                    FROM users u
                    LEFT JOIN admins a ON a.user_id = u.user_id
                    WHERE u.email = %s AND u.user_type = %s
                    LIMIT 1
                    """,
                    (email, target_type),
                )
                    account = cursor.fetchone()

                    if not account:
                        # Check if email exists with different user_type to provide helpful error
                        cursor.execute("""
                            SELECT user_type FROM users WHERE email = %s LIMIT 1
                        """, (email,))
                        existing_user = cursor.fetchone()
                        if existing_user:
                            actual_type = existing_user.get('user_type', '')
                            if actual_type == 'super_admin' and user_type != 'admin':
                                flash('Please select "Admin" as user type to login with this email.', 'error')
                            elif actual_type == 'hr' and user_type != 'hr':
                                flash('Please select "HR" as user type to login with this email.', 'error')
                            else:
                                flash('Invalid email or password.', 'error')
                        else:
                            flash('Invalid email or password.', 'error')
                        print(f'❌ Login failed: No account found for email={email}, user_type={target_type}')
                        _track_failed_login()
                        return render_template('login.html')

                    # Validate user_id - it should never be 0 or None
                    user_id = account.get('user_id')
                    if not user_id or user_id == 0 or not isinstance(user_id, int):
                        # Re-query to get the correct user_id directly from users table
                        print(f'⚠️ Invalid user_id={user_id} from query, re-querying users table...')
                        cursor.execute("""
                            SELECT user_id, email, password_hash, is_active, user_type
                            FROM users 
                            WHERE email = %s AND user_type = %s 
                            ORDER BY user_id DESC
                            LIMIT 1
                        """, (email, target_type))
                        user_record = cursor.fetchone()
                        if user_record:
                            found_user_id = user_record.get('user_id')
                            # Check if user_id is 0 - this is a data integrity issue
                            if found_user_id == 0:
                                print(f'⚠️ Found user_id=0 in database for email={email} - attempting to fix...')
                                # Try to find the correct user_id from admins table
                                cursor.execute("""
                                    SELECT a.user_id, a.admin_id
                                    FROM admins a
                                    WHERE a.email = %s
                                    LIMIT 1
                                """, (email,))
                                admin_record = cursor.fetchone()
                                if admin_record and admin_record.get('user_id') and admin_record.get('user_id') > 0:
                                    # Found valid user_id from admins table
                                    correct_user_id = admin_record.get('user_id')
                                    print(f'✅ Found valid user_id={correct_user_id} from admins table for email={email}')
                                    # Use the correct user_id from admins table for login
                                    # Note: We can't UPDATE user_id if it's a primary key, so we'll just use the correct one
                                    user_id = correct_user_id
                                    account['user_id'] = user_id
                                    # Also update password_hash and other fields from the user_record
                                    account['password_hash'] = user_record.get('password_hash') or account.get('password_hash')
                                    account['is_active'] = user_record.get('is_active', 1)
                                    account['user_type'] = user_record.get('user_type', target_type)
                                    print(f'✅ Using user_id={user_id} from admins table (users table has user_id=0 which is invalid)')
                                else:
                                    # No valid user_id in admins table either - try to find any valid user_id for this email
                                    cursor.execute("""
                                        SELECT user_id FROM users 
                                        WHERE email = %s AND user_id > 0 
                                        ORDER BY user_id DESC 
                                        LIMIT 1
                                    """, (email,))
                                    alt_user_record = cursor.fetchone()
                                    if alt_user_record and alt_user_record.get('user_id') and alt_user_record.get('user_id') > 0:
                                        user_id = alt_user_record.get('user_id')
                                        account['user_id'] = user_id
                                        print(f'✅ Found alternative user_id={user_id} for email={email}')
                                    else:
                                        print(f'❌ Cannot find valid user_id for email={email}')
                                        print(f'   User record: {user_record}')
                                        flash('Account configuration error. Invalid user ID in database. Please contact support.', 'error')
                                        _track_failed_login()
                                        return render_template('login.html')
                            elif found_user_id and isinstance(found_user_id, int) and found_user_id > 0:
                                user_id = found_user_id
                                # Update account dict with correct values from users table
                                account['user_id'] = user_id
                                account['password_hash'] = user_record.get('password_hash') or account.get('password_hash')
                                account['is_active'] = user_record.get('is_active', 1)
                                account['user_type'] = user_record.get('user_type', target_type)
                                print(f'✅ Corrected user_id={user_id} for email={email} from users table')
                            else:
                                print(f'❌ Invalid user_id={found_user_id} found in database for email={email}')
                                print(f'   User record: {user_record}')
                                flash('Account configuration error. Invalid user ID in database. Please contact support.', 'error')
                                _track_failed_login()
                                return render_template('login.html')
                        else:
                            # User doesn't exist in users table - this is a critical error
                            print(f'❌ Cannot find user record in users table for email={email}, user_type={target_type}')
                            print(f'   This means the user account does not exist in the users table.')
                            flash('Account not found. Please contact support to create your account.', 'error')
                            _track_failed_login()
                            return render_template('login.html')
                    else:
                        # user_id is valid, but let's verify it exists in the database
                        cursor.execute("""
                            SELECT user_id FROM users WHERE user_id = %s LIMIT 1
                        """, (user_id,))
                        verify_record = cursor.fetchone()
                        if not verify_record:
                            print(f'⚠️ Warning: user_id={user_id} from query does not exist in users table, re-querying...')
                            # Re-query by email
                            cursor.execute("""
                                SELECT user_id FROM users WHERE email = %s AND user_type = %s AND user_id > 0
                                ORDER BY user_id DESC
                                LIMIT 1
                            """, (email, target_type))
                            verify_record = cursor.fetchone()
                            if verify_record:
                                correct_user_id = verify_record.get('user_id')
                                if correct_user_id and correct_user_id > 0:
                                    user_id = correct_user_id
                                    account['user_id'] = user_id
                                    print(f'✅ Corrected user_id={user_id} for email={email}')
                                else:
                                    print(f'❌ Invalid user_id={correct_user_id} from verification query')
                                    flash('Account configuration error. Please contact support.', 'error')
                                    _track_failed_login()
                                    return render_template('login.html')
                            else:
                                print(f'❌ Cannot verify user_id={user_id} - user not found in database')
                                flash('Account configuration error. Please contact support.', 'error')
                                _track_failed_login()
                                return render_template('login.html')

                    if not account.get('admin_id'):
                        # Auto-fix: Create missing admin record for this user
                        print(f'⚠️ Auto-fixing: Creating missing admin record for email={email}, user_id={user_id}')
                        
                        try:
                            # Final validation: user_id must be a positive integer
                            if not user_id or not isinstance(user_id, int) or user_id <= 0:
                                raise Exception(f"Invalid user_id={user_id} (must be positive integer) before creating admin record")
                            
                            # Verify user exists - but use the user_id we already validated
                            # Don't re-query by email as it might return user_id=0 again
                            cursor.execute("""
                                SELECT user_id, email, password_hash, is_active, user_type 
                                FROM users 
                                WHERE user_id = %s
                                LIMIT 1
                            """, (user_id,))
                            user_info = cursor.fetchone()
                            if not user_info:
                                # User with this user_id doesn't exist in users table
                                # This is okay if we got user_id from admins table - we'll proceed anyway
                                print(f'⚠️ Warning: user_id={user_id} not found in users table, but proceeding with admin record creation')
                                # Get user info from the account dict or use defaults
                                user_email = email
                                password_hash = account.get('password_hash')
                                is_active = account.get('is_active', 1)
                            else:
                                # User exists, use the data from users table
                                user_email = user_info.get('email', email)
                                password_hash = user_info.get('password_hash') or account.get('password_hash')
                                is_active = user_info.get('is_active', 1)
                            
                            # Final validation before proceeding - user_id should already be validated
                            if not user_id or not isinstance(user_id, int) or user_id <= 0:
                                raise Exception(f"Invalid user_id={user_id} before creating admin record")
                            
                            print(f'✅ Proceeding with admin record creation for user_id={user_id}, email={email}')
                            
                            # Check which columns exist in admins table
                            cursor.execute('SHOW COLUMNS FROM admins')
                            admin_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in cursor.fetchall()}
                            
                            # Build INSERT statement based on available columns
                            fields = ['user_id', 'full_name', 'email']
                            values = ['%s', '%s', '%s']
                            params = [user_id, account.get('full_name') or 'Admin User', user_email]
                            
                            if 'password_hash' in admin_columns and password_hash:
                                fields.append('password_hash')
                                values.append('%s')
                                params.append(password_hash)
                            
                            # branch_id column has been removed from admins table
                            # HR accounts manage all branches
                            
                            # Create admin record
                            sql = f"INSERT INTO admins ({', '.join(fields)}) VALUES ({', '.join(values)})"
                            cursor.execute(sql, tuple(params))
                            admin_id = cursor.lastrowid
                            
                            if not admin_id or admin_id == 0:
                                raise Exception(f"Failed to create admin record - admin_id={admin_id}")
                            
                            db.commit()
                            
                            # Update account dict with new admin_id and ensure user_id is correct
                            account['admin_id'] = admin_id
                            account['user_id'] = user_id  # Ensure user_id is set correctly
                            print(f'✅ Created admin record: admin_id={admin_id} for user_id={user_id}, email={user_email}')
                        except Exception as fix_error:
                            log.exception(f'❌ Failed to auto-fix admin record: {fix_error}')
                            db.rollback()
                            flash('Account configuration error. Admin/HR account not properly set up. Please contact support.', 'error')
                            _track_failed_login()
                            return render_template('login.html')

                    if not account.get('password_hash'):
                        flash('Account error. Please contact support.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not account.get('is_active'):
                        flash('This account has been deactivated. Contact your administrator.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not check_password(account['password_hash'], password):
                        print(f'❌ Login failed: Password check failed for email={email}, user_type={target_type}')
                        flash('Invalid email or password.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    # For admin/HR, require 2FA verification
                    role = 'admin' if account['user_type'] == 'super_admin' else 'hr'
                    final_user_id = account.get('user_id')
                    final_admin_id = account.get('admin_id')
                    admin_name = account.get('full_name', 'Admin')
                    
                    # Create 2FA verification request (use 'hr' table for hr role, 'admin' for super_admin)
                    temp_token, verification_code = create_2fa_verification(final_user_id, email, admin_name, role)
                    if not temp_token:
                        flash('Unable to send verification code. Please try again.', 'error')
                        return render_template('login.html')
                    
                    # Store temp data in session for 2FA verification page
                    session['2fa_temp_token'] = temp_token
                    session['2fa_user_id'] = final_user_id
                    session['2fa_admin_id'] = final_admin_id
                    session['2fa_role'] = role
                    session['2fa_email'] = email
                    session['2fa_full_name'] = admin_name
                    
                    flash(f'Verification code sent to {email}. Please check your email.', 'info')
                    return redirect(url_for('verify_2fa_login', _external=True))

                elif target_type == 'applicant':
                    # Applicant authentication
                    cursor.execute(
                        """
                        SELECT
                            u.user_id,
                            u.password_hash,
                            u.user_type,
                            u.is_active,
                            u.email_verified,
                            ap.applicant_id,
                            ap.full_name,
                            ap.phone_number,
                            ap.verification_token
                        FROM users u
                        JOIN applicants ap ON ap.user_id = u.user_id
                        WHERE u.email = %s AND u.user_type = 'applicant'
                        LIMIT 1
                        """,
                        (email,),
                    )
                    applicant = cursor.fetchone()

                    if not applicant:
                        flash('Invalid email or password.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not applicant.get('password_hash'):
                        flash('Account error. Please contact support.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not check_password(applicant['password_hash'], password):
                        flash('Invalid email or password.', 'error')
                        _track_failed_login()
                        return render_template('login.html')

                    if not applicant.get('is_active'):
                        flash('This account has been deactivated. Please contact support.', 'error')
                        _track_failed_login()
                        return render_template('login.html')
                    
                    # Check email verification
                    email_verified = applicant.get('email_verified', 0)
                    if not email_verified:
                        session['pending_verification_email'] = email
                        flash('Please verify your email address before logging in. Check your inbox for the verification link.', 'error')
                        return immediate_redirect(url_for('resend_verification', _external=True))

                    if not applicant.get('applicant_id'):
                        flash('Account error. Please contact support.', 'error')
                        return render_template('login.html')

                    if not applicant.get('user_id'):
                        flash('Account error. Please contact support.', 'error')
                        return render_template('login.html')

                    # Applicant: skip 2FA and log in directly
                    final_user_id = applicant.get('user_id')
                    final_applicant_id = applicant.get('applicant_id')
                    applicant_name = applicant.get('full_name', 'Applicant')

                    # Log applicant in without 2FA
                    try:
                        login_user(final_applicant_id, 'applicant', email, applicant_name, auth_user_id=final_user_id)
                        flash('Logged in successfully.', 'success')
                        return immediate_redirect(url_for('applicant_dashboard', _external=True))
                    except Exception as e:
                        print(f'❌ Failed to log applicant in: {e}')
                        flash('An unexpected error occurred while logging in. Please try again.', 'error')
                        return render_template('login.html')

            except Exception as exc:
                log.exception('❌ Login error: %s', exc)
                if db:
                    try:
                        db.rollback()
                    except Exception:
                        pass
                flash('An unexpected error occurred. Please try again.', 'error')
                return render_template('login.html')
            finally:
                if cursor:
                    try:
                        cursor.close()
                    except Exception:
                        pass
        else:
            # Handle GET request
            # Only redirect to dashboard if user is actually logged in (not just session data)
            if is_logged_in():
                # Double-check that session is valid by checking if user exists
                # Use try-except to prevent blocking if database is slow
                try:
                    current_user = get_current_user()
                except Exception as user_err:
                    print(f'⚠️ Error getting current user: {user_err}')
                    current_user = None
                
                if current_user:
                    role = session.get('user_role')
                    # Redirect to appropriate dashboard based on role using immediate redirect
                    if role == 'applicant':
                        return immediate_redirect(url_for('applicant_dashboard', _external=True))
                    if role == 'hr':
                        return immediate_redirect(url_for('hr_dashboard', _external=True))
                    if role == 'admin':
                        return immediate_redirect(url_for('admin_dashboard', _external=True))
                else:
                    # Session data exists but user is invalid - clear it
                    session.clear()

            return render_template('login.html')
    except Exception as outer_exc:
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Login outer error: {outer_exc}')
        print(f'Full traceback: {error_details}')
        if db:
            try:
                db.rollback()
            except Exception:
                pass
        flash('An unexpected error occurred. Please try again.', 'error')
        return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Allow applicants to create accounts."""
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        phone_number = request.form.get('phone_number', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
            
        if not all([full_name, email, phone_number, password, confirm_password]):
            flash('Please complete all required fields.', 'error')
            return render_template('register.html')
            
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('register.html')
            
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('register.html')
            
        db = get_db()
        if not db:
            flash('Database connection error. Please try again later.', 'error')
            return render_template('register.html')
            
        cursor = db.cursor(dictionary=True)

        try:
            cursor.execute(
                'SELECT user_id FROM users WHERE email = %s LIMIT 1',
                (email,),
            )
            if cursor.fetchone():
                flash('Email address is already registered.', 'error')
                return render_template('register.html')

            password_hash = hash_password(password)
            verification_token = generate_token()
            
            # Set token expiration to 60 seconds from now
            from datetime import datetime, timedelta
            token_expires = datetime.now() + timedelta(seconds=60)
            
            # Insert user with email_verified = 0 (unverified)
            cursor.execute(
                """
                INSERT INTO users (email, password_hash, user_type, is_active, email_verified)
                VALUES (%s, %s, 'applicant', 1, 0)
                """,
                (email, password_hash),
            )
            user_id = cursor.lastrowid

            # Store password_hash in applicants table with verification token and expiration
            cursor.execute(
                """
                INSERT INTO applicants (user_id, full_name, email, phone_number, password_hash, verification_token, verification_token_expires)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (user_id, full_name, email, phone_number or None, password_hash, verification_token, token_expires),
            )
            applicant_id = cursor.lastrowid
            db.commit()
            
            # Send verification email
            try:
                send_verification_email(email, verification_token, applicant_name=full_name)
                flash('Registration successful! Please check your email to verify your account before logging in.', 'success')
            except Exception as email_error:
                print(f'⚠️ Failed to send verification email: {email_error}')
                flash('Registration successful! However, we could not send the verification email. Please contact support.', 'warning')
        except Exception as exc:
            db.rollback()
            print(f'❌ Registration error: {exc}')
            flash('An unexpected error occurred during registration.', 'error')
            return render_template('register.html')
        finally:
            cursor.close()

        return immediate_redirect(url_for('login', _external=True))
        
    return render_template('register.html')
    

@app.route('/verify-email/<token>')
def verify_email(token):
    """Verify email using token - checks for 60-second expiration."""
    db = get_db()
    if not db:
        flash('Database connection error. Please try again later.', 'error')
        return immediate_redirect(url_for('login', _external=True))

    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            SELECT a.applicant_id, a.email, a.user_id, a.verification_token_expires, u.email_verified
            FROM applicants a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.verification_token = %s
            LIMIT 1
            ''',
            (token,),
        )
        applicant = cursor.fetchone()

        if not applicant:
            flash('The verification link is invalid or has expired.', 'error')
            return immediate_redirect(url_for('login', _external=True))
        
        # Check if already verified
        if applicant.get('email_verified'):
            flash('Your email is already verified. You can log in.', 'info')
            return immediate_redirect(url_for('login', _external=True))
        
        # Check token expiration (60 seconds)
        expires_at = applicant.get('verification_token_expires')
        if expires_at:
            from datetime import datetime
            if isinstance(expires_at, str):
                expires_at = datetime.strptime(expires_at, '%Y-%m-%d %H:%M:%S')
            if datetime.now() > expires_at:
                flash('The verification link has expired. Please request a new verification email.', 'error')
                return immediate_redirect(url_for('resend_verification', _external=True))

        # Mark email as verified in users table
        cursor.execute(
            '''
            UPDATE users
            SET email_verified = 1
            WHERE user_id = %s
            ''',
            (applicant['user_id'],),
        )
        
        # Clear verification token in applicants table
        cursor.execute(
            '''
            UPDATE applicants
            SET verification_token = NULL, verification_token_expires = NULL
            WHERE applicant_id = %s
            ''',
            (applicant['applicant_id'],),
        )
        db.commit()
        flash('Email verified successfully! You can now log in.', 'success')
    except Exception as exc:
        db.rollback()
        log.exception('❌ Email verification error: %s', exc)
        flash('Unable to verify email at this time.', 'error')
    finally:
        cursor.close()

    return immediate_redirect(url_for('login', _external=True))


@app.route('/resend-verification', methods=['GET', 'POST'])
def resend_verification():
    """Resend verification email to user with 60-second token expiration."""
    preset_email = session.pop('pending_verification_email', None)

    if request.method == 'POST':
        email = (request.form.get('email') or preset_email or '').strip().lower()
        
        if not email:
            flash('Please enter your email address.', 'error')
            return render_template('resend_verification.html', preset_email=preset_email)

        db = get_db()
        if not db:
            flash('Database connection error. Please try again later.', 'error')
            return render_template('resend_verification.html', preset_email=preset_email)

        cursor = db.cursor(dictionary=True)
        try:
            # Check if user exists and is not verified
            cursor.execute(
                '''
                SELECT a.applicant_id, a.email, a.full_name, a.user_id, u.email_verified
                FROM applicants a
                JOIN users u ON u.user_id = a.user_id
                WHERE a.email = %s
                LIMIT 1
                ''',
                (email,),
            )
            applicant = cursor.fetchone()

            if not applicant:
                # Don't reveal if email exists for security
                flash('If this email is registered and not verified, a verification email will be sent.', 'info')
                return render_template('resend_verification.html', preset_email=preset_email)

            if applicant.get('email_verified'):
                flash('Your email is already verified. You can log in.', 'info')
                return immediate_redirect(url_for('login', _external=True))
            
            # Generate new token and set expiration (60 seconds)
            from datetime import datetime, timedelta
            verification_token = generate_token()
            token_expires = datetime.now() + timedelta(seconds=60)
            
            # Update verification token
            cursor.execute(
                '''
                UPDATE applicants
                SET verification_token = %s, verification_token_expires = %s
                WHERE applicant_id = %s
                ''',
                (verification_token, token_expires, applicant['applicant_id']),
            )
            db.commit()

            # Send verification email
            try:
                send_verification_email(email, verification_token, applicant_name=applicant.get('full_name'))
                flash('Verification email sent! Please check your inbox and click the link to verify your email. The link expires in 60 seconds.', 'success')
            except Exception as email_error:
                log.warning('⚠️ Failed to send verification email: %s', email_error)
                flash('Failed to send verification email. Please try again later or contact support.', 'error')
            
        except Exception as exc:
            db.rollback()
            log.exception('❌ Resend verification error: %s', exc)
            flash('An error occurred. Please try again later.', 'error')
        finally:
            cursor.close()

        return immediate_redirect(url_for('login', _external=True))

    # Get preset email from session if available (from login redirect)
    return render_template('resend_verification.html', preset_email=preset_email)


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user_type = request.form.get('user_type', 'applicant').strip()  # applicant, admin, hr
        
        if not email:
            flash('Please enter your email address.', 'error')
            return redirect(url_for('forgot_password'))

        db = get_db()
        if not db:
            flash('Database connection error. Please try again later.', 'error')
            return redirect(url_for('forgot_password'))

        cursor = db.cursor(dictionary=True)
        try:
            user_found = False
            role = 'applicant'
            
            # Check applicants
            if user_type == 'applicant':
                cursor.execute(
                    '''
                    SELECT applicant_id
                    FROM applicants
                    WHERE email = %s
                    LIMIT 1
                    ''',
                    (email,),
                )
                user = cursor.fetchone()
                if user:
                    user_found = True
                    role = 'applicant'
            
            # Check admins (both super_admin and hr)
            if not user_found:
                cursor.execute(
                    '''
                    SELECT a.admin_id, u.user_type, u.is_active
                    FROM admins a
                    JOIN users u ON u.user_id = a.user_id
                    WHERE a.email = %s AND u.is_active = 1
                    LIMIT 1
                    ''',
                    (email,),
                )
                admin = cursor.fetchone()
                if admin:
                    user_found = True
                    role = 'admin' if admin.get('user_type') == 'super_admin' else 'hr'

            if not user_found:
                flash('If an account exists for that email, you will receive a password reset link shortly.', 'info')
                return immediate_redirect(url_for('login', _external=True))

            token = generate_token()

            cursor.execute(
                'DELETE FROM password_resets WHERE user_email = %s',
                (email,),
            )
            cursor.execute(
                '''
                INSERT INTO password_resets (user_email, token, role, expired_at)
                VALUES (%s, %s, %s, DATE_ADD(NOW(), INTERVAL 30 MINUTE))
                ''',
                (email, token, role),
            )
            db.commit()

            send_password_reset_email(email, token)
            flash('If an account exists for that email, a password reset link has been sent.', 'info')
            return immediate_redirect(url_for('login', _external=True))
        except Exception as exc:
            db.rollback()
            print(f'❌ Forgot password error: {exc}')
            flash('Unable to process password reset right now.', 'error')
        finally:
            cursor.close()

    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    db = get_db()
    if not db:
        flash('Database connection error. Please try again later.', 'error')
        return redirect(url_for('forgot_password'))

    cursor = db.cursor(dictionary=True)
    try:
        # Fetch the reset record
        cursor.execute(
            '''
            SELECT user_email, role, expired_at
            FROM password_resets
            WHERE token = %s
            ORDER BY created_at DESC
            LIMIT 1
            ''',
            (token,),
        )
        reset_record = cursor.fetchone()

        if not reset_record:
            flash('The password reset link is invalid or has expired.', 'error')
            return redirect(url_for('forgot_password'))

        # Check if token has expired
        expired_at = reset_record.get('expired_at')
        if expired_at:
            # Always use naive datetime comparison since database stores naive datetimes
            current_time = datetime.now()
            
            if expired_at < current_time:
                flash('The password reset link has expired. Please request a new one.', 'error')
                return redirect(url_for('forgot_password'))

        # For GET requests, just show the reset form
        if request.method == 'GET':
            return render_template('reset_password.html', token=token)

        # For POST requests, process the password reset
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not new_password or not confirm_password:
            flash('Please enter and confirm your new password.', 'error')
            return redirect(url_for('reset_password', token=token))

        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return redirect(url_for('reset_password', token=token))

        if new_password != confirm_password:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('reset_password', token=token))

        role = reset_record.get('role', 'applicant')
        email = reset_record['user_email']
        
        # Update password based on role
        if role == 'applicant':
            cursor.execute(
                'SELECT applicant_id, user_id FROM applicants WHERE email = %s LIMIT 1',
                (email,),
            )
            user = cursor.fetchone()
            if not user:
                flash('Unable to locate the account for password reset.', 'error')
                return redirect(url_for('forgot_password'))
            
            # Hash the new password
            password_hash = hash_password(new_password)
            
            # Update both applicants and users table
            cursor.execute(
                '''
                UPDATE applicants
                SET password_hash = %s
                WHERE applicant_id = %s
                ''',
                (password_hash, user['applicant_id']),
            )
            
            # Also update the users table
            if user.get('user_id'):
                cursor.execute(
                    'UPDATE users SET password_hash = %s WHERE user_id = %s',
                    (password_hash, user['user_id']),
                )
            
            # Log the password change
            try:
                log_profile_change(user['applicant_id'], 'applicant', 'password', '[updated]', '[updated]')
            except Exception as log_exc:
                print(f'⚠️ Warning: Could not log password change: {log_exc}')
        else:  # admin or hr
            cursor.execute(
                '''
                SELECT a.admin_id, u.user_id
                FROM admins a
                JOIN users u ON u.user_id = a.user_id
                WHERE a.email = %s
                LIMIT 1
                ''',
                (email,),
            )
            admin = cursor.fetchone()
            if not admin:
                flash('Unable to locate the account for password reset.', 'error')
                return redirect(url_for('forgot_password'))
            
            cursor.execute(
                'UPDATE users SET password_hash = %s WHERE user_id = %s',
                (hash_password(new_password), admin['user_id']),
            )
        
        # Delete the reset token
        cursor.execute(
            'DELETE FROM password_resets WHERE token = %s',
            (token,),
        )
        db.commit()
        flash('Your password has been reset. You can now log in.', 'success')
        return immediate_redirect(url_for('login', _external=True))

    except Exception as exc:
        db.rollback()
        log.exception('❌ Reset password error: %s', exc)
        flash('Unable to reset password at this time.', 'error')
        return redirect(url_for('forgot_password'))
    finally:
        cursor.close()

@app.route('/verify-2fa-login', methods=['GET', 'POST'])
def verify_2fa_login():
    """Verify 2FA code for admin, HR, and applicant login."""
    # Check if user has initiated 2FA
    temp_token = session.get('2fa_temp_token')
    user_id = session.get('2fa_user_id')
    admin_id = session.get('2fa_admin_id')
    applicant_id = session.get('2fa_applicant_id')
    role = session.get('2fa_role')
    email = session.get('2fa_email')
    full_name = session.get('2fa_full_name')
    # If somehow an applicant reached this page, abort — applicants no longer use 2FA.
    if role == 'applicant':
        for k in ('2fa_temp_token','2fa_user_id','2fa_admin_id','2fa_applicant_id','2fa_role','2fa_email','2fa_full_name'):
            session.pop(k, None)
        flash('Two-factor authentication is not required for applicants. Please log in normally.', 'info')
        return redirect(url_for('login', _external=True))

    if not temp_token or not user_id:
        flash('No pending verification. Please log in again.', 'error')
        return redirect(url_for('login', _external=True))
    
    if request.method == 'POST':
        verification_code = request.form.get('verification_code', '').strip()
        
        if not verification_code:
            flash('Please enter the verification code.', 'error')
            return render_template('verify_2fa_login.html', role=role)
        
        # Verify the code with the appropriate role
        is_valid, result = verify_2fa_code(temp_token, verification_code, role)
        
        if is_valid:
            # Code verified successfully - complete the login
            try:
                # Clear 2FA session data
                session.pop('2fa_temp_token', None)
                session.pop('2fa_user_id', None)
                session.pop('2fa_admin_id', None)
                session.pop('2fa_applicant_id', None)
                session.pop('2fa_role', None)
                session.pop('2fa_email', None)
                session.pop('2fa_full_name', None)
                
                # Determine login ID based on role
                login_id = admin_id if role in ('admin', 'hr') else applicant_id
                
                # Log in the user
                login_user(login_id, role, email, full_name, auth_user_id=user_id)
                session.pop('branch_id', None)
                
                flash('Welcome back!', 'success')
                
                # Route to appropriate dashboard
                if role == 'applicant':
                    return immediate_redirect(url_for('applicant_dashboard', _external=True))
                elif role == 'hr':
                    return immediate_redirect(url_for('hr_dashboard', _external=True))
                else:  # admin
                    return immediate_redirect(url_for('admin_dashboard', _external=True))
            except Exception as exc:
                print(f'❌ 2FA login error: {exc}')
                import traceback
                traceback.print_exc()
                flash('Error during login. Please try again.', 'error')
                return redirect(url_for('login', _external=True))
        else:
            # Code is invalid
            remaining_attempts = result if isinstance(result, int) else 0
            if remaining_attempts > 0:
                flash(f'Invalid code. {remaining_attempts} attempt(s) remaining.', 'error')
            else:
                flash('Too many failed attempts. Please log in again.', 'error')
                # Clear 2FA data and redirect to login
                session.pop('2fa_temp_token', None)
                session.pop('2fa_user_id', None)
                session.pop('2fa_admin_id', None)
                session.pop('2fa_applicant_id', None)
                session.pop('2fa_role', None)
                session.pop('2fa_email', None)
                session.pop('2fa_full_name', None)
                return redirect(url_for('login', _external=True))
            return render_template('verify_2fa_login.html', role=role, email=email)
    
    return render_template('verify_2fa_login.html', role=role, email=email)


@app.route('/verify-password-change', methods=['GET', 'POST'])
@login_required()
def verify_password_change():
    """Verify OTP for a pending password change stored in session['pw_change']."""
    from datetime import datetime
    pw = session.get('pw_change')
    if not pw:
        flash('No pending password change request found.', 'error')
        return redirect(url_for('index'))

    # Ensure the session user matches the pending change
    current_role = session.get('user_role')
    # Basic check: if applicant role, match applicant; if admin/hr, ensure admin/hr
    if pw.get('role') == 'applicant' and current_role != 'applicant':
        flash('Unauthorized.', 'error')
        session.pop('pw_change', None)
        return redirect(url_for('login'))
    if pw.get('role') in ('admin', 'hr') and current_role not in ('admin', 'hr'):
        flash('Unauthorized.', 'error')
        session.pop('pw_change', None)
        return redirect(url_for('login'))

    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        if not code:
            flash('Please enter the verification code.', 'error')
            return render_template('verify_password_change.html')

        # Check expiry
        try:
            if float(pw.get('expires', 0)) < datetime.now().timestamp():
                session.pop('pw_change', None)
                flash('The verification code has expired. Please try changing your password again.', 'error')
                return redirect(url_for('applicant_profile') if pw.get('role') == 'applicant' else url_for('admin_profile'))
        except Exception:
            pass

        if code != pw.get('code'):
            flash('Invalid verification code.', 'error')
            return render_template('verify_password_change.html')

        # Code valid — apply the password change
        db = get_db()
        if not db:
            flash('Database error.', 'error')
            return redirect(url_for('index'))

        cursor = db.cursor()
        try:
            role = pw.get('role')
            new_hash = pw.get('hash')
            email = pw.get('email')
            if role == 'applicant':
                applicant_id = pw.get('applicant_id')
                cursor.execute('UPDATE applicants SET password_hash = %s WHERE applicant_id = %s', (new_hash, applicant_id))
                # also update users table if user_id present
                if pw.get('user_id'):
                    cursor.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', (new_hash, pw.get('user_id')))
            else:
                # admin/hr
                # find user_id for given email
                cursor.execute('SELECT user_id FROM users WHERE email = %s LIMIT 1', (email,))
                row = cursor.fetchone()
                if row:
                    uid = row[0]
                    cursor.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', (new_hash, uid))

            # cleanup
            db.commit()
            session.pop('pw_change', None)
            flash('Your password has been changed successfully.', 'success')
            return redirect(url_for('index'))
        except Exception as exc:
            db.rollback()
            print(f'❌ Error applying password change: {exc}')
            flash('Unable to change password now.', 'error')
            return redirect(url_for('index'))
        finally:
            cursor.close()

    return render_template('verify_password_change.html')

@app.route('/applicant/dashboard')
@login_required('applicant')
def applicant_dashboard():
    """Applicant dashboard powered by live application data."""
    try:
        applicant_id = session.get('user_id')
        if not applicant_id:
            flash('Unable to identify your account. Please log in again.', 'error')
            return immediate_redirect(url_for('login', _external=True))
        
        dashboard = build_applicant_dashboard_data(applicant_id)
        dashboard = dashboard if isinstance(dashboard, dict) else {}
        raw_stats = dashboard.get('stats', {}) if isinstance(dashboard.get('stats'), dict) else {}
        raw_applications = dashboard.get('applications', []) if isinstance(dashboard.get('applications'), list) else []
        raw_interviews = dashboard.get('upcoming_interviews', []) if isinstance(dashboard.get('upcoming_interviews'), list) else []
        
        applicant_info = {
            'name': session.get('user_name') or session.get('user_email') or 'Applicant'
        }
        
        # Map database statuses to display stats
        # Database uses: 'pending', 'interviewed', 'hired', 'rejected'
        stats_view = {
            'total_applications': raw_stats.get('total_applications') or len(raw_applications),
            'pending': raw_stats.get('pending') or raw_stats.get('under_review') or raw_stats.get('reviewed') or 0,
            'interviewed': raw_stats.get('interviewed') or raw_stats.get('interview') or 0,
            'hired': raw_stats.get('hired') or raw_stats.get('accepted') or 0,
            'rejected': raw_stats.get('rejected') or 0,
        }
        
        recent_applications = []
        for app_entry in raw_applications[:5]:
            recent_applications.append({
                'id': app_entry.get('application_id'),
                'job_title': app_entry.get('job_title') or 'Job Removed',
                'company_name': app_entry.get('branch_name') or 'Unassigned Branch',
                'location': app_entry.get('job_location') or app_entry.get('branch_name'),
                'status': (app_entry.get('status') or '').lower(),
                'applied_date': app_entry.get('applied_at') or 'Recently',
                'last_update': app_entry.get('updated_at') or '',
                'has_interview': bool(app_entry.get('has_interview')),
            })
        
        upcoming_interviews = []
        for interview in raw_interviews[:5]:
            scheduled_value = interview.get('date_time') or interview.get('scheduled_date') or 'To be determined'
            upcoming_interviews.append({
                'id': interview.get('interview_id'),
                'job_title': interview.get('job_title') or 'Interview',
                'company_name': interview.get('branch_name') or 'Hiring Team',
                'date': scheduled_value,
                'time': '',
                'time_remaining': '',
                'location': interview.get('location') or 'To be confirmed',
                'is_online': 'online' in ((interview.get('type') or interview.get('interview_mode') or '').lower()),
            })
        
        recent_activity = []
        for item in recent_applications:
            recent_activity.append({
                'type': 'application',
                'description': f"Applied to {item['job_title']}",
                'status': item['status'],
                'timestamp': item['applied_date'],
            })
        for interview in upcoming_interviews:
            recent_activity.append({
                'type': 'interview',
                'description': f"Interview for {interview['job_title']}",
                'status': 'interview',
                'timestamp': interview['date'],
            })
        
        # Fetch rejected applications for the modal
        rejected_applications = []
        try:
            db_rejected = get_db()
            if db_rejected:
                cursor = db_rejected.cursor(dictionary=True)
                try:
                    ensure_schema_compatibility()
                    _update_job_columns(cursor)
                    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
                    
                    cursor.execute(
                        f'''
                        SELECT a.application_id,
                               a.status,
                               a.applied_at,
                               {job_title_expr} AS job_title,
                               COALESCE(b.branch_name, 'Unassigned') AS branch_name
                        FROM applications a
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE a.applicant_id = %s AND a.status = 'rejected'
                        ORDER BY a.applied_at DESC
                        ''',
                        (applicant_id,),
                    )
                    rejected_rows = cursor.fetchall() or []
                    for row in rejected_rows:
                        rejected_applications.append({
                            'id': row.get('application_id'),
                            'job_title': row.get('job_title') or 'Job Removed',
                            'branch_name': row.get('branch_name') or 'Unassigned Branch',
                                'applied_at': format_human_datetime(row.get('applied_at')) if row.get('applied_at') else 'Recently',
                                'submitted_at': format_human_datetime(row.get('applied_at')) if row.get('applied_at') else 'Recently',
                        })
                finally:
                    cursor.close()
        except Exception as rejected_error:
            print(f'⚠️ Error fetching rejected applications: {rejected_error}')
        
        return render_template(
            'applicant/dashboard.html',
            applicant=applicant_info,
            stats=stats_view,
            recent_applications=recent_applications,
            recent_activity=recent_activity[:8],
            upcoming_interviews=upcoming_interviews,
            rejected_applications=rejected_applications,
        )
    except Exception as exc:
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant dashboard error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load dashboard. Please try again later.', 'error')
        # Return empty dashboard to prevent 500 error
        empty_dashboard = {
            'stats': {
                'total_applications': 0,
                'applied': 0,
                'under_review': 0,
                'interview': 0,
                'hired': 0,
                'rejected': 0,
            },
            'applications': [],
            'upcoming_interviews': [],
            'notifications': [],
        }
        return render_template('applicant/dashboard.html', dashboard=empty_dashboard)


@app.route('/applicant/applications')
@login_required('applicant')
def applicant_applications():
    """Applicant's application history with analytics."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/applications.html', applications=[], analytics={})


    
    
    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        # Get dynamic job column expressions
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        
        # Get status filter from request
        status_filter = request.args.get('status', '').strip().lower()
        
        # Build WHERE clause with status filter
        where_clauses = ['a.applicant_id = %s']
        params = [applicant_id]
        
        if status_filter:
            # Map display statuses to database statuses
            status_map = {
                'pending': ['pending', 'reviewed', 'applied', 'under_review'],
                'scheduled': ['scheduled'],
                'interviewed': ['interviewed', 'interview'],
                'hired': ['hired', 'accepted'],
                'rejected': ['rejected', 'withdrawn'],  # Include 'withdrawn' as it's normalized to 'rejected' for display
            }
            db_statuses = status_map.get(status_filter, [status_filter])
            
            # Build IN clause for multiple status mappings
            if db_statuses:
                placeholders = ','.join(['%s'] * len(db_statuses))
                where_clauses.append(f'a.status IN ({placeholders})')
                params.extend(db_statuses)
                print(f"🔍 Applicant status filter: '{status_filter}' -> WHERE a.status IN {db_statuses}")
        
        where_sql = ' AND '.join(where_clauses)
        
        # Fetch all applications with job details, including viewed_at status
        # Normalize withdrawn to rejected in SQL query - remove withdrawn status completely
        # Use subqueries for interviews to prevent duplicates from multiple interviews per application
        cursor.execute(
            f'''
            SELECT DISTINCT a.application_id,
                   CASE 
                       WHEN a.status = 'withdrawn' THEN 'rejected'
                       ELSE a.status
                   END AS status,
                   a.applied_at,
                   a.viewed_at,
                   j.job_id,
                   {job_title_expr} AS job_title,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   {job_title_expr} AS position_title,
                   (SELECT interview_id FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_id,
                   (SELECT scheduled_date FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS scheduled_date,
                   COALESCE((SELECT interview_mode FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1), 'in-person') AS interview_mode,
                   (IF(
                       (SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'interviews' AND COLUMN_NAME = 'location') > 0,
                       (SELECT location FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1),
                       NULL
                   )) AS interview_location,
                   a.viewed_at
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY a.applied_at DESC
            ''',
            tuple(params),
        )
        applications = cursor.fetchall()
        
        # Deduplicate by application_id to prevent any remaining duplicates
        seen_app_ids = set()
        unique_applications = []
        for app in applications:
            app_id = app.get('application_id')
            if app_id and app_id not in seen_app_ids:
                seen_app_ids.add(app_id)
                unique_applications.append(app)
        applications = unique_applications
        
        # Calculate analytics - fetch all applications for accurate stats
        # Normalize withdrawn to rejected in SQL query
        cursor.execute(
            f'''
            SELECT CASE 
                       WHEN a.status = 'withdrawn' THEN 'rejected'
                       ELSE a.status
                   END AS status
            FROM applications a
            WHERE a.applicant_id = %s
            ''',
            (applicant_id,),
        )
        all_apps = cursor.fetchall()
        
        total = len(all_apps)
        status_counts = {
            'pending': 0,
            'reviewed': 0,
            'interviewed': 0,
            'hired': 0,
            'accepted': 0,
            'rejected': 0,
            'applied': 0,
            'under_review': 0,
            'interview': 0
        }
        
        for app in all_apps:
            status = (app.get('status') or 'pending').lower()
            # Status is already normalized to rejected in SQL query (withdrawn -> rejected)
            if status in status_counts:
                status_counts[status] += 1
        
        # Map for display (combine similar statuses)
        # Note: withdrawn is already normalized to rejected in SQL query
        display_counts = {
            'pending': status_counts.get('pending', 0) + status_counts.get('reviewed', 0) + status_counts.get('applied', 0) + status_counts.get('under_review', 0),
            'interviewed': status_counts.get('interviewed', 0) + status_counts.get('interview', 0),
            'hired': status_counts.get('hired', 0) + status_counts.get('accepted', 0),
            'rejected': status_counts.get('rejected', 0),  # Withdrawn already normalized to rejected in SQL
        }
        
        analytics = {
            'total_applications': total,
            'status_breakdown': display_counts,
            'response_rate': round(
                (
                    display_counts.get('interviewed', 0)
                    + display_counts.get('hired', 0)
                    + display_counts.get('rejected', 0)
                )
                / total
                * 100,
                1,
            ) if total > 0 else 0,
            'interview_count': len([a for a in applications if a.get('interview_id')]),
        }
        
        # Format applications
        formatted_apps = []
        for app in applications:
            # Normalize withdrawn to rejected - remove withdrawn status completely
            app_status = (app.get('status') or 'pending').lower()
            if app_status == 'withdrawn':
                app_status = 'rejected'
            
            formatted_apps.append({
                'application_id': app.get('application_id'),
                'job_id': app.get('job_id'),
                'job_title': app.get('job_title'),
                'branch_name': app.get('branch_name'),
                'position_title': app.get('position_title'),
                'status': app_status,  # Use normalized status (withdrawn -> rejected)
                'applied_at': format_human_datetime(app.get('applied_at')),
                'submitted_at': format_human_datetime(app.get('applied_at')),
                'has_interview': app.get('interview_id') is not None,
                'interview_date': format_human_datetime(app.get('scheduled_date')) if app.get('scheduled_date') else None,
                'interview_mode': app.get('interview_mode'),
                'interview_location': app.get('interview_location'),
                'is_viewed': app.get('viewed_at') is not None,  # Check if application has been viewed by HR/Admin
            })
        
        return render_template('applicant/applications.html', applications=formatted_apps, analytics=analytics, status_filter=status_filter)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant applications error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Unable to load applications: {str(exc)}', 'error')
        return render_template('applicant/applications.html', applications=[], analytics={})
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/interviews', methods=['GET', 'POST'])
@login_required('applicant')
def applicant_interviews():
    """Applicant's interview schedule and management."""
    applicant_id = session.get('user_id')
    if not applicant_id:
        flash('Please log in to view your interviews.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    # Validate applicant_id is an integer
    try:
        applicant_id = int(applicant_id)
    except (ValueError, TypeError):
        print(f'⚠️ Invalid applicant_id in session: {applicant_id}')
        flash('Unable to identify your account. Please log in again.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/interviews.html', interviews=[], upcoming=[], past=[])
    
    cursor = None
    try:
        # Ensure schema compatibility - wrap in try-except to prevent crashes
        try:
            ensure_schema_compatibility()
        except Exception as schema_error:
            print(f'⚠️ Schema compatibility check failed (non-critical): {schema_error}')
            # Continue anyway - schema might still be compatible
        
        cursor = db.cursor(dictionary=True)
        # Handle POST actions (confirm/cancel interview)
        if request.method == 'POST':
            action = request.form.get('action')
            interview_id = request.form.get('interview_id', type=int)
            
            if action == 'confirm_interview' and interview_id:
                # Verify interview belongs to this applicant and get current status
                cursor.execute(
                    '''
                    SELECT i.interview_id, i.status, i.scheduled_date
                    FROM interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    WHERE i.interview_id = %s AND a.applicant_id = %s
                    LIMIT 1
                    ''',
                    (interview_id, applicant_id),
                )
                interview = cursor.fetchone()
                if not interview:
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Interview not found.'}), 404
                    flash('Interview not found.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                current_status = (interview.get('status') or 'scheduled').lower()
                
                # Check if already confirmed
                if current_status == 'confirmed':
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'This interview is already confirmed.'})
                    flash('This interview is already confirmed.', 'info')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if cancelled - cannot confirm a cancelled interview
                if current_status == 'cancelled':
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot confirm a cancelled interview. Please contact HR if you need to reschedule.'})
                    flash('Cannot confirm a cancelled interview. Please contact HR if you need to reschedule.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if already completed or no_show
                if current_status in ('completed', 'no_show'):
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot modify the status of a completed interview.'})
                    flash('Cannot modify the status of a completed interview.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if interview is in the past
                scheduled_date = interview.get('scheduled_date')
                if scheduled_date and isinstance(scheduled_date, datetime) and scheduled_date < datetime.now():
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot confirm past interviews.'})
                    flash('Cannot confirm past interviews.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Update interview status to confirmed
                # IMPORTANT: This action does TWO things:
                # 1. Updates the interview status in the database to 'confirmed'
                # 2. Notifies HR users about the confirmation
                # Both operations are committed together in a single transaction
                try:
                    # First, try to update status to 'confirmed'
                    try:
                        # First, ensure enum has 'confirmed' value
                        try:
                            cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                            print('✅ Ensured "confirmed" is in status enum')
                        except Exception as enum_check_err:
                            error_msg = str(enum_check_err).lower()
                            # If enum already has the value or table doesn't exist, that's okay
                            if 'duplicate' not in error_msg and 'already' not in error_msg:
                                print(f'⚠️ Could not modify enum (may already be correct): {enum_check_err}')
                        
                        # Now update the status - wrap in try-except to catch MySQL enum errors
                        try:
                            cursor.execute(
                                'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                ('confirmed', interview_id),
                            )
                            rows_updated = cursor.rowcount
                            
                            if rows_updated == 0:
                                # Verify interview exists
                                cursor.execute('SELECT interview_id FROM interviews WHERE interview_id = %s', (interview_id,))
                                exists = cursor.fetchone()
                                if not exists:
                                    raise Exception(f'Interview {interview_id} does not exist.')
                                else:
                                    raise Exception(f'Update returned 0 rows but interview exists.')
                            
                            print(f'✅ Interview {interview_id} status updated to confirmed. Rows updated: {rows_updated}')
                        except Exception as mysql_err:
                            error_msg = str(mysql_err).lower()
                            # If it's an enum/invalid value error, the ALTER should have fixed it
                            # But if it still fails, add note as fallback
                            if 'enum' in error_msg or 'invalid' in error_msg or 'value' in error_msg:
                                print(f'⚠️ MySQL enum error (unexpected after ALTER): {mysql_err}')
                                # Try one more time after ensuring enum
                                try:
                                    cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                                    cursor.execute(
                                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                        ('confirmed', interview_id),
                                    )
                                    rows_updated = cursor.rowcount
                                    if rows_updated > 0:
                                        print(f'✅ Interview {interview_id} status updated after retry. Rows updated: {rows_updated}')
                                    else:
                                        raise Exception('Retry also returned 0 rows')
                                except Exception as retry_err:
                                    print(f'⚠️ Retry also failed: {retry_err}')
                                    raise
                            else:
                                raise
                    except Exception as status_update_error:
                        # Final fallback: add a note
                        print(f'⚠️ Status update failed: {status_update_error}')
                        try:
                            cursor.execute(
                                'UPDATE interviews SET notes = CONCAT(COALESCE(notes, ""), "\n\n[Applicant Confirmed Attendance on ", NOW(), "]") WHERE interview_id = %s',
                                (interview_id,),
                            )
                            print(f'✅ Added confirmation note to interview {interview_id}')
                        except Exception as note_err:
                            if '1054' in str(note_err) or 'Unknown column' in str(note_err):
                                print(f'⚠️ Note column missing; confirmation recorded without notes')
                            else:
                                print(f'❌ Failed to add confirmation note: {note_err}')
                                raise
                    # Get interview details for HR notification
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.scheduled_date, i.interview_mode, i.interview_mode,
                               a.application_id, a.applicant_id,
                               ap.full_name AS applicant_name, ap.email AS applicant_email,
                               j.job_id, COALESCE(j.job_title, 'Untitled Job') AS job_title,
                               b.branch_id, b.branch_name
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE i.interview_id = %s
                        LIMIT 1
                        ''',
                        (interview_id,),
                    )
                    interview_details = cursor.fetchone()
                    
                    # Notify HR about the confirmation - ALWAYS notify HR when status is updated
                    if interview_details:
                        try:
                            branch_id = interview_details.get('branch_id')
                            applicant_name = interview_details.get('applicant_name') or 'Applicant'
                            job_title = interview_details.get('job_title') or 'Position'
                            scheduled_date = interview_details.get('scheduled_date')
                            scheduled_str = format_human_datetime(scheduled_date) if scheduled_date else 'TBD'
                            
                            notification_message = f'Applicant {applicant_name} has confirmed attendance for interview: {job_title} scheduled on {scheduled_str}.'
                            
                            # Check if notification already exists to prevent duplicates
                            cursor.execute(
                                '''
                                SELECT notification_id FROM notifications
                                WHERE application_id = %s AND message = %s
                                LIMIT 1
                                ''',
                                (interview_details.get('application_id'), notification_message),
                            )
                            existing_notification = cursor.fetchone()
                            
                            if not existing_notification:
                                # Get HR users to notify (for logging purposes)
                                # HR accounts manage all branches (branch_id column removed from admins table)
                                hr_users_to_notify = []
                                cursor.execute(
                                    '''
                                    SELECT a.admin_id, u.email, a.full_name
                                    FROM admins a
                                    JOIN users u ON u.user_id = a.user_id
                                    WHERE a.is_active = 1 AND u.user_type = 'hr' AND u.is_active = 1
                                    '''
                                )
                                all_hr = cursor.fetchall() or []
                                hr_users_to_notify.extend(all_hr)
                                
                                # Intentionally do NOT create an admin/HR notification for applicant confirmation
                                # This notification is applicant-facing only (applicant will be notified via notifications/email)
                                print(f'ℹ️ Applicant confirmation notification created (no HR notification): {notification_message}')
                            else:
                                print(f'⚠️ HR notification already exists for this confirmation - skipping duplicate notification')
                                print(f'   - Status updated to: confirmed')
                        except Exception as notify_err:
                            log.exception('⚠️ Error creating HR notification for interview confirmation: %s', notify_err)
                            # Don't fail the whole operation if notification fails, but log it
                    
                    # Commit the transaction
                    db.commit()
                    print(f'✅ Transaction committed for interview {interview_id} confirmation')
                    
                    # Verify the update was successful
                    cursor.execute(
                        'SELECT status FROM interviews WHERE interview_id = %s LIMIT 1',
                        (interview_id,)
                    )
                    verify_status = cursor.fetchone()
                    actual_status = (verify_status.get('status') or '').lower() if verify_status else None
                    print(f'✅ Verified interview {interview_id} status after update: {actual_status}')
                    
                    if request.accept_mimetypes.accept_json:
                        return jsonify({
                            'success': True, 
                            'message': 'Interview attendance confirmed successfully.', 
                            'status': actual_status or 'confirmed'
                        })
                    flash('Interview attendance confirmed successfully.', 'success')
                except Exception as update_exc:
                    db.rollback()
                    print(f'⚠️ Error updating interview status: {update_exc}')
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Unable to confirm interview. Please contact HR.'}), 500
                    flash('Unable to confirm interview. Please contact HR.', 'error')
            
            elif action == 'cancel_interview' and interview_id:
                # Verify interview belongs to this applicant and get current status
                cursor.execute(
                    '''
                    SELECT i.interview_id, i.status, i.scheduled_date
                    FROM interviews i
                    JOIN applications a ON i.application_id = a.application_id
                    WHERE i.interview_id = %s AND a.applicant_id = %s
                    LIMIT 1
                    ''',
                    (interview_id, applicant_id),
                )
                interview = cursor.fetchone()
                if not interview:
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Interview not found.'}), 404
                    flash('Interview not found.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                current_status = (interview.get('status') or 'scheduled').lower()
                
                # Check if already cancelled
                if current_status == 'cancelled':
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'This interview is already cancelled.'})
                    flash('This interview is already cancelled.', 'info')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if already completed or no_show
                if current_status in ('completed', 'no_show'):
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot cancel a completed interview.'})
                    flash('Cannot cancel a completed interview.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Check if interview is in the past
                scheduled_date = interview.get('scheduled_date')
                if scheduled_date and isinstance(scheduled_date, datetime) and scheduled_date < datetime.now():
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Cannot cancel past interviews.'})
                    flash('Cannot cancel past interviews.', 'error')
                    return immediate_redirect(url_for('applicant_interviews', _external=True))
                
                # Update interview status to cancelled
                # IMPORTANT: This action does TWO things:
                # 1. Updates the interview status in the database to 'cancelled'
                # 2. Notifies HR users about the cancellation
                # Both operations are committed together in a single transaction
                try:
                    # Try to update status to 'cancelled' (should be in enum already)
                    try:
                        cursor.execute(
                            'UPDATE interviews SET status = %s WHERE interview_id = %s',
                            ('cancelled', interview_id),
                        )
                        rows_updated = cursor.rowcount
                        
                        if rows_updated == 0:
                            # Check if interview exists
                            cursor.execute('SELECT interview_id FROM interviews WHERE interview_id = %s', (interview_id,))
                            exists = cursor.fetchone()
                            if not exists:
                                raise Exception(f'Interview {interview_id} does not exist.')
                            else:
                                raise Exception(f'Update failed but interview exists. Possible enum issue.')
                        
                        print(f'✅ Interview {interview_id} status updated to cancelled. Rows updated: {rows_updated}')
                    except Exception as status_update_error:
                        error_msg = str(status_update_error).lower()
                        print(f'⚠️ Status update error: {status_update_error}')
                        
                        # If it's an enum error, try to fix it
                        if 'enum' in error_msg or 'invalid' in error_msg or 'value' in error_msg:
                            try:
                                # Ensure 'cancelled' is in enum
                                cursor.execute("ALTER TABLE interviews MODIFY COLUMN status ENUM('scheduled', 'confirmed', 'rescheduled', 'completed', 'cancelled', 'no_show') DEFAULT 'scheduled'")
                                # Try update again
                                cursor.execute(
                                    'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                    ('cancelled', interview_id),
                                )
                                rows_updated = cursor.rowcount
                                if rows_updated > 0:
                                    print(f'✅ Interview {interview_id} status updated to cancelled after enum fix.')
                                else:
                                    raise Exception('Update still failed after enum fix')
                            except Exception as enum_fix_err:
                                print(f'⚠️ Could not fix enum: {enum_fix_err}')
                                # Fallback: add a note
                                try:
                                    cursor.execute(
                                        'UPDATE interviews SET notes = CONCAT(COALESCE(notes, ""), "\n\n[Applicant Cancelled on ", NOW(), "]") WHERE interview_id = %s',
                                        (interview_id,),
                                    )
                                    print(f'✅ Added cancellation note to interview {interview_id}')
                                except Exception as note_err:
                                    if '1054' in str(note_err) or 'Unknown column' in str(note_err):
                                        print(f'⚠️ Note column missing; cancellation recorded without notes')
                                    else:
                                        raise
                        else:
                            # Other error - try adding note
                            try:
                                cursor.execute(
                                    'UPDATE interviews SET notes = CONCAT(COALESCE(notes, ""), "\n\n[Applicant Cancelled on ", NOW(), "]") WHERE interview_id = %s',
                                    (interview_id,),
                                )
                                print(f'✅ Added cancellation note to interview {interview_id}')
                            except Exception as note_err:
                                if '1054' in str(note_err) or 'Unknown column' in str(note_err):
                                    print(f'⚠️ Note column missing; cancellation recorded without notes')
                                else:
                                    print(f'❌ Failed to add cancellation note: {note_err}')
                                    raise
                    # Get interview details for HR notification
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.scheduled_date, i.interview_mode, i.interview_mode,
                               a.application_id, a.applicant_id,
                               ap.full_name AS applicant_name, ap.email AS applicant_email,
                               j.job_id, COALESCE(j.job_title, 'Untitled Job') AS job_title,
                               b.branch_id, b.branch_name
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE i.interview_id = %s
                        LIMIT 1
                        ''',
                        (interview_id,),
                    )
                    interview_details = cursor.fetchone()
                    
                    # Notify HR about the cancellation - ALWAYS notify HR when status is updated
                    if interview_details:
                        try:
                            branch_id = interview_details.get('branch_id')
                            applicant_name = interview_details.get('applicant_name') or 'Applicant'
                            job_title = interview_details.get('job_title') or 'Position'
                            scheduled_date = interview_details.get('scheduled_date')
                            scheduled_str = format_human_datetime(scheduled_date) if scheduled_date else 'TBD'
                            
                            notification_message = f'Applicant {applicant_name} has cancelled the interview: {job_title} scheduled on {scheduled_str}.'
                            
                            # Check if notification already exists to prevent duplicates
                            cursor.execute(
                                '''
                                SELECT notification_id FROM notifications
                                WHERE application_id = %s AND message = %s
                                LIMIT 1
                                ''',
                                (interview_details.get('application_id'), notification_message),
                            )
                            existing_notification = cursor.fetchone()
                            
                            if not existing_notification:
                                # Create notification for HR users in the same branch
                                hr_users_to_notify = []
                                
                                if branch_id:
                                    # HR accounts manage all branches (branch_id column removed from admins table)
                                    cursor.execute(
                                        '''
                                        SELECT a.admin_id, u.email, a.full_name
                                        FROM admins a
                                        JOIN users u ON u.user_id = a.user_id
                                        WHERE a.is_active = 1 AND u.user_type = 'hr' AND u.is_active = 1
                                        '''
                                    )
                                    all_hr = cursor.fetchall() or []
                                    hr_users_to_notify.extend(all_hr)
                                
                                # Intentionally avoid creating an admin/HR notification for applicant cancellations
                                # Applicant-facing notification already created above; HR need not receive a duplicate.
                                print(f'ℹ️ Applicant cancellation notification created (no HR notification): {notification_message}')
                            else:
                                print(f'⚠️ HR notification already exists for this cancellation - skipping duplicate notification')
                                print(f'   - Status updated to: cancelled')
                        except Exception as notify_err:
                            log.exception('⚠️ Error creating HR notification for interview cancellation: %s', notify_err)
                            # Don't fail the whole operation if notification fails, but log it
                    
                    # Commit the transaction
                    db.commit()
                    print(f'✅ Transaction committed for interview {interview_id} cancellation')
                    
                    # Verify the update was successful
                    cursor.execute(
                        'SELECT status FROM interviews WHERE interview_id = %s LIMIT 1',
                        (interview_id,)
                    )
                    verify_status = cursor.fetchone()
                    actual_status = (verify_status.get('status') or '').lower() if verify_status else None
                    print(f'✅ Verified interview {interview_id} status after update: {actual_status}')
                    
                    if request.accept_mimetypes.accept_json:
                        return jsonify({
                            'success': True, 
                            'message': 'Interview cancellation requested.', 
                            'status': actual_status or 'cancelled'
                        })
                    flash('Interview cancellation requested.', 'success')
                except Exception as update_exc:
                    db.rollback()
                    print(f'⚠️ Error updating interview status: {update_exc}')
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Unable to cancel interview. Please contact HR.'}), 500
                    flash('Unable to cancel interview. Please contact HR.', 'error')
            
            elif action == 'delete_all_interviews':
                # Delete all interviews for this applicant
                try:
                    # Get all interview IDs for this applicant before deletion (for notification)
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.application_id, i.scheduled_date,
                               j.job_id, COALESCE(j.job_title, 'Untitled Job') AS job_title,
                               b.branch_id, b.branch_name
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        LEFT JOIN branches b ON j.branch_id = b.branch_id
                        WHERE a.applicant_id = %s
                        ''',
                        (applicant_id,)
                    )
                    interviews_to_delete = cursor.fetchall()
                    deleted_count = len(interviews_to_delete)
                    
                    if deleted_count == 0:
                        if request.accept_mimetypes.accept_json:
                            return jsonify({'success': False, 'message': 'No interviews found to delete.'})
                        flash('No interviews found to delete.', 'info')
                        return immediate_redirect(url_for('applicant_interviews', _external=True))
                    
                    # Delete all interviews for this applicant
                    cursor.execute(
                        '''
                        DELETE i FROM interviews i
                        INNER JOIN applications a ON i.application_id = a.application_id
                        WHERE a.applicant_id = %s
                        ''',
                        (applicant_id,)
                    )
                    deleted_rows = cursor.rowcount
                    
                    # Notify HR about the deletion
                    if interviews_to_delete:
                        try:
                            # Get unique branch IDs
                            branch_ids = set()
                            for interview in interviews_to_delete:
                                branch_id = interview.get('branch_id')
                                if branch_id:
                                    branch_ids.add(branch_id)
                            
                            # Get applicant info
                            cursor.execute(
                                'SELECT full_name, email FROM applicants WHERE applicant_id = %s LIMIT 1',
                                (applicant_id,)
                            )
                            applicant_data = cursor.fetchone()
                            applicant_name = applicant_data.get('full_name') if applicant_data else 'Applicant'
                            
                            notification_message = f'Applicant {applicant_name} has deleted all {deleted_count} interview(s).'
                            
                            # Do NOT create an admin/HR notification for applicant-initiated deletion of all interviews
                            # Applicant-facing notification has already been recorded; avoid notifying HR.
                            print(f'ℹ️ Applicant deleted all interviews notification recorded (no HR notification): {notification_message}')
                        except Exception as notify_err:
                            print(f'⚠️ Error creating HR notification for deleting all interviews: {notify_err}')
                            # Don't fail the whole operation if notification fails
                    
                    db.commit()
                    print(f'✅ Deleted {deleted_rows} interview(s) for applicant {applicant_id}')
                    
                    if request.accept_mimetypes.accept_json:
                        return jsonify({
                            'success': True,
                            'message': f'All {deleted_rows} interview(s) deleted successfully.',
                            'deleted_count': deleted_rows
                        })
                    flash(f'All {deleted_rows} interview(s) deleted successfully.', 'success')
                except Exception as delete_exc:
                    db.rollback()
                    log.exception('⚠️ Error deleting all interviews: %s', delete_exc)
                    if request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'message': 'Unable to delete all interviews. Please try again.'}), 500
                    flash('Unable to delete all interviews. Please try again.', 'error')
            
            if request.method == 'POST':
                if request.accept_mimetypes.accept_json:
                    # Return JSON response for AJAX requests
                    return jsonify({'success': True, 'redirect': url_for('applicant_interviews')})
                return immediate_redirect(url_for('applicant_interviews', _external=True))
        
        # GET request - fetch interviews
        interviews = []
        try:
            # Check if interviews table exists
            table_exists = False
            try:
                cursor.execute("SHOW TABLES LIKE 'interviews'")
                if cursor.fetchone():
                    table_exists = True
            except Exception as table_check_error:
                print(f'⚠️ Error checking for interviews table: {table_check_error}')
                # Assume table exists and try to query anyway
                table_exists = True
            
            # Check if applications table exists
            applications_table_exists = False
            try:
                cursor.execute("SHOW TABLES LIKE 'applications'")
                if cursor.fetchone():
                    applications_table_exists = True
            except Exception as app_table_check_error:
                print(f'⚠️ Error checking for applications table: {app_table_check_error}')
                applications_table_exists = True
            
            if table_exists and applications_table_exists and cursor:
                # Use the simplest possible query first - no dynamic functions, no complex joins
                # This query should work regardless of schema variations
                # Determine if the interviews table has a `location` column
                try:
                    cursor.execute("SHOW COLUMNS FROM interviews LIKE 'location'")
                    has_location = bool(cursor.fetchone())
                except Exception:
                    has_location = False

                location_select = 'i.location' if has_location else "'' AS location"

                query = f'''
                    SELECT DISTINCT i.interview_id,
                           i.scheduled_date,
                           COALESCE(i.interview_mode, 'in-person') AS interview_mode,
                           i.interview_mode,
                           i.notes,
                           COALESCE(i.status, 'scheduled') AS interview_status,
                           a.application_id,
                           COALESCE(a.status, 'pending') AS application_status,
                           a.job_id,
                           'Untitled Job' AS job_title,
                           'Unassigned' AS branch_name,
                           {location_select} AS location
                    FROM interviews i
                    INNER JOIN applications a ON i.application_id = a.application_id
                    WHERE a.applicant_id = %s
                    ORDER BY i.scheduled_date DESC
                '''
                
                try:
                    cursor.execute(query, (applicant_id,))
                    interviews = cursor.fetchall() or []
                    
                    # Deduplicate by interview_id to prevent duplicates
                    seen_ids = set()
                    unique_interviews = []
                    for interview in interviews:
                        interview_id = interview.get('interview_id')
                        if interview_id and interview_id not in seen_ids:
                            seen_ids.add(interview_id)
                            unique_interviews.append(interview)
                    interviews = unique_interviews
                    
                    # Try to enrich with job title and branch name if possible
                    if interviews:
                        try:
                            # Get job IDs from interviews
                            job_ids = [int(i.get('job_id')) for i in interviews if i.get('job_id')]
                            if job_ids:
                                # Try to get job titles and branch names
                                jobs_query = '''
                                    SELECT j.job_id,
                                           COALESCE(j.job_title, 'Untitled Job') AS job_title,
                                           COALESCE(b.branch_name, 'Unassigned') AS branch_name
                                    FROM jobs j
                                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                                    WHERE j.job_id IN (%s)
                                ''' % ','.join(['%s'] * len(job_ids))
                                cursor.execute(jobs_query, tuple(job_ids))
                                jobs_data = {row['job_id']: row for row in cursor.fetchall()}
                                
                                # Update interviews with job data
                                for interview in interviews:
                                    job_id = interview.get('job_id')
                                    if job_id and job_id in jobs_data:
                                        job_data = jobs_data[job_id]
                                        interview['job_title'] = job_data.get('job_title', 'Untitled Job')
                                        interview['branch_name'] = job_data.get('branch_name', 'Unassigned')
                        except Exception as enrich_error:
                            log.warning('⚠️ Error enriching interview data: %s', enrich_error)
                            # Continue with default values - not critical
                            pass
                except Exception as execute_error:
                    log.exception('⚠️ Error executing interview query: %s', execute_error)
                    # Last resort: absolute minimal query
                    try:
                        # Minimal query fallback - include location if available
                        try:
                            cursor.execute("SHOW COLUMNS FROM interviews LIKE 'location'")
                            has_location_min = bool(cursor.fetchone())
                        except Exception:
                            has_location_min = False
                        location_select_min = 'i.location' if has_location_min else "'' AS location"
                        minimal_query = f'''
                            SELECT DISTINCT i.interview_id,
                                   i.scheduled_date,
                                   i.interview_mode,
                                   i.interview_mode,
                                   i.notes,
                                   i.status AS interview_status,
                                   a.application_id,
                                   a.status AS application_status,
                                   a.job_id,
                                   {location_select_min} AS location
                            FROM interviews i
                            INNER JOIN applications a ON i.application_id = a.application_id
                            WHERE a.applicant_id = %s
                            ORDER BY i.scheduled_date DESC
                        '''
                        cursor.execute(minimal_query, (applicant_id,))
                        interviews = cursor.fetchall() or []
                        
                        # Deduplicate by interview_id
                        seen_ids = set()
                        unique_interviews = []
                        for interview in interviews:
                            interview_id = interview.get('interview_id')
                            if interview_id and interview_id not in seen_ids:
                                seen_ids.add(interview_id)
                                unique_interviews.append(interview)
                        interviews = unique_interviews
                        
                        # Add default values for missing fields
                        for interview in interviews:
                            interview['job_title'] = 'Untitled Job'
                            interview['branch_name'] = 'Unassigned'
                            interview['application_status'] = interview.get('application_status') or 'pending'
                            interview['interview_status'] = interview.get('interview_status') or 'scheduled'
                            interview['interview_mode'] = interview.get('interview_mode') or 'in-person'
                    except Exception as minimal_error:
                        log.exception('⚠️ Error executing minimal interview query: %s', minimal_error)
                        interviews = []
        except Exception as query_error:
            print(f'⚠️ Error fetching interviews: {query_error}')
            import traceback
            traceback.print_exc()
            interviews = []
        
        now = datetime.now()
        upcoming = []
        past = []
        
        for interview in interviews:
            try:
                scheduled = interview.get('scheduled_date')
                interview_status = interview.get('interview_status') or 'scheduled'
                
                if not scheduled:
                    # Skip interviews without a scheduled date
                    continue
                
                # Handle both datetime objects and date strings
                scheduled_dt = None
                if isinstance(scheduled, datetime):
                    scheduled_dt = scheduled
                elif isinstance(scheduled, date):
                    # date but not datetime
                    scheduled_dt = datetime.combine(scheduled, datetime.min.time())
                elif isinstance(scheduled, str):
                    try:
                        scheduled_dt = datetime.strptime(scheduled, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        try:
                            scheduled_dt = datetime.strptime(scheduled, '%Y-%m-%d')
                        except ValueError:
                            try:
                                scheduled_dt = datetime.strptime(scheduled, '%Y-%m-%d %H:%M:%S.%f')
                            except ValueError:
                                log.warning('⚠️ Could not parse scheduled_date: %s', scheduled)
                                continue
                else:
                    print(f'⚠️ Unexpected scheduled_date type: {type(scheduled)}')
                    continue
                
                if not scheduled_dt:
                    continue
                
                # Format the date safely
                try:
                    formatted_date = format_human_datetime(scheduled_dt)
                except Exception as format_error:
                    print(f'⚠️ Error formatting date: {format_error}')
                    formatted_date = str(scheduled_dt)
                
                interview_data = {
                    'interview_id': interview.get('interview_id'),
                    'application_id': interview.get('application_id'),
                    'job_title': interview.get('job_title') or 'Unknown Job Position',
                    'branch_name': interview.get('branch_name') or 'Unassigned',
                    'scheduled_date': formatted_date,
                    'interview_mode': interview.get('interview_mode') or 'in-person',
                    'location': interview.get('location'),
                    'notes': interview.get('notes'),
                    'application_status': interview.get('application_status') or 'pending',
                    'interview_status': interview_status,
                }
                
                if scheduled_dt >= now:
                    upcoming.append(interview_data)
                else:
                    past.append(interview_data)
            except Exception as process_error:
                log.exception('⚠️ Error processing interview %s: %s', interview.get('interview_id'), process_error)
                continue
            
        # Ensure we always return valid data
        return render_template(
            'applicant/interviews.html', 
            interviews=upcoming + past, 
            upcoming=upcoming or [], 
            past=past or []
        )
    except Exception as exc:
        if db:
            try:
                if hasattr(db, 'is_connected'):
                    if db.is_connected():
                        db.rollback()
                else:
                    try:
                        db.rollback()
                    except Exception:
                        # Ignore rollback errors
                        pass
            except Exception:
                # Ignore DB connection checks errors
                pass
        log.exception(f'❌ Applicant interviews error: {exc}')
        
        # Ensure cursor is closed if it exists
        if cursor:
            try:
                cursor.close()
            except Exception:
                # Ignore cursor close errors
                pass
        
        # Try to flash the error, but don't fail if flashing fails
        try:
            flash('Unable to load interviews. Please try again later.', 'error')
        except Exception:
            # Ignore flash errors
            pass
        
        # Always return valid template data - return empty lists to show no interviews instead of error
        try:
            return render_template(
                'applicant/interviews.html', 
                interviews=[], 
                upcoming=[], 
                past=[]
            )
        except Exception as template_error:
            # If template rendering fails, return a simple text response
            log.exception(f'❌ Error rendering template: {template_error}')
            from flask import Response
            return Response('Unable to load interviews. Please try again later.', status=200, mimetype='text/plain')
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                # Ignore cursor close errors
                pass

@app.route('/applicant/notifications')
@login_required('applicant')
def applicant_notifications():
    """Applicant notification center with preferences."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/communications.html', notifications=[], unread_count=0, preferences={})
    
    cursor = db.cursor(dictionary=True)
    try:
        # Fetch notifications
        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        # Verify notifications table exists before continuing
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        table_exists = cursor.fetchone()
        if not table_exists:
            return render_template(
                'applicant/communications.html',
                notifications=[],
                unread_count=0,
                preferences={
                    'email_enabled': True,
                    'email_frequency': 'immediate',
                    'categories': ['application_updates', 'interview_alerts', 'system_messages'],
                },
            )
        
        # Check notifications table columns
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        
        has_application_fk = 'application_id' in notification_columns
        if not has_application_fk:
            print('⚠️ Notifications table missing application_id column; skipping applicant notifications view.')
            preferences = {
                'email_enabled': True,
                'email_frequency': 'immediate',
                'categories': ['application_updates', 'interview_alerts', 'system_messages'],
            }
            return render_template(
                'applicant/communications.html',
                notifications=[],
                unread_count=0,
                preferences=preferences,
            )
        
        # Build sent_at expression
        if 'sent_at' in notification_columns:
            sent_at_expr = 'n.sent_at'
        elif 'created_at' in notification_columns:
            sent_at_expr = 'n.created_at'
        else:
            sent_at_expr = 'NOW()'
        
        # Build is_read expression
        if 'is_read' in notification_columns:
            is_read_expr = 'COALESCE(n.is_read, 0)'
        else:
            is_read_expr = '0'
        
        # Get dynamic job column expressions
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'System Notification'")
        
        select_fields = [
            'n.notification_id',
            'n.message',
            f'{sent_at_expr} AS sent_at',
            f'{is_read_expr} AS is_read',
            'n.application_id',
            'COALESCE(a.status, \'\') AS application_status',
            f'COALESCE({job_title_expr}, \'System Notification\') AS job_title',
        ]
        query = f'''
            SELECT DISTINCT {', '.join(select_fields)}
            FROM notifications n
            JOIN applications a ON n.application_id = a.application_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            WHERE a.applicant_id = %s
            AND (
                n.message LIKE 'You applied for%'
                OR n.message LIKE 'Your application status%'
                OR n.message LIKE 'Congratulations! You have been hired%'
            )
            ORDER BY {sent_at_expr} DESC
            LIMIT 50
        '''
        cursor.execute(query, (applicant_id,))
        notifications = cursor.fetchall()
        
        unread_count = len([n for n in notifications if not n.get('is_read')])
        
        # Fetch notification preferences (if table exists)
        preferences = {
            'email_enabled': True,
            'email_frequency': 'immediate',  # immediate, daily, weekly
            'categories': ['application_updates', 'interview_alerts', 'system_messages'],
        }
        
        formatted_notifications = []
        for notif in notifications:
            formatted_notifications.append({
                'notification_id': notif.get('notification_id'),
                'message': notif.get('message'),
                'sent_at': format_human_datetime(notif.get('sent_at')),
                'is_read': notif.get('is_read', False),
                'application_id': notif.get('application_id'),
                'job_title': notif.get('job_title'),
                'application_status': notif.get('application_status'),
            })
        
        return render_template(
            'applicant/communications.html',
            notifications=formatted_notifications,
            unread_count=unread_count,
            preferences=preferences,
        )
    except Exception as exc:
        if db:
            db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant notifications error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load notifications. Please try again later.', 'error')
        return render_template('applicant/communications.html', notifications=[], unread_count=0, preferences={})
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/notifications/read-all', methods=['POST'])
@login_required('applicant')
def mark_all_notifications_read():
    """Mark all notifications as read for the applicant."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        # Fallback to HTML flow
        if request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('applicant_notifications'))
    
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            UPDATE notifications n
            JOIN applications a ON n.application_id = a.application_id
            SET n.is_read = 1
            WHERE a.applicant_id = %s AND n.is_read = 0
            ''',
            (applicant_id,),
        )
        db.commit()
        # Content negotiation: JSON for AJAX, redirect with flash otherwise
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'All notifications marked as read'})
        flash('All notifications marked as read.', 'success')
        return redirect(url_for('applicant_notifications'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Mark all notifications read error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': str(exc)}), 500
        flash('Failed to mark notifications as read.', 'error')
        return redirect(url_for('applicant_notifications'))
    finally:
        cursor.close()


@app.route('/applicant/notifications/<int:notification_id>/read', methods=['POST'])
@login_required('applicant')
def mark_notification_read(notification_id):
    """Mark a notification as read."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        if request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('applicant_notifications'))
    
    cursor = db.cursor()
    try:
        # Verify ownership
        cursor.execute(
            '''
            UPDATE notifications n
            JOIN applications a ON n.application_id = a.application_id
            SET n.is_read = TRUE
            WHERE n.notification_id = %s AND a.applicant_id = %s
            ''',
            (notification_id, applicant_id),
        )
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True})
        flash('Notification marked as read.', 'success')
        return redirect(url_for('applicant_notifications'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Mark notification read error: {exc}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': str(exc)}), 500
        flash('Failed to mark notification as read.', 'error')
        return redirect(url_for('applicant_notifications'))
    finally:
        cursor.close()


@app.route('/applicant/notifications/preferences', methods=['POST'])
@login_required('applicant')
def update_notification_preferences():
    """Update notification preferences."""
    applicant_id = session.get('user_id')
    email_enabled = request.form.get('email_enabled') == 'on'
    email_frequency = request.form.get('email_frequency', 'immediate')
    
    # Store preferences (could be in a separate table)
    # For now, just flash success
    flash('Notification preferences updated successfully.', 'success')
    return redirect(url_for('applicant_notifications'))


@app.route('/applicant/notifications/delete-all', methods=['POST'])
@login_required('applicant')
def delete_all_applicant_notifications():
    """Delete all notifications for the current applicant."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicant_notifications'))
    
    applicant_id = session.get('user_id')
    cursor = db.cursor()
    try:
        # Ensure notifications table exists
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        if not cursor.fetchone():
            flash('No notifications to delete.', 'info')
            return redirect(url_for('applicant_notifications'))

        # Introspect columns to decide deletion strategy
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row[0] if isinstance(row, (list, tuple)) else row.get('Field') for row in (cursor.fetchall() or []) if row}

        # Delete notifications explicitly targeted to this applicant
        if 'applicant_id' in notification_columns:
            cursor.execute('DELETE FROM notifications WHERE applicant_id = %s', (applicant_id,))
        # Also delete notifications linked via application -> applicant
        if 'application_id' in notification_columns:
            cursor.execute(
                '''
                DELETE n FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                WHERE a.applicant_id = %s
                ''',
                (applicant_id,)
            )

        deleted = cursor.rowcount
        db.commit()

        # Final cleanup: remove any JSON-response notifications or leftover messages related to deletion
        try:
            if 'application_id' in notification_columns:
                cursor.execute(
                    '''
                    DELETE n FROM notifications n
                    LEFT JOIN applications a ON n.application_id = a.application_id
                    WHERE (n.applicant_id = %s OR a.applicant_id = %s OR n.application_id IS NULL)
                    AND (
                        (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                        OR n.message = 'Notifications deleted.'
                        OR n.message LIKE '%Notifications deleted%'
                        OR n.message LIKE '%All notifications deleted%'
                        OR n.message LIKE '%Notification deleted successfully%'
                    )
                    ''',
                    (applicant_id, applicant_id),
                )
            else:
                cursor.execute(
                    '''
                    DELETE FROM notifications
                    WHERE applicant_id = %s
                    AND (
                        (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                        OR message = 'Notifications deleted.'
                        OR message LIKE '%Notifications deleted%'
                        OR message LIKE '%All notifications deleted%'
                        OR message LIKE '%Notification deleted successfully%'
                    )
                    ''',
                    (applicant_id,)
                )
            final_cleaned = cursor.rowcount
            if final_cleaned > 0:
                db.commit()
        except Exception as cleanup_err:
            print(f'⚠️ Error cleaning up JSON notifications for applicant {applicant_id}: {cleanup_err}')

        flash('All notifications deleted.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        print(f'❌ Delete all applicant notifications error: {exc}')
        print(traceback.format_exc())
        flash('Failed to delete notifications.', 'error')
    finally:
        cursor.close()
    return redirect(url_for('applicant_notifications'))


@app.route('/applicant/account')
@login_required('applicant')
def applicant_account():
    """Account settings and security overview for applicants."""
    applicant_id = session.get('user_id')
    auth_user_id = session.get('auth_user_id')  # Get user_id from users table
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/account.html', account={}, login_history=[])

    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            '''
            SELECT full_name,
                   email,
                   phone_number,
                   created_at,
                   last_login
            FROM applicants
            WHERE applicant_id = %s
            LIMIT 1
            ''',
            (applicant_id,),
        )
        account = cursor.fetchone() or {}

        for key in ['created_at', 'last_login']:
            if account.get(key):
                account[key] = format_human_datetime(account.get(key))

        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        # Check auth_sessions table columns
        cursor.execute('SHOW COLUMNS FROM auth_sessions')
        session_columns_raw = cursor.fetchall() or []
        session_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in session_columns_raw if row}
        
        # Build logout_time expression
        if 'last_activity' in session_columns and 'logout_time' in session_columns:
            logout_expr = 'COALESCE(last_activity, logout_time)'
        elif 'logout_time' in session_columns:
            logout_expr = 'logout_time'
        elif 'last_activity' in session_columns:
            logout_expr = 'last_activity'
        else:
            logout_expr = 'NULL'
        
        # Use auth_user_id (from users table) for auth_sessions query
        # If auth_user_id is not available, try to get it from applicants table
        if not auth_user_id:
            cursor.execute(
                'SELECT user_id FROM applicants WHERE applicant_id = %s LIMIT 1',
                (applicant_id,),
            )
            user_record = cursor.fetchone()
            if user_record:
                auth_user_id = user_record.get('user_id')
        
        login_history = []
        active_sessions = 0
        
        if auth_user_id:
            try:
                # Check which columns exist in auth_sessions
                has_ip_address = 'ip_address' in session_columns
                has_user_agent = 'user_agent' in session_columns
                
                # Determine login_time expression (some schemas use 'login_time', others use 'created_at')
                if 'login_time' in session_columns:
                    login_expr = 'login_time'
                elif 'created_at' in session_columns:
                    login_expr = 'created_at'
                elif 'last_login' in session_columns:
                    login_expr = 'last_login'
                else:
                    login_expr = 'NULL'

                # Build SELECT statement dynamically based on available columns
                select_fields = [f'{login_expr} AS login_time', f'{logout_expr} AS logout_time', 'COALESCE(is_active, 1) AS is_active']
                if has_ip_address:
                    select_fields.append("COALESCE(ip_address, 'Unknown') AS ip_address")
                else:
                    select_fields.append("'Unknown' AS ip_address")
                
                if has_user_agent:
                    select_fields.append("COALESCE(user_agent, 'Unknown') AS user_agent")
                else:
                    select_fields.append("'Unknown' AS user_agent")
                
                cursor.execute(
                    f'''
                    SELECT {', '.join(select_fields)}
                    FROM auth_sessions
                    WHERE user_id = %s
                    ORDER BY {login_expr} DESC
                    LIMIT 10
                    ''',
                    (auth_user_id,),
                )
                sessions = cursor.fetchall() or []
                
                for row in sessions:
                    is_active = bool(row.get('is_active', 1))
                    if is_active:
                        active_sessions += 1
                    logout_value = format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None
                    if is_active:
                        logout_value = None
                    login_history.append(
                        {
                            'login_time': format_human_datetime(row.get('login_time')),
                            'logout_time': logout_value,
                            'ip_address': row.get('ip_address') or '—',
                            'user_agent': row.get('user_agent') or 'Unknown device',
                            'status': 'Active' if is_active else 'Signed out',
                            'is_active': is_active,
                        }
                    )
            except Exception as session_error:
                log.exception('⚠️ Error fetching login history: %s', session_error)
                login_history = []
                active_sessions = 0
        
        # Check notifications table for is_read column
        try:
            cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
            has_is_read = cursor.fetchone() is not None
            
            if has_is_read:
                notif_query = '''
                    SELECT 
                        COUNT(*) AS total,
                        SUM(CASE WHEN COALESCE(n.is_read, 0) = 0 THEN 1 ELSE 0 END) AS unread
                    FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    WHERE a.applicant_id = %s
                '''
            else:
                notif_query = '''
                    SELECT 
                        COUNT(*) AS total,
                        0 AS unread
                    FROM notifications n
                    JOIN applications a ON n.application_id = a.application_id
                    WHERE a.applicant_id = %s
                '''
            cursor.execute(notif_query, (applicant_id,))
        except Exception:
            # If notifications table doesn't exist or query fails, set defaults
            pass
        communications_stats = cursor.fetchone() or {'total': 0, 'unread': 0}

        account_overview = {
            'full_name': account.get('full_name'),
            'email': account.get('email'),
            'phone_number': account.get('phone_number'),
            'joined_at': account.get('created_at'),
            'last_login': account.get('last_login'),
            'active_sessions': active_sessions,
            'login_history': login_history,
            'communications': communications_stats,
        }

        # Fetch notification preferences (default values if table doesn't exist)
        preferences = {
            'email_enabled': True,
            'email_notifications': True,
            'email_frequency': 'immediate',
            'categories': ['application_updates', 'interview_alerts', 'system_messages'],
        }

        return render_template('applicant/account.html', account=account_overview, preferences=preferences)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant account settings error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Unable to load account settings: {str(exc)}', 'error')
        preferences = {
            'email_enabled': True,
            'email_notifications': True,
            'email_frequency': 'immediate',
        }
        return render_template('applicant/account.html', account={}, login_history=[], preferences=preferences)
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/upload-resume', methods=['POST'])
@login_required('applicant')
def upload_resume_before_apply():
    """Upload resume before submitting application - allows applicants to upload resume first."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        return jsonify({'success': False, 'error': 'Database connection error.'}), 500
    
    cursor = db.cursor(dictionary=True)
    try:
        resume_file = request.files.get('resume_file')
        if not resume_file or not resume_file.filename:
            return jsonify({'success': False, 'error': 'No file provided.'}), 400

        # If job-specific file rules exist, fetch them and pass to the saver
        job_id = request.form.get('job_id', type=int)
        allowed_exts = None
        max_size_mb = None
        try:
            if job_id:
                ensure_schema_compatibility()
                _update_job_columns(cursor)
                # Whitelist allowed column names to prevent SQL injection
                ALLOWED_JOB_COLUMNS = {'allowed_extensions', 'max_file_size_mb'}
                sel = []
                for col in ['allowed_extensions', 'max_file_size_mb']:
                    if col in JOB_COLUMNS and col in ALLOWED_JOB_COLUMNS:
                        sel.append(col)
                if sel:
                    # Use backticks to safely quote column names
                    columns_str = ', '.join([f'`{col}`' for col in sel])
                    cursor.execute(f"SELECT {columns_str} FROM jobs WHERE job_id = %s LIMIT 1", (job_id,))
                    job_row = cursor.fetchone()
                    if job_row:
                        # job_row may be dict-like
                        allowed_exts = job_row.get('allowed_extensions') if isinstance(job_row, dict) else job_row[0] if sel and sel[0] == 'allowed_extensions' else None
                        if 'max_file_size_mb' in sel:
                            if isinstance(job_row, dict):
                                max_size_mb = job_row.get('max_file_size_mb')
                            else:
                                max_size_mb = job_row[1] if len(job_row) > 1 else None
        except Exception:
            # Non-fatal - fallback to defaults
            allowed_exts = None
            max_size_mb = None

        file_info, error = save_uploaded_file(resume_file, applicant_id, allowed_extensions=allowed_exts, max_file_size_mb=max_size_mb)
        if not file_info:
            return jsonify({'success': False, 'error': error or 'Unable to process the uploaded resume file.'}), 400
        
        # Respect file_type provided by the client (resume, letter, license)
        file_type = (request.form.get('file_type') or 'resume').strip().lower()
        if file_type not in ('resume', 'letter', 'license'):
            file_type = 'resume'

        cursor.execute(
            '''
            INSERT INTO resumes (applicant_id, file_name, file_path, file_type)
            VALUES (%s, %s, %s, %s)
            ''',
            (
                applicant_id,
                file_info['original_filename'],
                file_info.get('storage_path') or file_info.get('file_path', ''),
                file_type,
            ),
        )
        resume_id = cursor.lastrowid
        db.commit()
        
        return jsonify({
            'success': True,
            'resume_id': resume_id,
            'file_name': file_info['original_filename'],
            'message': 'Resume uploaded successfully!'
        }), 200
    except Exception as exc:
        db.rollback()
        print(f'❌ Error uploading resume: {exc}')
        return jsonify({'success': False, 'error': str(exc)}), 500
    finally:
        cursor.close()


@app.route('/applicant/apply/<int:job_id>', methods=['GET', 'POST'])
@app.route('/applicant/apply', methods=['GET', 'POST'])
@login_required('applicant')
def apply_to_job(job_id=None):
    """Show application form (GET) or submit application (POST) for a job."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('jobs'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Check if editing existing application - accept edit id from querystring OR form POST
        edit_application_id = request.args.get('edit', type=int)
        if not edit_application_id:
            # fallback to form value when submitting the edit form
            try:
                form_edit = request.form.get('edit_application_id')
                if form_edit:
                    edit_application_id = int(form_edit)
            except Exception:
                edit_application_id = None

        existing_app = None
        if edit_application_id:
            # Verify application belongs to applicant and hasn't been viewed
            cursor.execute(
                'SELECT application_id, job_id, viewed_at, resume_id FROM applications WHERE application_id = %s AND applicant_id = %s LIMIT 1',
                (edit_application_id, applicant_id),
            )
            existing_app = cursor.fetchone()
            if not existing_app:
                flash('Application not found.', 'error')
                return redirect(url_for('applicant_applications'))
            if existing_app.get('viewed_at'):
                flash('This application cannot be edited because it has already been viewed by HR/Admin.', 'error')
                return redirect(url_for('applicant_applications'))
            # Use the job_id from existing application if no job_id provided
            if not job_id:
                job_id = existing_app.get('job_id')
        
        # When editing, job cannot be changed - always use the existing job_id
        # (This prevents applicants from changing the job when editing their application)
        if edit_application_id and existing_app:
            # Force use of existing job_id - ignore any job_id from form
            job_id = existing_app.get('job_id')
        
        # Check if already applied (for new applications only) - run during GET or when route-enforced job_id is final
        if request.method == 'GET' and not edit_application_id and job_id:
            cursor.execute(
                'SELECT application_id FROM applications WHERE applicant_id = %s AND job_id = %s LIMIT 1',
                (applicant_id, job_id),
            )
            if cursor.fetchone():
                flash('You have already applied for this position.', 'warning')
                return redirect(url_for('jobs'))
        
        # Get all available jobs for dropdown (when editing)
        ensure_schema_compatibility()
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        job_description_expr = job_column_expr('job_description', alternatives=['description'])
        job_requirements_expr = job_column_expr('job_requirements', alternatives=['requirements'])
        job_allowed_ext_expr = job_column_expr('allowed_extensions')
        job_max_size_expr = job_column_expr('max_file_size_mb')
        
        # Fetch all open/active jobs for the job selector
        status_placeholders = ','.join(['%s'] * len(PUBLISHABLE_JOB_STATUSES))
        cursor.execute(
            f'''
            SELECT 
                j.job_id,
                {job_title_expr} AS job_title,
                {job_description_expr} AS job_description,
                {job_requirements_expr} AS job_requirements,
                {job_allowed_ext_expr} AS allowed_extensions,
                {job_max_size_expr} AS max_file_size_mb,
                COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                {job_title_expr} AS position_title
            FROM jobs j
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE j.status IN ({status_placeholders})
            ORDER BY j.created_at DESC, j.job_id DESC
            ''',
            tuple(PUBLISHABLE_JOB_STATUSES),
        )
        all_jobs = cursor.fetchall()
        
        # Get specific job details if job_id is provided
        job = None
        if job_id:
            job = next((j for j in all_jobs if j.get('job_id') == job_id), None)
            if not job:
                flash('Job not found or no longer available.', 'error')
                return redirect(url_for('jobs'))
        elif edit_application_id and existing_app:
            # If editing and no job_id, use the existing job
            job_id = existing_app.get('job_id')
            job = next((j for j in all_jobs if j.get('job_id') == job_id), None)
        
        # If no specific job is provided and not editing, render the form with the dropdown
        if not job and not edit_application_id:
            # Informational only; allow applicant to choose from the dropdown on the page
            flash('Please choose a job from the dropdown.', 'info')
            
            # GET: Show application form with jobs list
            cursor.execute(
                '''
                SELECT resume_id, file_name, uploaded_at
                FROM resumes
                WHERE applicant_id = %s
                ORDER BY uploaded_at DESC
                ''',
                (applicant_id,),
            )
            resumes = cursor.fetchall()
            formatted_resumes = []
            for resume in resumes:
                file_size_bytes = 0
                try:
                    fp = (resume.get('file_path') or '').replace('\\', '/')
                    if fp:
                        abs_fp = os.path.join(app.root_path, fp)
                        if os.path.exists(abs_fp):
                            file_size_bytes = os.path.getsize(abs_fp)
                except Exception:
                    file_size_bytes = 0
                formatted_resumes.append({
                    'resume_id': resume.get('resume_id'),
                    'file_name': resume.get('file_name'),
                    'uploaded_at': format_human_datetime(resume.get('uploaded_at')),
                    'file_size': format_file_size(file_size_bytes),
                })

            return render_template(
                'applicant/apply.html',
                job=None,
                jobs=all_jobs,
                resumes=formatted_resumes,
                edit_application_id=edit_application_id,
            )
        
        # GET: Show application form
        if request.method == 'GET':
            # Get applicant's existing resumes
            cursor.execute(
                '''
                SELECT resume_id, file_name, uploaded_at
                FROM resumes
                WHERE applicant_id = %s
                ORDER BY uploaded_at DESC
                ''',
                (applicant_id,),
            )
            resumes = cursor.fetchall()
            
            # If applicant has no saved resumes, allow showing the apply form
            # so they can upload a resume as part of the application flow.
            # Previously this redirected applicants to their profile which
            # prevented using the on-form upload button. Keep an informational
            # message but render the form so the upload flow works.
            if not resumes:
                flash('You have no saved resumes yet. You can upload one below before submitting your application.', 'info')
            
            # Format resume data
            formatted_resumes = []
            for resume in resumes:
                # Compute file size from stored file path if available
                file_size_bytes = 0
                try:
                    fp = (resume.get('file_path') or '').replace('\\', '/')
                    if fp:
                        abs_fp = os.path.join(app.root_path, fp)
                        if os.path.exists(abs_fp):
                            file_size_bytes = os.path.getsize(abs_fp)
                except Exception:
                    file_size_bytes = 0

                formatted_resumes.append({
                    'resume_id': resume.get('resume_id'),
                    'file_name': resume.get('file_name'),
                    'uploaded_at': format_human_datetime(resume.get('uploaded_at')),
                    'file_size': format_file_size(file_size_bytes),
                })

            # If editing an application, fetch files already attached to that application
            attached_files = []
            if edit_application_id and existing_app:
                try:
                    cursor.execute(
                        '''
                        SELECT r.resume_id, r.file_name, r.file_path, r.file_type, r.uploaded_at
                        FROM application_attachments aa
                        JOIN resumes r ON aa.resume_id = r.resume_id
                        WHERE aa.application_id = %s
                        ORDER BY r.uploaded_at DESC
                        ''',
                        (edit_application_id,)
                    )
                    arows = cursor.fetchall() or []
                    for a in arows:
                        fp = (a.get('file_path') or '').replace('\\', '/')
                        file_size_bytes = 0
                        try:
                            if fp:
                                abs_fp = os.path.join(app.root_path, fp)
                                if os.path.exists(abs_fp):
                                    file_size_bytes = os.path.getsize(abs_fp)
                        except Exception:
                            file_size_bytes = 0

                        attached_files.append({
                            'resume_id': a.get('resume_id'),
                            'file_name': a.get('file_name'),
                            'uploaded_at': format_human_datetime(a.get('uploaded_at')),
                            'file_size_bytes': file_size_bytes,
                            'file_size': format_file_size(file_size_bytes),
                            'file_type': a.get('file_type') or 'resume',
                                'preview_url': url_for('preview_resume', resume_id=a.get('resume_id'))
                        })
                except Exception as _e:
                    print(f'⚠️ Could not fetch attached files for edit application {edit_application_id}: {_e}')

            application_viewed = bool(existing_app.get('viewed_at')) if existing_app else False
            return render_template(
                'applicant/apply.html',
                job=job,
                jobs=all_jobs,  # Pass all jobs for dropdown
                resumes=formatted_resumes,
                attached_files=attached_files,
                edit_application_id=edit_application_id,
                application_viewed=application_viewed,
            )
        
        # POST: Submit application
        resume_id = None
        # Collect IDs of resumes we create from multi-file upload so they can be linked
        newly_uploaded_ids = []

        # Support client-side multi-file uploads named file_0, file_1, ... with a form field `total_files`
        try:
            total_files = int(request.form.get('total_files') or 0)
        except Exception:
            total_files = 0

        if total_files and total_files > 0:
            # Save each uploaded file and create resume records
            for i in range(total_files):
                fkey = f'file_{i}'
                resume_file = request.files.get(fkey)
                if not resume_file or not getattr(resume_file, 'filename', None):
                    continue
                try:
                    allowed_exts = job.get('allowed_extensions') if job else None
                    max_size_mb = job.get('max_file_size_mb') if job else None
                except Exception:
                    allowed_exts = None
                    max_size_mb = None

                file_info, error = save_uploaded_file(resume_file, applicant_id, allowed_extensions=allowed_exts, max_file_size_mb=max_size_mb)
                if not file_info:
                    # Non-fatal for bulk uploads: show error and continue with others
                    print(f'⚠️ Error saving uploaded file {resume_file.filename}: {error}')
                    continue

                # Default to 'resume' file_type for multi-file uploads (client may not indicate resume/letter/license)
                file_type = 'resume'
                try:
                    cursor.execute(
                        '''
                        INSERT INTO resumes (applicant_id, file_name, file_path, file_type)
                        VALUES (%s, %s, %s, %s)
                        ''',
                        (
                            applicant_id,
                            file_info['original_filename'],
                            file_info.get('storage_path') or file_info.get('file_path', ''),
                            file_type,
                        ),
                    )
                    new_id = cursor.lastrowid
                    newly_uploaded_ids.append(str(new_id))
                    # Commit per-file to ensure uploaded_at is set and auto-linking works if needed
                    db.commit()
                except Exception as exc:
                    db.rollback()
                    print(f'⚠️ Failed to insert resume record for {resume_file.filename}: {exc}')

            # Use the first newly uploaded file as the primary resume for application.resume_id
            if newly_uploaded_ids:
                try:
                    resume_id = int(newly_uploaded_ids[0])
                except Exception:
                    resume_id = None

        else:
            # Handle single-file upload under legacy field name 'resume_file'
            resume_file = request.files.get('resume_file')
            if resume_file and getattr(resume_file, 'filename', None):
                # Use job-specific rules for validation when available
                try:
                    allowed_exts = job.get('allowed_extensions') if job else None
                    max_size_mb = job.get('max_file_size_mb') if job else None
                except Exception:
                    allowed_exts = None
                    max_size_mb = None
                file_info, error = save_uploaded_file(resume_file, applicant_id, allowed_extensions=allowed_exts, max_file_size_mb=max_size_mb)
                if file_info:
                    # Get file_type from request (resume, letter, license)
                    file_type = request.form.get('file_type', 'resume').strip().lower()
                    if file_type not in ['resume', 'letter', 'license']:
                        file_type = 'resume'

                    cursor.execute(
                        '''
                        INSERT INTO resumes (applicant_id, file_name, file_path, file_type)
                        VALUES (%s, %s, %s, %s)
                        ''',
                        (
                            applicant_id,
                            file_info['original_filename'],
                            file_info.get('storage_path') or file_info.get('file_path', ''),
                            file_type,
                        ),
                    )
                    resume_id = cursor.lastrowid
                else:
                    flash(error or 'Unable to process the uploaded resume file.', 'error')
                    return redirect(url_for('apply_to_job', job_id=job_id))
        
        # If no file uploaded, use resume_id from form or get latest resume
        if not resume_id:
            resume_id_input = request.form.get('resume_id', '').strip()
            if resume_id_input:
                # Verify resume belongs to applicant and matches job allowed extensions
                cursor.execute(
                    'SELECT resume_id, file_name FROM resumes WHERE resume_id = %s AND applicant_id = %s',
                    (resume_id_input, applicant_id),
                )
                row = cursor.fetchone()
                if not row:
                    flash('Invalid resume selection.', 'error')
                    return redirect(url_for('apply_to_job', job_id=job_id))

                # Check extension against job rules (if any)
                allowed_raw = job.get('allowed_extensions') if job else None
                if allowed_raw:
                    allowed_set = {p.strip().lstrip('.').lower() for p in str(allowed_raw).split(',') if p.strip()}
                else:
                    allowed_set = set()

                file_name = row.get('file_name') or ''
                file_ext = file_name.rsplit('.', 1)[-1].lower() if '.' in file_name else ''
                if allowed_set and file_ext and file_ext not in allowed_set:
                    flash('Selected resume is not an allowed file type for this job.', 'error')
                    return redirect(url_for('apply_to_job', job_id=job_id))

                resume_id = int(resume_id_input)
            else:
                # Get latest resume if no selection
                cursor.execute(
                    '''
                    SELECT resume_id FROM resumes 
                    WHERE applicant_id = %s 
                    ORDER BY uploaded_at DESC 
                    LIMIT 1
                    ''',
                    (applicant_id,),
                )
                resume_record = cursor.fetchone()
                resume_id = resume_record.get('resume_id') if resume_record else None
        
        # If editing an existing application and no new resume was provided, keep the original resume
        if not resume_id and edit_application_id and existing_app:
            try:
                existing_resume_id = existing_app.get('resume_id') if isinstance(existing_app, dict) else (existing_app[3] if len(existing_app) > 3 else None)
                if existing_resume_id:
                    resume_id = existing_resume_id
            except Exception:
                pass

        # Require at least one resume/attachment to apply
        if not resume_id:
            flash('You must upload or select a resume/attachment before applying.', 'error')
            return redirect(url_for('apply_to_job', job_id=job_id))
        
        # If job_id was not provided in the route, or a dropdown selection was made, get it from the form
        form_job_id_raw = request.form.get('job_id')
        if not job_id or (form_job_id_raw and not edit_application_id):
            try:
                job_id = int((form_job_id_raw or '').strip())
            except Exception:
                job_id = None
            if not job_id:
                flash('Please select a job position before submitting.', 'error')
                return redirect(url_for('apply_to_job'))

            # Validate job exists and is available
            try:
                ensure_schema_compatibility()
                _update_job_columns(cursor)
                status_placeholders = ','.join(['%s'] * len(PUBLISHABLE_JOB_STATUSES))
                cursor.execute(
                    f'SELECT job_id, allowed_extensions, max_file_size_mb FROM jobs WHERE job_id = %s AND status IN ({status_placeholders}) LIMIT 1',
                    (job_id, *tuple(PUBLISHABLE_JOB_STATUSES))
                )
                job_row = cursor.fetchone()
                if not job_row:
                    flash('Selected job is not available.', 'error')
                    return redirect(url_for('apply_to_job'))
                # Minimal job dict for extension checks later
                job = job_row
            except Exception:
                job = {'job_id': job_id}

            # Prevent duplicate applications (for new applications only)
            if not edit_application_id:
                cursor.execute('SELECT application_id FROM applications WHERE applicant_id = %s AND job_id = %s LIMIT 1', (applicant_id, job_id))
                if cursor.fetchone():
                    flash('You have already applied for this position.', 'warning')
                    return redirect(url_for('jobs'))

        # When editing, always use the existing job_id - job cannot be changed
        if edit_application_id and existing_app:
            # Force use of existing job_id - prevent changing job when editing
            job_id = existing_app.get('job_id')
        
        # Create or update application
        if edit_application_id:
            # Update existing application (only if not viewed) - job cannot be changed
            # job_id is already set to existing_app.job_id above, so it will remain unchanged
            cursor.execute(
                '''
                UPDATE applications 
                SET job_id = %s, resume_id = %s, applied_at = NOW() 
                WHERE application_id = %s AND applicant_id = %s AND viewed_at IS NULL
                ''',
                (job_id, resume_id, edit_application_id, applicant_id),
            )
            if cursor.rowcount == 0:
                flash('Application cannot be updated because it has been viewed by HR/Admin.', 'error')
                return redirect(url_for('applicant_applications'))
            application_id = edit_application_id
        else:
            # Create new application - status must be 'pending' (database enum: 'pending', 'scheduled', 'interviewed', 'hired', 'rejected')
            cursor.execute(
                '''
                INSERT INTO applications (applicant_id, job_id, resume_id, status, applied_at)
                VALUES (%s, %s, %s, 'pending', NOW())
                ''',
                (applicant_id, job_id, resume_id),
            )
            application_id = cursor.lastrowid
            print(f'✅ Application created - ID: {application_id}, Job ID: {job_id}, Applicant ID: {applicant_id}, Status: pending')

        # If multiple resumes were provided (multiple hidden inputs named 'resume_id', 'letter_id', 'license_id'),
        # store them in application_attachments for full record keeping.
        try:
            # Collect all attachment IDs from form
            resume_ids = request.form.getlist('resume_id') or []
            letter_ids = request.form.getlist('letter_id') or []
            license_ids = request.form.getlist('license_id') or []
            # Debug: print raw form keys and values to help diagnose missing attachments
            try:
                print('--- apply_to_job form keys ---')
                for k in request.form.keys():
                    print(f"form[{k}] = {request.form.getlist(k)}")
            except Exception as _e:
                print(f'⚠️ Could not print form keys: {_e}')
            all_ids = resume_ids + letter_ids + license_ids
            print(f'DEBUG: resume_ids={resume_ids}, letter_ids={letter_ids}, license_ids={license_ids}, combined all_ids={all_ids}')
            all_ids = resume_ids + letter_ids + license_ids
            
            # Ensure we include resume_id from earlier single-file upload path if present
            if resume_id and str(resume_id) not in [str(r) for r in all_ids]:
                all_ids.insert(0, str(resume_id))

            # Include any newly uploaded resume IDs (from multi-file upload) so they get linked
            try:
                if newly_uploaded_ids:
                    for nid in newly_uploaded_ids:
                        if str(nid) not in [str(x) for x in all_ids]:
                            all_ids.append(str(nid))
            except Exception:
                pass

            # Insert attachments linking each file to this application
            # If no explicit attachment ids were provided via form, try to auto-link
            # recently uploaded resumes for this applicant that are not yet linked
            # to any application (helps when the client uploaded via AJAX but
            # hidden inputs didn't get included in the POST).
            if not all_ids:
                try:
                    cursor.execute(
                        '''
                        SELECT r.resume_id
                        FROM resumes r
                        LEFT JOIN application_attachments aa ON aa.resume_id = r.resume_id
                        WHERE r.applicant_id = %s
                          AND aa.attachment_id IS NULL
                          AND r.uploaded_at >= (NOW() - INTERVAL 15 MINUTE)
                        ORDER BY r.uploaded_at DESC
                        LIMIT 5
                        ''',
                        (applicant_id,)
                    )
                    recent_unlinked = [str(r.get('resume_id')) for r in (cursor.fetchall() or [])]
                    if recent_unlinked:
                        print(f'DEBUG: auto-linking recent unlinked resumes: {recent_unlinked}')
                        all_ids = recent_unlinked
                except Exception as _e:
                    print(f'⚠️ Auto-link recent resumes failed: {_e}')

            if all_ids:
                # Precompute allowed set for attachments
                allowed_raw = job.get('allowed_extensions') if job else None
                if allowed_raw:
                    allowed_set = {p.strip().lstrip('.').lower() for p in str(allowed_raw).split(',') if p.strip()}
                else:
                    allowed_set = set()

                for rid in all_ids:
                    try:
                        rid_int = int(rid)
                    except Exception:
                        continue

                    # Verify the resume belongs to this applicant and get file_name
                    cursor.execute('SELECT resume_id, file_name, applicant_id FROM resumes WHERE resume_id = %s LIMIT 1', (rid_int,))
                    rinfo = cursor.fetchone()
                    if not rinfo:
                        print(f'WARN: resume id {rid_int} not found when linking attachments')
                        continue
                    if rinfo.get('applicant_id') != applicant_id:
                        print(f'WARN: resume id {rid_int} does not belong to applicant {applicant_id} - skipping')
                        continue

                    # Check extension matches allowed set (if defined)
                    fname = rinfo.get('file_name') or ''
                    ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''
                    if allowed_set and ext and ext not in allowed_set:
                        print(f'WARN: resume id {rid_int} has disallowed extension for job {job_id} - skipping')
                        continue

                    # Avoid duplicate entries
                    cursor.execute(
                        'SELECT attachment_id FROM application_attachments WHERE application_id = %s AND resume_id = %s LIMIT 1',
                        (application_id, rid_int),
                    )
                    if not cursor.fetchone():
                        cursor.execute(
                            'INSERT INTO application_attachments (application_id, resume_id) VALUES (%s, %s)',
                            (application_id, rid_int),
                        )
                    else:
                        print(f'INFO: attachment already exists for application {application_id}, resume {rid_int}')
                print(f'✅ Linked {len(all_ids)} file(s) to application {application_id} via application_attachments')
                # Debug: list attachments actually in DB for this application
                try:
                    cursor.execute('SELECT resume_id FROM application_attachments WHERE application_id = %s', (application_id,))
                    linked = [r.get('resume_id') for r in cursor.fetchall()]
                    print(f'DEBUG: linked resume_ids in DB for application {application_id}: {linked}')
                except Exception as _e:
                    print(f'⚠️ Could not query application_attachments: {_e}')
        except Exception as attach_err:
            print(f'⚠️ Could not save application attachments for application {application_id}: {attach_err}')        # AUTOMATIC: Notify applicant about successful submission
        if application_id:
            # Get applicant and job info for email and HR notification
            cursor.execute(
                '''
                SELECT ap.email, ap.full_name, ap.applicant_id,
                       j.job_title AS job_title, j.branch_id,
                       b.branch_name
                FROM applicants ap
                JOIN jobs j ON j.job_id = %s
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                WHERE ap.applicant_id = %s
                LIMIT 1
                ''',
                (job_id, applicant_id)
            )
            info = cursor.fetchone()
            
            # Send email to applicant (but don't create notification - HR will see their own notification)
            job_title = info.get('job_title') or info.get('job_title_alt') or 'the position'
            applicant_email = info.get('email')
            applicant_name = info.get('full_name') or 'Applicant'
            
            if applicant_email:
                email_subject = f'Application Submitted - {job_title}'
                email_body = f"""Dear {applicant_name},

Thank you for applying to the position: {job_title}

Your application has been successfully submitted. Our team will review your application and contact you soon.

Best regards,
J&T Express Recruitment Team
            """.strip()
            
                try:
                    send_email(applicant_email, email_subject, email_body)
                    print(f'✅ Confirmation email sent to applicant {applicant_email}')
                except Exception as email_err:
                    print(f'⚠️ Error sending confirmation email: {email_err}')
            
            # AUTOMATIC: Notify HR about new application (system notification + email)
            # HR will see this notification when they view branch-scoped notifications
            try:
                branch_id = info.get('branch_id')
                branch_name = info.get('branch_name') or 'a branch'
                hr_notification_message = f'{applicant_name} applied for {job_title} at {branch_name}.'
                
                # Create HR notification linked to application (HR will see it through branch-scoped queries)
                # HR notifications are fetched by joining notifications with applications and jobs by branch_id
                notification_columns = set()
                try:
                    cursor.execute('SHOW COLUMNS FROM notifications')
                    notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                except Exception:
                    pass
                
                # Create notification linked to application for HR visibility
                # HR will see this through branch-scoped queries (joining through applications -> jobs -> branch_id)
                # Applicants won't see this notification because they only see notifications starting with "You applied"
                # HR notifications use third-person format "{applicant_name} applied" which is filtered out in applicant queries
                if 'application_id' in notification_columns and 'message' in notification_columns:
                    # Check if notification already exists to prevent duplicates
                    cursor.execute(
                        '''
                        SELECT notification_id FROM notifications
                        WHERE application_id = %s AND message = %s
                        LIMIT 1
                        ''',
                        (application_id, hr_notification_message)
                    )
                    existing_notification = cursor.fetchone()
                    
                    if not existing_notification:
                        # Only create notification if it doesn't already exist
                        if 'sent_at' in notification_columns:
                            cursor.execute(
                                '''
                                INSERT INTO notifications (application_id, message, sent_at, is_read)
                                VALUES (%s, %s, NOW(), 0)
                                ''',
                                (application_id, hr_notification_message)
                            )
                        else:
                            cursor.execute(
                                '''
                                INSERT INTO notifications (application_id, message, is_read)
                                VALUES (%s, %s, 0)
                                ''',
                                (application_id, hr_notification_message)
                            )
                        print(f'✅ HR system notification created for application {application_id}')
                    else:
                        print(f'⚠️ HR notification already exists for application {application_id} - skipping duplicate')
                else:
                    print(f'⚠️ Notifications table missing required columns')
                
                # Send email to HR users managing this branch
                try:
                    # HR accounts manage all branches (branch_id column removed from admins table)
                    hr_query = '''
                        SELECT DISTINCT u.email, a.full_name
                        FROM users u
                        JOIN admins a ON a.user_id = u.user_id
                        WHERE u.user_type = 'hr' AND u.is_active = 1 AND a.is_active = 1
                    '''
                    cursor.execute(hr_query)
                    hr_users = cursor.fetchall()
                    
                    if hr_users:
                        # Send email to each HR user
                        email_subject = f'New Application Received - {job_title}'
                        email_body = f"""Dear HR Team,

A new application has been submitted:

Applicant: {applicant_name}
Position: {job_title}
Branch: {branch_name}
Application ID: {application_id}

Please log in to the HR portal to review this application.

Best regards,
J&T Express Recruitment System
                        """.strip()
                        
                        for hr_user in hr_users:
                            hr_email = hr_user.get('email')
                            if hr_email:
                                try:
                                    send_email(hr_email, email_subject, email_body)
                                    print(f'✅ HR notification email sent to {hr_email} for application {application_id}')
                                except Exception as email_err:
                                    print(f'⚠️ Error sending HR notification email to {hr_email}: {email_err}')
                        print(f'✅ HR notification emails sent to {len(hr_users)} HR user(s) for application {application_id}')
                    else:
                        print(f'⚠️ No HR users found for branch {branch_id} - no emails sent')
                except Exception as hr_email_err:
                    print(f'⚠️ Error sending HR notification emails: {hr_email_err}')
                    import traceback
                    traceback.print_exc()
                    
            except Exception as hr_notify_err:
                print(f'⚠️ Error creating HR notification: {hr_notify_err}')
                import traceback
                traceback.print_exc()
            
            # AUTOMATIC: Notify Admin about new application (system notification)
            # NOTE: Admin notification is skipped since HR notification above already covers this
            # Both HR and Admin users see the same notification, so we don't need to create a duplicate
            # The HR notification created above is sufficient for both HR and Admin users
            # This prevents duplicate notifications with the same message
        
        db.commit()
        flash('Application submitted successfully! You have been automatically notified.', 'success')
        return redirect(url_for('applicant_applications'))
    except Exception as exc:
        db.rollback()
        log.exception('❌ Apply to job error: %s', exc)
        try:
            log.debug('--- REQUEST DEBUG START ---')
            log.debug('Request method: %s, remote_addr: %s, applicant_session_user: %s', request.method, request.remote_addr, session.get('user_id'))
        except Exception:
            log.debug('Could not read request metadata for debug')
        try:
            for k in request.form.keys():
                log.debug('FORM[%s] = %s', k, request.form.getlist(k))
        except Exception as _e:
            log.debug('Could not dump form keys: %s', _e)
        try:
            for fkey in request.files.keys():
                f = request.files.get(fkey)
                fname = getattr(f, 'filename', None)
                log.debug('FILE[%s] -> filename=%s', fkey, fname)
        except Exception as _e:
            log.debug('Could not dump files: %s', _e)
        # Also write debug information to a file under instance/ for later inspection
        try:
            log_dir = app.instance_path
            os.makedirs(log_dir, exist_ok=True)
            log_file = os.path.join(log_dir, 'apply_error.log')
            with open(log_file, 'a', encoding='utf-8') as fh:
                fh.write('\n--- APPLY ERROR ' + datetime.now().isoformat() + ' ---\n')
                fh.write(f'Error: {str(exc)}\n')
                fh.write('Traceback:\n')
                fh.write(''.join(traceback.format_exception(None, exc, exc.__traceback__)))
                fh.write('\nRequest form keys and values:\n')
                try:
                    for k in request.form.keys():
                        fh.write(f'FORM[{k}] = {request.form.getlist(k)}\n')
                except Exception:
                    fh.write('Could not dump form keys\n')
                fh.write('Request files:\n')
                try:
                    for fkey in request.files.keys():
                        f = request.files.get(fkey)
                        fname = getattr(f, 'filename', None)
                        fh.write(f'FILE[{fkey}] -> filename={fname}\n')
                except Exception:
                    fh.write('Could not dump files\n')
                fh.write('--- END APPLY ERROR ---\n')
        except Exception as file_err:
            log.exception('⚠️ Could not write apply error log: %s', file_err)
        
        # Check if it's a CSRF error
        error_str = str(exc).lower()
        if 'csrf' in error_str or 'token' in error_str:
            flash('Security token expired. Please refresh the page and try again.', 'error')
        else:
            flash('Unable to submit application. Please try again.', 'error')
        return redirect(url_for('apply_to_job', job_id=job_id))
    finally:
        cursor.close()


@app.route('/applicant/applications/<int:application_id>/delete', methods=['POST'])
@login_required('applicant')
def delete_application(application_id):
    """Permanently delete a job application and all related data."""
    applicant_id = session.get('user_id')
    if not applicant_id:
        flash('Please log in to delete applications.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    # Validate applicant_id is an integer
    try:
        applicant_id = int(applicant_id)
    except (ValueError, TypeError):
        flash('Unable to identify your account. Please log in again.', 'error')
        return immediate_redirect(url_for('login', _external=True))
    
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return immediate_redirect(url_for('applicant_applications', _external=True))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Verify application belongs to applicant
        cursor.execute(
            '''
            SELECT application_id, status, viewed_at, job_id
            FROM applications
            WHERE application_id = %s AND applicant_id = %s
            LIMIT 1
            ''',
            (application_id, applicant_id),
        )
        application = cursor.fetchone()
        
        if not application:
            flash('Application not found or you do not have permission to delete it.', 'error')
            return immediate_redirect(url_for('applicant_applications', _external=True))
        
        # Check if application has been viewed by HR/Admin - warn but allow deletion
        if application.get('viewed_at'):
            # Still allow deletion but warn the user
            pass
        
        # Delete related records first to avoid foreign key constraints
        try:
            # Delete interviews related to this application
            cursor.execute('DELETE FROM interviews WHERE application_id = %s', (application_id,))
            print(f'✅ Deleted interviews for application {application_id}')
        except Exception as interview_error:
            print(f'⚠️ Error deleting interviews: {interview_error}')
            # Continue anyway
        
        try:
            # Delete notifications related to this application
            cursor.execute('DELETE FROM notifications WHERE application_id = %s', (application_id,))
            print(f'✅ Deleted notifications for application {application_id}')
        except Exception as notif_error:
            print(f'⚠️ Error deleting notifications: {notif_error}')
            # Continue anyway
        
        
        # Finally, delete the application itself
        cursor.execute('DELETE FROM applications WHERE application_id = %s AND applicant_id = %s', (application_id, applicant_id))
        
        deleted_count = cursor.rowcount
        if deleted_count > 0:
            db.commit()
            flash('Application permanently deleted successfully.', 'success')
            print(f'✅ Application {application_id} deleted by applicant {applicant_id}')
        else:
            db.rollback()
            flash('Failed to delete application. Please try again.', 'error')
        
        return immediate_redirect(url_for('applicant_applications', _external=True))
    except Exception as exc:
        db.rollback()
        print(f'❌ Delete application error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to delete application. Please try again later.', 'error')
        return immediate_redirect(url_for('applicant_applications', _external=True))
    finally:
        cursor.close()


@app.route('/applicant/jobs/<int:job_id>/save', methods=['POST'])
@login_required('applicant')
def save_job(job_id):   
    """Save a job posting."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('jobs'))


    
    
    cursor = db.cursor(dictionary=True)
    try:
        # Check if saved_jobs table exists, if not create it
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_jobs (
                saved_job_id INT AUTO_INCREMENT PRIMARY KEY,
                applicant_id INT NOT NULL,
                job_id INT NOT NULL,
                saved_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY unique_saved_job (applicant_id, job_id),
                FOREIGN KEY (applicant_id) REFERENCES applicants(applicant_id) ON DELETE CASCADE,
                FOREIGN KEY (job_id) REFERENCES jobs(job_id) ON DELETE CASCADE
            )
        """)
        
        # Check if already saved - use applicant_id and job_id instead of saved_job_id
        cursor.execute(
            'SELECT applicant_id FROM saved_jobs WHERE applicant_id = %s AND job_id = %s',
            (applicant_id, job_id),
        )
        if cursor.fetchone():
            flash('Job already saved.', 'info')
        else:
            cursor.execute(
                'INSERT INTO saved_jobs (applicant_id, job_id) VALUES (%s, %s)',
                (applicant_id, job_id),
            )
            db.commit()
            flash('Job saved successfully.', 'success')
        
        return redirect(url_for('saved_jobs', just_saved=job_id))
    except Exception as exc:
        db.rollback()
        print(f'❌ Save job error: {exc}')
        flash('Unable to save job. Please try again.', 'error')
        return redirect(url_for('jobs'))
    finally:
        cursor.close()


@app.route('/applicant/jobs/<int:job_id>/unsave', methods=['POST'])
@login_required('applicant')
def unsave_job(job_id):
    """Unsave a job posting."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('jobs'))
    
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(
            'DELETE FROM saved_jobs WHERE applicant_id = %s AND job_id = %s',
            (applicant_id, job_id),
        )
        db.commit()
        flash('Job removed from saved list.', 'success')
        return redirect(request.referrer or url_for('jobs'))
    except Exception as exc:
        db.rollback()
        print(f'❌ Unsave job error: {exc}')
        flash('Unable to remove saved job. Please try again.', 'error')
        return redirect(url_for('jobs'))
    finally:
        cursor.close()


@app.route('/hr/archived-applicants')
@login_required('hr', 'admin')
def archived_applicants():
    """HR view: list archived applicants. Reuses fetch_archived_applicants_data() to ensure consistent behavior with admin."""
    try:
        archived, branches = fetch_archived_applicants_data()
        return render_template('hr/archived_applicants.html', archived=archived, branches=branches)
    except Exception as exc:
        print(f'❌ HR archived applicants error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to load archived applicants. Please try again later.', 'error')
        return redirect(url_for('hr_dashboard'))


@app.route('/hr/archived-applicants/restore', methods=['POST'])
@login_required('hr', 'admin')
def restore_applicant():
    """Restore an archived application back to pending status."""
    application_id = request.form.get('application_id')
    if not application_id:
        flash('Invalid request.', 'error')
        return redirect(url_for('archived_applicants'))

    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('archived_applicants'))

    # Get current user for role-scoped redirects and permission checks
    user = get_current_user()

    cursor = db.cursor()
    try:
        # Update status back to pending and clear archived_at if present
        try:
            cursor.execute("UPDATE applications SET status = 'pending', archived_at = NULL WHERE application_id = %s", (application_id,))
        except Exception:
            # Fallback if archived_at does not exist
            cursor.execute("UPDATE applications SET status = 'pending' WHERE application_id = %s", (application_id,))
        db.commit()
        flash('Application restored successfully.', 'success')
    except Exception as exc:
        db.rollback()
        print(f'❌ Restore applicant error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to restore application. Please try again later.', 'error')
    finally:
        cursor.close()

    # Redirect to appropriate archived page depending on caller role/referrer
    ref = (request.referrer or '')
    try:
        # Prefer the already-fetched `user` object from earlier in this function
        role = (user.get('role') if isinstance(user, dict) else None) or session.get('user_role') or (get_current_user() or {}).get('role')
    except Exception:
        role = None

    # If current user is HR, always send them to the HR archived list.
    if role and str(role).lower() == 'hr':
        return redirect(url_for('archived_applicants'))

    # If the referrer or current role indicates admin, go to admin archived page
    if ref and '/admin/' in ref:
        return redirect(url_for('admin_archived_applicants'))
    if role and str(role).lower() == 'admin':
        return redirect(url_for('admin_archived_applicants'))

    # Fallback to HR archived list
    return redirect(url_for('archived_applicants'))


@app.route('/hr/archived-applicants/delete', methods=['POST'])
@login_required('hr', 'admin')
def delete_applicant_permanent():
    """Permanently delete an archived application and related records."""
    application_id = request.form.get('application_id')
    if not application_id:
        flash('Invalid request.', 'error')
        return redirect(url_for('archived_applicants'))

    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('archived_applicants'))

    cursor = db.cursor()
    try:
        # Instead of hard-deleting the application row, mark it as removed so the applicant's record remains
        try:
            # Nullify archived_at (if present) and set status to 'archived'
            try:
                cursor.execute("UPDATE applications SET status = %s, archived_at = NULL, updated_at = NOW() WHERE application_id = %s", ('archived', application_id))
            except Exception:
                # Fallback if archived_at or updated_at doesn't exist
                cursor.execute("UPDATE applications SET status = %s WHERE application_id = %s", ('archived', application_id))

            if cursor.rowcount > 0:
                db.commit()
                flash('Application archived (soft-archived).', 'success')
            else:
                db.rollback()
                flash('Failed to update application status.', 'error')
        except Exception:
            db.rollback()
            raise
    except Exception as exc:
        db.rollback()
        print(f'❌ Permanent delete applicant error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to delete application. Please try again later.', 'error')
    finally:
        cursor.close()

    # Redirect appropriately depending on admin vs hr
    ref = (request.referrer or '')
    try:
        role = session.get('user_role') or (get_current_user() or {}).get('role')
    except Exception:
        role = None

    if ref and '/admin/' in ref:
        return redirect(url_for('admin_archived_applicants'))
    if role == 'admin':
        return redirect(url_for('admin_archived_applicants'))
    return redirect(url_for('archived_applicants'))


@app.route('/hr/applicants/<int:application_id>/archive', methods=['POST'])
@app.route('/admin/applicants/<int:application_id>/archive', methods=['POST'])
@login_required('hr', 'admin')
def archive_applicant(application_id):
    """Archive an application by setting status to 'archived' and archived_at timestamp."""
    user = get_current_user()
    db = get_db()
    wants_json = (
        request.is_json
        or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or request.accept_mimetypes.accept_json
        or ('application/json' in (request.headers.get('Accept') or '').lower())
    )
    # Capture optional redirect preference from the client so both AJAX and non-AJAX flows
    posted_redirect = None
    try:
        posted_redirect = (request.values.get('redirect') or None)
    except Exception:
        posted_redirect = None
    if not db:
        if wants_json:
            return jsonify({'success': False, 'error': 'Database connection error.'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    cursor = db.cursor(dictionary=True)
    # Log context for debugging
    try:
        print(f"➡️ Archive request: user={user and user.get('id')}, role={user and user.get('role')}, application_id={application_id}")
        try:
            print(f"   Session keys: {list(session.keys())}")
            print(f"   session.user_role={session.get('user_role')}, session.user_id={session.get('user_id')}")
            print(f"   Request cookies: {dict(request.cookies)}")
            print(f"   X-Requested-With: {request.headers.get('X-Requested-With')}, is_json: {request.is_json}, Accept: {request.headers.get('Accept')}")
            print(f"   posted_redirect initial: {posted_redirect}")
        except Exception:
            pass
    except Exception:
        pass
    try:
        # If HR, ensure the application belongs to their branch
        branch_id = get_branch_scope(user)
        if branch_id:
            try:
                # Also fetch the application's job branch for better diagnostics
                cursor.execute(
                    '''
                    SELECT j.branch_id AS job_branch_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.application_id = %s
                    LIMIT 1
                    ''',
                    (application_id,)
                )
                app_branch_row = cursor.fetchone() or {}
                app_branch_id = app_branch_row.get('job_branch_id') if isinstance(app_branch_row, dict) else None

                cursor.execute(
                    '''
                    SELECT a.application_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.application_id = %s AND j.branch_id = %s
                    LIMIT 1
                    ''',
                    (application_id, branch_id),
                )
                if not cursor.fetchone():
                    # When called via AJAX, return a 403 JSON instead of redirecting
                    try:
                        print(f"⚠️ Branch scope check failed: user.branch_id={branch_id} app.job_branch_id={app_branch_id} application_id={application_id}")
                    except Exception:
                        pass
                    if wants_json:
                        return jsonify({'success': False, 'error': 'Not authorized', 'role': session.get('user_role'), 'user_branch_id': session.get('branch_id'), 'app_branch_id': app_branch_id}), 403
                    flash('You do not have permission to archive this application.', 'error')
                    return redirect(url_for('applicants'))
            except Exception as e:
                print(f'⚠️ Error checking branch scope for archive: {e}')
        
        # Check if `archived_at` column exists and choose update accordingly
        try:
            cursor.execute("SHOW COLUMNS FROM applications LIKE 'archived_at'")
            has_archived_at = bool(cursor.fetchone())
        except Exception:
            has_archived_at = False

        if has_archived_at:
            sql = "UPDATE applications SET status = 'archived', archived_at = NOW() WHERE application_id = %s"
            params = (application_id,)
        else:
            sql = "UPDATE applications SET status = 'archived' WHERE application_id = %s"
            params = (application_id,)

        try:
            cursor.execute(sql, params)
            if cursor.rowcount > 0:
                db.commit()
                # Double-check the status was updated
                try:
                    cursor.execute('SELECT status, archived_at FROM applications WHERE application_id = %s LIMIT 1', (application_id,))
                    post = cursor.fetchone() or {}
                    print(f"✅ Archive DB check: application_id={application_id}, status={post.get('status')}, archived_at={post.get('archived_at')}")
                except Exception:
                    pass
                flash('Application archived successfully.', 'success')
                # Determine redirect target depending on caller (prefer explicit role over referrer)
                ref_inner = (request.referrer or '')
                try:
                    role_inner = session.get('user_role') or (get_current_user() or {}).get('role')
                except Exception:
                    role_inner = None
                role_inner = (role_inner or '').lower()
                # Choose redirect purely by role to avoid unauthorized admin page access
                if role_inner == 'admin':
                    redirect_target = url_for('admin_archived_applicants', _external=False)
                else:
                    redirect_target = url_for('archived_applicants', _external=False)
                try:
                    print(f"➡️ Archive redirect decision: role_inner={role_inner}, ref_inner={ref_inner}, posted_redirect={posted_redirect}, redirect_target={redirect_target}")
                except Exception:
                    pass
                # Allow client to override redirect by posting a 'redirect' value, but keep role safety:
                # - Admin can only override to admin archived page
                # - HR can only override to HR archived page
                if posted_redirect and (posted_redirect.startswith('/') or posted_redirect.startswith(url_for('archived_applicants')) or posted_redirect.startswith(url_for('admin_archived_applicants'))):
                    if role_inner == 'admin' and posted_redirect.startswith(url_for('admin_archived_applicants')):
                        redirect_target = posted_redirect
                    elif role_inner != 'admin' and posted_redirect.startswith(url_for('archived_applicants')):
                        redirect_target = posted_redirect
                if (
                    request.is_json
                    or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                    or request.accept_mimetypes.accept_json
                    or ('application/json' in (request.headers.get('Accept') or '').lower())
                ):
                    return jsonify({'success': True, 'message': 'Application archived successfully.', 'redirect': redirect_target})
            else:
                db.rollback()
                flash('Application not found.', 'error')
                if (
                    request.is_json
                    or request.headers.get('X-Requested-With') == 'XMLHttpRequest'
                    or request.accept_mimetypes.accept_json
                    or ('application/json' in (request.headers.get('Accept') or '').lower())
                ):
                    return jsonify({'success': False, 'error': 'Application not found.'}), 404
        except Exception as e:
            db.rollback()
            print(f'⚠️ Error executing archive query: {e} SQL={sql} PARAMS={params}')
            raise
    except Exception as exc:
        db.rollback()
        import traceback
        tb = traceback.format_exc()
        print(f'❌ Archive applicant error: {exc}')
        print(tb)
        # Provide a clearer message for admins while keeping user-friendly text
        flash('Unable to archive application. Please try again later.', 'error')
    finally:
        cursor.close()

    # After archiving, redirect appropriately:
    # - If the request came from the admin UI (referrer contains '/admin/'), go to admin archived page
    # - Otherwise, if current user role is admin, go to admin archived page
    # - Else, go to HR archived page
    ref = (request.referrer or '')
    try:
        role = session.get('user_role') or (get_current_user() or {}).get('role')
    except Exception:
        role = None

    # Honor posted redirect for non-AJAX flows as well (sanitized)
    if posted_redirect and (posted_redirect.startswith('/') or posted_redirect.startswith(url_for('archived_applicants')) or posted_redirect.startswith(url_for('admin_archived_applicants'))):
        if (role or '').lower() == 'admin' and posted_redirect.startswith(url_for('admin_archived_applicants')):
            return redirect(posted_redirect)
        if (role or '').lower() != 'admin' and posted_redirect.startswith(url_for('archived_applicants')):
            return redirect(posted_redirect)
    # Choose redirect purely by role to avoid unauthorized admin page access
    if (role or '').lower() == 'admin':
        return redirect(url_for('admin_archived_applicants'))
    return redirect(url_for('archived_applicants'))


@app.route('/hr/applicants/<int:application_id>/delete', methods=['POST'])
@login_required('hr', 'admin')
def delete_applicant_now(application_id):
    """Permanently delete an application from the Applicants view (no redirect to Archived page)."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))

    cursor = db.cursor()
    try:
        # Soft-archive application so applicant's record remains intact.
        try:
            # Prefer setting archived_at to NULL when supported
            cursor.execute(
                "UPDATE applications SET status = %s, archived_at = NULL, updated_at = NOW() WHERE application_id = %s",
                ('archived', application_id)
            )
        except Exception:
            # Fallback for older schemas without archived_at
            cursor.execute(
                "UPDATE applications SET status = %s WHERE application_id = %s",
                ('archived', application_id)
            )

        if cursor.rowcount > 0:
            db.commit()
            flash('Application archived (soft-archived).', 'success')
        else:
            db.rollback()
            flash('Failed to update application status.', 'error')

    except Exception as exc:
        db.rollback()
        print(f'❌ Delete applicant error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to delete application. Please try again later.', 'error')
    finally:
        try:
            cursor.close()
        except Exception:
            pass

    # If admin performed delete from admin UI, keep them in admin applicants view
    ref = (request.referrer or '')
    try:
        role = session.get('user_role') or (get_current_user() or {}).get('role')
    except Exception:
        role = None

    if ref and '/admin/' in ref:
        return redirect(url_for('applicants'))
    if role == 'admin':
        return redirect(url_for('applicants'))
    return redirect(url_for('applicants'))


@app.route('/applicant/jobs/saved')
@login_required('applicant')
def saved_jobs():
    """View saved job postings."""
    filters = {
        'keyword': request.args.get('keyword', '').strip(),
        'branch_id': request.args.get('branch_id', type=int),
        'position_id': request.args.get('position_id', type=int),
    }
    filters = {k: v for k, v in filters.items() if v}

    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/jobs.html', jobs=[], branches=[], positions=[], current_filters=filters, saved_mode=True)
    
    cursor = db.cursor(dictionary=True)
    try:
        # Check if saved_jobs table exists
        cursor.execute("SHOW TABLES LIKE 'saved_jobs'")
        if not cursor.fetchone():
            flash('No saved jobs found.', 'info')
            return render_template('applicant/jobs.html', jobs=[], branches=[], positions=[], current_filters=filters, saved_mode=True)
        
        # Check the schema of saved_jobs table
        cursor.execute("DESCRIBE saved_jobs")
        saved_columns = [col['Field'] for col in cursor.fetchall()]
        print(f"Saved jobs columns: {saved_columns}")  # Debug info
        
        # Ensure schema compatibility for jobs table
        ensure_schema_compatibility()
        
        # Fetch saved jobs - use dynamic column checking
        _update_job_columns(cursor)
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        salary_currency_expr = job_column_expr('salary_currency', default="'PHP'")
        salary_min_expr = job_column_expr('salary_min', default='NULL')
        salary_max_expr = job_column_expr('salary_max', default='NULL')
        
        job_location_expr = job_column_expr('job_location', alternatives=['location'], default='NULL')
        
        job_description_expr = job_column_expr('job_description', alternatives=['description'], default='NULL')
        employment_type_expr = job_column_expr('employment_type', default='NULL')
        work_arrangement_expr = job_column_expr('work_arrangement', default='NULL')
        experience_level_expr = job_column_expr('experience_level', default='NULL')
        application_deadline_expr = job_column_expr('application_deadline', default='NULL')
        status_expr = job_column_expr('status', default="'open'")
        branch_id_expr = job_column_expr('branch_id', default='NULL')
        position_id_expr = job_column_expr('position_id', default='NULL')
        
        # Query for saved jobs - removed all references to saved_job_id
        cursor.execute(
            f'''
            SELECT j.job_id, {job_title_expr} AS job_title, {job_description_expr} AS job_description, 
                   {employment_type_expr} AS employment_type, 
                   {work_arrangement_expr} AS work_arrangement,
                   {experience_level_expr} AS experience_level, 
                   {job_location_expr} AS job_location, 
                   {salary_min_expr} AS salary_min, 
                   {salary_max_expr} AS salary_max, 
                   {salary_currency_expr} AS salary_currency,
                   {application_deadline_expr} AS application_deadline, 
                   {status_expr} AS status, 
                   {branch_id_expr} AS branch_id,
                   {position_id_expr} AS position_id,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   {job_title_expr} AS position_name,
                   'General' AS department,
                   sj.saved_at
            FROM saved_jobs sj
            JOIN jobs j ON sj.job_id = j.job_id
            LEFT JOIN branches b ON {branch_id_expr} = b.branch_id
            WHERE sj.applicant_id = %s
            ORDER BY sj.saved_at DESC
            ''',
            (applicant_id,),
        )
        saved_jobs_list = cursor.fetchall()
        
        # Format jobs similar to fetch_open_jobs
        jobs = []
        for job in saved_jobs_list:
            jobs.append({
                'job_id': job.get('job_id'),
                'title': job.get('job_title'),
                'job_title': job.get('job_title'),
                'summary': (job.get('job_description') or '')[:200] if job.get('job_description') else '',
                'description': job.get('job_description'),
                'employment_type': job.get('employment_type'),
                'work_arrangement': job.get('work_arrangement'),
                'experience_level': job.get('experience_level'),
                'location': job.get('job_location'),
                'job_location': job.get('job_location'),
                'salary_min': job.get('salary_min'),
                'salary_max': job.get('salary_max'),
                'salary_currency': job.get('salary_currency'),
                'salary_display': format_salary_range(job.get('salary_currency'), job.get('salary_min'), job.get('salary_max')),
                'application_deadline': job.get('application_deadline'),
                'status': job.get('status'),
                'branch_name': job.get('branch_name'),
                'branch_id': job.get('branch_id'),
                'position_name': job.get('position_name'),
                'position_id': job.get('position_id'),
                'department': job.get('department'),
                'saved_at': job.get('saved_at'),
                'is_saved': True,
            })

        def matches_filters(job_data):
            if filters.get('keyword'):
                keyword = filters['keyword'].lower()
                haystacks = [
                    (job_data.get('title') or '').lower(),
                    (job_data.get('job_title') or '').lower(),
                    (job_data.get('summary') or '').lower(),
                    (job_data.get('description') or '').lower(),
                    (job_data.get('branch_name') or '').lower(),
                    (job_data.get('position_name') or '').lower(),
                ]
                if not any(keyword in text for text in haystacks):
                    return False
            if filters.get('branch_id') and str(filters['branch_id']) != str(job_data.get('branch_id') or ''):
                return False
            if filters.get('position_id') and str(filters['position_id']) != str(job_data.get('position_id') or ''):
                return False
            return job_data.get('status') in PUBLISHABLE_JOB_STATUSES

        jobs = [job for job in jobs if matches_filters(job)]
        
        branches = fetch_branches()
        positions = fetch_positions()
        
        return render_template(
            'applicant/jobs.html',
            jobs=jobs,
            branches=branches,
            positions=positions,
            current_filters=filters,
            saved_mode=True,
        )
    except Exception as exc:
        db.rollback()
        print(f'❌ Saved jobs error: {exc}')
        flash('Unable to load saved jobs.', 'error')
        return render_template('applicant/jobs.html', jobs=[], branches=[], positions=[], current_filters=filters, saved_mode=True)
    finally:
        cursor.close()


@app.route('/applicant/profile', methods=['GET', 'POST'])
@login_required('applicant')
def applicant_profile():
    """Applicant profile management with resume upload."""
    applicant_id = session.get('user_id')
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('applicant/profile.html', applicant=None, resumes=[], login_history=[], profile_history=[])

    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility
        ensure_schema_compatibility()
        
        cursor.execute(
            '''
            SELECT applicant_id, full_name, email, phone_number, password_hash,
                   last_login, created_at
            FROM applicants
            WHERE applicant_id = %s
            LIMIT 1
            ''',
            (applicant_id,),
        )
        applicant_record = cursor.fetchone()

        if not applicant_record:
            flash('Unable to load your profile. Please contact support.', 'error')
            return redirect(url_for('logout'))

        stored_password_hash = applicant_record.get('password_hash')

        if request.method == 'POST':
            action = request.form.get('action', 'update_profile')

            if action == 'update_profile':
                full_name = request.form.get('full_name', '').strip() or applicant_record.get('full_name')
                phone_number = request.form.get('phone_number', '').strip() or applicant_record.get('phone_number')
                email = request.form.get('email', '').strip().lower() or applicant_record.get('email')
                resume = request.files.get('resume')

                if not all([full_name, phone_number, email]):
                    flash('Full name, phone number, and email are required.', 'error')
                    return redirect(url_for('applicant_profile'))

                email_changed = email != applicant_record.get('email')
                verification_token = None
                token_expires = None

                if email_changed:
                    cursor.execute(
                        'SELECT applicant_id FROM applicants WHERE email = %s AND applicant_id <> %s LIMIT 1',
                        (email, applicant_id),
                    )
                    if cursor.fetchone():
                        flash('That email address is already in use. Please choose another.', 'error')
                        return redirect(url_for('applicant_profile'))
                    from datetime import datetime, timedelta
                    verification_token = generate_token()
                    token_expires = datetime.now() + timedelta(seconds=60)

                # Log changes before updating
                if full_name != applicant_record.get('full_name'):
                    log_profile_change(applicant_id, 'applicant', 'full_name', applicant_record.get('full_name'), full_name)
                if phone_number != applicant_record.get('phone_number'):
                    log_profile_change(applicant_id, 'applicant', 'phone_number', applicant_record.get('phone_number'), phone_number)
                if email_changed:
                    log_profile_change(applicant_id, 'applicant', 'email', applicant_record.get('email'), email)

                # Always update all fields to ensure changes are saved
                # Check if last_profile_update column exists before using it
                cursor.execute('SHOW COLUMNS FROM applicants LIKE "last_profile_update"')
                has_last_profile_update = cursor.fetchone() is not None
                
                if email_changed:
                    if has_last_profile_update:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s,
                                email = %s,
                                verification_token = %s,
                                verification_token_expires = %s,
                                last_profile_update = NOW()
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, email, verification_token, token_expires, applicant_id),
                        )
                    else:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s,
                                email = %s,
                                verification_token = %s,
                                verification_token_expires = %s
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, email, verification_token, token_expires, applicant_id),
                        )
                else:
                    if has_last_profile_update:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s,
                                last_profile_update = NOW()
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, applicant_id),
                        )
                    else:
                        cursor.execute(
                            '''
                            UPDATE applicants
                            SET full_name = %s,
                                phone_number = %s
                            WHERE applicant_id = %s
                            ''',
                            (full_name, phone_number, applicant_id),
                        )
                
                # Verify the update was successful
                rows_affected = cursor.rowcount
                if rows_affected == 0:
                    flash('Failed to update profile. Please try again.', 'error')
                    db.rollback()
                    return redirect(url_for('applicant_profile'))

                resume_uploaded = False
                if resume and resume.filename:
                    file_info, error = save_uploaded_file(resume, applicant_id)
                    if file_info:
                        # Delete old resumes to ensure only one resume exists
                        cursor.execute(
                            'SELECT resume_id, file_path FROM resumes WHERE applicant_id = %s',
                            (applicant_id,)
                        )
                        old_resumes = cursor.fetchall()
                        for old_resume in old_resumes:
                            old_file_path = old_resume.get('file_path') if isinstance(old_resume, dict) else old_resume[1]
                            try:
                                import os
                                if old_file_path and os.path.exists(old_file_path):
                                    os.remove(old_file_path)
                            except Exception as del_err:
                                print(f'⚠️ Error deleting old resume file: {del_err}')
                        # Delete old resume records from database
                        cursor.execute(
                            'DELETE FROM resumes WHERE applicant_id = %s',
                            (applicant_id,)
                        )
                        # Insert new resume
                        cursor.execute(
                            '''
                            INSERT INTO resumes (applicant_id, file_name, file_path)
                            VALUES (%s, %s, %s)
                            ''',
                            (
                                applicant_id,
                                file_info['original_filename'],
                                file_info.get('storage_path') or file_info.get('file_path', ''),
                            ),
                        )
                        log_profile_change(
                            applicant_id,
                            'applicant',
                            'resume',
                            'previous_upload',
                            file_info.get('storage_path') or file_info.get('file_path', ''),
                        )
                        resume_uploaded = True
                    else:
                        flash(error or 'Unable to process the uploaded file.', 'error')

                # Commit the transaction to ensure changes are saved
                try:
                    db.commit()
                    print(f'✅ Profile updated successfully for applicant {applicant_id}: name={full_name}, phone={phone_number}, email={email}')
                except Exception as commit_error:
                    db.rollback()
                    print(f'❌ Error committing profile update: {commit_error}')
                    flash('Failed to save profile changes. Please try again.', 'error')
                    return redirect(url_for('applicant_profile'))

                # Update session with new values immediately
                session['user_name'] = full_name
                session['user_email'] = email

                # AUTOMATIC: Notify applicant via system notification and email about profile update
                try:
                    # Get any application_id for this applicant (for notification linking)
                    cursor.execute(
                        'SELECT application_id FROM applications WHERE applicant_id = %s LIMIT 1',
                        (applicant_id,)
                    )
                    app_record = cursor.fetchone()
                    application_id = app_record.get('application_id') if app_record else None
                    
                    # Create notification message and email
                    notification_message = 'Your profile information has been updated successfully.'
                    email_subject = 'Profile Updated - J&T Express'
                    email_body = f"""Dear {full_name},

Your profile information has been successfully updated.

Updated Information:
- Name: {full_name}
- Email: {email}
- Phone: {phone_number}
{'• Resume: Uploaded' if resume_uploaded else ''}

If you did not make this change, please contact our support team immediately.

Best regards,
J&T Express Recruitment Team
                    """.strip()
                    
                    # Use auto_notify_and_email function if we have an application_id
                    if application_id:
                        auto_notify_and_email(
                            cursor, application_id, notification_message,
                            email_subject, email_body,
                            email, full_name
                        )
                    else:
                        # Just send email if no application_id
                        send_email(email, email_subject, email_body)
                    
                    print(f'✅ Profile update notification sent to applicant {email}')
                except Exception as notify_err:
                    print(f'⚠️ Error sending profile update notification: {notify_err}')

                if resume_uploaded:
                    flash('Resume uploaded successfully.', 'success')

                if email_changed and verification_token:
                    # Get applicant name for personalized email
                    cursor.execute(
                        '''
                        SELECT full_name FROM applicants WHERE applicant_id = %s LIMIT 1
                        ''',
                        (applicant_id,),
                    )
                    applicant_name_record = cursor.fetchone()
                    name = applicant_name_record.get('full_name') if applicant_name_record else None
                    send_verification_email(email, verification_token, applicant_name=name)
                    session['pending_verification_email'] = email
                    flash('Profile updated successfully. Please verify your new email address from your inbox.', 'success')
                else:
                    flash('Profile updated successfully. You have been notified via email.', 'success')

                # Reload the page to show updated information
                return redirect(url_for('applicant_profile'))

            if action == 'change_password':
                current_password = request.form.get('current_password', '').strip()
                new_password = request.form.get('new_password', '').strip()
                confirm_password = request.form.get('confirm_password', '').strip()

                if not all([current_password, new_password, confirm_password]):
                    flash('Please complete all password fields.', 'error')
                    return redirect(url_for('applicant_profile'))

                if not check_password(stored_password_hash, current_password):
                    flash('Current password is incorrect.', 'error')
                    return redirect(url_for('applicant_profile'))

                if len(new_password) < 6:
                    flash('New password must be at least 6 characters.', 'error')
                    return redirect(url_for('applicant_profile'))

                if new_password != confirm_password:
                    flash('New passwords do not match.', 'error')
                    return redirect(url_for('applicant_profile'))

                # Immediately apply password change (no OTP verification)
                try:
                    new_hash = hash_password(new_password)
                    cursor.execute('UPDATE applicants SET password_hash = %s WHERE applicant_id = %s', (new_hash, applicant_id))
                    # also update users table if linked
                    if applicant_record.get('user_id'):
                        cursor.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', (new_hash, applicant_record.get('user_id')))
                    db.commit()
                    # Notify user via email about password change
                    try:
                        applicant_email = applicant_record.get('email')
                        applicant_name = applicant_record.get('full_name') or ''
                        from datetime import datetime
                        subject = 'Your account password was changed'
                        body = f"Hi {applicant_name},\n\nYour account password was successfully changed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\nIf you did not perform this action, please contact support or reset your password immediately.\n"
                        if applicant_email:
                            send_email(applicant_email, subject, body)
                    except Exception as _:
                        pass
                    flash('Your password has been changed successfully.', 'success')
                    return redirect(url_for('applicant_profile'))
                except Exception as exc:
                    db.rollback()
                    print(f'❌ Error changing applicant password: {exc}')
                    flash('Unable to change password now.', 'error')
                    return redirect(url_for('applicant_profile'))

            if action == 'delete_account':
                try:
                    # Get user_id and applicant info from applicants table before deleting (needed to delete from users table and create notification)
                    cursor.execute(
                        'SELECT user_id, email, full_name FROM applicants WHERE applicant_id = %s LIMIT 1',
                        (applicant_id,),
                    )
                    applicant_user_record = cursor.fetchone()
                    user_id = applicant_user_record.get('user_id') if applicant_user_record else None
                    applicant_email = applicant_user_record.get('email') if applicant_user_record else None
                    applicant_name = applicant_user_record.get('full_name') if applicant_user_record else 'Unknown Applicant'
                    
                    if not user_id:
                        print(f'⚠️ WARNING: No user_id found for applicant {applicant_id} (email: {applicant_email})')
                        print(f'⚠️ This applicant may not be linked to a users table record')
                    else:
                        print(f'🔍 Found user_id={user_id} for applicant {applicant_id} (email: {applicant_email})')
                    
                    # Create admin notification before deleting applicant
                    try:
                        admin_notification_msg = f'Applicant "{applicant_name}" (Email: {applicant_email}) has deleted their account.'
                        create_admin_notification(cursor, admin_notification_msg)
                        print(f'✅ Admin notification created for applicant account deletion: {applicant_name}')
                    except Exception as notify_err:
                        print(f'⚠️ Error creating admin notification for applicant deletion: {notify_err}')
                    
                    # Delete related data first (in order to avoid foreign key constraints)
                    # 1. Delete notifications linked to applicant's applications
                    cursor.execute("SHOW TABLES LIKE 'notifications'")
                    if cursor.fetchone():
                        cursor.execute(
                            '''
                            DELETE FROM notifications
                            WHERE application_id IN (
                                SELECT application_id FROM applications WHERE applicant_id = %s
                            )
                            ''',
                            (applicant_id,),
                        )
                    
                    # 2. Delete password reset tokens
                    applicant_email = applicant_record.get('email') if applicant_record else None
                    if applicant_email:
                        cursor.execute(
                            'DELETE FROM password_resets WHERE user_email = %s AND role = %s',
                            (applicant_email, 'applicant'),
                        )
                    
                    # 3. Delete auth sessions (use user_id if available, otherwise use applicant_id)
                    cursor.execute("SHOW TABLES LIKE 'auth_sessions'")
                    if cursor.fetchone():
                        if user_id:
                            cursor.execute(
                                'DELETE FROM auth_sessions WHERE user_id = %s',
                                (user_id,),
                            )
                        else:
                            cursor.execute(
                                'DELETE FROM auth_sessions WHERE user_id = %s AND role = %s',
                                (applicant_id, 'applicant'),
                            )
                    
                    # 4. Delete profile changes history
                    cursor.execute("SHOW TABLES LIKE 'profile_changes'")
                    if cursor.fetchone():
                        if user_id:
                            cursor.execute(
                                'DELETE FROM profile_changes WHERE user_id = %s',
                                (user_id,),
                            )
                        else:
                            cursor.execute(
                                'DELETE FROM profile_changes WHERE applicant_id = %s',
                                (applicant_id,),
                            )
                    
                    # 5. Delete saved jobs
                    cursor.execute("SHOW TABLES LIKE 'saved_jobs'")
                    if cursor.fetchone():
                        cursor.execute(
                            'DELETE FROM saved_jobs WHERE applicant_id = %s',
                            (applicant_id,),
                        )
                    
                    # 6. Delete interviews related to applicant's applications
                    cursor.execute("SHOW TABLES LIKE 'interviews'")
                    if cursor.fetchone():
                        cursor.execute(
                            'DELETE FROM interviews WHERE application_id IN (SELECT application_id FROM applications WHERE applicant_id = %s)',
                            (applicant_id,),
                        )
                        print(f'✅ Deleted interviews for applicant {applicant_id}')
                    
                    # 7. Delete applications explicitly
                    cursor.execute(
                        'DELETE FROM applications WHERE applicant_id = %s',
                        (applicant_id,),
                    )
                    applications_deleted = cursor.rowcount
                    print(f'✅ Deleted {applications_deleted} application(s) for applicant {applicant_id}')
                    
                    # 8. Delete resumes explicitly
                    cursor.execute("SHOW TABLES LIKE 'resumes'")
                    if cursor.fetchone():
                        cursor.execute(
                            'DELETE FROM resumes WHERE applicant_id = %s',
                            (applicant_id,),
                        )
                        resumes_deleted = cursor.rowcount
                        print(f'✅ Deleted {resumes_deleted} resume(s) for applicant {applicant_id}')
                    
                    # 9. Delete the applicant record from database
                    cursor.execute(
                        'DELETE FROM applicants WHERE applicant_id = %s',
                        (applicant_id,),
                    )
                    applicant_deleted = cursor.rowcount > 0
                    print(f'✅ Deleted applicant record {applicant_id} from applicants table' if applicant_deleted else f'⚠️ No applicant record found with ID {applicant_id}')
                    
                    # 10. Finally, delete the user record from users table to remove from system users
                    # IMPORTANT: Delete from users table BEFORE committing to ensure it's removed from system users
                    if user_id:
                        try:
                            # First verify the user exists
                            cursor.execute(
                                'SELECT user_id, email, user_type FROM users WHERE user_id = %s AND user_type = %s',
                                (user_id, 'applicant'),
                            )
                            user_check = cursor.fetchone()
                            if user_check:
                                print(f'🔍 Found user record to delete: user_id={user_id}, email={user_check.get("email")}')
                                # Delete from users table
                                cursor.execute(
                                    'DELETE FROM users WHERE user_id = %s AND user_type = %s',
                                    (user_id, 'applicant'),
                                )
                                user_deleted = cursor.rowcount > 0
                                if user_deleted:
                                    print(f'✅ Successfully deleted user record {user_id} from users table')
                                else:
                                    print(f'⚠️ DELETE query executed but no rows affected for user_id {user_id}')
                            else:
                                print(f'⚠️ User record not found with user_id={user_id} and user_type=applicant')
                        except Exception as user_delete_error:
                            log.exception(f'❌ Error deleting user record {user_id}: {user_delete_error}')
                            # Rollback and re-raise to prevent partial deletion
                            db.rollback()
                            raise
                    else:
                        print(f'⚠️ No user_id found for applicant {applicant_id} - cannot delete from users table')
                    
                    # Commit all deletions together
                    db.commit()
                    print(f'✅ Account deletion completed and committed for applicant {applicant_id}')
                    
                    # Close database cursor before clearing session
                    cursor.close()
                    
                    # Clear session completely before logout
                    logout_user()
                    
                    # Force clear ALL session data to prevent redirect loops
                    session.clear()
                    for key in list(session.keys()):
                        session.pop(key, None)
                    
                    # Return an HTML page with immediate JavaScript redirect to prevent redirect page
                    login_url = url_for('login', _external=True)
                    html_content = f'''<!DOCTYPE html>
<html>
<head>
    <meta http-equiv="refresh" content="0; url={login_url}">
    <script>window.location.replace("{login_url}");</script>
</head>
<body>
    <script>window.location.replace("{login_url}");</script>
</body>
</html>'''
                    from flask import Response
                    return Response(html_content, status=302, headers={'Location': login_url, 'Content-Type': 'text/html; charset=utf-8'})
                except Exception as delete_exc:
                    db.rollback()
                    import traceback
                    print(f'❌ Delete account error: {delete_exc}')
                    traceback.print_exc()
                    flash('Failed to delete account. Please contact support.', 'error')
                    return redirect(url_for('applicant_profile'))

            if action == 'terminate_sessions':
                current_session_id = session.get('auth_session_id')
                if current_session_id:
                    cursor.execute(
                        '''
                        UPDATE auth_sessions
                        SET logout_time = NOW(), is_active = 0
                        WHERE user_id = %s
                          AND role = 'applicant'
                          AND session_id <> %s
                        ''',
                        (applicant_id, current_session_id),
                    )
                else:
                    cursor.execute(
                        '''
                        UPDATE auth_sessions
                        SET logout_time = NOW(), is_active = 0
                        WHERE user_id = %s
                          AND role = 'applicant'
                        ''',
                        (applicant_id,),
                    )
                db.commit()
                flash('All other active sessions have been signed out.', 'success')
                return redirect(url_for('applicant_profile'))
            
            if action == 'delete_resume':
                resume_id = request.form.get('resume_id')
                if not resume_id:
                    flash('Invalid resume selection.', 'error')
                    return redirect(url_for('applicant_profile'))

                cursor.execute(
                    '''
                    SELECT file_path
                    FROM resumes
                    WHERE resume_id = %s AND applicant_id = %s
                    LIMIT 1
                    ''',
                    (resume_id, applicant_id),
                )
                resume_record = cursor.fetchone()

                if not resume_record:
                    flash('Resume not found or already removed.', 'warning')
                    return redirect(url_for('applicant_profile'))

                # Detach resume from any applications to avoid FK constraints
                try:
                    cursor.execute(
                        '''
                        UPDATE applications
                        SET resume_id = NULL
                        WHERE applicant_id = %s AND resume_id = %s
                        ''',
                        (applicant_id, resume_id),
                    )
                except Exception as fk_exc:
                    # Proceed even if this optional step fails; deletion might still succeed without FK
                    print(f'⚠️ Could not detach resume {resume_id} from applications: {fk_exc}')

                cursor.execute('DELETE FROM resumes WHERE resume_id = %s', (resume_id,))
                db.commit()

                file_path = resume_record.get('file_path')
                if file_path:
                    try:
                        if os.path.isabs(file_path):
                            absolute_path = os.path.realpath(file_path)
                        else:
                            absolute_path = os.path.realpath(os.path.join(app.instance_path, file_path))
                        if os.path.exists(absolute_path):
                            os.remove(absolute_path)
                    except OSError as err:
                        print(f'⚠️ Failed to delete resume file {file_path}: {err}')

                log_profile_change(applicant_id, 'applicant', 'resume', file_path or '', 'deleted')
                flash('Resume removed successfully.', 'success')
                return redirect(url_for('applicant_profile'))

        # GET and post-processing context
        # Re-fetch applicant data to ensure we have the latest information (important after updates)
        # Check if last_profile_update column exists before selecting it
        cursor.execute('SHOW COLUMNS FROM applicants LIKE "last_profile_update"')
        has_last_profile_update = cursor.fetchone() is not None
        
        if has_last_profile_update:
            cursor.execute(
                '''
                SELECT applicant_id, full_name, email, phone_number, password_hash,
                       last_login, created_at, last_profile_update
                FROM applicants
                WHERE applicant_id = %s
                LIMIT 1
                ''',
                (applicant_id,),
            )
        else:
            cursor.execute(
                '''
                SELECT applicant_id, full_name, email, phone_number, password_hash,
                       last_login, created_at
                FROM applicants
                WHERE applicant_id = %s
                LIMIT 1
                ''',
                (applicant_id,),
            )
        fresh_applicant_record = cursor.fetchone() or applicant_record
        
        cursor.execute(
            '''
            SELECT resume_id,
                   file_name,
                   file_path,
                   uploaded_at
            FROM resumes
            WHERE applicant_id = %s
            ORDER BY uploaded_at DESC
            ''',
            (applicant_id,),
        )
        resumes = cursor.fetchall() or []

        # Use fresh data for display
        applicant = {key: value for key, value in fresh_applicant_record.items() if key != 'password_hash'}
        for date_field in ['last_login', 'created_at']:
            if date_field in applicant:
                applicant[date_field] = format_human_datetime(applicant.get(date_field))

        # Check auth_sessions table columns
        cursor.execute('SHOW COLUMNS FROM auth_sessions')
        session_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
        
        # Build logout_time expression
        if 'last_activity' in session_columns and 'logout_time' in session_columns:
            logout_expr = 'COALESCE(last_activity, logout_time)'
        elif 'last_activity' in session_columns:
            logout_expr = 'last_activity'
        elif 'logout_time' in session_columns:
            logout_expr = 'logout_time'
        else:
            logout_expr = 'NULL'

        # Determine login_time expression (some schemas use 'login_time', others use 'created_at')
        if 'login_time' in session_columns:
            login_expr = 'login_time'
        elif 'created_at' in session_columns:
            login_expr = 'created_at'
        elif 'last_login' in session_columns:
            login_expr = 'last_login'
        else:
            login_expr = 'NULL'

        login_rows = fetch_rows(
            f'''
            SELECT session_id, {login_expr} AS login_time, {logout_expr} AS logout_time, COALESCE(is_active, 1) AS is_active
            FROM auth_sessions
            WHERE user_id = %s
            ORDER BY {login_expr} DESC
            LIMIT 10
            ''',
            (applicant_id,),
        )
        active_session_id = session.get('auth_session_id')
        login_history = []
        for row in login_rows:
            is_active = bool(row.get('is_active', 1))
            logout_value = None if is_active else (format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None)
            login_history.append({
                'session_id': row.get('session_id'),
                'login_time': format_human_datetime(row.get('login_time')),
                'logout_time': logout_value,
                'active': is_active,
                'is_current': row.get('session_id') == active_session_id,
            })

        # Fetch profile change history. The schema uses 'change_type' and 'applicant_id'.
        cursor.execute("SHOW TABLES LIKE 'profile_changes'")
        if cursor.fetchone():
            profile_rows = fetch_rows(
                '''
                SELECT change_type, old_value, new_value, changed_at
                FROM profile_changes
                WHERE applicant_id = %s
                ORDER BY changed_at DESC
                LIMIT 10
                ''',
                (applicant_id,),
            )
        else:
            profile_rows = []

        profile_history = [
            {
                'field': row.get('change_type') or row.get('field_changed'),
                'old': row.get('old_value'),
                'new': row.get('new_value'),
                'changed_at': format_human_datetime(row.get('changed_at')),
            }
            for row in profile_rows
        ]

        resumes_context = []
        for item in resumes:
            file_path = (item.get('file_path') or '').replace('\\', '/')
            file_name = item.get('file_name') or os.path.basename(file_path) or 'Resume'
            # Compute file size from filesystem when possible
            try:
                if file_path:
                    if os.path.isabs(file_path):
                        abs_path = os.path.realpath(file_path)
                    else:
                        abs_path = os.path.realpath(os.path.join(app.instance_path, file_path))
                else:
                    abs_path = None
                file_size_bytes = os.path.getsize(abs_path) if abs_path and os.path.exists(abs_path) else 0
            except Exception:
                file_size_bytes = 0
            file_size = format_file_size(file_size_bytes)

            resumes_context.append(
                {
                    'resume_id': item.get('resume_id'),
                    'file_name': file_name,
                    'file_size': file_size,
                    'uploaded_at': format_human_datetime(item.get('uploaded_at')),
                    'download_url': url_for('download_resume', resume_id=item.get('resume_id')),
                    'view_url': url_for('preview_resume', resume_id=item.get('resume_id')),
                    'is_pdf': file_name.lower().endswith('.pdf'),
                    'file_path': file_path,
                }
            )

        return render_template(
            'applicant/profile.html',
            applicant=applicant,
            resumes=resumes_context,
            login_history=login_history,
            profile_history=profile_history,
            active_session_id=active_session_id,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applicant profile error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Unable to load profile: {str(exc)}', 'error')
        return render_template('applicant/profile.html', applicant=None, resumes=[], login_history=[], profile_history=[])
    finally:
        if cursor:
            cursor.close()


@app.route('/applicant/resumes/<int:resume_id>/download')
@login_required('applicant')
def download_resume(resume_id):
    applicant_id = session.get('user_id')
    record = execute_query(
        '''
        SELECT file_path, file_name
        FROM resumes
        WHERE resume_id = %s AND applicant_id = %s
        LIMIT 1
        ''',
        (resume_id, applicant_id),
        fetch_one=True,
    )

    if not record or not record.get('file_path'):
        html = '<!doctype html><meta charset="utf-8"><title>Preview Error</title><div style="background:#0f172a;color:#e2e8f0;padding:20px;font-family:system-ui"><h1 style="font-size:18px;margin:0 0 10px">Preview Error</h1><p>Resume not found.</p></div>'
        resp = Response(html, status=404, mimetype='text/html')
        resp.headers['X-Content-Type-Options'] = 'nosniff'
        resp.headers['Content-Security-Policy'] = "frame-ancestors *"
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        resp.headers['Pragma'] = 'no-cache'
        return resp

    file_path = record['file_path']
    absolute_path = os.path.join(app.root_path, file_path)

    if not os.path.exists(absolute_path):
        flash('Resume file is no longer available.', 'error')
        return redirect(url_for('applicant_profile'))

    download_name = record.get('file_name') or os.path.basename(file_path)
    mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'
    return send_file(absolute_path, as_attachment=True, download_name=download_name, mimetype=mimetype_value)





@app.route('/admin/applicant/<int:applicant_id>/attachments', methods=['GET'])
@login_required('admin', 'hr')
def admin_applicant_attachments(applicant_id):
    """Debug endpoint: return JSON of all attachments for an applicant grouped by file_type."""
    db = get_db()
    if not db:
        return jsonify({'success': False, 'error': 'Database connection error.'}), 500
    cursor = db.cursor(dictionary=True)
    try:
        # Get attachments linked to any application for this applicant
        cursor.execute(
            '''
            SELECT aa.attachment_id, aa.application_id, r.resume_id, r.file_name, r.file_path, r.file_type, r.uploaded_at
            FROM application_attachments aa
            JOIN resumes r ON aa.resume_id = r.resume_id
            JOIN applications a ON aa.application_id = a.application_id
            WHERE a.applicant_id = %s
            ORDER BY r.uploaded_at DESC
            ''',
            (applicant_id,)
        )
        rows = cursor.fetchall() or []
        grouped = {'resume': [], 'letter': [], 'license': [], 'other': []}
        for r in rows:
            ftype = (r.get('file_type') or 'resume').lower()
            entry = {
                'attachment_id': r.get('attachment_id'),
                'application_id': r.get('application_id'),
                'resume_id': r.get('resume_id'),
                'file_name': r.get('file_name'),
                'uploaded_at': r.get('uploaded_at').isoformat() if r.get('uploaded_at') else None,
                'file_path': r.get('file_path'),
            }
            if ftype in grouped:
                grouped[ftype].append(entry)
            else:
                grouped['other'].append(entry)

        return jsonify({'success': True, 'applicant_id': applicant_id, 'attachments': grouped}), 200
    except Exception as e:
        print(f'⚠️ Debug attachments error: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cursor.close()


@app.route('/applicant/resumes/<int:resume_id>/view')
@login_required('applicant')
def preview_resume(resume_id):
    applicant_id = session.get('user_id')
    record = execute_query(
        '''
        SELECT file_path, file_name
        FROM resumes
        WHERE resume_id = %s AND applicant_id = %s
        LIMIT 1
        ''',
        (resume_id, applicant_id),
        fetch_one=True,
    )

    if not record or not record.get('file_path'):
        flash('Resume not found.', 'error')
        return redirect(url_for('applicant_profile'))

    file_path = record['file_path']

    # Resolve secure upload base inside Flask instance folder (uploads must live under instance uploads)
    upload_base = os.path.join(app.instance_path, app.config.get('UPLOAD_FOLDER', 'uploads/resumes'))

    # Resolve absolute path only within upload_base; do NOT fall back to static/web-root locations
    if os.path.isabs(file_path):
        absolute_path = os.path.realpath(file_path)
    else:
        absolute_path = os.path.realpath(os.path.join(app.instance_path, file_path))

    upload_base_real = os.path.realpath(upload_base)
    # Only allow files that are inside the configured secure upload directory
    # Use both os.sep and '/' for cross-platform compatibility (Windows uses \, Linux uses /)
    if not (absolute_path == upload_base_real or absolute_path.startswith(upload_base_real + os.sep) or absolute_path.startswith(upload_base_real + '/')):
        html = '<!doctype html><meta charset="utf-8"><title>Preview Error</title><div style="background:#0f172a;color:#e2e8f0;padding:20px;font-family:system-ui"><h1 style="font-size:18px;margin:0 0 10px">Preview Error</h1><p>Resume file is not stored in a secure location.</p></div>'
        resp = Response(html, status=403, mimetype='text/html')
        resp.headers['X-Content-Type-Options'] = 'nosniff'
        resp.headers['Content-Security-Policy'] = "frame-ancestors *"
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        resp.headers['Pragma'] = 'no-cache'
        return resp

    if not os.path.exists(absolute_path):
        html = '<!doctype html><meta charset="utf-8"><title>Preview Error</title><div style="background:#0f172a;color:#e2e8f0;padding:20px;font-family:system-ui"><h1 style="font-size:18px;margin:0 0 10px">Preview Error</h1><p>Resume file is no longer available.</p></div>'
        resp = Response(html, status=404, mimetype='text/html')
        resp.headers['X-Content-Type-Options'] = 'nosniff'
        resp.headers['Content-Security-Policy'] = "frame-ancestors *"
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        resp.headers['Pragma'] = 'no-cache'
        return resp

    # Always send inline for viewing (opens in browser, not download)
    download_name = record.get('file_name') or os.path.basename(absolute_path)

    # Determine mimetype based on file extension
    file_ext = os.path.splitext(download_name.lower())[1]
    # If Word document, attempt to convert to PDF for inline preview (requires soffice)
    if file_ext in ['.doc', '.docx']:
        try:
            import subprocess, tempfile, shutil, threading

            tmpdir = tempfile.mkdtemp(prefix='preview_')
            converted_name = os.path.splitext(download_name)[0] + '.pdf'
            converted_path = os.path.join(tmpdir, converted_name)

            # Run LibreOffice headless conversion; ignore output unless error
            subprocess.check_call([
                'soffice', '--headless', '--convert-to', 'pdf', '--outdir', tmpdir, absolute_path
            ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            if os.path.exists(converted_path):
                resp = send_file(converted_path, as_attachment=False, download_name=converted_name, mimetype='application/pdf')
                # set inline disposition and security headers later
                def _cleanup(path=tmpdir):
                    try:
                        shutil.rmtree(path)
                    except Exception:
                        pass
                # cleanup after 30 seconds to allow client to fetch
                threading.Timer(30.0, _cleanup).start()

                resp.headers['X-Content-Type-Options'] = 'nosniff'
                resp.headers['Content-Security-Policy'] = "frame-ancestors *"
                resp.headers['Referrer-Policy'] = 'no-referrer'
                resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                resp.headers['Pragma'] = 'no-cache'
                resp.headers['Content-Type'] = 'application/pdf'
                resp.headers['Content-Disposition'] = 'inline; filename="' + converted_name + '"'
                resp.headers['Accept-Ranges'] = 'bytes'
                return resp
        except Exception:
            # conversion via soffice failed; attempt Windows COM conversion if on Windows
            try:
                import sys
                if sys.platform.startswith('win'):
                    try:
                        import tempfile as _tempfile, shutil as _shutil, threading as _threading
                        # Attempt Microsoft Word COM automation (requires pywin32 and MS Word installed)
                        try:
                            import pythoncom
                            import win32com.client
                            tmpdir2 = _tempfile.mkdtemp(prefix='preview_win_')
                            converted_name = os.path.splitext(download_name)[0] + '.pdf'
                            converted_path = os.path.join(tmpdir2, converted_name)

                            pythoncom.CoInitialize()
                            word = win32com.client.DispatchEx('Word.Application')
                            word.Visible = False
                            doc = word.Documents.Open(absolute_path, ReadOnly=1)
                            # 17 = wdFormatPDF
                            doc.SaveAs(converted_path, FileFormat=17)
                            doc.Close(False)
                            word.Quit()

                            if os.path.exists(converted_path):
                                resp = send_file(converted_path, as_attachment=False, download_name=converted_name, mimetype='application/pdf')
                                def _cleanup2(path=tmpdir2):
                                    try:
                                        _shutil.rmtree(path)
                                    except Exception:
                                        pass
                                _threading.Timer(30.0, _cleanup2).start()

                                resp.headers['X-Content-Type-Options'] = 'nosniff'
                                resp.headers['Content-Security-Policy'] = "frame-ancestors *"
                                resp.headers['Referrer-Policy'] = 'no-referrer'
                                resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                                resp.headers['Pragma'] = 'no-cache'
                                resp.headers['Content-Type'] = 'application/pdf'
                                resp.headers['Content-Disposition'] = 'inline; filename="' + converted_name + '"'
                                resp.headers['Accept-Ranges'] = 'bytes'
                                return resp
                        except Exception:
                            pass
                    except Exception:
                        pass
            except Exception:
                pass
    if file_ext == '.pdf':
        mimetype_value = 'application/pdf'
    elif file_ext in ['.doc', '.docx']:
        try:
            if file_ext == '.docx':
                import zipfile, re, html as _html
                with zipfile.ZipFile(absolute_path) as z:
                    xml = z.read('word/document.xml').decode('utf-8', 'ignore')
                paras = re.findall(r'<w:p[\\s\\S]*?</w:p>', xml)
                parts = []
                for p in paras:
                    runs = re.findall(r'<w:t[^>]*>([\\s\\S]*?)</w:t>', p)
                    if runs:
                        parts.append('<p>' + _html.escape(''.join(runs)) + '</p>')
                html_doc = '<!doctype html><html><head><meta charset="utf-8"><style>body{font-family:system-ui;background:#0f172a;color:#e2e8f0;padding:20px;margin:0}p{margin:0 0 10px}</style></head><body>' + ''.join(parts) + '</body></html>'
                response = Response(html_doc, mimetype='text/html')
                response.headers['X-Content-Type-Options'] = 'nosniff'
                response.headers['Referrer-Policy'] = 'no-referrer'
                response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                response.headers['Pragma'] = 'no-cache'
                response.headers['Content-Security-Policy'] = "frame-ancestors *"
                response.headers['Content-Disposition'] = 'inline; filename="' + download_name + '"'
                return response
        except Exception:
            pass
        mimetype_value = 'application/msword' if file_ext == '.doc' else 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    elif file_ext in ['.txt']:
        mimetype_value = 'text/plain'
    else:
        mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'

    response = send_file(absolute_path, as_attachment=False, download_name=download_name, mimetype=mimetype_value)

    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'no-referrer'
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
    response.headers['Pragma'] = 'no-cache'

    response.headers['Content-Security-Policy'] = "frame-ancestors *"

    if file_ext in ['.pdf', '.jpg', '.jpeg', '.png', '.txt']:
        if file_ext == '.pdf':
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Accept-Ranges'] = 'bytes'
        response.headers['Content-Disposition'] = 'inline; filename="' + download_name + '"'
        return response
    else:
        name_safe = _html.escape(download_name)
        wrapper = f'<!doctype html><meta charset="utf-8"><title>{name_safe}</title><div style="background:#0f172a;color:#e2e8f0;padding:20px;font-family:system-ui"><h1 style="font-size:18px;margin:0 0 10px">{name_safe}</h1><p>Preview is not available for this file type.</p><a href="{url_for("download_resume", resume_id=resume_id)}" style="display:inline-block;margin-top:10px;padding:8px 12px;border-radius:8px;background:#1f2937;color:#fff;text-decoration:none">Download</a></div>'
        resp = Response(wrapper, mimetype='text/html')
        resp.headers['X-Content-Type-Options'] = 'nosniff'
        resp.headers['Content-Security-Policy'] = "frame-ancestors *"
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        resp.headers['Pragma'] = 'no-cache'
        return resp


@app.route('/admin/resumes/<int:resume_id>/download')
@login_required('admin', 'hr')
def admin_download_resume(resume_id):
    """Allow admin and HR to download applicant resumes."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Verify resume exists and get file info
        cursor.execute(
            '''
            SELECT r.file_path, r.file_name, a.applicant_id
            FROM resumes r
            JOIN applicants a ON r.applicant_id = a.applicant_id
            WHERE r.resume_id = %s
            LIMIT 1
            ''',
            (resume_id,),
        )
        record = cursor.fetchone()
        
        if not record or not record.get('file_path'):
            flash('Resume not found.', 'error')
            return redirect(url_for('applicants'))
        
        # For HR users, verify the resume belongs to their branch
        if user.get('role') == 'hr':
            branch_id = get_branch_scope(user)
            if branch_id:
                cursor.execute(
                    '''
                    SELECT a.application_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND j.branch_id = %s
                    LIMIT 1
                    ''',
                    (record['applicant_id'], branch_id),
                )
                if not cursor.fetchone():
                    flash('You can only access resumes from your branch.', 'error')
                    return redirect(url_for('applicants'))
        
        file_path = record['file_path']

        # Resolve secure upload base inside Flask instance folder
        upload_base = os.path.join(app.instance_path, app.config.get('UPLOAD_FOLDER', 'uploads/resumes'))

        # Support both absolute and instance-relative stored paths
        if os.path.isabs(file_path):
            absolute_path = os.path.realpath(file_path)
        else:
            absolute_path = os.path.realpath(os.path.join(app.instance_path, file_path))

        upload_base_real = os.path.realpath(upload_base)
        # Ensure the resolved path is inside our secure upload directory
        if not (absolute_path == upload_base_real or absolute_path.startswith(upload_base_real + os.sep) or absolute_path.startswith(upload_base_real + '/')):
            flash('Resume file is not in a secure storage location.', 'error')
            return redirect(url_for('applicants'))

        if not os.path.exists(absolute_path):
            flash('Resume file is no longer available.', 'error')
            return redirect(url_for('applicants'))

        download_name = record.get('file_name') or os.path.basename(absolute_path)
        mimetype_value = mimetypes.guess_type(download_name)[0] or 'application/octet-stream'

        response = send_file(absolute_path, as_attachment=True, download_name=download_name, mimetype=mimetype_value)
        # Security headers to reduce content-sniffing and clickjacking
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['Referrer-Policy'] = 'no-referrer'
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'
        return response
    except Exception as exc:
        print(f'❌ Admin download resume error: {exc}')
        flash('Unable to download resume.', 'error')
        return redirect(url_for('applicants'))
    finally:
        cursor.close()


@app.route('/admin/resumes/<int:resume_id>/view')
@login_required('admin', 'hr')
def admin_view_resume(resume_id):
    """Allow admin and HR to view applicant resumes."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    
    cursor = db.cursor(dictionary=True)
    try:
        # Verify resume exists and get file info
        cursor.execute(
            '''
            SELECT r.file_path, r.file_name, a.applicant_id
            FROM resumes r
            JOIN applicants a ON r.applicant_id = a.applicant_id
            WHERE r.resume_id = %s
            LIMIT 1
            ''',
            (resume_id,),
        )
        record = cursor.fetchone()
        
        if not record or not record.get('file_path'):
            print(f'🔍 admin_view_resume: resume_id={resume_id} record={record}')
            flash('Resume not found.', 'error')
            return redirect(url_for('applicants'))

        # For HR users, verify the resume belongs to their branch
        if user.get('role') == 'hr':
            branch_id = get_branch_scope(user)
            if branch_id:
                cursor.execute(
                    '''
                    SELECT a.application_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND j.branch_id = %s
                    LIMIT 1
                    ''',
                    (record['applicant_id'], branch_id),
                )
                if not cursor.fetchone():
                    flash('You can only access resumes from your branch.', 'error')
                    return redirect(url_for('applicants'))
        
        file_path = record['file_path']

        # Resolve secure upload base inside Flask instance folder
        upload_base = os.path.join(app.instance_path, app.config.get('UPLOAD_FOLDER', 'uploads/resumes'))

        if os.path.isabs(file_path):
            absolute_path = os.path.realpath(file_path)
        else:
            absolute_path = os.path.realpath(os.path.join(app.instance_path, file_path))

        upload_base_real = os.path.realpath(upload_base)
        # Only allow files that are inside our secure upload directory
        if not (absolute_path == upload_base_real or absolute_path.startswith(upload_base_real + os.sep) or absolute_path.startswith(upload_base_real + '/')):
            flash('Resume file is not in a secure storage location.', 'error')
            return redirect(url_for('applicants'))

        if not os.path.exists(absolute_path):
            flash('Resume file is no longer available.', 'error')
            return redirect(url_for('applicants'))

        download_name = record.get('file_name') or os.path.basename(absolute_path)
        file_ext = os.path.splitext(download_name.lower())[1]
        if file_ext in ['.doc', '.docx']:
            try:
                import subprocess, tempfile, shutil, threading
                tmpdir = tempfile.mkdtemp(prefix='preview_admin_')
                converted_name = os.path.splitext(download_name)[0] + '.pdf'
                converted_path = os.path.join(tmpdir, converted_name)
                subprocess.check_call(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', tmpdir, absolute_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                if os.path.exists(converted_path):
                    resp = send_file(converted_path, as_attachment=False, download_name=converted_name, mimetype='application/pdf')
                    def _cleanup(path=tmpdir):
                        try: shutil.rmtree(path)
                        except Exception: pass
                    threading.Timer(30.0, _cleanup).start()
                    resp.headers['X-Content-Type-Options'] = 'nosniff'
                    resp.headers['X-Frame-Options'] = 'DENY'
                    resp.headers['Referrer-Policy'] = 'no-referrer'
                    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                    resp.headers['Pragma'] = 'no-cache'
                    resp.headers['Content-Type'] = 'application/pdf'
                    resp.headers['Content-Disposition'] = 'inline; filename="' + converted_name + '"'
                    resp.headers['Accept-Ranges'] = 'bytes'
                    return resp
            except Exception:
                try:
                    import sys
                    if sys.platform.startswith('win'):
                        import tempfile as _tempfile, shutil as _shutil, threading as _threading
                        import pythoncom, win32com.client
                        tmpdir2 = _tempfile.mkdtemp(prefix='preview_admin_win_')
                        converted_name = os.path.splitext(download_name)[0] + '.pdf'
                        converted_path = os.path.join(tmpdir2, converted_name)
                        pythoncom.CoInitialize()
                        word = win32com.client.DispatchEx('Word.Application')
                        word.Visible = False
                        doc = word.Documents.Open(absolute_path, ReadOnly=1)
                        doc.SaveAs(converted_path, FileFormat=17)
                        doc.Close(False)
                        word.Quit()
                        if os.path.exists(converted_path):
                            resp = send_file(converted_path, as_attachment=False, download_name=converted_name, mimetype='application/pdf')
                            def _cleanup2(path=tmpdir2):
                                try: _shutil.rmtree(path)
                                except Exception: pass
                            _threading.Timer(30.0, _cleanup2).start()
                            resp.headers['X-Content-Type-Options'] = 'nosniff'
                            resp.headers['X-Frame-Options'] = 'DENY'
                            resp.headers['Referrer-Policy'] = 'no-referrer'
                            resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                            resp.headers['Pragma'] = 'no-cache'
                            resp.headers['Content-Type'] = 'application/pdf'
                            resp.headers['Content-Disposition'] = 'inline; filename="' + converted_name + '"'
                            resp.headers['Accept-Ranges'] = 'bytes'
                            return resp
                except Exception:
                    pass
        mimetype_value = 'application/pdf' if file_ext == '.pdf' else ('text/plain' if file_ext == '.txt' else (mimetypes.guess_type(download_name)[0] or 'application/octet-stream'))

        # Read file safely and return Response with security headers
        from flask import Response
        with open(absolute_path, 'rb') as f:
            file_data = f.read()

        as_attachment = False if mimetype_value == 'application/pdf' else True
        disposition = 'inline' if not as_attachment else f'attachment; filename="{download_name}"'

        response = Response(
            file_data,
            mimetype=mimetype_value,
            headers={
                'Content-Disposition': disposition,
                'Content-Type': mimetype_value,
                'Content-Length': str(len(file_data)),
                'X-Content-Type-Options': 'nosniff',
                'X-Frame-Options': 'DENY',
                'Referrer-Policy': 'no-referrer',
                'Cache-Control': 'no-store, no-cache, must-revalidate, private',
                'Pragma': 'no-cache'
            }
        )

        if mimetype_value == 'application/pdf':
            response.headers['Accept-Ranges'] = 'bytes'

        return response
    except Exception as exc:
        print(f'❌ Admin view resume error: {exc}')
        flash('Unable to view resume.', 'error')
        return redirect(url_for('applicants'))
    finally:
        cursor.close()


@app.route('/admin/resumes/view-by-path')
@login_required('admin', 'hr')
def admin_view_resume_by_path():
    """Allow viewing a resume by its stored file_path when no resume_id is available.
    This attempts to resolve the file by matching `file_path` in the `resumes` table for
    branch-scoping/security checks; otherwise denies access.
    """
    path_b64 = request.args.get('path')
    if not path_b64:
        flash('Invalid resume reference.', 'error')
        return redirect(url_for('applicants'))

    try:
        import base64
        file_path = base64.urlsafe_b64decode(path_b64.encode()).decode()
    except Exception:
        flash('Invalid resume reference.', 'error')
        return redirect(url_for('applicants'))

    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))

    cursor = db.cursor(dictionary=True)
    try:
        # Try to find a resume record matching this path to enforce branch scoping
        cursor.execute('SELECT resume_id, file_name, applicant_id FROM resumes WHERE file_path = %s LIMIT 1', (file_path,))
        record = cursor.fetchone()
        if not record:
            flash('Resume not found.', 'error')
            return redirect(url_for('applicants'))

        # Reuse admin_view_resume logic: verify HR branch scope
        user = get_current_user()
        if user.get('role') == 'hr':
            branch_id = get_branch_scope(user)
            if branch_id:
                cursor.execute(
                    '''
                    SELECT a.application_id
                    FROM applications a
                    JOIN jobs j ON a.job_id = j.job_id
                    WHERE a.applicant_id = %s AND j.branch_id = %s
                    LIMIT 1
                    ''',
                    (record['applicant_id'], branch_id),
                )
                if not cursor.fetchone():
                    flash('You can only access resumes from your branch.', 'error')
                    return redirect(url_for('applicants'))

        # Build absolute path and stream inline similar to admin_view_resume
        upload_base = os.path.join(app.instance_path, app.config.get('UPLOAD_FOLDER', 'uploads/resumes'))
        if os.path.isabs(file_path):
            absolute_path = os.path.realpath(file_path)
        else:
            absolute_path = os.path.realpath(os.path.join(app.instance_path, file_path))

        upload_base_real = os.path.realpath(upload_base)
        if not (absolute_path == upload_base_real or absolute_path.startswith(upload_base_real + os.sep) or absolute_path.startswith(upload_base_real + '/')):
            flash('Resume file is not in a secure storage location.', 'error')
            return redirect(url_for('applicants'))

        if not os.path.exists(absolute_path):
            flash('Resume file is no longer available.', 'error')
            return redirect(url_for('applicants'))

        download_name = record.get('file_name') or os.path.basename(absolute_path)
        file_ext = os.path.splitext(download_name.lower())[1]
        if file_ext in ['.doc', '.docx']:
            try:
                import subprocess, tempfile, shutil, threading
                tmpdir = tempfile.mkdtemp(prefix='preview_admin_path_')
                converted_name = os.path.splitext(download_name)[0] + '.pdf'
                converted_path = os.path.join(tmpdir, converted_name)
                subprocess.check_call(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', tmpdir, absolute_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                if os.path.exists(converted_path):
                    resp = send_file(converted_path, as_attachment=False, download_name=converted_name, mimetype='application/pdf')
                    def _cleanup(path=tmpdir):
                        try: shutil.rmtree(path)
                        except Exception: pass
                    threading.Timer(30.0, _cleanup).start()
                    resp.headers['X-Content-Type-Options'] = 'nosniff'
                    resp.headers['X-Frame-Options'] = 'DENY'
                    resp.headers['Referrer-Policy'] = 'no-referrer'
                    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                    resp.headers['Pragma'] = 'no-cache'
                    resp.headers['Content-Type'] = 'application/pdf'
                    resp.headers['Content-Disposition'] = 'inline; filename="' + converted_name + '"'
                    resp.headers['Accept-Ranges'] = 'bytes'
                    return resp
            except Exception:
                try:
                    import sys
                    if sys.platform.startswith('win'):
                        import tempfile as _tempfile, shutil as _shutil, threading as _threading
                        import pythoncom, win32com.client
                        tmpdir2 = _tempfile.mkdtemp(prefix='preview_admin_path_win_')
                        converted_name = os.path.splitext(download_name)[0] + '.pdf'
                        converted_path = os.path.join(tmpdir2, converted_name)
                        pythoncom.CoInitialize()
                        word = win32com.client.DispatchEx('Word.Application')
                        word.Visible = False
                        doc = word.Documents.Open(absolute_path, ReadOnly=1)
                        doc.SaveAs(converted_path, FileFormat=17)
                        doc.Close(False)
                        word.Quit()
                        if os.path.exists(converted_path):
                            resp = send_file(converted_path, as_attachment=False, download_name=converted_name, mimetype='application/pdf')
                            def _cleanup2(path=tmpdir2):
                                try: _shutil.rmtree(path)
                                except Exception: pass
                            _threading.Timer(30.0, _cleanup2).start()
                            resp.headers['X-Content-Type-Options'] = 'nosniff'
                            resp.headers['X-Frame-Options'] = 'DENY'
                            resp.headers['Referrer-Policy'] = 'no-referrer'
                            resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
                            resp.headers['Pragma'] = 'no-cache'
                            resp.headers['Content-Type'] = 'application/pdf'
                            resp.headers['Content-Disposition'] = 'inline; filename="' + converted_name + '"'
                            resp.headers['Accept-Ranges'] = 'bytes'
                            return resp
                except Exception:
                    pass
        mimetype_value = 'application/pdf' if file_ext == '.pdf' else (mimetypes.guess_type(download_name)[0] or 'application/octet-stream')

        from flask import Response
        with open(absolute_path, 'rb') as f:
            file_data = f.read()

        as_attachment = False if mimetype_value == 'application/pdf' else True
        disposition = 'inline' if not as_attachment else f'attachment; filename="{download_name}"'

        response = Response(
            file_data,
            mimetype=mimetype_value,
            headers={
                'Content-Disposition': disposition,
                'Content-Type': mimetype_value,
                'Content-Length': str(len(file_data)),
                'X-Content-Type-Options': 'nosniff',
                'X-Frame-Options': 'DENY',
                'Referrer-Policy': 'no-referrer',
                'Cache-Control': 'no-store, no-cache, must-revalidate, private',
                'Pragma': 'no-cache'
            }
        )

        if mimetype_value == 'application/pdf':
            response.headers['Accept-Ranges'] = 'bytes'

        return response
    except Exception as exc:
        print(f'❌ Admin view resume by path error: {exc}')
        flash('Unable to view resume.', 'error')
        return redirect(url_for('applicants'))
    finally:
        cursor.close()



@app.route('/admin/dashboard')
@login_required('admin', 'hr')
def admin_dashboard():
    dashboard_data = build_admin_dashboard_data(get_current_user())
    # Add theme CSS for branch-specific UI
    branch_info = dashboard_data.get('branch_info', {})
    if branch_info:
        dashboard_data['theme_css'] = get_branch_theme_css(branch_info)
        dashboard_data['branch_banner_style'] = get_branch_banner_style(branch_info)
        dashboard_data['branch_logo_html'] = get_branch_logo_html(branch_info)
    else:
        dashboard_data['theme_css'] = get_branch_theme_css(None)
        dashboard_data['branch_banner_style'] = get_branch_banner_style(None)
        dashboard_data['branch_logo_html'] = get_branch_logo_html(None)
    return render_template('admin/admin_dashboard.html', dashboard_data=dashboard_data)





@app.route('/hr/dashboard')
@login_required('hr')
def hr_dashboard():
    """HR Dashboard with multi-branch support."""
    try:
        user = get_current_user()
        if not user:
            flash('Please log in to access the dashboard.', 'error')
            return immediate_redirect(url_for('login', _external=True))
        
        # Get HR user's assigned branch (if any)
        hr_branch_id = user.get('branch_id')
        
        # If HR is assigned to a specific branch, force that branch
        # Otherwise, get selected branch from query parameter for filtering
        if hr_branch_id:
            # HR is assigned to a specific branch - can only see that branch
            selected_branch_id = hr_branch_id
        else:
            # HR has all branches access - allow filtering
            selected_branch_id = request.args.get('branch_id', type=int)
        
        # Build main dashboard data (filtered by HR's assigned branch)
        try:
            dashboard_data = build_admin_dashboard_data(user, branch_id=selected_branch_id)
        except Exception as build_error:
            print(f'❌ Error building dashboard data: {build_error}')
            import traceback
            traceback.print_exc()
            dashboard_data = {
                'user': {'full_name': user.get('name', 'HR User'), 'email': user.get('email', ''), 'role': 'hr'},
                'stats': {},
                'metrics': {},
                'branch_info': {},
                'notifications': [],
                'upcoming_interviews': [],
                'recent_applications': [],
                'recent_activity': [],
                'recent_jobs': [],
            }
        
        # Get branch information for selected branch (if filtering)
        branch_info = {}
        if selected_branch_id:
            try:
                # Some deployments may not have theme columns in the branches table
                # to avoid SQL errors, only select core branch fields here and
                # let the theme helper supply defaults when theme columns are missing.
                branch_rows = fetch_rows('''SELECT branch_id, branch_name, address
                    FROM branches WHERE branch_id = %s''', (selected_branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
            except Exception as branch_error:
                print(f'⚠️ Error fetching branch info: {branch_error}')

            dashboard_data['branch_info'] = branch_info
            # Ensure theme CSS and helpers are present even if DB lacks theme columns
            try:
                dashboard_data['theme_css'] = get_branch_theme_css(branch_info)
                dashboard_data['branch_banner_style'] = get_branch_banner_style(branch_info)
                dashboard_data['branch_logo_html'] = get_branch_logo_html(branch_info)
            except Exception:
                # If theme helpers fail, fall back to empty/defaults
                dashboard_data['theme_css'] = ''
                dashboard_data['branch_banner_style'] = ''
                dashboard_data['branch_logo_html'] = ''
        dashboard_data['selected_branch_id'] = selected_branch_id
        dashboard_data['is_all_branches_hr'] = (hr_branch_id is None)  # True if HR has all-branches access
        # If HR user has All-Branches access and applied a branch filter, record it for UI only
        if not hr_branch_id and selected_branch_id:
            dashboard_data['filtered_branch_id'] = selected_branch_id
            try:
                rows = fetch_rows('SELECT branch_id, branch_name FROM branches WHERE branch_id = %s LIMIT 1', (selected_branch_id,))
                dashboard_data['filtered_branch_name'] = rows[0].get('branch_name') if rows else None
            except Exception:
                dashboard_data['filtered_branch_name'] = None
        
        # Get branches list for HR
        all_branches = []
        if hr_branch_id:
            # HR is assigned to a specific branch - only show that branch
            try:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address, operating_hours, is_active FROM branches WHERE branch_id = %s', (hr_branch_id,))
                all_branches = branch_rows or []
            except Exception as branch_fetch_error:
                print(f'⚠️ Error fetching assigned branch: {branch_fetch_error}')
                all_branches = []
        else:
            # HR has all branches access - get all branches
            try:
                all_branches = fetch_branches() or []
            except Exception as branch_fetch_error:
                print(f'⚠️ Error fetching branches: {branch_fetch_error}')
                all_branches = []
        
        # Get branch-specific statistics for each branch
        branch_stats = []
        for branch in all_branches:
            try:
                bid = branch.get('branch_id')
                if not bid:
                    continue
                    
                branch_open_jobs = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM jobs j
                    WHERE j.branch_id = %s AND j.status IN ('open', 'published', 'active')
                    """,
                    (bid,)
                ) or 0
                branch_applicants = fetch_count(
                    """
                    SELECT COUNT(DISTINCT a.applicant_id) AS count
                    FROM applications a
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s
                    """,
                    (bid,)
                ) or 0
                branch_interviews = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM interviews i
                    JOIN applications a ON a.application_id = i.application_id
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s AND DATE(i.scheduled_date) = CURDATE()
                    """,
                    (bid,)
                ) or 0
                branch_hires = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM applications a
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s AND a.status = 'hired'
                    """,
                    (bid,)
                ) or 0
                branch_pending = fetch_count(
                    """
                    SELECT COUNT(*) AS count
                    FROM applications a
                    JOIN jobs j ON j.job_id = a.job_id
                    WHERE j.branch_id = %s AND a.status = 'pending'
                    """,
                    (bid,)
                ) or 0
                
                branch_stats.append({
                    'branch_id': bid,
                    'branch_name': branch.get('branch_name', 'Unknown'),
                    'address': branch.get('address', ''),
                    'open_jobs': branch_open_jobs,
                    'applicants': branch_applicants,
                    'interviews_today': branch_interviews,
                    'hires': branch_hires,
                    'pending_applications': branch_pending,
                    'status': 'active',
                })
            except Exception as stat_error:
                print(f'⚠️ Error calculating stats for branch {branch.get("branch_id")}: {stat_error}')
                continue
        
        dashboard_data['branch_stats'] = branch_stats
        dashboard_data['all_branches'] = all_branches
        
        # Get recent job postings with branch info (filtered by selected branch if provided)
        try:
            if selected_branch_id:
                recent_jobs_with_branch = fetch_rows(
                    """
                    SELECT j.job_id,
                           COALESCE(j.job_title, 'Untitled Job') AS job_title,
                           j.status,
                           j.created_at AS posted_at,
                           b.branch_id,
                           b.branch_name,
                           (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
                    FROM jobs j
                    LEFT JOIN branches b ON b.branch_id = j.branch_id
                    WHERE j.branch_id = %s
                    ORDER BY j.created_at DESC
                    LIMIT 10
                    """,
                    (selected_branch_id,)
                ) or []
            else:
                recent_jobs_with_branch = fetch_rows(
                    """
                    SELECT j.job_id,
                           COALESCE(j.job_title, 'Untitled Job') AS job_title,
                           j.status,
                           j.created_at AS posted_at,
                           b.branch_id,
                           b.branch_name,
                           (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
                    FROM jobs j
                    LEFT JOIN branches b ON b.branch_id = j.branch_id
                    ORDER BY j.created_at DESC
                    LIMIT 10
                    """
                ) or []
            dashboard_data['recent_jobs_with_branch'] = recent_jobs_with_branch
        except Exception as jobs_error:
            print(f'⚠️ Error fetching recent jobs: {jobs_error}')
            dashboard_data['recent_jobs_with_branch'] = []
        
        # Get recent applicants with branch info (filtered by selected branch if provided)
        try:
            if selected_branch_id:
                recent_applicants_with_branch = fetch_rows(
                    """
                    SELECT ap.applicant_id,
                           ap.full_name AS applicant_name,
                           ap.email,
                           COALESCE(j.job_title, 'N/A') AS job_title,
                           b.branch_id,
                           b.branch_name,
                           a.status,
                           a.applied_at
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    WHERE j.branch_id = %s
                    ORDER BY a.applied_at DESC
                    LIMIT 10
                    """,
                    (selected_branch_id,)
                ) or []
            else:
                recent_applicants_with_branch = fetch_rows(
                    """
                    SELECT ap.applicant_id,
                           ap.full_name AS applicant_name,
                           ap.email,
                           COALESCE(j.job_title, 'N/A') AS job_title,
                           b.branch_id,
                           b.branch_name,
                           a.status,
                           a.applied_at
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    ORDER BY a.applied_at DESC
                    LIMIT 10
                    """
                ) or []
            dashboard_data['recent_applicants_with_branch'] = recent_applicants_with_branch
        except Exception as applicants_error:
            print(f'⚠️ Error fetching recent applicants: {applicants_error}')
            dashboard_data['recent_applicants_with_branch'] = []
        
        # Theme CSS disabled for HR dashboard
        dashboard_data['theme_css'] = ''
        dashboard_data['branch_banner_style'] = ''
        dashboard_data['branch_logo_html'] = ''
        
        return render_template('hr/dashboard.html', dashboard_data=dashboard_data)
    except Exception as exc:
        print(f'❌ HR Dashboard error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to load dashboard. Please try again later.', 'error')
        # Return minimal dashboard data
        return render_template('hr/dashboard.html', dashboard_data={
            'user': {'full_name': 'HR User', 'email': '', 'role': 'hr'},
            'stats': {},
            'metrics': {},
            'branch_info': {},
            'notifications': [],
            'upcoming_interviews': [],
            'recent_applications': [],
            'recent_activity': [],
            'recent_jobs': [],
            'recent_jobs_with_branch': [],
            'recent_applicants_with_branch': [],
            'all_branches': [],
            'branch_stats': [],
            'selected_branch_id': None,
        })


# HR Communications route removed - feature disabled
# @app.route('/hr/communications', methods=['GET', 'POST'])
# @login_required('hr')
# def hr_communications():
#     """Branch-scoped communications center for HR."""
#     ... (route code removed)


# HR-Specific Routes (Branch-Scoped)
# Note: HR routes redirect to admin routes which already handle branch filtering
# The admin routes check user role and filter by branch_id for HR users
# HR templates will be updated to extend admin base template for consistency


# Admin Navigation Routes
@app.route('/api/admin/notifications')
@login_required('admin', 'hr')
def api_admin_notifications():
    """API endpoint to fetch admin notifications in JSON format."""
    user = get_current_user()
    if not user or user.get('role') not in ('admin', 'hr'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        formatted, unread = fetch_notifications_for({'system_only': True}, limit=5)
        return jsonify({
            'success': True,
            'notifications': formatted,
            'unread_count': unread,
            'unread_display': '99+' if unread > 99 else unread
        })
    except Exception as e:
        print(f'⚠️ Error fetching admin notifications API: {e}')
        return jsonify({
            'success': False,
            'error': str(e),
            'notifications': [],
            'unread_count': 0,
            'unread_display': 0
        }), 500


@app.route('/api/recent-applicants', methods=['GET'])
@login_required('admin', 'hr')
def api_recent_applicants():
    """Return recent applicants with branch information.

    Defaults branch_name to 'All Branches' when branch is None.
    """
    branch_id = request.args.get('branch_id', type=int)
    try:
        if branch_id:
            rows = fetch_rows(
                '''
                SELECT ap.applicant_id, ap.full_name AS applicant_name, ap.email,
                       COALESCE(j.job_title, 'N/A') AS job_title,
                       b.branch_id, b.branch_name, a.status, a.applied_at
                FROM applicants ap
                JOIN applications a ON ap.applicant_id = a.applicant_id
                JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                WHERE j.branch_id = %s
                ORDER BY a.applied_at DESC
                LIMIT 10
                ''',
                (branch_id,)
            ) or []
        else:
            rows = fetch_rows(
                '''
                SELECT ap.applicant_id, ap.full_name AS applicant_name, ap.email,
                       COALESCE(j.job_title, 'N/A') AS job_title,
                       b.branch_id, b.branch_name, a.status, a.applied_at
                FROM applicants ap
                JOIN applications a ON ap.applicant_id = a.applicant_id
                JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                ORDER BY a.applied_at DESC
                LIMIT 10
                ''',
            ) or []

        formatted = []
        for r in rows:
            formatted.append({
                'applicant_id': r.get('applicant_id'),
                'applicant_name': r.get('applicant_name'),
                'email': r.get('email'),
                'job_title': r.get('job_title'),
                'branch_id': r.get('branch_id'),
                'branch_name': r.get('branch_name') or 'All Branches',
                'status': r.get('status'),
                'applied_at': format_human_datetime(r.get('applied_at')) if r.get('applied_at') else None,
            })
        return jsonify({'success': True, 'recent_applicants': formatted})
    except Exception as exc:
        print(f'❌ API recent applicants error: {exc}')
        return jsonify({'success': False, 'recent_applicants': []}), 500

@app.route('/admin/notifications')
@login_required('admin', 'hr')
def admin_notifications():
    """System-wide notification management."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/notifications.html', notifications=[], unread_count=0)
    
    cursor = db.cursor(dictionary=True)
    try:
        ensure_schema_compatibility()
        
        # Check if is_read and sent_at columns exist
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns_raw = cursor.fetchall()
        notification_columns = [col.get('Field') if isinstance(col, dict) else col[0] for col in notification_columns_raw]
        
        # Build dynamic expressions
        sent_at_expr = 'COALESCE(n.sent_at, n.created_at, NOW())' if 'sent_at' in notification_columns else 'COALESCE(n.created_at, NOW())'
        is_read_expr = 'COALESCE(n.is_read, 0)' if 'is_read' in notification_columns else '0'
        
        # Show only system-level notifications for Admin page (exclude applicant/application-specific entries)
        params = []
        where_sql = 'WHERE n.application_id IS NULL'

        cursor.execute(
            f'''
            SELECT n.notification_id,
                   n.message,
                   {sent_at_expr} AS sent_at,
                   {is_read_expr} AS is_read,
                   a.application_id,
                   a.status AS application_status,
                   COALESCE(j.job_title, 'N/A') AS job_title,
                   COALESCE(ap.full_name, 'Unknown') AS applicant_name
            FROM notifications n
            LEFT JOIN applications a ON n.application_id = a.application_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
            {where_sql}
            ORDER BY {sent_at_expr} DESC
            LIMIT 200
            ''',
            tuple(params)
        )
        notifications = cursor.fetchall() or []
        
        unread_count = len([n for n in notifications if not n.get('is_read')])
        system_count = len(notifications)
        application_count = 0
        
        formatted_notifications = []
        for notif in notifications:
            formatted_notifications.append({
                'notification_id': notif.get('notification_id'),
                'message': notif.get('message'),
                'sent_at': format_human_datetime(notif.get('sent_at')),
                'is_read': notif.get('is_read', False),
                'application_id': notif.get('application_id'),
                'job_title': notif.get('job_title'),
                'applicant_name': notif.get('applicant_name'),
                'application_status': notif.get('application_status'),
                'type': 'system',
            })
        
        return render_template(
            'admin/notifications.html',
            notifications=formatted_notifications,
            unread_count=unread_count,
            system_count=system_count,
            application_count=application_count
        )
    except Exception as exc:
        if db:
            db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Admin notifications error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load notifications. Please try again later.', 'error')
        return render_template('admin/notifications.html', notifications=[], unread_count=0)
    finally:
        cursor.close()


@app.route('/admin/notifications/read-all', methods=['POST'])
@login_required('admin', 'hr')
def mark_all_admin_notifications_read():
    """Mark all notifications as read."""
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        
        # Clean up any JSON responses that might have been stored as notifications
        try:
            cursor.execute("""
                DELETE FROM notifications 
                WHERE message LIKE '{%' 
                AND (message LIKE '%"success"%' OR message LIKE '%"message"%')
            """)
            deleted_count = cursor.rowcount
            if deleted_count > 0:
                print(f'✅ Cleaned up {deleted_count} JSON response notifications from database')
        except Exception as cleanup_error:
            print(f'⚠️ Error cleaning up JSON notifications: {cleanup_error}')
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if has_is_read:
            cursor.execute('UPDATE notifications SET is_read = 1 WHERE is_read = 0')
            db.commit()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': True, 'message': 'All notifications marked as read'})
            flash('All notifications marked as read.', 'success')
        else:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': 'Notification read status not available.'}), 400
            flash('Notification read status not available.', 'error')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Mark all notifications read error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark all notifications as read.'}), 500
        flash('Failed to mark all notifications as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('admin_notifications'))


@app.route('/admin/notifications/<int:notification_id>/read', methods=['POST'])
@login_required('admin', 'hr')
def mark_admin_notification_read(notification_id):
    """Mark a notification as read."""
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        
        # Check if this notification is a JSON response and delete it if so
        cursor.execute('SELECT message FROM notifications WHERE notification_id = %s', (notification_id,))
        notif_record = cursor.fetchone()
        if notif_record:
            message = notif_record.get('message', '') if isinstance(notif_record, dict) else (notif_record[0] if notif_record else '')
            if message and message.strip().startswith('{') and '"success"' in message:
                # Delete JSON response notifications instead of marking as read
                cursor.execute('DELETE FROM notifications WHERE notification_id = %s', (notification_id,))
                db.commit()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': True, 'message': 'Invalid notification removed', 'notification_id': notification_id})
                flash('Invalid notification removed.', 'success')
                return redirect(url_for('admin_notifications'))
        
        # Check if is_read column exists
        cursor.execute('SHOW COLUMNS FROM notifications LIKE %s', ('is_read',))
        has_is_read = cursor.fetchone() is not None
        
        if has_is_read:
            cursor.execute(
                'UPDATE notifications SET is_read = 1 WHERE notification_id = %s',
                (notification_id,)
            )
        else:
            # If column doesn't exist, just return success (column will be created by ensure_schema_compatibility)
            pass
        
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'Notification marked as read', 'notification_id': notification_id})
        flash('Notification marked as read.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Mark notification read error: {exc}')
        print(f'Full traceback: {error_details}')
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to mark notification as read.'}), 500
        flash('Failed to mark notification as read.', 'error')
    finally:
        cursor.close()
    
    return redirect(url_for('admin_notifications'))


@app.route('/admin/notifications/delete-all', methods=['POST'])
@login_required('admin', 'hr')
def delete_all_admin_notifications():
    """Delete notifications. Admin: all. HR with branch: only notifications for their branch. HR without branch: all."""
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_notifications'))
    
    cursor = db.cursor()
    try:
        ensure_schema_compatibility()
        user = get_current_user()
        role = user.get('role') if user else session.get('user_role')
        branch_id = get_branch_scope(user) if user else (session.get('branch_id') if role == 'hr' else None)
        
        # Verify notifications table and needed columns
        cursor.execute("SHOW TABLES LIKE 'notifications'")
        if not cursor.fetchone():
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': True, 'message': 'No notifications to delete'}), 200
            flash('No notifications to delete.', 'info')
            return redirect(url_for('admin_notifications'))
        
        cursor.execute('SHOW COLUMNS FROM notifications')
        notification_columns = {row[0] if isinstance(row, (list, tuple)) else row.get('Field') for row in (cursor.fetchall() or []) if row}
        has_application_fk = 'application_id' in notification_columns
        
        if role == 'admin' or (role == 'hr' and not branch_id):
            # Admin or HR managing all branches: delete all notifications
            cursor.execute('DELETE FROM notifications')
        else:
            # HR with specific branch: delete only branch-scoped notifications (requires application_id)
            if not has_application_fk:
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': False, 'error': 'Unable to scope notifications by branch.'}), 400
                flash('Unable to scope notifications by branch.', 'error')
                return redirect(url_for('admin_notifications'))
            cursor.execute(
                '''
                DELETE n FROM notifications n
                JOIN applications a ON n.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                WHERE j.branch_id = %s
                ''',
                (branch_id,),
            )
        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            # Final cleanup: Remove any JSON response notifications that might have been created
            try:
                if branch_id:
                    cursor.execute("""
                        DELETE n FROM notifications n
                        LEFT JOIN applications a ON n.application_id = a.application_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE (j.branch_id = %s OR n.application_id IS NULL)
                        AND (
                            (n.message LIKE '{%' AND (n.message LIKE '%"success"%' OR n.message LIKE '%"message"%' OR n.message LIKE '%"error"%'))
                            OR n.message = 'Notifications deleted.'
                            OR n.message LIKE '%Notifications deleted%'
                            OR n.message LIKE '%All notifications deleted%'
                            OR n.message LIKE '%Notification deleted successfully%'
                        )
                    """, (branch_id,))
                else:
                    cursor.execute("""
                        DELETE FROM notifications 
                        WHERE (
                            (message LIKE '{%' AND (message LIKE '%"success"%' OR message LIKE '%"message"%' OR message LIKE '%"error"%'))
                            OR message = 'Notifications deleted.'
                            OR message LIKE '%Notifications deleted%'
                            OR message LIKE '%All notifications deleted%'
                            OR message LIKE '%Notification deleted successfully%'
                        )
                    """)
                final_cleaned = cursor.rowcount
                if final_cleaned > 0:
                    print(f'✅ Final cleanup: Removed {final_cleaned} JSON response notification(s) after admin delete-all')
                    db.commit()
            except Exception as final_cleanup_err:
                print(f'⚠️ Error in final cleanup: {final_cleanup_err}')
                db.rollback()
            
            return jsonify({'success': True, 'message': 'Notifications deleted.'})
        flash('Notifications deleted.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        print(f'❌ Delete all admin notifications error: {exc}')
        print(traceback.format_exc())
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to delete notifications.'}), 500
        flash('Failed to delete notifications.', 'error')
    finally:
        cursor.close()
    return redirect(url_for('admin_notifications'))


@app.route('/admin/manage-branches', methods=['GET', 'POST'])
@login_required('admin')
def manage_branches():
    """Comprehensive branch management with CRUD operations."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/manage_branches.html', branches=[], total_branches=0)
    
    cursor = db.cursor(dictionary=True)
    
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if action == 'add':
                branch_name = request.form.get('branch_name', '').strip()
                address = request.form.get('address', '').strip()
                operating_hours = request.form.get('operating_hours', '').strip() or None
                is_active = request.form.get('is_active')
                # Stamp who created the branch (current admin)
                current_user = get_current_user()
                created_by_admin = None
                if current_user and current_user.get('role') in ('admin', 'hr'):
                    # current_user['id'] corresponds to admin_id for admins/hr
                    created_by_admin = current_user.get('id')
                
                # Validation
                if not branch_name or not address:
                    error_msg = 'Branch name and address are required.'
                    flash(error_msg, 'error')
                    if is_ajax:
                        return jsonify({'success': False, 'error': error_msg}), 400
                else:
                    try:
                        # Check which columns exist in branches table
                        cursor.execute('SHOW COLUMNS FROM branches')
                        columns = {row.get('Field') for row in cursor.fetchall() if row}
                        
                        # Build INSERT query dynamically based on available columns
                        insert_fields = ['branch_name', 'address']
                        insert_values = [branch_name, address]
                        
                        # Add operating_hours if column exists
                        if 'operating_hours' in columns:
                            insert_fields.append('operating_hours')
                            insert_values.append(operating_hours)
                        
                        # Add is_active if column exists (dropdown sends '1' for Active, '0' for Inactive)
                        if 'is_active' in columns:
                            # Handle dropdown values: '1' (Active) or '0' (Inactive)
                            # Also handle checkbox legacy: 'on' (checked) or None/False (unchecked)
                            if is_active == '1' or is_active == 'on' or is_active is True:
                                is_active_bool = True
                            elif is_active == '0' or is_active is False:
                                is_active_bool = False
                            else:
                                # Default to Active if not explicitly set
                                is_active_bool = True
                            insert_fields.append('is_active')
                            insert_values.append(is_active_bool)
                        
                        # Add created_by_admin_id if column exists and we have an admin id
                        if 'created_by_admin_id' in columns and created_by_admin is not None:
                            insert_fields.append('created_by_admin_id')
                            insert_values.append(created_by_admin)

                        # Build and execute INSERT query (compute fields/placeholders after finalizing insert_fields)
                        placeholders = ', '.join(['%s'] * len(insert_fields))
                        fields = ', '.join(insert_fields)

                        cursor.execute(
                            f'''
                            INSERT INTO branches ({fields})
                            VALUES ({placeholders})
                            ''',
                            tuple(insert_values),
                        )
                        db.commit()
                        # Log in admin notifications
                        try:
                            admin_name = (current_user.get('name') if current_user else 'System')
                            create_admin_notification(cursor, f"Branch added: {branch_name} (by {admin_name})")
                        except Exception as notify_err:
                            print(f'⚠️ Error creating branch notification: {notify_err}')
                        flash('Branch added successfully.', 'success')
                        if is_ajax:
                            return jsonify({'success': True, 'redirect': url_for('manage_branches', _external=False)})
                    except Exception as add_error:
                        db.rollback()
                        print(f"❌ ERROR adding branch: {add_error}")
                        import traceback
                        traceback.print_exc()
                        error_msg = f'Error adding branch: {str(add_error)}'
                        flash(error_msg, 'error')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500
            
            elif action == 'update':
                branch_id = request.form.get('branch_id')
                branch_name = request.form.get('branch_name', '').strip()
                address = request.form.get('address', '').strip()
                # Get operating_hours from form (field name is 'operating_hours' with underscore)
                # Strip whitespace and convert empty/whitespace-only strings to None (same as add action)
                operating_hours = request.form.get('operating_hours', '').strip() or None
                is_active = request.form.get('is_active')
                
                if not branch_id or not branch_name or not address:
                    error_msg = 'Branch ID, name, and address are required.'
                    flash(error_msg, 'error')
                    if is_ajax:
                        return jsonify({'success': False, 'error': error_msg}), 400
                else:
                    # AUTOMATIC: Always update operating_hours (exists in schema)
                    # Check which columns exist in branches table for is_active
                    cursor.execute('SHOW COLUMNS FROM branches')
                    columns = {row.get('Field') for row in cursor.fetchall() if row}
                    
                    # Build UPDATE query with operating_hours always included
                    update_fields = [
                        'branch_name = %s',
                        'address = %s',
                        'operating_hours = %s',  # AUTOMATIC: Always included
                    ]
                    update_values = [branch_name, address, operating_hours]
                    
                    # Add is_active if column exists (dropdown sends '1' for Active, '0' for Inactive)
                    if 'is_active' in columns:
                        # Handle dropdown values: '1' (Active) or '0' (Inactive)
                        # Also handle checkbox legacy: 'on' (checked) or None/False (unchecked)
                        if is_active == '1' or is_active == 'on' or is_active is True:
                            is_active_bool = True
                        elif is_active == '0' or is_active is False:
                            is_active_bool = False
                        else:
                            # Default to Active if not explicitly set
                            is_active_bool = True
                        update_fields.append('is_active = %s')
                        update_values.append(is_active_bool)
                    
                    update_values.append(branch_id)
                    
                    # Execute UPDATE
                    try:
                        cursor.execute(
                            f'''
                            UPDATE branches
                            SET {', '.join(update_fields)}
                            WHERE branch_id = %s
                            ''',
                            tuple(update_values),
                        )
                        
                        # Verify the update was successful
                        rows_affected = cursor.rowcount
                        
                        if rows_affected > 0:
                            db.commit()
                            success_msg = f'Branch updated successfully. Operating hours: {operating_hours or "Not set"}'
                            flash(success_msg, 'success')
                            if is_ajax:
                                return jsonify({'success': True, 'redirect': url_for('manage_branches', _external=False)})
                        else:
                            db.rollback()
                            warning_msg = 'Branch not found or no changes made.'
                            flash(warning_msg, 'warning')
                            if is_ajax:
                                return jsonify({'success': False, 'error': warning_msg}), 404
                    except Exception as update_error:
                        db.rollback()
                        print(f"❌ ERROR updating branch: {update_error}")
                        import traceback
                        traceback.print_exc()
                        error_msg = f'Error updating branch: {update_error}'
                        flash(error_msg, 'error')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500

            elif action == 'delete':
                branch_id = request.form.get('branch_id') or request.form.get('id')
                if not branch_id:
                    error_msg = 'Branch ID is required.'
                    flash(error_msg, 'error')
                    if is_ajax:
                        return jsonify({'success': False, 'error': error_msg}), 400
                else:
                    try:
                        # Ensure branch exists
                        cursor.execute('SELECT branch_id, branch_name FROM branches WHERE branch_id = %s LIMIT 1', (branch_id,))
                        branch = cursor.fetchone()
                        if not branch:
                            error_msg = 'Branch not found.'
                            flash(error_msg, 'error')
                            if is_ajax:
                                return jsonify({'success': False, 'error': error_msg}), 404
                        else:
                            # Prevent deletion if there are dependent jobs
                            cursor.execute('SELECT COUNT(*) AS cnt FROM jobs WHERE branch_id = %s', (branch_id,))
                            job_count_row = cursor.fetchone() or { 'cnt': 0 }
                            job_count = job_count_row.get('cnt', 0)
                            if job_count and int(job_count) > 0:
                                error_msg = 'Cannot delete branch with existing job postings. Remove associated jobs first.'
                                flash(error_msg, 'error')
                                if is_ajax:
                                    return jsonify({'success': False, 'error': error_msg}), 400
                            else:
                                cursor.execute('DELETE FROM branches WHERE branch_id = %s', (branch_id,))
                                db.commit()
                                try:
                                    create_admin_notification(cursor, f"Branch deleted: {branch.get('branch_name')}")
                                except Exception:
                                    pass
                                flash('Branch deleted successfully.', 'success')
                                if is_ajax:
                                    return jsonify({'success': True, 'redirect': url_for('manage_branches', _external=False)})
                    except Exception as del_err:
                        db.rollback()
                        import traceback
                        traceback.print_exc()
                        print(f'❌ ERROR deleting branch: {del_err}')
                        error_msg = f'Error deleting branch: {del_err}'
                        flash(error_msg, 'error')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500
            
            # For non-AJAX requests, redirect normally
            return redirect(url_for('manage_branches', _external=False))
        
        # Get filters from query parameters
        keyword = request.args.get('keyword', '').strip()
        status_filter = request.args.get('status', '').strip()
        
        # Build WHERE clauses
        where_clauses = []
        params = []
        
        # Keyword search (case-insensitive, searches branch_name, address)
        if keyword:
            keyword_pattern = f"%{keyword}%"
            where_clauses.append('(LOWER(b.branch_name) LIKE LOWER(%s) OR LOWER(b.address) LIKE LOWER(%s))')
            params.extend([keyword_pattern, keyword_pattern])
        
        # Status filter
        if status_filter and status_filter != 'all':
            if status_filter == 'active':
                where_clauses.append('b.is_active = 1')
            elif status_filter == 'inactive':
                where_clauses.append('b.is_active = 0')
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Fetch branches with comprehensive metrics
        rows = fetch_rows(
            f"""
            SELECT 
                b.branch_id AS id,
                b.branch_id,
                b.branch_name,
                b.address,
                b.operating_hours,
                b.is_active,
                (SELECT COUNT(*) FROM jobs j WHERE j.branch_id = b.branch_id AND j.status IN ('published', 'active', 'open')) AS active_jobs,
                (SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = b.branch_id) AS total_applications,
                (SELECT COUNT(*) FROM applications a JOIN jobs j ON a.job_id = j.job_id WHERE j.branch_id = b.branch_id AND a.status = 'hired') AS accepted_applications
            FROM branches b
            WHERE {where_sql}
            ORDER BY b.branch_name ASC
            """,
            tuple(params) if params else None
        )
        
        total_branches = len(rows)
        return render_template(
            'admin/manage_branches.html',
            branches=rows,
            total_branches=total_branches,
            current_filters={'keyword': keyword, 'status': status_filter},
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Branch management error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        return render_template('admin/manage_branches.html', branches=[], total_branches=0)
    finally:
        cursor.close()


@app.route('/admin/hr-accounts', methods=['GET', 'POST'])
@login_required('admin')
def hr_accounts():
    """Comprehensive HR account management with CRUD operations - Admin only."""
    user = get_current_user()
    # Explicitly deny HR users
    if user.get('role') == 'hr':
        flash('Access denied. HR account management is only available to administrators.', 'error')
        return redirect(url_for('hr_dashboard'))
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/hr_accounts_management.html', accounts=[], branches=[])
    
    cursor = db.cursor(dictionary=True)
    
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'add':
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                password = request.form.get('password', '').strip()
                # Branch scope: empty value or "" => All Branches (NULL), otherwise numeric branch_id
                raw_branch = request.form.get('branch_id', '')
                try:
                    branch_id = int(raw_branch) if raw_branch not in (None, '', 'all') else None
                except ValueError:
                    branch_id = None
                is_active = request.form.get('is_active') == 'on'
                
                if not all([full_name, email, password]):
                    flash('Full name, email, and password are required.', 'error')
                elif len(password) < 6:
                    flash('Password must be at least 6 characters.', 'error')
                else:
                    
                    # Check if email already exists in users table
                    cursor.execute(
                        'SELECT user_id FROM users WHERE email = %s LIMIT 1',
                        (email,),
                    )
                    if cursor.fetchone():
                        flash('Email address is already registered.', 'error')
                    else:
                        # First, create user in users table
                        # Admin/HR accounts are automatically verified (no email verification required)
                        password_hash = hash_password(password)
                        cursor.execute(
                            '''
                            INSERT INTO users (email, password_hash, user_type, is_active, email_verified)
                            VALUES (%s, %s, 'hr', %s, 1)
                            ''',
                            (email, password_hash, is_active),
                        )
                        user_id = cursor.lastrowid
                        
                        # Get password_hash from users table to copy to admins table
                        cursor.execute(
                            'SELECT password_hash FROM users WHERE user_id = %s LIMIT 1',
                            (user_id,)
                        )
                        user_record = cursor.fetchone()
                        admin_password_hash = user_record.get('password_hash') if user_record else password_hash
                        
                        # Then, create admin record linked to the user with password_hash
                        # Insert branch_id (NULL = All Branches) and set role/is_active
                        cursor.execute(
                            '''
                            INSERT INTO admins (user_id, full_name, email, password_hash, role, is_active, branch_id)
                            VALUES (%s, %s, %s, %s, 'hr', %s, %s)
                            ''',
                            (user_id, full_name, email, admin_password_hash, is_active, branch_id),
                        )
                        admin_id = cursor.lastrowid
                        
                        # HR accounts manage all branches
                        
                        # Create admin notification for HR account creation
                        try:
                            admin_msg = f'New HR account created: {full_name} ({email}).'
                            create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating notification for HR account creation: {notify_err}')
                        
                        db.commit()
                        flash('HR account created successfully. Account can manage all branches.', 'success')
            
            elif action == 'update':
                admin_id = request.form.get('admin_id')
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                raw_branch = request.form.get('branch_id', '')
                try:
                    branch_id = int(raw_branch) if raw_branch not in (None, '', 'all') else None
                except ValueError:
                    branch_id = None
                is_active = request.form.get('is_active') == 'on'
                
                if not admin_id or not full_name or not email:
                    flash('Admin ID, full name, and email are required.', 'error')
                else:
                    
                    # Get current user_id from admins table
                    cursor.execute(
                        'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                        (admin_id,),
                    )
                    admin_record = cursor.fetchone()
                    if not admin_record:
                        flash('HR account not found.', 'error')
                    else:
                        user_id = admin_record['user_id']
                        
                        # Check if email is already in use by another user
                        cursor.execute(
                            'SELECT user_id FROM users WHERE email = %s AND user_id <> %s LIMIT 1',
                            (email, user_id),
                        )
                        if cursor.fetchone():
                            flash('Email address is already in use.', 'error')
                        else:
                            # Update users table
                            cursor.execute(
                                '''
                                UPDATE users
                                SET email = %s, is_active = %s
                                WHERE user_id = %s
                                ''',
                                (email, is_active, user_id),
                            )
                            
                            # Update admins table with branch scope, ensure role remains 'hr' and is_active is synced
                            cursor.execute(
                                '''
                                UPDATE admins
                                SET full_name = %s,
                                    email = %s,
                                    role = 'hr',
                                    is_active = %s,
                                    branch_id = %s
                                WHERE admin_id = %s
                                ''',
                                (full_name, email, is_active, branch_id, admin_id),
                            )
                            
                            # HR accounts manage all branches
                            
                            db.commit()
                            flash('HR account updated successfully. Account can manage all branches.', 'success')
            
            elif action == 'reset_password':
                admin_id = request.form.get('admin_id')
                new_password = request.form.get('new_password', '').strip()
                
                if not admin_id or not new_password or len(new_password) < 6:
                    flash('Admin ID and password (min 6 characters) are required.', 'error')
                else:
                    try:
                        # Get user_id from admins table
                        cursor.execute(
                            'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                            (admin_id,),
                        )
                        admin_record = cursor.fetchone()
                        if not admin_record:
                            flash('HR account not found.', 'error')
                        else:
                            user_id = admin_record['user_id']
                            # Hard update password in users table - permanently changes password in database
                            cursor.execute(
                                'UPDATE users SET password_hash = %s WHERE user_id = %s',
                                (hash_password(new_password), user_id),
                            )
                            db.commit()
                            flash('Password reset successfully in system and database.', 'success')
                    except Exception as exc:
                        db.rollback()
                        flash(f'Failed to reset password: {exc}', 'error')
            
            return redirect(url_for('hr_accounts'))
        
        accounts = fetch_hr_accounts()
        branches = fetch_branches()
        
        # Ensure accounts is a list
        if accounts is None:
            accounts = []
        
        print(f'🔍 HR Accounts Route: Fetched {len(accounts)} accounts')
        if accounts and len(accounts) > 0:
            print(f'🔍 First account: {accounts[0]}')
        else:
            print('⚠️ No accounts returned from fetch_hr_accounts()')
            # Try a direct query to see what's in the database
            try:
                cursor.execute("""
                    SELECT a.admin_id, a.full_name, u.email, u.user_type, u.is_active
                    FROM admins a
                    LEFT JOIN users u ON u.user_id = a.user_id
                    LIMIT 10
                """)
                all_admins = cursor.fetchall()
                print(f'🔍 All admins in database (first 10): {len(all_admins) if all_admins else 0}')
                if all_admins:
                    for admin in all_admins[:3]:
                        print(f'   - Admin ID: {admin.get("admin_id")}, Name: {admin.get("full_name")}, User Type: {admin.get("user_type")}')
                
                # Also check specifically for HR accounts
                cursor.execute("""
                    SELECT a.admin_id, a.full_name, u.email, u.user_type
                    FROM admins a
                    INNER JOIN users u ON u.user_id = a.user_id
                    WHERE u.user_type = 'hr'
                """)
                hr_admins = cursor.fetchall()
                print(f'🔍 HR admins found directly: {len(hr_admins) if hr_admins else 0}')
                if hr_admins:
                    for hr in hr_admins[:3]:
                        print(f'   - HR Admin ID: {hr.get("admin_id")}, Name: {hr.get("full_name")}, Email: {hr.get("email")}')
            except Exception as debug_error:
                print(f'⚠️ Debug query error: {debug_error}')
                import traceback
                traceback.print_exc()
        
        # Check auth_sessions table columns first
        cursor.execute('SHOW COLUMNS FROM auth_sessions')
        session_columns_raw = cursor.fetchall() or []
        session_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in session_columns_raw if row}
        
        # Get comprehensive login history for each account (Admin and HR)
        for account in accounts:
            # Get user_id from admins table for this account
            cursor.execute(
                'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                (account.get('admin_id'),),
            )
            admin_record = cursor.fetchone()
            if admin_record:
                user_id = admin_record['user_id']
                # Fetch login history from auth_sessions table
                cursor.execute(
                    '''
                    SELECT 
                        created_at AS login_time,
                        logout_time,
                        1 AS is_active
                    FROM auth_sessions
                    WHERE user_id = %s
                    ORDER BY created_at DESC
                    LIMIT 20
                    ''',
                    (user_id,),
                )
                login_rows = cursor.fetchall() or []
                # Format login history
                account['login_history'] = []
                for row in login_rows:
                    account['login_history'].append({
                        'login_time': format_human_datetime(row.get('login_time')) if row.get('login_time') else 'Never',
                        'logout_time': format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None,
                        'is_active': row.get('is_active', 1),
                    })
            else:
                account['login_history'] = []
        
        # Ensure accounts is always a list, even if empty
        if not accounts:
            accounts = []
        
        print(f'🔍 Rendering template with {len(accounts)} accounts')
        return render_template('admin/hr_accounts_management.html', accounts=accounts, branches=branches)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ HR accounts management error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        branches = fetch_branches() or []
        return render_template('admin/hr_accounts_management.html', accounts=[], branches=branches)
    finally:
        cursor.close()


@app.route('/admin/job-postings', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def job_postings():
    """Enhanced job postings management with bulk operations and advanced filtering."""
    user = get_current_user()
    if not user:
        flash('Please login to access this page.', 'error')
        return immediate_redirect(url_for('login', _external=True))

    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        branches = fetch_branches() or []
        positions = fetch_positions() or []
        metadata = {
            'status_options': VALID_JOB_STATUSES,
            'employment_types': VALID_EMPLOYMENT_TYPES,
            'work_arrangements': VALID_WORK_ARRANGEMENTS,
            'experience_levels': VALID_EXPERIENCE_LEVELS,
        }
        template = 'hr/job_postings.html' if user.get('role') == 'hr' else 'admin/job_postings.html'
        return render_template(
            template,
            jobs=[],
            branches=branches,
            positions=positions,
            user=user or {},
            current_branch=None,
            current_filters={},
            branch_info=None,
            job_meta=metadata,
        )

    cursor = db.cursor(dictionary=True)

    try:
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        branch_scope = get_branch_scope(user)

        def extract_job_payload(form, admin_id=None):
            """Collect and sanitize job form fields."""
            errors = []
            job_title = (form.get('job_title') or form.get('title') or '').strip()
            job_description = (form.get('job_description') or form.get('description') or '').strip()
            job_requirements = (form.get('job_requirements') or form.get('requirements') or '').strip()
            # Optional per-job file rules
            allowed_extensions = (form.get('allowed_extensions') or '').strip()
            max_file_size_mb = form.get('max_file_size_mb') or None
            # Removed fields: employment_type, work_arrangement, experience_level, job_location, salary_currency, salary_min, salary_max, application_deadline, position_name, position_id (not in actual schema)

            # Handle status - database uses 'open' but code can use 'active' (they're equivalent for visibility)
            # IMPORTANT: Default to 'open' so jobs are immediately visible to applicants from all branches
            status_input = form.get('status', '').strip().lower()
            # Map 'active' to 'open' for database compatibility, or use 'open' directly
            if status_input == 'active':
                status = 'open'  # Database uses 'open' as the active status
            elif status_input == 'open':
                status = 'open'
            elif status_input == 'closed':
                status = 'closed'  # Only explicitly closed jobs are hidden
            else:
                # Default to 'open' for any other value or empty - ensures jobs are visible by default
                status = 'open'
            
            # Ensure that when posting a new job, if status is not explicitly set, default to 'open' (visible)
            # This ensures jobs are immediately visible to applicants from ALL branches
            if not status_input or status_input == '':
                status = 'open'
            
            # Enforce branch scope for HR users assigned to specific branches
            # If HR user is assigned to a branch, they can ONLY post jobs to that branch
            if branch_scope is not None:
                # HR is assigned to a specific branch - force that branch
                branch_id = branch_scope
            else:
                # HR has no branch assignment - can select any branch from form or assign to all branches
                branch_raw = (form.get('branch_id') or '').strip()
                # Empty value means "assign to all branches" (NULL in database)
                branch_id = int(branch_raw) if branch_raw.isdigit() else None
                # branch_id can be None (all branches) or an integer (specific branch)

            if not job_title:
                errors.append('Job title is required.')
            # job_summary removed - not in schema
            if not job_description:
                errors.append('Job description is required.')
            if not job_requirements:
                errors.append('Job requirements are required.')

            # Use actual schema columns: title, description, requirements, status, branch_id, posted_by, posted_at
            # posted_at will be set by MySQL NOW() function in the INSERT/UPDATE query for accurate server time
            # Set posted_at to None - MySQL NOW() will be used in SQL for accurate server timezone
            posted_at = None
            
            payload = {
                'title': job_title,  # Actual column name
                'description': job_description,  # Actual column name
                'requirements': job_requirements,  # Actual column name
                'allowed_extensions': allowed_extensions,
                'max_file_size_mb': max_file_size_mb,
                'status': status,
                'branch_id': branch_id,
                'posted_by': admin_id if admin_id else None,
                'posted_at': posted_at,
            }

            return payload, errors

        if request.method == 'POST':
            action = request.form.get('action')
            
            # Block admins from posting/editing/deleting jobs - they can only view
            if user.get('role') == 'admin' and action in ('add', 'edit', 'duplicate', 'bulk_update', 'bulk_delete'):
                flash('Admins can only view job postings. Job management is restricted to HR staff.', 'error')
                return redirect(url_for('job_postings'))
            
            actor_admin_id_raw = session.get('user_id')
            # Validate admin_id exists in admins table to satisfy foreign key constraint
            actor_admin_id = get_valid_admin_id(actor_admin_id_raw)

            if action == 'add':
                payload, errors = extract_job_payload(request.form, actor_admin_id)

                if errors:
                    for message in errors:
                        flash(message, 'error')
                    # Redirect back to form on validation errors
                    return redirect(url_for('job_postings'))
                else:
                    try:
                        # Build INSERT statement using actual schema columns
                        ensure_schema_compatibility()
                        _update_job_columns(cursor)
                        
                        # Use actual schema column names (detect what's available)
                        ensure_schema_compatibility()
                        _update_job_columns(cursor)

                        # Determine actual column names to use for INSERT
                        title_col = job_column_name('job_title', alternatives=['title'], default=None)
                        desc_col = job_column_name('job_description', alternatives=['description'], default=None)
                        req_col = job_column_name('job_requirements', alternatives=['requirements'], default=None)
                        posted_by_col = 'posted_by' if 'posted_by' in JOB_COLUMNS else None
                        branch_col = 'branch_id' if 'branch_id' in JOB_COLUMNS else None
                        status_col = 'status' if 'status' in JOB_COLUMNS else None
                        created_col = 'created_at' if 'created_at' in JOB_COLUMNS else ('posted_at' if 'posted_at' in JOB_COLUMNS else None)

                        # Ensure title is not None or empty
                        job_title = payload.get('title') or 'Untitled Job'
                        if not str(job_title).strip():
                            job_title = 'Untitled Job'
                        
                        # Build dynamic INSERT using detected columns
                        cols = []
                        placeholders = []
                        params = []

                        # title
                        if title_col:
                            cols.append(title_col)
                            placeholders.append('%s')
                            params.append(job_title)

                        # description
                        if desc_col:
                            cols.append(desc_col)
                            placeholders.append('%s')
                            params.append(payload.get('description'))

                        # requirements
                        if req_col:
                            cols.append(req_col)
                            placeholders.append('%s')
                            params.append(payload.get('requirements'))

                        # status
                        if status_col:
                            cols.append(status_col)
                            placeholders.append('%s')
                            params.append(payload.get('status'))

                        # branch
                        if branch_col:
                            cols.append(branch_col)
                            placeholders.append('%s')
                            params.append(payload.get('branch_id'))

                        # per-job file rules (optional)
                        allowed_ext_col = job_column_name('allowed_extensions') if 'allowed_extensions' in JOB_COLUMNS else None
                        max_size_col = job_column_name('max_file_size_mb') if 'max_file_size_mb' in JOB_COLUMNS else None
                        if allowed_ext_col:
                            cols.append(allowed_ext_col)
                            placeholders.append('%s')
                            params.append(payload.get('allowed_extensions'))
                        if max_size_col:
                            cols.append(max_size_col)
                            placeholders.append('%s')
                            params.append(payload.get('max_file_size_mb'))
                        
                        # location - get from branch address if available
                        if 'location' in JOB_COLUMNS and payload.get('branch_id'):
                            try:
                                cursor.execute('SELECT address FROM branches WHERE branch_id = %s LIMIT 1', (payload.get('branch_id'),))
                                branch_row = cursor.fetchone()
                                branch_address = branch_row.get('address') if branch_row else None
                                if branch_address:
                                    cols.append('location')
                                    placeholders.append('%s')
                                    params.append(branch_address)
                            except Exception as loc_err:
                                print(f'⚠️ Could not fetch branch address: {loc_err}')

                        # posted_by
                        if posted_by_col:
                            cols.append(posted_by_col)
                            placeholders.append('%s')
                            params.append(payload.get('posted_by') if payload.get('posted_by') else None)

                        # created_at / posted_at handling: use NOW() when status open/active
                        if created_col:
                            cols.append(created_col)
                            if payload.get('status') in ('active', 'open'):
                                placeholders.append('NOW()')
                            else:
                                placeholders.append('%s')
                                params.append(None)

                        if not cols:
                            raise Exception('No writable job columns detected on jobs table')

                        cols_sql = ', '.join(cols)
                        vals_sql = ', '.join(placeholders)
                        sql = f'INSERT INTO jobs ({cols_sql}) VALUES ({vals_sql})'
                        cursor.execute(sql, tuple(params))
                        job_id = cursor.lastrowid
                        print(f'✅ Job {job_id} inserted successfully')
                        
                        # AUTOMATIC: Handle job status (posted_at, etc.)
                        if job_id:
                            auto_handle_job_status(cursor, job_id, payload['status'])
                        
                        # Create admin notification for new job posting
                        try:
                            hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                            branch_name = 'Unknown Branch'
                            if payload.get('branch_id'):
                                cursor.execute('SELECT branch_name FROM branches WHERE branch_id = %s LIMIT 1', (payload['branch_id'],))
                                branch_row = cursor.fetchone()
                                if branch_row:
                                    branch_name = branch_row.get('branch_name', 'Unknown Branch')
                            admin_msg = f'HR {hr_name} posted a new job: "{payload.get("title", "Untitled")}" at {branch_name}.'
                            create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating notification for job posting: {notify_err}')
                        
                        db.commit()
                        # Verify the job was saved with correct status
                        posted_at_col = job_column_name('posted_at', alternatives=['created_at']) or 'created_at'
                        cursor.execute(f'SELECT job_id, status, {posted_at_col} FROM jobs WHERE job_id = %s', (job_id,))
                        saved_job = cursor.fetchone()
                        print(f'✅ Job posted successfully - ID: {job_id}, Status in DB: {saved_job["status"] if saved_job else "N/A"}, Posted by: {payload.get("posted_by")}')
                        if payload['status'] in ('active', 'open'):
                            flash('Job posting created successfully and is now automatically visible to applicants.', 'success')
                        else:
                            flash('Job posting created successfully. Set status to "Active" to make it visible to applicants.', 'success')
                        # Redirect after successful insert
                        return redirect(url_for('job_postings'))
                    except Exception as db_error:
                        db.rollback()
                        import traceback
                        error_details = traceback.format_exc()
                        print(f'❌ Database insert error: {db_error}')
                        print(f'Full traceback: {error_details}')
                        print(f'Payload: {payload}')
                        # Check for specific database errors
                        error_msg = str(db_error)
                        if 'foreign key constraint' in error_msg.lower():
                            if 'branch_id' in error_msg.lower():
                                flash('Invalid branch selected. Please select a valid branch.', 'error')
                            elif 'position_id' in error_msg.lower():
                                flash('Invalid position selected. Please select a valid position.', 'error')
                            elif 'posted_by' in error_msg.lower() or 'admin' in error_msg.lower():
                                flash('Invalid user account. Please log out and log back in.', 'error')
                            else:
                                flash(f'Database constraint error: {error_msg}', 'error')
                        elif 'cannot be null' in error_msg.lower() or 'not null' in error_msg.lower():
                            flash('Missing required information. Please fill in all required fields.', 'error')
                        else:
                            flash(f'Failed to create job posting: {error_msg}. Please check the console for details.', 'error')
                        # Redirect back to form on database error
                        return redirect(url_for('job_postings'))

            elif action == 'bulk_update':
                job_ids = request.form.getlist('job_ids')
                bulk_status_input = normalize_choice(request.form.get('bulk_status'), VALID_JOB_STATUSES, None)
                
                # Map 'active' to 'open' for database compatibility (database enum: 'open', 'closed')
                if bulk_status_input == 'active':
                    bulk_status = 'open'
                elif bulk_status_input == 'closed':
                    bulk_status = 'closed'
                else:
                    bulk_status = bulk_status_input

                if job_ids and bulk_status:
                    placeholders = ','.join(['%s'] * len(job_ids))
                    
                    # Check if posted_at column exists
                    posted_at_col = job_column_name('posted_at')
                    
                    # Build UPDATE statement dynamically
                    update_parts = ['status = %s']
                    params = [bulk_status]
                    
                    # Only add posted_at if column exists
                    if posted_at_col:
                        update_parts.append(f'{posted_at_col} = %s')
                        if bulk_status == 'open':
                            # Use MySQL NOW() for accurate server time
                            cursor.execute('SELECT NOW() as current_time')
                            result = cursor.fetchone()
                            posted_at_value = result[0] if result else None
                            params.append(posted_at_value)
                        else:
                            params.append(None)
                    
                    params.extend(job_ids)
                    
                    branch_clause = ''
                    if branch_scope is not None:
                        branch_clause = ' AND branch_id = %s'
                        params.append(branch_scope)

                    update_sql = ', '.join(update_parts)
                    cursor.execute(
                        f'''
                        UPDATE jobs
                        SET {update_sql}
                        WHERE job_id IN ({placeholders}){branch_clause}
                        ''',
                        tuple(params),
                    )
                    
                    # AUTOMATIC: Handle status for each job
                    for job_id_val in job_ids:
                        try:
                            auto_handle_job_status(cursor, job_id_val, bulk_status)
                        except Exception:
                            pass  # Non-blocking
                    
                    # Create admin notification for bulk update
                    try:
                        hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                        admin_msg = f'HR {hr_name} bulk updated {len(job_ids)} job posting(s) to status: {bulk_status}.'
                        create_admin_notification(cursor, admin_msg)
                    except Exception as notify_err:
                        print(f'⚠️ Error creating notification for bulk update: {notify_err}')
                    
                    db.commit()
                    flash(f'{len(job_ids)} job posting(s) updated successfully. Status changes are automatically handled.', 'success')
                else:
                    flash('Select job postings and a valid status for bulk update.', 'warning')

            elif action == 'delete':
                job_id = request.form.get('job_id')
                if not job_id:
                    flash('Job ID is required.', 'error')
                else:
                    try:
                        # Verify job exists and belongs to user's branch (if HR)
                        cursor.execute(
                            'SELECT job_id, job_title, branch_id FROM jobs WHERE job_id = %s LIMIT 1',
                            (job_id,)
                        )
                        job = cursor.fetchone()
                        
                        if not job:
                            flash('Job posting not found.', 'error')
                        else:
                            # Check branch access for HR users
                            if branch_scope is not None:
                                if job.get('branch_id') != branch_scope:
                                    flash('You do not have permission to delete this job posting.', 'error')
                                    if is_ajax:
                                        return jsonify({'success': False, 'error': 'Permission denied.'}), 403
                                    return redirect(url_for('job_postings'))
                            
                            job_title = job.get('title') or 'Job Posting'
                            
                            # Hard delete from database - permanently removes job posting
                            cursor.execute('DELETE FROM jobs WHERE job_id = %s', (job_id,))
                            
                            # Create admin notification for job deletion
                            try:
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                admin_msg = f'HR {hr_name} deleted job posting: "{job_title}"'
                                create_admin_notification(cursor, admin_msg)
                            except Exception as notify_err:
                                print(f'⚠️ Error creating notification for job deletion: {notify_err}')
                            
                            db.commit()
                            
                            if is_ajax:
                                return jsonify({'success': True, 'message': 'Job posting deleted successfully.'})
                            flash('Job posting deleted successfully from system and database.', 'success')
                    except Exception as exc:
                        db.rollback()
                        error_msg = f'Failed to delete job posting: {str(exc)}'
                        print(f'❌ Error deleting job: {exc}')
                        if is_ajax:
                            return jsonify({'success': False, 'error': error_msg}), 500
                        flash(error_msg, 'error')
            
            elif action == 'bulk_delete':
                job_ids = request.form.getlist('job_ids')
                if job_ids:
                    try:
                        placeholders = ','.join(['%s'] * len(job_ids))
                        params = [*job_ids]
                        branch_clause = ''
                        if branch_scope is not None:
                            branch_clause = ' AND branch_id = %s'
                            params.append(branch_scope)

                        # Hard delete from database - permanently removes job postings
                        cursor.execute(
                            f'DELETE FROM jobs WHERE job_id IN ({placeholders}){branch_clause}',
                            tuple(params),
                        )
                        
                        # Create admin notification for bulk delete
                        try:
                            hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                            admin_msg = f'HR {hr_name} deleted {len(job_ids)} job posting(s) from the system.'
                            create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating notification for bulk delete: {notify_err}')
                        
                        db.commit()
                        flash(f'{len(job_ids)} job posting(s) deleted successfully from system and database.', 'success')
                    except Exception as exc:
                        db.rollback()
                        flash(f'Failed to delete job postings: {exc}', 'error')
                else:
                    flash('Select at least one job to delete.', 'warning')

            elif action == 'edit':
                # Treat edit same as add (update flow) for consistency
                action = 'add'
                # Fall through to the 'add' handler below

            elif action == 'duplicate':
                job_id_raw = request.form.get('job_id', '').strip()
                if not job_id_raw or not job_id_raw.isdigit():
                    flash('Invalid job ID.', 'error')
                    return redirect(url_for('job_postings'))
                
                job_id = int(job_id_raw)
                branch_clause = ''
                params = [job_id]
                if branch_scope is not None:
                    branch_clause = ' AND branch_id = %s'
                    params.append(branch_scope)

                # Use actual schema columns
                cursor.execute(
                    f'''
                    SELECT job_title AS title, description, requirements, status, branch_id, position_id, posted_by
                    FROM jobs
                    WHERE job_id = %s{branch_clause}
                    ''',
                    tuple(params),
                )
                original_job = cursor.fetchone()
                
                if original_job:
                    # Build dynamic INSERT for duplicate using actual jobs columns
                    ensure_schema_compatibility()
                    _update_job_columns(cursor)
                    title_col = job_column_name('job_title', alternatives=['title'], default=None)
                    desc_col = job_column_name('job_description', alternatives=['description'], default=None)
                    req_col = job_column_name('job_requirements', alternatives=['requirements'], default=None)
                    branch_col = 'branch_id' if 'branch_id' in JOB_COLUMNS else None
                    posted_by_col = 'posted_by' if 'posted_by' in JOB_COLUMNS else None
                    created_col = 'created_at' if 'created_at' in JOB_COLUMNS else ('posted_at' if 'posted_at' in JOB_COLUMNS else None)

                    cols = []
                    placeholders = []
                    params = []

                    if title_col:
                        cols.append(title_col)
                        placeholders.append('%s')
                        params.append(f"{original_job.get('title') or original_job.get('job_title', 'Untitled')} (Copy)")

                    if desc_col:
                        cols.append(desc_col)
                        placeholders.append('%s')
                        params.append(original_job.get('description') or original_job.get('job_description', ''))

                    if req_col:
                        cols.append(req_col)
                        placeholders.append('%s')
                        params.append(original_job.get('requirements') or original_job.get('job_requirements', ''))

                    if 'status' in JOB_COLUMNS:
                        cols.append('status')
                        placeholders.append('%s')
                        params.append('open')

                    if branch_col:
                        cols.append(branch_col)
                        placeholders.append('%s')
                        params.append(original_job.get('branch_id'))

                    # location - get from branch address if available
                    if 'location' in JOB_COLUMNS and original_job.get('branch_id'):
                        try:
                            cursor.execute('SELECT address FROM branches WHERE branch_id = %s LIMIT 1', (original_job.get('branch_id'),))
                            branch_row = cursor.fetchone()
                            branch_address = branch_row.get('address') if branch_row else None
                            if branch_address:
                                cols.append('location')
                                placeholders.append('%s')
                                params.append(branch_address)
                        except Exception as loc_err:
                            print(f'⚠️ Could not fetch branch address: {loc_err}')

                    if posted_by_col:
                        cols.append(posted_by_col)
                        placeholders.append('%s')
                        params.append(actor_admin_id)

                    if created_col:
                        cols.append(created_col)
                        placeholders.append('NOW()')

                    if not cols:
                        flash('Cannot duplicate job: no writable job columns detected.', 'error')
                    else:
                        cols_sql = ', '.join(cols)
                        vals_sql = ', '.join(placeholders)
                        sql = f'INSERT INTO jobs ({cols_sql}) VALUES ({vals_sql})'
                        cursor.execute(sql, tuple(params))
                        db.commit()
                        flash('Job posting duplicated successfully.', 'success')
                else:
                    flash('Job not found.', 'error')

            return redirect(url_for('job_postings'))

        # Apply filters - extract and validate all filter parameters
        filters = {}
        
        # Keyword filter (case-insensitive, trim whitespace)
        keyword_raw = (request.args.get('keyword', '') or request.args.get('search', '')).strip()
        if keyword_raw:
            filters['keyword'] = keyword_raw
        
        # Branch filter - HR users can filter by branch (they manage all branches)
        # Admin users can also filter by branch
        branch_id_raw = request.args.get('branch_id', '').strip()
        if branch_id_raw and branch_id_raw.isdigit():
            filters['branch_id'] = int(branch_id_raw)
        
        # Position filter
        position_id_raw = request.args.get('position_id', '').strip()
        if position_id_raw and position_id_raw.isdigit():
            filters['position_id'] = int(position_id_raw)
        
        # Status filter (validate against valid statuses)
        status_raw = request.args.get('status', '').strip().lower()
        if status_raw and status_raw in VALID_JOB_STATUSES:
            filters['status'] = status_raw
        
        # Employment type filter
        employment_type_raw = request.args.get('employment_type', '').strip().lower()
        if employment_type_raw and employment_type_raw in VALID_EMPLOYMENT_TYPES:
            filters['employment_type'] = employment_type_raw
        
        # Work arrangement filter
        work_arrangement_raw = request.args.get('work_arrangement', '').strip().lower()
        if work_arrangement_raw and work_arrangement_raw in VALID_WORK_ARRANGEMENTS:
            filters['work_arrangement'] = work_arrangement_raw
        
        # Experience level filter
        experience_level_raw = request.args.get('experience_level', '').strip().lower()
        if experience_level_raw and experience_level_raw in VALID_EXPERIENCE_LEVELS:
            filters['experience_level'] = experience_level_raw
        
        # Department filter
        department_raw = request.args.get('department', '').strip()
        if department_raw:
            filters['department'] = department_raw
        
        # Location filter
        location_raw = request.args.get('location', '').strip()
        if location_raw:
            filters['location'] = location_raw

        # Build WHERE clauses
        where_clauses = []
        params = []

        # Branch filter - apply branch_scope if set, otherwise use branch_id filter
        # HR users can filter by branch_id even though they manage all branches
        if branch_scope is not None:
            # If HR has a specific branch scope, use it (but this should be None for HR now)
            where_clauses.append('j.branch_id = %s')
            params.append(branch_scope)
        elif filters.get('branch_id'):
            # Apply branch filter if specified (for both HR and Admin)
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])

        # Keyword search (case-insensitive, searches title, description, and requirements)
        if filters.get('keyword'):
            keyword = f"%{filters['keyword']}%"
            # Use dynamic column checking for keyword fields
            job_title_col = job_column('job_title', 'title')
            job_desc_col = job_column('job_description', 'description')
            job_req_col = job_column('job_requirements', 'requirements')
            
            keyword_fields = [col for col in [job_title_col, job_desc_col, job_req_col] if col]
            if keyword_fields:
                # Use LOWER() for case-insensitive search
                like_clauses = [f"LOWER(j.{column}) LIKE LOWER(%s)" for column in keyword_fields]
                where_clauses.append(f"({' OR '.join(like_clauses)})")
                params.extend([keyword] * len(like_clauses))

        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])

        # Status filter - map 'active' to show all active statuses ('open', 'published', 'active')
        if filters.get('status'):
            status_filter = filters['status']
            if status_filter == 'active':
                # Show all active statuses: 'open', 'published', 'active' (database may have different status values)
                where_clauses.append('j.status IN (%s, %s, %s)')
                params.extend(['open', 'published', 'active'])
            else:
                # For 'closed' or other statuses, use exact match
                where_clauses.append('j.status = %s')
                params.append(status_filter)

        # Department filter removed (positions table no longer exists)
        # if filters.get('department'):
        #     where_clauses.append('p.department = %s')
        #     params.append(filters['department'])

        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'

        # Ensure job columns are updated before building expressions
        _update_job_columns(cursor)
        
        # Define all column expressions for SELECT clause
        # Use COALESCE to handle NULL values and ensure we always get a title
        job_title_col = job_column('job_title', 'title')
        if job_title_col:
            job_title_expr = f"COALESCE(j.{job_title_col}, 'Untitled Job')"
        else:
            job_title_expr = "'Untitled Job'"
        job_description_expr = job_column_expr('job_description', alternatives=['description'])
        job_requirements_expr = job_column_expr('job_requirements', alternatives=['requirements'])
        
        # Build position_name expression - use job_title as fallback since position_name doesn't exist in schema
        position_name_expr = f'{job_title_expr}'
        created_at_expr = job_column_expr('created_at')
        posted_at_expr = job_column_expr('posted_at', alternatives=['created_at'])
        position_id_expr = job_column_expr('position_id', default='NULL')
        branch_id_expr = job_column_expr('branch_id', default='NULL')
        status_expr = job_column_expr('status', default="'open'")  # Database uses 'open', not 'active'
        
        # Build admin join conditions dynamically for both created_by and posted_by
        if 'created_by_admin_id' in JOB_COLUMNS:
            created_by_join = 'LEFT JOIN admins a_created ON j.created_by_admin_id = a_created.admin_id'
        else:
            created_by_join = 'LEFT JOIN admins a_created ON NULL = a_created.admin_id'
        
        if 'posted_by' in JOB_COLUMNS:
            posted_by_join = 'LEFT JOIN admins a_posted ON j.posted_by = a_posted.admin_id'
        else:
            posted_by_join = 'LEFT JOIN admins a_posted ON NULL = a_posted.admin_id'

        # Sort order handling
        sort_order = request.args.get('sort', 'newest').strip().lower()
        order_by_clause = f'COALESCE({posted_at_expr}, {created_at_expr}) DESC'  # Default: newest first
        
        if sort_order == 'oldest':
            order_by_clause = f'COALESCE({posted_at_expr}, {created_at_expr}) ASC'
        elif sort_order == 'title_asc':
            order_by_clause = f'{job_title_expr} ASC'
        elif sort_order == 'title_desc':
            order_by_clause = f'{job_title_expr} DESC'
        elif sort_order == 'applications_desc':
            order_by_clause = '(SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) DESC'
        elif sort_order == 'applications_asc':
            order_by_clause = '(SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) ASC'
        
        # Store sort in filters for template
        if sort_order != 'newest':
            filters['sort'] = sort_order

        query = f'''
            SELECT
                j.job_id,
                COALESCE({job_title_expr}, 'Untitled Job') AS job_title,
                {job_description_expr} AS job_description,
                {job_requirements_expr} AS job_requirements,
                {status_expr} AS status,
                {branch_id_expr} AS branch_id,
                {position_id_expr} AS position_id,
                {created_at_expr} AS created_at,
                {posted_at_expr} AS posted_at,
                COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                {position_name_expr} AS position_name,
                'General' AS department,
                COALESCE(a_posted.full_name, 'System') AS posted_by_name,
                (SELECT COUNT(*) FROM applications apps WHERE apps.job_id = j.job_id) AS application_count
            FROM jobs j
            LEFT JOIN branches b ON {branch_id_expr} = b.branch_id
            {posted_by_join}
            WHERE {where_sql}
            ORDER BY {order_by_clause}
        '''

        cursor.execute(query, tuple(params) if params else None)
        jobs_raw = cursor.fetchall()
        
        # Debug: Check if position_name is in results
        if jobs_raw and len(jobs_raw) > 0:
            sample_job = jobs_raw[0]
            if isinstance(sample_job, dict):
                print(f'🔍 Sample job position_name from query: "{sample_job.get("position_name")}"')
            else:
                print(f'🔍 Sample job (non-dict): {sample_job}')
        
        # Ensure position_name exists in all job results
        for job in jobs_raw:
            if isinstance(job, dict):
                if 'position_name' not in job:
                    job['position_name'] = ''
                # Debug each job's position_name
                if job.get('job_id'):
                    print(f'🔍 Job ID {job.get("job_id")} position_name: "{job.get("position_name")}"')

        def build_option_list(values):
            return [{'value': value, 'label': value.replace('_', ' ').title()} for value in values]

        formatted_jobs = []
        for job in jobs_raw:
            salary_display = format_salary_range(job.get('salary_currency'), job.get('salary_min'), job.get('salary_max'))
            posted_ts = job.get('posted_at') or job.get('created_at')
            
            # Ensure job_requirements is available (check both possible column names)
            requirements_value = job.get('job_requirements') or job.get('requirements') or ''
            
            formatted_jobs.append(
                {
                    'job_id': job.get('job_id'),
                    'job_title': job.get('job_title'),
                    'title': job.get('job_title'),
                    'job_summary': (job.get('job_description') or '')[:200] if job.get('job_description') else '',
                    'job_description': job.get('job_description'),
                    'job_requirements': requirements_value,
                    'requirements': requirements_value,
                    'employment_type': job.get('employment_type'),
                    'work_arrangement': job.get('work_arrangement'),
                    'experience_level': job.get('experience_level'),
                    'job_location': job.get('job_location'),
                    'salary_currency': job.get('salary_currency'),
                    'salary_min': job.get('salary_min'),
                    'salary_max': job.get('salary_max'),
                    'salary_display': salary_display,
                    'application_deadline': job.get('application_deadline'),
                    'application_deadline_display': format_human_datetime(job.get('application_deadline')),
                    'status': job.get('status'),
                    'branch_id': job.get('branch_id'),
                    'branch_name': job.get('branch_name'),
                    'position_id': job.get('position_id'),
                    'position_name': (job.get('position_name') or '').strip() if job.get('position_name') else '',  # Include position_name from database, trim whitespace
                    'position_title': job.get('position_name') or job.get('job_title') or job.get('title'),
                    'department': job.get('department'),
                    'created_by_name': job.get('created_by_name'),
                    'posted_by_name': job.get('posted_by_name'),
                    'application_count': job.get('application_count', 0),
                    'created_at': format_human_datetime(job.get('created_at')),
                    'updated_at': format_human_datetime(job.get('updated_at')),
                    'posted_at': format_human_datetime(posted_ts),
                }
            )

        branches = fetch_branches()
        positions = fetch_positions()

        current_branch = None
        if branch_scope is not None:
            current_branch = next((branch for branch in branches if branch.get('branch_id') == branch_scope), None)

        # Get unique departments and locations for filters
        # Positions table removed - no departments available
        departments = []

        # Location filter removed (not in actual schema)
        locations = []

        job_meta = {
            'status_options': build_option_list(VALID_JOB_STATUSES),
            'departments': departments,
            'locations': locations,
        }

        template = 'hr/job_postings.html' if user.get('role') == 'hr' else 'admin/job_postings.html'
        branch_info = current_branch

        # Normalize filters for template comparison (ensure IDs are integers)
        normalized_filters = {}
        for key, value in filters.items():
            if key in ['branch_id', 'position_id']:
                try:
                    normalized_filters[key] = int(value) if value is not None else None
                except (ValueError, TypeError):
                    normalized_filters[key] = None
            else:
                normalized_filters[key] = value

        return render_template(
            template,
            jobs=formatted_jobs,
            branches=branches,
            positions=positions,
            user=user,
            current_branch=current_branch,
            current_filters=normalized_filters,
            branch_info=branch_info,
            job_meta=job_meta,
        )
    except Exception as exc:
        if db:
            db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Job postings error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        try:
            branches = fetch_branches() or []
            positions = fetch_positions() or []
        except Exception:
            branches = []
            positions = []

        job_meta = {
            'status_options': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_JOB_STATUSES],
            'employment_types': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_EMPLOYMENT_TYPES],
            'work_arrangements': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_WORK_ARRANGEMENTS],
            'experience_levels': [{'value': value, 'label': value.replace('_', ' ').title()} for value in VALID_EXPERIENCE_LEVELS],
        }
        template = 'hr/job_postings.html' if (user or {}).get('role') == 'hr' else 'admin/job_postings.html'
        return render_template(
            template,
            jobs=[],
            branches=branches,
            positions=positions,
            user=user or {},
            current_branch=None,
            current_filters={},
            branch_info=None,
            job_meta=job_meta,
        )
    finally:
        cursor.close()


@app.route('/admin/job-postings/<int:job_id>/update', methods=['POST'])
@login_required('admin', 'hr')
def update_job_posting(job_id):
    user = get_current_user()
    
    # Block admins from updating jobs - they can only view
    if user.get('role') == 'admin':
        flash('Admins can only view job postings. Job updates are restricted to HR staff.', 'error')
        return redirect(url_for('job_postings'))
    
    branch_scope = get_branch_scope(user)
    actor_admin_id = session.get('user_id')

    payload, errors = None, []

    def extract_payload(form):
        local_payload, local_errors = {}, []

        job_title = (form.get('job_title') or form.get('title') or '').strip()
        job_summary = ''  # job_summary column doesn't exist in schema
        job_description = (form.get('job_description') or form.get('description') or '').strip()
        job_requirements = (form.get('job_requirements') or form.get('requirements') or '').strip()
        position_name = (form.get('position') or form.get('position_name') or '').strip()
        # Keep as empty string if empty (don't convert to None) - matches schema DEFAULT NULL but allows empty strings
        print(f'🔍 Extracted position_name from form (update): "{position_name}"')
        
        # Get position_id from form if position was selected from dropdown
        position_id_raw = form.get('position_id', '').strip()
        if position_id_raw:
            try:
                position_id = int(position_id_raw)
            except (ValueError, TypeError):
                position_id = None
        else:
            position_id = None

        employment_type = normalize_choice(
            form.get('employment_type'),
            VALID_EMPLOYMENT_TYPES,
            VALID_EMPLOYMENT_TYPES[0],
        )

        # Handle status - database uses 'open' but code can use 'active' (they're equivalent for visibility)
        status_input = form.get('status', '').strip()
        # Map 'active' to 'open' for database compatibility, or use 'open' directly
        if status_input == 'active':
            status = 'open'  # Database uses 'open' as the active status
        elif status_input == 'open':
            status = 'open'
        else:
            status = normalize_choice(status_input, VALID_JOB_STATUSES, 'open')
        
        # Ensure default to 'open' if not set
        if not status_input or status_input == '':
            status = 'open'

        if branch_scope is not None:
            branch_id = branch_scope
        else:
            branch_raw = (form.get('branch_id') or '').strip()
            branch_id = int(branch_raw) if branch_raw.isdigit() else None
            if branch_id is None:
                local_errors.append('Branch is required for job postings.')

        # Position field removed - always set to None
        position_id = None

        if not job_title:
            local_errors.append('Job title is required.')
        # job_summary removed - not in schema
        if not job_description:
            local_errors.append('Job description is required.')
        if not job_requirements:
            local_errors.append('Job requirements are required.')

        # Automatic status handling - posted_at automatically set based on status
        # Note: posted_at is handled in the UPDATE query, not here

        local_payload = {
            'job_title': job_title,
            'title': job_title,  # Also include as 'title' for compatibility
            'job_description': job_description,
            'description': job_description,  # Also include as 'description' for compatibility
            'job_requirements': job_requirements,
            'requirements': job_requirements,  # Also include as 'requirements' for compatibility
            'status': status,
            'branch_id': branch_id,
        }

        return local_payload, local_errors

    payload, errors = extract_payload(request.form)
    
    # Debug: Print form data
    import traceback
    print(f'🔍 Update job posting - Job ID: {job_id}')
    print(f'🔍 Form data received: {dict(request.form)}')
    print(f'🔍 Payload extracted: {payload}')
    print(f'🔍 Errors: {errors}')

    if errors:
        for message in errors:
            flash(message, 'error')
        return redirect(url_for('job_postings'))

    try:
        # Validate admin_id exists in admins table to satisfy foreign key constraint
        actor_admin_id = get_valid_admin_id(actor_admin_id)
        
        # Build UPDATE statement dynamically based on available columns
        ensure_schema_compatibility()
        db = get_db()
        if db:
            cursor_temp = db.cursor()
            try:
                _update_job_columns(cursor_temp)
            finally:
                cursor_temp.close()
        
        # Determine which columns to use
        use_updated_by = 'updated_by_admin_id' in JOB_COLUMNS
        
        # AUTOMATIC: Status handling - get status from payload (already mapped to 'open' if 'active')
        status = payload.get('status', 'open')  # Default to 'open' (database value)
        print(f'🔍 Update - Status from payload: {status}')
        
        # Use actual schema columns
        # Map payload keys to actual column values
        job_title = payload.get('job_title') or payload.get('title') or ''
        job_description = payload.get('job_description') or payload.get('description') or ''
        job_requirements = payload.get('job_requirements') or payload.get('requirements') or ''
        
        # More robust update flow:
        # 1) Load current job row (respecting branch restriction for branch-scoped HR users)
        # 2) Compute which columns actually differ from payload
        # 3) Only issue UPDATE with the differing columns
        # This avoids permission/WHERE ordering issues and MySQL returning 0 affected rows
        if not db:
            flash('Database connection error.', 'error')
            return redirect(url_for('job_postings'))

        cursor = db.cursor(dictionary=True)
        try:
            # Load current job unconditionally (we'll enforce branch permission after we have the row)
            cursor.execute('SELECT * FROM jobs WHERE job_id = %s', (job_id,))
            current = cursor.fetchone()
            if not current:
                flash('Job posting not found.', 'error')
                return redirect(url_for('job_postings'))

            # Enforce branch-scoped HR permission: if HR is assigned to a branch, they can only update jobs in that branch
            branch_scope_val = get_branch_scope(user)
            if (user.get('role') == 'hr') and (branch_scope_val is not None):
                current_branch_id = current.get('branch_id') if isinstance(current, dict) else None
                if current_branch_id != branch_scope_val:
                    flash('You do not have permission to update this job. It belongs to a different branch.', 'error')
                    return redirect(url_for('job_postings'))

            # Compute candidate columns and their desired values (only for columns that exist)
            candidate_columns = []
            desired_values = []
            
            # Debug: Print current row keys
            print(f'🔍 Current row keys: {list(current.keys()) if isinstance(current, dict) else "Not a dict"}')
            print(f'🔍 Extracted values - job_title: "{job_title}", job_description: "{job_description}", job_requirements: "{job_requirements}", status: "{status}"')

            # Title / job_title handling
            if 'title' in current:
                candidate_columns.append('title')
                desired_values.append(job_title)
            elif 'job_title' in current:
                candidate_columns.append('job_title')
                desired_values.append(job_title)

            # Description / job_description - check actual current row keys
            if 'job_description' in current:
                candidate_columns.append('job_description')
                desired_values.append(job_description)
            elif 'description' in current:
                candidate_columns.append('description')
                desired_values.append(job_description)

            # Requirements / job_requirements — ALWAYS include these for updates
            # to ensure they're persisted even if empty
            if 'job_requirements' in current:
                candidate_columns.append('job_requirements')
                desired_values.append(job_requirements)
            elif 'requirements' in current:
                candidate_columns.append('requirements')
                desired_values.append(job_requirements)
            else:
                # Column doesn't exist in current row; check schema and try anyway
                req_col = job_column_name('job_requirements', alternatives=['requirements'], default=None)
                if req_col:
                    candidate_columns.append(req_col)
                    desired_values.append(job_requirements)
                    print(f'🔍 job_requirements not in fetched row, but column "{req_col}" exists in schema; including for update')

            # Status
            if 'status' in current:
                candidate_columns.append('status')
                desired_values.append(status)

            # Position name
            # position_name is not in actual schema, skip it

            # branch_id change allowed only for admin or HR all-branches
            if ('branch_id' in current) and ((user.get('role') == 'admin') or (user.get('role') == 'hr' and branch_scope_val is None)):
                # Only include branch change if provided in payload (not None)
                if payload.get('branch_id') is not None:
                    candidate_columns.append('branch_id')
                    desired_values.append(payload.get('branch_id'))

            # Optionally updated_by_admin_id and updated_at — we'll append these for any permitted update
            # so that an explicit submit by an authorized user records who made the change.
            # IMPORTANT: For HR and admin updates, we ALWAYS force updated_at = NOW() to ensure the row is actually modified.
            if 'updated_by_admin_id' in current:
                # Avoid duplicate
                if 'updated_by_admin_id' not in candidate_columns:
                    candidate_columns.append('updated_by_admin_id')
                    desired_values.append(actor_admin_id)
            
            # Always add updated_at for HR/admin so UPDATE actually modifies the row
            if 'updated_at' in current:
                if 'updated_at' not in candidate_columns:
                    candidate_columns.append('updated_at')
                    desired_values.append('__NOW__')
                    print(f'🔍 Forcing updated_at = NOW() for authorized user update')

            # posted_at automatic handling when status is open/active
            if status in ('active', 'open') and ('posted_at' in current):
                candidate_columns.append('posted_at')
                desired_values.append('__NOW__')

            # Compare desired_values to current row to determine which columns actually changed
            set_clauses = []
            params = []
            
            # Core fields that should ALWAYS be included in updates (force diff detection)
            core_update_fields = {'title', 'job_title', 'description', 'job_description', 'requirements', 'job_requirements', 'status'}
            
            for col, desired in zip(candidate_columns, desired_values):
                if desired == '__NOW__':
                    # Always update NOW() without a parameter
                    set_clauses.append(f"{col} = NOW()")
                    print(f'🔍 Column {col}: setting to NOW() (no parameter)')
                    continue

                current_val = current.get(col) if isinstance(current, dict) else None
                # Normalize None/empty-string comparisons for textual fields
                norm_current = '' if current_val is None else str(current_val).strip()
                norm_desired = '' if desired is None else str(desired).strip()
                print(f'🔍 Column {col}: current="{norm_current}", desired="{norm_desired}", match={norm_current == norm_desired}')
                
                # For core fields, ALWAYS include them in the update regardless of whether they changed
                # This ensures all key fields are properly set
                if col in core_update_fields or norm_current != norm_desired:
                    set_clauses.append(f"{col} = %s")
                    params.append(desired)
                    if col in core_update_fields and norm_current == norm_desired:
                        print(f'  ✅ Core field - forcing update even though values match')
                    elif norm_current != norm_desired:
                        print(f'  ✅ Different - adding to SET')

            if not set_clauses:
                # For HR and admin users, force updated_at = NOW() to ensure the update applies
                is_admin_user = (user.get('role') == 'admin')
                is_hr_all_branches = (user.get('role') == 'hr') and (get_branch_scope(user) is None)
                
                if (is_admin_user or is_hr_all_branches) and ('updated_at' in current):
                    # Force updated_at to ensure row is modified
                    set_clauses.append('updated_at = NOW()')
                    print(f'🔍 No field changes, but forcing updated_at = NOW() for authorized user')
                else:
                    # Nothing to update
                    print(f'ℹ️ No changed fields detected for job {job_id}; no UPDATE executed')
                    flash('No changes detected; job posting remains unchanged.', 'info')
                    return redirect(url_for('job_postings'))

            # Build WHERE clause: if HR assigned to a branch, restrict by branch_id as well
            where_clause = 'job_id = %s'
            where_params = [job_id]
            if (user.get('role') == 'hr') and (branch_scope_val is not None):
                where_clause += ' AND branch_id = %s'
                where_params.append(branch_scope_val)

            full_params = tuple(params + where_params)

            print(f'🔍 Computed SET clauses: {set_clauses}')
            print(f'🔍 Computed Params: {list(full_params)}')

            try:
                cursor.execute(
                    f"UPDATE jobs SET {', '.join(set_clauses)} WHERE {where_clause}",
                    full_params,
                )
            except Exception as sql_err:
                print(f'❌ SQL execution error: {sql_err}')
                print(f'🔍 SQL: UPDATE jobs SET {', '.join(set_clauses)} WHERE {where_clause}')
                print(f'🔍 Params: {full_params}')
                raise

            print(f'🔍 Update executed, rows affected: {cursor.rowcount}')

            if cursor.rowcount == 0:
                # If zero rows, the job likely still exists but values might have been changed concurrently.
                # For admins and HR users with All-Branches scope, treat this as a successful update
                # (they have permission and their submit should be accepted even if values are identical).
                is_admin_user = (user.get('role') == 'admin')
                is_hr_all_branches = (user.get('role') == 'hr') and (get_branch_scope(user) is None)
                if is_admin_user or is_hr_all_branches:
                    try:
                        auto_handle_job_status(cursor, job_id, status)
                    except Exception:
                        pass
                    db.commit()
                    print(f'ℹ️ Job {job_id} update submitted by privileged user; treating as success despite 0 rows affected')
                    flash('Job posting updated successfully. Changes are now visible in the job postings list.', 'success')
                    return redirect(url_for('job_postings'))

                # Non-privileged users: report no changes detected
                db.commit()
                flash('No changes detected or update could not be applied.', 'info')
                return redirect(url_for('job_postings'))

            # AUTOMATIC: Handle job status changes
            auto_handle_job_status(cursor, job_id, status)
            db.commit()
            print(f'✅ Job posting {job_id} updated successfully in database')
            flash('Job posting updated successfully. Changes are now visible in the job postings list.', 'success')
            return redirect(url_for('job_postings'))
        except Exception as db_exc:
            db.rollback()
            import traceback
            error_details = traceback.format_exc()
            print(f'❌ Database update error: {db_exc}')
            print(f'Full traceback: {error_details}')
            flash(f'Failed to update job posting: {db_exc}. Please check the console for details.', 'error')
        finally:
            cursor.close()
            
    except Exception as exc:
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Update job posting error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Failed to update job posting: {exc}. Please check the console for details.', 'error')

    return redirect(url_for('job_postings'))


@app.route('/admin/applicants', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def applicants():
    """Enhanced candidate tracking with filters, stats, and comprehensive data."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        return render_template(template, applications=[], branch_info=None, stats={}, filters={}, positions=[], jobs=[], branches=[], view_mode='list')
    
    cursor = db.cursor(dictionary=True)
    try:
        # Ensure schema compatibility before querying (best-effort)
        try:
            ensure_schema_compatibility()
        except Exception as e:
            print(f'⚠️ Schema compatibility check failed but continuing: {e}')
        branch_id = get_branch_scope(user)
        
        # Handle POST actions (bulk updates, etc.)
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'delete':
                # Only admin can delete applicants from system users
                if user.get('role') != 'admin':
                    flash('Access denied. Only administrators can delete applicants.', 'error')
                    return redirect(url_for('applicants'))
                
                applicant_id = request.form.get('applicant_id')
                if applicant_id:
                    try:
                        # Get user_id from applicants table before deleting
                        cursor.execute(
                            'SELECT user_id, full_name, email FROM applicants WHERE applicant_id = %s LIMIT 1',
                            (applicant_id,),
                        )
                        applicant_record = cursor.fetchone()
                        if applicant_record:
                            user_id = applicant_record['user_id']
                            applicant_name = applicant_record.get('full_name', 'Unknown')
                            applicant_email = applicant_record.get('email', 'Unknown')
                            
                            print(f'🔍 Deleting applicant: applicant_id={applicant_id}, user_id={user_id}, name={applicant_name}, email={applicant_email}')
                            
                            # Create admin notification before deleting applicant
                            try:
                                admin_notification_msg = f'Admin deleted applicant account: "{applicant_name}" (Email: {applicant_email}).'
                                create_admin_notification(cursor, admin_notification_msg)
                                print(f'✅ Admin notification created for applicant account deletion: {applicant_name}')
                            except Exception as notify_err:
                                print(f'⚠️ Error creating admin notification for applicant deletion: {notify_err}')
                            
                            # Delete all related data (complete deletion from database and system)
                            # 1. Delete notifications linked to applicant's applications
                            cursor.execute('DELETE FROM notifications WHERE application_id IN (SELECT application_id FROM applications WHERE applicant_id = %s)', (applicant_id,))
                            
                            # 2. Delete interviews related to applicant's applications
                            cursor.execute("SHOW TABLES LIKE 'interviews'")
                            if cursor.fetchone():
                                cursor.execute('DELETE FROM interviews WHERE application_id IN (SELECT application_id FROM applications WHERE applicant_id = %s)', (applicant_id,))
                                print(f'✅ Deleted interviews for applicant {applicant_id}')
                            
                            # 3. Delete saved jobs
                            cursor.execute('DELETE FROM saved_jobs WHERE applicant_id = %s', (applicant_id,))
                            
                            # 4. Delete applications
                            cursor.execute('DELETE FROM applications WHERE applicant_id = %s', (applicant_id,))
                            applications_deleted = cursor.rowcount
                            print(f'✅ Deleted {applications_deleted} application(s) for applicant {applicant_id}')
                            
                            # 5. Delete resumes
                            cursor.execute('DELETE FROM resumes WHERE applicant_id = %s', (applicant_id,))
                            resumes_deleted = cursor.rowcount
                            print(f'✅ Deleted {resumes_deleted} resume(s) for applicant {applicant_id}')
                            
                            # 6. Delete password resets
                            if applicant_email:
                                cursor.execute('DELETE FROM password_resets WHERE user_email = %s', (applicant_email,))
                            
                            # 7. Delete auth sessions
                            cursor.execute("SHOW TABLES LIKE 'auth_sessions'")
                            if cursor.fetchone():
                                if user_id:
                                    cursor.execute('DELETE FROM auth_sessions WHERE user_id = %s', (user_id,))
                            
                            # 8. Delete profile changes history
                            cursor.execute("SHOW TABLES LIKE 'profile_changes'")
                            if cursor.fetchone():
                                if user_id:
                                    cursor.execute('DELETE FROM profile_changes WHERE user_id = %s', (user_id,))
                                else:
                                    cursor.execute('DELETE FROM profile_changes WHERE applicant_id = %s', (applicant_id,))
                            
                            # 9. Delete applicant record
                            cursor.execute('DELETE FROM applicants WHERE applicant_id = %s', (applicant_id,))
                            applicant_deleted = cursor.rowcount > 0
                            print(f'✅ Deleted applicant record {applicant_id} from applicants table' if applicant_deleted else f'⚠️ No applicant record found with ID {applicant_id}')
                            
                            # 8. ALWAYS delete user record from users table to ensure removal from system users
                            if user_id:
                                # Try with user_type check first
                                cursor.execute('DELETE FROM users WHERE user_id = %s AND user_type = %s', (user_id, 'applicant'))
                                user_deleted = cursor.rowcount > 0
                                
                                # If no rows affected with user_type check, try without user_type (in case of data inconsistency)
                                if not user_deleted:
                                    print(f'⚠️ No user deleted with user_type check, trying without user_type check...')
                                    cursor.execute('DELETE FROM users WHERE user_id = %s', (user_id,))
                                    user_deleted = cursor.rowcount > 0
                                
                                if user_deleted:
                                    print(f'✅ Successfully deleted user record {user_id} from users table')
                                else:
                                    print(f'⚠️ DELETE query executed but no rows affected for user_id {user_id} - user may not exist in users table')
                            else:
                                print(f'⚠️ No user_id found for applicant {applicant_id} - cannot delete from users table')
                            
                            db.commit()
                            print(f'✅ Applicant deletion completed for applicant_id {applicant_id}')
                            flash(f'Applicant {applicant_name} deleted successfully from system and database.', 'success')
                        else:
                            flash('Applicant not found.', 'error')
                    except Exception as exc:
                        db.rollback()
                        log.exception(f'❌ Error deleting applicant: {exc}')
                        flash(f'Failed to delete applicant: {exc}', 'error')
                
                return redirect(url_for('applicants'))
            
            elif action == 'bulk_update_status':
                # Restrict admin from changing status - only HR can change status
                if user.get('role') == 'admin':
                    flash('Access denied. Only HR users can change application status.', 'error')
                    return redirect(url_for('applicants'))
                
                application_ids = request.form.getlist('application_ids')
                new_status = request.form.get('bulk_status', '').strip()
                
                # Note: All statuses can now be manually changed via bulk update
                # HR has full control over application statuses
                
                # Map simplified statuses to database statuses
                # ALL statuses can now be manually changed by HR: pending, scheduled, interviewed, hired, rejected
                CANONICAL_STATUS_MAP = {
                    'pending': 'pending',
                    'scheduled': 'scheduled',
                    'interview': 'interviewed',
                    'interviewed': 'interviewed',
                    'hired': 'hired',
                    'rejected': 'rejected',
                    # Legacy mappings for backward compatibility
                    'applied': 'pending',
                    'under_review': 'pending',
                    'reviewed': 'pending',
                    'shortlisted': 'pending',
                    'accepted': 'hired',  # Map old 'accepted' to new 'hired'
                }
                new_status = CANONICAL_STATUS_MAP.get(new_status, new_status)
                
                if application_ids and new_status:
                    # AUTOMATIC: Update status and notify applicants for each application
                    updated = 0
                    branch_id = get_branch_scope(user)  # Get user's assigned branch
                    
                    for app_id in application_ids:
                        # SECURITY: Verify application belongs to user's branch (if HR with assigned branch)
                        if user.get('role') == 'hr' and branch_id:
                            cursor.execute(
                                '''
                                SELECT a.application_id
                                FROM applications a
                                JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s AND j.branch_id = %s
                                ''',
                                (app_id, branch_id),
                            )
                        else:
                            cursor.execute(
                                '''
                                SELECT a.application_id
                                FROM applications a
                                WHERE a.application_id = %s
                                ''',
                                (app_id,),
                            )
                        if not cursor.fetchone():
                            continue
                        if auto_update_application_status(cursor, app_id, new_status):
                            updated += 1
                    
                    # Create admin notification for bulk update
                    try:
                        if user.get('role') == 'hr':
                            hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                            status_display = new_status.replace('_', ' ').title()
                            admin_msg = f'HR {hr_name} bulk updated {updated} candidate(s) to {status_display}.'
                            create_admin_notification(cursor, admin_msg)
                    except Exception as notify_err:
                        print(f'⚠️ Error creating admin notification: {notify_err}')
                    
                    db.commit()
                    # Special message for hired status
                    if new_status.lower() == 'hired':
                        flash(f'Congratulations! {updated} candidate(s) have been marked as HIRED. All applicants have been automatically notified via email and notification.', 'success')
                    else:
                        flash(f'{updated} candidate(s) status updated across the system. Applicants have been automatically notified via email and notification.', 'success')
            
            elif action == 'add_to_talent_pool':
                application_ids = request.form.getlist('application_ids')
                # Talent pool is essentially candidates with status in ['interviewed', 'hired']
                # This is handled by filtering in the view
                flash(f'{len(application_ids)} candidate(s) added to talent pool.', 'success')
            
            return redirect(url_for('applicants'))
        
        # Get filters from request
        status_param = request.args.get('status', '').strip()
        branch_filter = request.args.get('branch_id', type=int)
        filters = {
            'status': status_param.lower() if status_param else '',
            'branch_id': branch_filter,
            'position_id': request.args.get('position_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'source': request.args.get('source', '').strip(),
            'date_from': request.args.get('date_from', '').strip(),
            'date_to': request.args.get('date_to', '').strip(),
            'search': request.args.get('search', '').strip(),
            'view_mode': request.args.get('view_mode', 'list').strip(),
        }
        # Keep status filter even if empty (empty means "All Status")
        # Keep branch_id if it's not None
        # Filter out other empty values but ALWAYS keep status (even if empty string)
        filters = {k: v for k, v in filters.items() if (v or k == 'status') and (v is not None or k != 'branch_id')}
        
        # Debug: Print received status filter
        print(f"🔍 Received status filter from request: '{status_param}' -> processed: '{filters.get('status', '')}'")
        print(f"🔍 Status filter type: {type(filters.get('status'))}, value: '{filters.get('status')}', empty check: {not filters.get('status')}")
        
        # Build WHERE clauses
        where_clauses = []
        params = []

        # Apply filters provided by the request
        if filters.get('status'):
            where_clauses.append('a.status = %s')
            params.append(filters['status'])
        if filters.get('branch_id'):
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])
        elif branch_id:
            # Enforce user's branch scope when not explicitly filtering by branch
            where_clauses.append('j.branch_id = %s')
            params.append(branch_id)
        if filters.get('job_id'):
            where_clauses.append('a.job_id = %s')
            params.append(filters['job_id'])
        # Only apply position_id filter if the jobs table actually has that column
        if filters.get('position_id') and 'position_id' in JOB_COLUMNS:
            where_clauses.append('j.position_id = %s')
            params.append(filters['position_id'])
        if filters.get('source'):
            where_clauses.append('a.source = %s')
            params.append(filters['source'])
        if filters.get('date_from'):
            where_clauses.append('DATE(a.applied_at) >= %s')
            params.append(filters['date_from'])
        if filters.get('date_to'):
            where_clauses.append('DATE(a.applied_at) <= %s')
            params.append(filters['date_to'])
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            where_clauses.append('(ap.full_name LIKE %s OR ap.email LIKE %s OR ap.phone_number LIKE %s)')
            params.extend([search_term, search_term, search_term])

        # Detect whether the applications table has an `archived_at` column
        try:
            cursor.execute('SHOW COLUMNS FROM applications')
            _cols = {row.get('Field') for row in cursor.fetchall() or []}
            has_archived_at = 'archived_at' in _cols
        except Exception:
            has_archived_at = False

        # Exclude archived applicants; if `archived_at` is missing, avoid referencing it
        if has_archived_at:
            where_clauses.append('(a.status != "archived" AND a.archived_at IS NULL)')
        else:
            where_clauses.append('(a.status != "archived")')

        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'

        query = f'''
             SELECT a.application_id,
                 a.applicant_id,
                 ap.full_name AS applicant_name,
                 COALESCE(j.job_title, '') AS job_title,
                 a.status,
                 a.applied_at,
                 (SELECT COUNT(*) FROM interviews i WHERE i.application_id = a.application_id) AS interview_count,
                 (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id) AS last_interview_date,
                 ap.email,
                 ap.phone_number,
                 COALESCE(b.branch_name, '') AS branch_name,
                 a.job_id,
                 j.branch_id,
                 a.resume_id,
                 r.file_name AS resume_file_name,
                 ap.created_at AS applicant_created_at,
                 (SELECT COUNT(*) FROM resumes r2 WHERE r2.applicant_id = ap.applicant_id) AS total_documents
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
             LEFT JOIN jobs j ON a.job_id = j.job_id
             LEFT JOIN resumes r ON a.resume_id = r.resume_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY a.applied_at DESC
        '''
        cursor.execute(query, tuple(params) if params else None)
        applications = cursor.fetchall() or []
        
        # Debug: Check status distribution in results
        status_filter_value = filters.get('status', '').strip().lower() if filters else ''
        db_status = status_filter_value
        if status_filter_value and applications:
            status_counts = {}
            for app in applications:
                status = app.get('status', 'unknown')
                status_counts[status] = status_counts.get(status, 0) + 1
            print(f"🔍 Results status distribution: {status_counts}")
            print(f"🔍 Expected status: '{db_status}', Found statuses: {list(status_counts.keys())}")
            if db_status not in status_counts or status_counts.get(db_status, 0) != len(applications):
                print(f"⚠️ WARNING: Filter mismatch! Expected all '{db_status}', but found: {status_counts}")
        
        # Calculate quick stats (without status filter for accurate counts)
        stats_where_clauses = []
        stats_params = []
        if branch_id:
            stats_where_clauses.append('j.branch_id = %s')
            stats_params.append(branch_id)
        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     stats_where_clauses.append('j.position_id = %s')
        #     stats_params.append(filters['position_id'])
        if filters.get('date_from'):
            stats_where_clauses.append('DATE(a.applied_at) >= %s')
            stats_params.append(filters['date_from'])
        if filters.get('date_to'):
            stats_where_clauses.append('DATE(a.applied_at) <= %s')
            stats_params.append(filters['date_to'])
        if filters.get('search'):
            search_term = f"%{filters['search']}%"
            stats_where_clauses.append('(ap.full_name LIKE %s OR ap.email LIKE %s OR ap.phone_number LIKE %s)')
            stats_params.extend([search_term, search_term, search_term])
        
        # EXCLUDE archived applicants from stats (respect schema)
        if has_archived_at:
            stats_where_clauses.append('(a.status != "archived" AND a.archived_at IS NULL)')
        else:
            stats_where_clauses.append('(a.status != "archived")')
        
        stats_where_sql = ' AND '.join(stats_where_clauses) if stats_where_clauses else '1=1'
        
        cursor.execute(
            f'''
            SELECT 
                COUNT(*) AS total_candidates,
                COALESCE(SUM(CASE WHEN DATE(a.applied_at) = CURDATE() THEN 1 ELSE 0 END), 0) AS new_today,
                COALESCE(SUM(CASE WHEN a.status = 'reviewed' THEN 1 ELSE 0 END), 0) AS in_review,
                COALESCE(SUM(CASE WHEN a.status = 'interviewed' THEN 1 ELSE 0 END), 0) AS interview_stage
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE {stats_where_sql}
            ''',
            tuple(stats_params) if stats_params else None,
        )
        stats_row = cursor.fetchone() or {}
        stats = {
            'total_candidates': stats_row.get('total_candidates', 0) or 0,
            'new_today': stats_row.get('new_today', 0) or 0,
            'in_review': stats_row.get('in_review', 0) or 0,
            'interview_stage': stats_row.get('interview_stage', 0) or 0,
        }
        
        # Get unique positions for filter
        # Positions table removed - return empty list
        positions = []
        
        # Get all branches for filter dropdown
        branches = fetch_branches()
        
        # Get unique jobs for filter dropdown
        job_params = []
        # Build job dropdown query while respecting archived_at schema
        job_query = '''
            SELECT DISTINCT j.job_id, j.job_title AS job_title
            FROM jobs j
            JOIN applications a ON a.job_id = j.job_id
            WHERE ''' + ('(a.status != "archived" AND a.archived_at IS NULL)' if has_archived_at else '(a.status != "archived")') + "\n"
        if filters.get('branch_id'):
            job_query += ' AND j.branch_id = %s'
            job_params.append(filters['branch_id'])
        elif branch_id:
            job_query += ' AND j.branch_id = %s'
            job_params.append(branch_id)
        job_query += ' ORDER BY j.job_title'
        cursor.execute(job_query, tuple(job_params) if job_params else None)
        jobs = cursor.fetchall() or []
        
        # Format applications data
        formatted_applications = []
        for app in applications:
            try:
                interview_count = app.get('interview_count', 0) or 0
                status_value = (app.get('status') or 'pending').strip().lower()
                # Normalize withdrawn to rejected - remove withdrawn status completely
                if status_value == 'withdrawn':
                    status_value = 'rejected'
                # Note: Status is already managed by interview scheduling/completion logic
                # 'scheduled' when interview is scheduled, 'interviewed' when interview is completed
                # So we don't auto-override status based on interview_count
                formatted_applications.append({
                    'application_id': app.get('application_id'),
                    'applicant_id': app.get('applicant_id'),
                    'applicant_name': app.get('applicant_name') or '—',
                    'applicant_email': app.get('applicant_email') or '—',
                    'applicant_phone': app.get('applicant_phone') or '—',
                    'email_verified': app.get('email_verified', False),
                    'position_name': app.get('position_name') or 'Unassigned',
                    'position_id': app.get('position_id'),
                    'job_title': app.get('job_title') or '—',
                    'job_id': app.get('job_id'),
                    'branch_id': app.get('branch_id'),  # Add branch_id for filtering
                    'branch_name': app.get('branch_name') or 'Unassigned',
                    'status': status_value,  # Use normalized status (withdrawn -> rejected)
                    'applied_at': format_human_datetime(app.get('applied_at')) or '—',
                    'submitted_at': format_human_datetime(app.get('applied_at')) or '—',
                    'submitted_at_raw': app.get('applied_at'),
                    'resume_id': app.get('resume_id'),
                    'resume_file_name': app.get('resume_file_name'),
                    # Mark has_resume true when either the application links a resume
                    # or the applicant has uploaded documents (total_documents > 0).
                    'has_resume': (app.get('resume_id') is not None) or (app.get('total_documents', 0) or 0) > 0,
                    'total_documents': app.get('total_documents', 0) or 0,
                    'interview_count': app.get('interview_count', 0),
                    'last_interview_date': format_human_datetime(app.get('last_interview_date')) if app.get('last_interview_date') else None,
                    'applicant_created_at': format_human_datetime(app.get('applicant_created_at')),
                })
            except Exception as format_exc:
                print(f'⚠️ Error formatting application {app.get("application_id")}: {format_exc}')
                continue
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id_session = session.get('branch_id')
            if branch_id_session:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id_session,))
                if branch_rows:
                    branch_info = branch_rows[0]
            # If no session-based branch_info, try deriving from the user's assigned branch
            if branch_info is None:
                branch_scope = get_branch_scope(user)
                if branch_scope is not None:
                    branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_scope,))
                    if branch_rows:
                        branch_info = branch_rows[0]
        
        return render_template(
            template,
            applications=formatted_applications or [],
            branch_info=branch_info,
            dashboard_data={'branch_info': branch_info},
            stats=stats,
            filters=filters,
            positions=positions,
            jobs=jobs,
            branches=branches,
            view_mode=filters.get('view_mode', 'list'),
        )
    except Exception as exc:
        db.rollback()
        print(f'❌ Applicants tracking error: {exc}')
        import traceback
        traceback.print_exc()
        flash('An error occurred while loading candidates.', 'error')
        template = 'hr/applicants.html' if user.get('role') == 'hr' else 'admin/applicants.html'
        return render_template(template, applications=[], branch_info=None, dashboard_data={}, stats={}, filters={}, positions=[], jobs=[], branches=[], view_mode='list')
    finally:
        cursor.close()


@app.route('/admin/applications', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def applications():
    """Applications management - Disabled."""
    user = get_current_user()
    # Applications management
    # Note: previously this endpoint was temporarily disabled by redirecting users.
    # Removing the redirect so admins and HR can access the applications management UI again.
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/applications.html', applications=[], analytics={}, branches=[], jobs=[], current_filters={}, user=user or {})
    
    cursor = db.cursor(dictionary=True)
    
    try:
        allowed_statuses = set(APPLICATION_STATUSES)
        
        # Capture current filters from request (for preserving after POST)
        current_filters = {
            'keyword': request.args.get('keyword', '').strip(),
            'branch_id': request.args.get('branch_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'status': request.args.get('status', '').strip(),
        }

        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_status':
                # Restrict admin from changing status - only HR can change status
                if user.get('role') == 'admin':
                    flash('Access denied. Only HR users can change application status.', 'error')
                    params = {}
                    if current_filters.get('keyword'):
                        params['keyword'] = current_filters['keyword']
                    if current_filters.get('branch_id'):
                        params['branch_id'] = current_filters['branch_id']
                    if current_filters.get('job_id'):
                        params['job_id'] = current_filters['job_id']
                    if current_filters.get('status'):
                        params['status'] = current_filters['status']
                    return redirect(url_for('applications', **params))
                
                application_id = request.form.get('application_id')
                new_status_raw = request.form.get('status', '').strip()
                
                # Build redirect URL with preserved filters
                def build_redirect_url():
                    params = {}
                    if current_filters.get('keyword'):
                        params['keyword'] = current_filters['keyword']
                    if current_filters.get('branch_id'):
                        params['branch_id'] = current_filters['branch_id']
                    if current_filters.get('job_id'):
                        params['job_id'] = current_filters['job_id']
                    if current_filters.get('status'):
                        params['status'] = current_filters['status']
                    return url_for('applications', **params)
                
                # Validate application_id first
                if not application_id:
                    flash('Application ID is required.', 'error')
                    return redirect(build_redirect_url())
                
                # Validate status is provided
                if not new_status_raw:
                    error_msg = 'Status is required. Please select a status from the dropdown.'
                    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'error': error_msg}), 400
                    flash(error_msg, 'error')
                    return redirect(build_redirect_url())
                
                # Map simplified statuses to database statuses
                # ALL statuses can now be manually changed by HR: pending, scheduled, interviewed, hired, rejected
                CANONICAL_STATUS_MAP = {
                    'pending': 'pending',
                    'scheduled': 'scheduled',
                    'interview': 'interviewed',
                    'interviewed': 'interviewed',
                    'hired': 'hired',
                    'rejected': 'rejected',
                    # Legacy mappings for backward compatibility
                    'applied': 'pending',
                    'under_review': 'pending',
                    'reviewed': 'pending',
                    'shortlisted': 'pending',
                    'accepted': 'hired',  # Map old 'accepted' to new 'hired'
                }
                new_status = CANONICAL_STATUS_MAP.get(new_status_raw.lower(), new_status_raw.lower())
                
                # Validate mapped status is not empty
                if not new_status:
                    error_msg = f'Invalid status value: "{new_status_raw}". Please select a valid status.'
                    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'error': error_msg}), 400
                    flash(error_msg, 'error')
                    return redirect(build_redirect_url())
                
                if new_status not in allowed_statuses:
                    error_msg = f'Invalid status: {new_status}. Allowed statuses: {", ".join(allowed_statuses)}'
                    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                        return jsonify({'success': False, 'error': error_msg}), 400
                    flash(error_msg, 'error')
                    return redirect(build_redirect_url())
                
                if application_id and new_status in allowed_statuses:
                    # Verify application belongs to HR's branch (if HR is branch-scoped)
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT a.application_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id = %s AND j.branch_id = %s
                            ''',
                            (application_id, branch_id),
                        )
                        if not cursor.fetchone():
                            error_msg = 'You can only update applications for your branch.'
                            # Check if this is an AJAX request
                            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                                return jsonify({'success': False, 'error': error_msg}), 403
                            flash(error_msg, 'error')
                            # Build redirect URL with preserved filters
                            params = {}
                            if current_filters.get('keyword'):
                                params['keyword'] = current_filters['keyword']
                            if current_filters.get('branch_id'):
                                params['branch_id'] = current_filters['branch_id']
                            if current_filters.get('job_id'):
                                params['job_id'] = current_filters['job_id']
                            if current_filters.get('status'):
                                params['status'] = current_filters['status']
                            return redirect(url_for('applications', **params))
                    # Get applicant email before updating status
                    # Use job_column_expr to handle both job_title and title columns
                    job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
                    cursor.execute(
                        f'''
                        SELECT ap.email, ap.full_name, a.status AS old_status, COALESCE({job_title_expr}, 'Untitled Job') AS job_title
                        FROM applicants ap
                        JOIN applications a ON ap.applicant_id = a.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        WHERE a.application_id = %s
                        LIMIT 1
                        ''',
                        (application_id,)
                    )
                    applicant_info = cursor.fetchone()
                    
                    if not applicant_info:
                        error_msg = 'Application not found or applicant information is missing.'
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            return jsonify({'success': False, 'error': error_msg}), 404
                        flash(error_msg, 'error')
                        return redirect(build_redirect_url())
                    
                    old_status = applicant_info.get('old_status') if applicant_info else None
                    
                    # AUTOMATIC: Update status and notify applicant
                    try:
                        update_success = auto_update_application_status(cursor, application_id, new_status)
                    except Exception as update_err:
                        db.rollback()
                        error_msg = f'Failed to update application status: {str(update_err)}'
                        print(f'❌ Error updating application status: {update_err}')
                        import traceback
                        traceback.print_exc()
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            return jsonify({'success': False, 'error': error_msg}), 500
                        flash(error_msg, 'error')
                        return redirect(build_redirect_url())
                    
                    if update_success:
                        # If HR performed the action, add Admin system notification
                        try:
                            if user.get('role') == 'hr':
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                status_display = new_status.replace('_', ' ').title()
                                applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                admin_msg = f'HR {hr_name} updated application status to {status_display} for {applicant_name} ({job_title}).'
                                create_admin_notification(cursor, admin_msg)
                                print(f'✅ Admin notification created: {admin_msg}')
                        except Exception as notify_err:
                            print(f'⚠️ Error creating admin notification: {notify_err}')
                        
                        # Commit the transaction
                        db.commit()
                        print(f'✅ Application {application_id} status updated to "{new_status}" - transaction committed')
                        
                        # Check if this is an AJAX request
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            # Return JSON response for AJAX requests
                            status_display = new_status.replace('_', ' ').title()
                            if new_status.lower() == 'hired':
                                message = f'Congratulations! {applicant_info.get("full_name", "Applicant")} has been marked as HIRED. They have been automatically notified via email and notification.'
                            else:
                                message = 'Application status updated successfully. Applicant has been automatically notified via email and notification.'
                            return jsonify({
                                'success': True,
                                'message': message,
                                'status': new_status,
                                'status_display': status_display,
                                'application_id': application_id
                            })
                        
                        # Special success message for hired status (for regular form submissions)
                        if new_status.lower() == 'hired':
                            flash(f'Congratulations! {applicant_info.get("full_name", "Applicant")} has been marked as HIRED. They have been automatically notified via email and notification.', 'success')
                        else:
                            flash('Application status updated successfully. Applicant has been automatically notified via email and notification.', 'success')
                    else:
                        # Fallback to manual update if auto function fails
                        # Fallback manual update - guard against missing `updated_at` column
                        cursor.execute("SHOW COLUMNS FROM applications LIKE 'updated_at'")
                        _col = cursor.fetchone()
                        if _col:
                            cursor.execute(
                                'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s',
                                (new_status, application_id),
                            )
                        else:
                            cursor.execute(
                                'UPDATE applications SET status = %s WHERE application_id = %s',
                                (new_status, application_id),
                            )
                        db.commit()  # Ensure commit happens even in fallback

                        # If hired, try to sync interviews (non-blocking) and commit again
                        try:
                            if str(new_status).lower() == 'hired':
                                _sync_interviews_on_application_hired(cursor, application_id)
                                db.commit()
                        except Exception as _sync_err:
                            print(f'⚠️ Error syncing interviews after fallback hire update: {_sync_err}')
                        
                        # Check if this is an AJAX request
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            status_display = new_status.replace('_', ' ').title()
                            return jsonify({
                                'success': True,
                                'message': f'Application status updated to {status_display}.',
                                'status': new_status,
                                'status_display': status_display,
                                'application_id': application_id,
                                'note': 'Status updated but notification may not have been sent.'
                            })
                        # Still try to create notification even if auto function failed
                        try:
                            cursor.execute(
                                '''
                                SELECT ap.applicant_id, ap.email, ap.full_name, j.job_title AS job_title
                                FROM applicants ap
                                JOIN applications a ON ap.applicant_id = a.applicant_id
                                LEFT JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s
                                LIMIT 1
                                ''',
                                (application_id,)
                            )
                            applicant_info = cursor.fetchone()
                            if applicant_info:
                                status_display = new_status.replace('_', ' ').title()
                                job_title = applicant_info.get('job_title') or 'Your Application'
                                message = f'Your application status for "{job_title}" has been updated to: {status_display}'
                                
                                # Create notification - this notification goes to the APPLICANT (not HR)
                                # The notification is linked to the application, which is associated with the applicant
                                notification_columns = set()
                                try:
                                    cursor.execute('SHOW COLUMNS FROM notifications')
                                    notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                                except Exception:
                                    pass
                                
                                if 'sent_at' in notification_columns:
                                    cursor.execute(
                                        'INSERT INTO notifications (application_id, message, sent_at, is_read) VALUES (%s, %s, NOW(), 0)',
                                        (application_id, message)
                                    )
                                else:
                                    cursor.execute(
                                        'INSERT INTO notifications (application_id, message, is_read) VALUES (%s, %s, 0)',
                                        (application_id, message)
                                    )
                                
                                # Send email
                                try:
                                    email_subject = f'Application Status Update - {applicant_info.get("job_title") or "Your Application"}'
                                    email_body = f"""Dear {applicant_info.get('full_name') or 'Applicant'},

Your application status for the position "{applicant_info.get('job_title') or 'the position'}" has been updated.

New Status: {status_display}

Please log in to your account to view more details.

Best regards,
J&T Express Recruitment Team
                                    """.strip()
                                    send_email(applicant_info.get('email'), email_subject, email_body)
                                except Exception as email_err:
                                    print(f"⚠️ Email error (non-blocking): {email_err}")
                        except Exception as notify_err:
                            print(f"⚠️ Notification creation error (non-blocking): {notify_err}")
                        
                        # If HR performed the action, add Admin system notification
                        try:
                            if user.get('role') == 'hr':
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                status_display = new_status.replace('_', ' ').title()
                                applicant_name = applicant_info.get('full_name') if applicant_info else 'applicant'
                                job_title = applicant_info.get('job_title') if applicant_info else 'position'
                                admin_msg = f'HR {hr_name} updated application status to {status_display} for {applicant_name} ({job_title}).'
                                create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating admin notification: {notify_err}')
                        db.commit()
                        
                        # Check if this is an AJAX request before redirecting
                        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                            status_display = new_status.replace('_', ' ').title()
                            return jsonify({
                                'success': True,
                                'message': 'Application status updated successfully. Notification sent to applicant.',
                                'status': new_status,
                                'status_display': status_display,
                                'application_id': application_id
                            })
                        
                        flash('Application status updated successfully. Notification sent to applicant.', 'success')
                
                # Build redirect URL with preserved filters (only for non-AJAX requests)
                params = {}
                if current_filters.get('keyword'):
                    params['keyword'] = current_filters['keyword']
                if current_filters.get('branch_id'):
                    params['branch_id'] = current_filters['branch_id']
                if current_filters.get('job_id'):
                    params['job_id'] = current_filters['job_id']
                if current_filters.get('status'):
                    params['status'] = current_filters['status']
                return redirect(url_for('applications', **params))
            
            elif action == 'bulk_update_status':
                # Restrict admin from changing status - only HR can change status
                if user.get('role') == 'admin':
                    flash('Access denied. Only HR users can change application status.', 'error')
                    params = {}
                    if current_filters.get('keyword'):
                        params['keyword'] = current_filters['keyword']
                    if current_filters.get('branch_id'):
                        params['branch_id'] = current_filters['branch_id']
                    if current_filters.get('job_id'):
                        params['job_id'] = current_filters['job_id']
                    if current_filters.get('status'):
                        params['status'] = current_filters['status']
                    return redirect(url_for('applications', **params))
                
                application_ids = request.form.getlist('application_ids')
                bulk_status = request.form.get('bulk_status', '').strip()
                
                # Map simplified statuses to database statuses
                CANONICAL_STATUS_MAP = {
                    'pending': 'pending',
                    'interview': 'interviewed',
                    'interviewed': 'interviewed',
                    'hired': 'hired',
                    'rejected': 'rejected',
                    'applied': 'pending',
                    'under_review': 'pending',
                    'reviewed': 'pending',
                    'shortlisted': 'pending',
                    'accepted': 'hired',
                }
                bulk_status = CANONICAL_STATUS_MAP.get(bulk_status, bulk_status)
                
                if application_ids and bulk_status in allowed_statuses:
                    # Verify all applications belong to HR's branch
                    branch_id = get_branch_scope(user)
                    valid_ids = []
                    if branch_id:
                        placeholders = ','.join(['%s'] * len(application_ids))
                        cursor.execute(
                            f'''
                            SELECT a.application_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id IN ({placeholders}) AND j.branch_id = %s
                            ''',
                            (*application_ids, branch_id),
                        )
                        valid_ids = [row['application_id'] for row in cursor.fetchall()]
                        if len(valid_ids) != len(application_ids):
                            flash('Some applications do not belong to your branch. Only valid applications were updated.', 'warning')
                    else:
                        valid_ids = application_ids
                    
                    if valid_ids:
                        # AUTOMATIC: Update status and notify applicants for each application
                        updated_count = 0
                        for app_id in valid_ids:
                            if auto_update_application_status(cursor, app_id, bulk_status):
                                updated_count += 1
                        
                        # Create admin notification for bulk update
                        try:
                            if user.get('role') == 'hr':
                                hr_name = user.get('full_name') or user.get('name') or 'HR Staff'
                                status_display = bulk_status.replace('_', ' ').title()
                                admin_msg = f'HR {hr_name} bulk updated {updated_count} application(s) to {status_display}.'
                                create_admin_notification(cursor, admin_msg)
                        except Exception as notify_err:
                            print(f'⚠️ Error creating admin notification: {notify_err}')
                        
                        db.commit()
                        flash(f'{updated_count} application(s) updated successfully. Applicants have been automatically notified via email and notification.', 'success')
            
            # Build redirect URL with preserved filters for bulk update
            params = {}
            if current_filters.get('keyword'):
                params['keyword'] = current_filters['keyword']
            if current_filters.get('branch_id'):
                params['branch_id'] = current_filters['branch_id']
            if current_filters.get('job_id'):
                params['job_id'] = current_filters['job_id']
            if current_filters.get('status'):
                params['status'] = current_filters['status']
            return redirect(url_for('applications', **params))
        
        # Apply filters
        filters = {
            'keyword': request.args.get('keyword', '').strip(),
            'branch_id': request.args.get('branch_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'status': request.args.get('status', '').strip(),
            'date_from': request.args.get('date_from', '').strip(),
            'date_to': request.args.get('date_to', '').strip(),
        }
        # Keep keyword even if empty string to track search attempts
        # But only add to filters dict if it has a value
        keyword_value = filters.pop('keyword', '')
        filters = {k: v for k, v in filters.items() if v}
        if keyword_value:
            filters['keyword'] = keyword_value
        
        branch_id = get_branch_scope(user)
        where_clauses = []
        params = []
        
        if branch_id:
            where_clauses.append('j.branch_id = %s')
            params.append(branch_id)
        
        if filters.get('keyword'):
            keyword = filters['keyword'].strip()
            if keyword:
                keyword_pattern = f"%{keyword}%"
                # Search in applicant name, email, and job title - only for applicants who have applications
                # Use TRIM and LOWER() for case-insensitive search and to handle whitespace
                # Use COALESCE to handle NULL job titles
                where_clauses.append('(LOWER(TRIM(ap.full_name)) LIKE LOWER(%s) OR LOWER(TRIM(ap.email)) LIKE LOWER(%s) OR LOWER(TRIM(COALESCE(j.job_title, \'\'))) LIKE LOWER(%s))')
                params.extend([keyword_pattern, keyword_pattern, keyword_pattern])
        
        if filters.get('branch_id'):
            where_clauses.append('j.branch_id = %s')
            params.append(filters['branch_id'])
        
        if filters.get('job_id'):
            where_clauses.append('a.job_id = %s')
            params.append(filters['job_id'])
        
        if filters.get('status'):
            # Normalize status filter - map display statuses to database statuses
            status_filter = filters['status'].strip().lower()
            status_map = {
                'pending': ['pending', 'reviewed', 'applied', 'under_review', 'shortlisted'],
                'scheduled': ['scheduled'],
                'interviewed': ['interviewed', 'interview'],
                'hired': ['hired', 'accepted'],
                'rejected': ['rejected'],
            }
            db_statuses = status_map.get(status_filter, [status_filter])
            # Only apply filter if it's a valid status
            if db_statuses:
                placeholders = ','.join(['%s'] * len(db_statuses))
                where_clauses.append(f'a.status IN ({placeholders})')
                params.extend(db_statuses)
                print(f"🔍 Status filter applied: '{status_filter}' -> WHERE a.status IN {db_statuses}")
            else:
                print(f"⚠️ Invalid status filter: '{status_filter}' - not in status_map")
        
        if filters.get('date_from'):
            where_clauses.append('DATE(a.applied_at) >= %s')
            params.append(filters['date_from'])
        
        if filters.get('date_to'):
            where_clauses.append('DATE(a.applied_at) <= %s')
            params.append(filters['date_to'])
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Get dynamic column expressions for job title
        job_title_expr = job_column_expr('job_title', alternatives=['title'], default="'Untitled Job'")
        
        # Check if position_name column exists in jobs table
        cursor.execute('SHOW COLUMNS FROM jobs LIKE "position_name"')
        has_position_name_col = cursor.fetchone() is not None
        
        # Build position_title expression conditionally
        if has_position_name_col:
            position_title_expr = f'COALESCE(j.position_name, {job_title_expr})'
        else:
            position_title_expr = job_title_expr
        
        # Fetch applications
        # IMPORTANT: 
        # 1. Query starts with applications table, ensuring only applicants with applications are returned
        # 2. INNER JOIN with applicants ensures applicant data exists
        # 3. HR users only see applications for jobs in their branch (enforced by branch_id filter above)
        # 4. Admin users see all applications from all branches
        # 5. Keyword search only searches within applicants who have submitted applications
        # Use simpler approach: get ANY interview for each application, and use EXISTS to check
        # Normalize withdrawn to rejected in SQL query - remove withdrawn status completely
        cursor.execute(
            f'''
            SELECT a.application_id,
                   CASE 
                       WHEN a.status = 'withdrawn' THEN 'rejected'
                       ELSE a.status
                   END AS status,
                   a.applied_at,
                   ap.applicant_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   ap.phone_number AS applicant_phone,
                   -- Count applicant uploads and pick latest resume for preview/download
                   (SELECT COUNT(*) FROM resumes WHERE applicant_id = ap.applicant_id) AS total_documents,
                   (SELECT resume_id FROM resumes WHERE applicant_id = ap.applicant_id ORDER BY uploaded_at DESC LIMIT 1) AS resume_id,
                   COALESCE((SELECT file_name FROM resumes WHERE applicant_id = ap.applicant_id ORDER BY uploaded_at DESC LIMIT 1), '') AS resume_file_name,
                   COALESCE((SELECT file_path FROM resumes WHERE applicant_id = ap.applicant_id ORDER BY uploaded_at DESC LIMIT 1), '') AS resume_path,
                   j.job_id,
                   COALESCE(j.branch_id, b.branch_id) AS branch_id,
                   {job_title_expr} AS job_title,
                   COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                   {position_title_expr} AS position_title,
                   (SELECT interview_id FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_id,
                   (SELECT scheduled_date FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS scheduled_date,
                   (SELECT interview_mode FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1) AS interview_mode,
                   (IF(
                       (SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'interviews' AND COLUMN_NAME = 'location') > 0,
                       (SELECT location FROM interviews WHERE application_id = a.application_id ORDER BY scheduled_date DESC LIMIT 1),
                       NULL
                   )) AS interview_location
            FROM applications a
            INNER JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY a.applied_at DESC
            ''',
            tuple(params) if params else None,
        )
        applications = cursor.fetchall()
        
        # Calculate analytics - use correct status values from database enum (pending, scheduled, interviewed, hired, rejected)
        # Normalize withdrawn to rejected in SQL query - remove withdrawn status completely
        # Use DISTINCT to prevent duplicates from LEFT JOIN with interviews
        cursor.execute(
            f'''
            SELECT 
                COUNT(DISTINCT a.application_id) AS total,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'pending' THEN a.application_id END) AS pending,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'scheduled' THEN a.application_id END) AS scheduled,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'interviewed' THEN a.application_id END) AS interviewed,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'hired' THEN a.application_id END) AS hired,
                COUNT(DISTINCT CASE WHEN CASE WHEN a.status = 'withdrawn' THEN 'rejected' ELSE a.status END = 'rejected' THEN a.application_id END) AS rejected,
                COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                COUNT(DISTINCT CASE WHEN DATE(a.applied_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN a.application_id END) AS this_month,
                COUNT(DISTINCT CASE WHEN DATE(a.applied_at) >= DATE_SUB(CURDATE(), INTERVAL 7 DAY) THEN a.application_id END) AS this_week
            FROM applications a
            INNER JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN interviews i ON i.application_id = a.application_id
            WHERE {where_sql}
            ''',
            tuple(params) if params else None,
        )
        analytics = cursor.fetchone()
        
        # Note: We no longer auto-update status based on interview existence
        # Status flow: pending -> scheduled (when interview scheduled) -> interviewed (when interview completed) -> hired/rejected
        # Applications with 'scheduled' status should remain 'scheduled' until interview is marked as completed
        
        # Format applications
        formatted_apps = []
        status_filter_value = filters.get('status', '').strip().lower() if filters else ''
        for app in applications:
            # Get the original status from database - CRITICAL: preserve rejected/hired status
            app_status = app.get('status')
            if app_status:
                app_status = app_status.lower().strip()
            else:
                app_status = 'pending'
            
            # Normalize withdrawn to rejected - remove withdrawn status completely
            if app_status == 'withdrawn':
                app_status = 'rejected'
            
            application_id = app.get('application_id')
            
            # CRITICAL: Never override final statuses (hired, rejected) - preserve them exactly as stored
            # Status is already set to 'scheduled' when interview is scheduled, and 'interviewed' when interview is completed
            # So we don't need to auto-override based on has_interview anymore
            has_interview = app.get('interview_id') is not None
            
            # If status is rejected or hired, NEVER override it, regardless of interviews
            if app_status in ('rejected', 'hired'):
                # Preserve the original status - do not override
                pass
            # Note: Status is managed by interview scheduling/completion logic, so we preserve it as-is
            # For any other status, keep the original status
            
            formatted_apps.append({
                'application_id': app.get('application_id'),
                'applicant_id': app.get('applicant_id'),
                'applicant_name': app.get('applicant_name'),
                'applicant_email': app.get('applicant_email'),
                'applicant_phone': app.get('applicant_phone'),
                'job_id': app.get('job_id'),
                'job_title': app.get('job_title'),
                'branch_id': app.get('branch_id'),
                'branch_name': app.get('branch_name'),
                'position_title': app.get('position_title'),
                'status': app_status,  # Use normalized status (withdrawn -> rejected)
                'applied_at': format_human_datetime(app.get('applied_at')),
                'submitted_at': format_human_datetime(app.get('applied_at')),
                'resume_id': app.get('resume_id'),
                'has_resume': (app.get('resume_id') is not None) or (app.get('total_documents', 0) or 0) > 0,
                'has_interview': has_interview,
                'interview_date': format_human_datetime(app.get('scheduled_date')) if app.get('scheduled_date') else None,
                'interview_mode': app.get('interview_mode'),
                'interview_location': app.get('interview_location'),
            })
        
        branches = fetch_branches()
        jobs = fetch_jobs_for_user(user)
        
        # Check if JSON format is requested (for AJAX/modals)
        if request.args.get('format') == 'json' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            modal_type = request.args.get('modal', '').strip().lower()
            status_filter = request.args.get('status', '').strip()
            
            # Handle modal-specific requests
            if modal_type:
                if modal_type == 'total':
                    # Return all applications for total modal
                    return jsonify({
                        'applications': formatted_apps,
                        'analytics': analytics or {},
                        'total': len(formatted_apps)
                    })
                elif modal_type == 'pending':
                    # Return only pending applications
                    pending_apps = [app for app in formatted_apps if app.get('status') == 'pending']
                    return jsonify({
                        'applications': pending_apps,
                        'analytics': analytics or {},
                        'total': len(pending_apps)
                    })
                elif modal_type == 'hired':
                    # Return only hired applications
                    hired_apps = [app for app in formatted_apps if app.get('status') == 'hired']
                    return jsonify({
                        'applications': hired_apps,
                        'analytics': analytics or {},
                        'total': len(hired_apps)
                    })
                elif modal_type == 'rejected':
                    # Return only rejected applications
                    rejected_apps = [app for app in formatted_apps if app.get('status') == 'rejected']
                    return jsonify({
                        'applications': rejected_apps,
                        'analytics': analytics or {},
                        'total': len(rejected_apps)
                    })
                elif modal_type == 'interviewed':
                    # Return only interviewed applications
                    interviewed_apps = [app for app in formatted_apps if app.get('status') == 'interviewed']
                    return jsonify({
                        'applications': interviewed_apps,
                        'analytics': analytics or {},
                        'total': len(interviewed_apps)
                    })
                elif modal_type == 'scheduled':
                    # Return only scheduled applications
                    scheduled_apps = [app for app in formatted_apps if app.get('status') == 'scheduled']
                    return jsonify({
                        'applications': scheduled_apps,
                        'analytics': analytics or {},
                        'total': len(scheduled_apps)
                    })
            
            # Default: return filtered applications by status
            print(f"📊 JSON response for status='{status_filter}': {len(formatted_apps)} applications returned")
            if status_filter == 'rejected':
                rejected_apps = [app for app in formatted_apps if app.get('status') == 'rejected']
                print(f"📊 Rejected applications count: {len(rejected_apps)} out of {len(formatted_apps)} total")
                print(f"📊 Rejected application IDs: {[app.get('application_id') for app in rejected_apps]}")
            return jsonify({
                'applications': formatted_apps,
                'analytics': analytics or {},
                'total': len(formatted_apps)
            })
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/applications.html' if user.get('role') == 'hr' else 'admin/applications.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id = session.get('branch_id')
            if branch_id:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
        return render_template(
            template,
            applications=formatted_apps,
            analytics=analytics or {},
            branches=branches,
            jobs=jobs,
            current_filters=filters,
            user=user,
            branch_info=branch_info,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Applications management error: {exc}')
        print(f'Full traceback: {error_details}')
        
        # Check if this is an AJAX request
        if request.method == 'POST' and (request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json):
            return jsonify({
                'success': False,
                'error': f'An error occurred: {str(exc)}. Please check the console for details.'
            }), 500
        
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        template = 'hr/applications.html' if (user or {}).get('role') == 'hr' else 'admin/applications.html'
        return render_template(template, applications=[], analytics={}, branches=[], jobs=[], current_filters={}, user=user or {}, branch_info=None)
    finally:
        cursor.close()


@app.route('/admin/interviews/get-jobs', methods=['GET'])
@login_required('admin', 'hr')
def get_applicant_jobs():
    """Get jobs for a specific applicant."""
    user = get_current_user()
    applicant_id = request.args.get('applicant_id', type=int)
    
    if not applicant_id:
        return jsonify({'jobs': []})
    
    db = get_db()
    if not db:
        return jsonify({'jobs': []})
    
    cursor = db.cursor(dictionary=True)
    try:
        where_clause = 'a.applicant_id = %s'
        params = [applicant_id]
        branch_id = get_branch_scope(user)
        if branch_id:
            where_clause += ' AND j.branch_id = %s'
            params.append(branch_id)
        
        cursor.execute(
            f'''
            SELECT DISTINCT j.job_id, j.job_title, a.status AS application_status
            FROM applications a
            JOIN jobs j ON a.job_id = j.job_id
            WHERE {where_clause}
            ORDER BY j.job_title ASC
            ''',
            tuple(params),
        )
        jobs = cursor.fetchall()
        return jsonify({'jobs': jobs})
    except Exception as exc:
        print(f'❌ Get applicant jobs error: {exc}')
        return jsonify({'jobs': []})
    finally:
        cursor.close()


def _render_interviews():
    user = get_current_user()
    target_endpoint = 'admin_interviews' if user.get('role') == 'admin' else 'hr_interviews'
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        template = 'admin/interviews.html' if user.get('role') == 'admin' else 'hr/interviews.html'
        return render_template(template, user=user, interviews=[], upcoming=[], past=[], applications=[], applicants_for_schedule=[], current_filters={})
    
    cursor = db.cursor(dictionary=True)
    schedule_applications = []

    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            # Block scheduling for admins - they can only view
            if action == 'schedule' and user.get('role') == 'admin':
                flash('Admins can only view interviews. Scheduling is restricted to HR staff.', 'error')
                return redirect(url_for(target_endpoint))
            
            if action == 'schedule':
                application_id = request.form.get('application_id', '').strip()
                applicant_id = request.form.get('applicant_id', '').strip()
                job_id = request.form.get('job_id', '').strip()
                scheduled_date = request.form.get('scheduled_date', '').strip()
                scheduled_time = request.form.get('scheduled_time', '').strip()
                interview_mode = (request.form.get('interview_mode') or 'in_person').strip()
                location = (request.form.get('location') or '').strip()
                notes = (request.form.get('notes') or '').strip()
                
                print(f'🔍 Interview scheduling POST data: application_id={application_id}, applicant_id={applicant_id}, scheduled_date={scheduled_date}, scheduled_time={scheduled_time}')
                
                # Basic input validation
                if not scheduled_date or not scheduled_time:
                    flash('Interview date and time are required.', 'error')
                    return redirect(url_for(target_endpoint))
                elif not application_id and not applicant_id:
                    flash('Please select an applicant.', 'error')
                    return redirect(url_for(target_endpoint))
                else:
                    # Validate combined datetime is parseable and not in the past (grace: allow 5-minute window)
                    try:
                        scheduled_dt = datetime.strptime(f"{scheduled_date} {scheduled_time}", "%Y-%m-%d %H:%M")
                        # Allow 5-minute grace period for scheduling (accounting for seconds precision)
                        grace_period = timedelta(minutes=5)
                        if scheduled_dt < datetime.now() - grace_period:
                            flash('Interview time must be in the future.', 'error')
                            return redirect(url_for(target_endpoint))
                    except Exception:
                        flash('Invalid date/time format.', 'error')
                        return redirect(url_for(target_endpoint))
                    # Normalize notes length
                    if len(notes) > 1000:
                        notes = notes[:1000]
                    if len(location) > 255:
                        location = location[:255]
                    # Attempt to resolve the target application via application_id first
                    # For admin users, don't filter by branch - allow scheduling for all branches
                    branch_id = None if user.get('role') == 'admin' else get_branch_scope(user)
                    application = None
                    
                    if application_id:
                        lookup_clause = 'a.application_id = %s'
                        lookup_params = [application_id]
                        if branch_id:
                            lookup_clause += ' AND j.branch_id = %s'
                            lookup_params.append(branch_id)
                        cursor.execute(
                            f'''
                            SELECT a.application_id, a.applicant_id, a.job_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE {lookup_clause}
                            LIMIT 1
                            ''',
                            tuple(lookup_params),
                        )
                        application = cursor.fetchone()
                    
                    # Fallback: attempt to locate by applicant_id only (auto-select most recent eligible application)
                    if not application and applicant_id:
                        # Find the most recent application for this applicant that's eligible for interview
                        # Priority: reviewed > pending > interviewed (using actual database enum values)
                        # Exclude rejected and withdrawn applications - cannot schedule interview for rejected/withdrawn
                        lookup_clause = 'a.applicant_id = %s AND a.status NOT IN (%s, %s)'
                        lookup_params = [applicant_id, 'rejected', 'withdrawn']
                        if branch_id:
                            lookup_clause += ' AND j.branch_id = %s'
                            lookup_params.append(branch_id)
                        cursor.execute(
                            f'''
                            SELECT a.application_id, a.applicant_id, a.job_id, a.status
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE {lookup_clause}
                            ORDER BY 
                                CASE a.status
                                    WHEN 'reviewed' THEN 1
                                    WHEN 'pending' THEN 2
                                    WHEN 'interviewed' THEN 3
                                    ELSE 5
                                END,
                                a.applied_at DESC
                            LIMIT 1
                            ''',
                            tuple(lookup_params),
                        )
                        application = cursor.fetchone()
                    
                    # Fallback: attempt to locate by applicant + job if job_id was provided (for backward compatibility)
                    if not application and applicant_id and job_id:
                        lookup_clause = 'a.applicant_id = %s AND a.job_id = %s'
                        lookup_params = [applicant_id, job_id]
                        if branch_id:
                            lookup_clause += ' AND j.branch_id = %s'
                            lookup_params.append(branch_id)
                        cursor.execute(
                            f'''
                            SELECT a.application_id, a.applicant_id, a.job_id
                            FROM applications a
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE {lookup_clause}
                            LIMIT 1
                            ''',
                            tuple(lookup_params),
                        )
                        application = cursor.fetchone()
                    
                    if not application:
                        flash('No eligible application found for this applicant. Please ensure the applicant has an application that is not rejected or withdrawn (status must be: pending, reviewed, or interviewed).', 'error')
                        return redirect(url_for(target_endpoint))
                    else:
                        resolved_application_id = application.get('application_id')
                        resolved_applicant_id = application.get('applicant_id')
                        resolved_job_id = application.get('job_id')
                        
                        # Check if application status is 'rejected' or 'withdrawn' - prevent scheduling
                        application_status = (application.get('status') or '').lower()
                        if application_status in ('rejected', 'withdrawn'):
                            status_label = 'rejected' if application_status == 'rejected' else 'withdrawn'
                            flash(f'Cannot schedule interview for a {status_label} application. Please select an applicant with an eligible application (pending, reviewed, or interviewed).', 'error')
                            return redirect(url_for(target_endpoint))
                        
                        # Double-check by querying the current status from database
                        cursor.execute(
                            'SELECT status FROM applications WHERE application_id = %s LIMIT 1',
                            (resolved_application_id,)
                        )
                        current_status_record = cursor.fetchone()
                        if current_status_record:
                            current_status = (current_status_record.get('status') or '').lower()
                            if current_status in ('rejected', 'withdrawn'):
                                status_label = 'rejected' if current_status == 'rejected' else 'withdrawn'
                                flash(f'Cannot schedule interview for a {status_label} application. The application status is {status_label}.', 'error')
                                return redirect(url_for(target_endpoint))
                        
                        scheduled_datetime = f"{scheduled_date} {scheduled_time}"
                        
                        # Get current admin_id for reference (not stored in interviews table per schema)
                        admin_id = user.get('id')
                        
                        # Force interview mode to 'in-person' only
                        interview_mode = 'in-person'
                        
                        try:
                            # Step 1: Detect available columns in interviews table dynamically
                            cursor.execute("SHOW COLUMNS FROM interviews")
                            available_columns_result = cursor.fetchall()
                            available_columns = {col['Field'].lower() for col in available_columns_result}
                            
                            # Build dynamic INSERT statement based on available columns
                            insert_fields = ['application_id', 'scheduled_date', 'interview_mode', 'status']
                            insert_values = [resolved_application_id, scheduled_datetime, interview_mode, 'scheduled']
                            
                            if 'location' in available_columns:
                                insert_fields.insert(3, 'location')
                                insert_values.insert(3, location)
                            if 'notes' in available_columns:
                                insert_fields.insert(-1, 'notes')
                                insert_values.insert(-1, notes)
                            
                            fields_sql = ', '.join(insert_fields)
                            placeholders_sql = ', '.join(['%s'] * len(insert_values))
                            insert_sql = f'INSERT INTO interviews ({fields_sql}) VALUES ({placeholders_sql})'
                            
                            cursor.execute(insert_sql, tuple(insert_values))
                            interview_id = cursor.lastrowid
                            
                            columns_used = [f for f in insert_fields if f not in ('application_id', 'scheduled_date', 'interview_mode', 'status')]
                            if columns_used:
                                print(f'✅ Interview saved to database - ID: {interview_id}, Application ID: {resolved_application_id}, Extra fields: {columns_used}')
                            else:
                                print(f'✅ Interview saved to database - ID: {interview_id}, Application ID: {resolved_application_id} (core fields only)')
                            
                            # Step 2: AUTOMATIC: Update application status to 'scheduled' when interview is scheduled
                            # Update status to 'scheduled' unless it's already 'hired' or 'rejected' (final states)
                            # Guard update in case `updated_at` column is missing in the database schema
                            cursor.execute("SHOW COLUMNS FROM applications LIKE 'updated_at'")
                            _col = cursor.fetchone()
                            if _col:
                                cursor.execute(
                                    'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status NOT IN (%s, %s)',
                                    ('scheduled', resolved_application_id, 'hired', 'rejected'),
                                )
                            else:
                                cursor.execute(
                                    'UPDATE applications SET status = %s WHERE application_id = %s AND status NOT IN (%s, %s)',
                                    ('scheduled', resolved_application_id, 'hired', 'rejected'),
                                )
                            rows_updated = cursor.rowcount
                            if rows_updated > 0:
                                print(f'✅ Application status updated to "scheduled" for application {resolved_application_id}')
                            else:
                                print(f'⚠️ Application status not updated (may already be hired/rejected) for application {resolved_application_id}')
                            
                            # Step 3: Get applicant email for automatic notification
                            cursor.execute(
                                '''
                                SELECT ap.applicant_id, ap.email, ap.full_name, COALESCE(j.job_title, 'Position') AS job_title
                                FROM applicants ap
                                JOIN applications a ON ap.applicant_id = a.applicant_id
                                LEFT JOIN jobs j ON a.job_id = j.job_id
                                WHERE a.application_id = %s
                                LIMIT 1
                                ''',
                                (resolved_application_id,)
                            )
                            applicant_info = cursor.fetchone()
                            
                            if not applicant_info:
                                print(f'⚠️ Warning: No applicant info found for application {resolved_application_id}')
                                flash('Interview scheduled, but could not find applicant information for notification.', 'warning')
                            else:
                                # Step 4: AUTOMATIC: Notify and email applicant
                                # Format notification message to match applicant notification query pattern
                                job_title_display = applicant_info.get('job_title') or 'the position'
                                notification_message = f'You applied for {job_title_display}. Interview scheduled for {scheduled_date} at {scheduled_time}. Mode: {interview_mode.replace("_", " ").title()}. Location: {location or "TBD"}.'
                                email_subject = f'Interview Scheduled - {applicant_info.get("job_title") or "Your Application"}'
                                email_body = f"""Dear {applicant_info.get('full_name') or 'Applicant'},

Your interview has been scheduled for the position: {applicant_info.get('job_title') or 'the position'}

Interview Details:
- Date: {scheduled_date}
- Time: {scheduled_time}
- Mode: {interview_mode.replace('_', ' ').title()}
- Location: {location or 'To be determined'}

Please arrive on time and bring any required documents.

Best regards,
J&T Express Recruitment Team
                                """.strip()
                                
                                # Check if we already sent an email for this interview (prevent duplicates)
                                # Check for any recent notification about interview scheduling for this application
                                cursor.execute(
                                    '''
                                    SELECT notification_id FROM notifications
                                    WHERE application_id = %s 
                                    AND message LIKE %s
                                    AND sent_at >= DATE_SUB(NOW(), INTERVAL 5 MINUTE)
                                    LIMIT 1
                                    ''',
                                    (resolved_application_id, f'%Interview scheduled for {scheduled_date}%'),
                                )
                                recent_notification = cursor.fetchone()
                                
                                if recent_notification:
                                    print(f'⚠️ Recent interview scheduling notification found for application {resolved_application_id}. Skipping duplicate notification and email.')
                                else:
                                    # Create notification and send email (only if no recent duplicate)
                                    auto_notify_and_email(
                                        cursor, resolved_application_id, notification_message,
                                        email_subject, email_body,
                                        applicant_info.get('email'),
                                        applicant_info.get('full_name')
                                    )
                                    print(f'✅ Notification and email sent to applicant {applicant_info.get("applicant_id")} ({applicant_info.get("email")})')
                                # Intentionally do NOT create an admin/HR notification for scheduling — this is applicant-facing only
                                # (HR scheduling already notifies the applicant via notifications/email above)
                            
                            # Step 5: Commit all changes to database
                            db.commit()
                            log.info('✅ All changes committed to database successfully')
                            flash('Interview scheduled successfully. Applicant has been automatically notified via email and notification.', 'success')
                            return redirect(url_for(target_endpoint))
                        except Exception as schedule_error:
                            db.rollback()
                            log.exception(f'❌ Error scheduling interview: {schedule_error}')
                            flash(f'Error scheduling interview: {str(schedule_error)}', 'error')
                            return redirect(url_for(target_endpoint))
            
            elif action == 'update' or action == 'reschedule':
                # Block updating/rescheduling for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Updates are restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                scheduled_date = request.form.get('scheduled_date', '').strip()
                scheduled_time = request.form.get('scheduled_time', '').strip()
                interview_mode = request.form.get('interview_mode', 'in-person').strip()
                location = request.form.get('location', '').strip()
                notes = request.form.get('notes', '').strip()
                
                # Normalize interview_mode to match schema ENUM values
                normalized_mode = interview_mode.lower().replace('-', '_')
                if normalized_mode in ('in_person', 'in person', 'inperson'):
                    interview_mode = 'in-person'
                elif normalized_mode in ('online', 'remote', 'video', 'virtual', 'phone'):
                    interview_mode = 'remote'
                else:
                    interview_mode = 'in-person'  # Default
                
                if not interview_id or not scheduled_date or not scheduled_time:
                    flash('Interview ID, date, and time are required.', 'error')
                else:
                    # Validate datetime for reschedule (allow 5-minute grace period)
                    try:
                        scheduled_dt = datetime.strptime(f"{scheduled_date} {scheduled_time}", "%Y-%m-%d %H:%M")
                        # Allow 5-minute grace period for rescheduling (accounting for seconds precision)
                        grace_period = timedelta(minutes=5)
                        if scheduled_dt < datetime.now() - grace_period:
                            flash('New interview time must be in the future.', 'error')
                            return redirect(url_for(target_endpoint))
                    except Exception:
                        flash('Invalid date/time format.', 'error')
                        return redirect(url_for(target_endpoint))
                    if len(notes) > 1000:
                        notes = notes[:1000]
                    if len(location) > 255:
                        location = location[:255]
                    scheduled_datetime = f"{scheduled_date} {scheduled_time}"
                    # Verify interview belongs to HR's branch
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id, i.application_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        interview_record = cursor.fetchone()
                        if not interview_record:
                            flash('You can only update interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                        application_id = interview_record['application_id']
                    else:
                        cursor.execute('SELECT application_id FROM interviews WHERE interview_id = %s', (interview_id,))
                        interview_record = cursor.fetchone()
                        application_id = interview_record['application_id'] if interview_record else None
                    
                    # Build dynamic UPDATE statement based on available columns
                    try:
                        # Detect available columns in interviews table
                        cursor.execute("SHOW COLUMNS FROM interviews")
                        available_columns_result = cursor.fetchall()
                        available_columns = {col['Field'].lower() for col in available_columns_result}
                        
                        # Build UPDATE fields dynamically
                        update_fields = ['scheduled_date = %s', 'interview_mode = %s']
                        update_values = [scheduled_datetime, interview_mode]
                        
                        if 'location' in available_columns:
                            update_fields.append('location = %s')
                            update_values.append(location)
                        if 'notes' in available_columns:
                            update_fields.append('notes = %s')
                            update_values.append(notes)
                        
                        update_fields.append('interview_id = %s')
                        update_values.append(interview_id)
                        
                        update_sql = 'UPDATE interviews SET ' + ', '.join(update_fields[:len(update_fields)-1]) + ' WHERE interview_id = %s'
                        cursor.execute(update_sql, tuple(update_values))
                        
                        columns_used = [f.split(' = ')[0] for f in update_fields[:-1] if f.split(' = ')[0] not in ('scheduled_date', 'interview_mode')]
                        if columns_used:
                            print(f'✅ Interview {interview_id} updated with extra fields: {columns_used}')
                        else:
                            print(f'✅ Interview {interview_id} updated (core fields only)')
                    except Exception as update_err:
                        print(f'❌ Failed to update interview {interview_id}: {update_err}')
                        raise
                    
                    # Get applicant email for notification
                    if action == 'reschedule' and application_id:
                        cursor.execute(
                            '''
                            SELECT ap.email, ap.full_name, j.job_title
                            FROM applicants ap
                            JOIN applications a ON ap.applicant_id = a.applicant_id
                            LEFT JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id = %s
                            LIMIT 1
                            ''',
                            (application_id,)
                        )
                        applicant_info = cursor.fetchone()
                        
                        # Create notification for reschedule
                        try:
                            ensure_schema_compatibility()
                            notification_columns = set()
                            try:
                                cursor.execute('SHOW COLUMNS FROM notifications')
                                notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                            except Exception:
                                pass
                            
                            reschedule_message = f'Interview rescheduled to {scheduled_date} at {scheduled_time}. Please check your interview schedule.'
                            if 'sent_at' in notification_columns:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, sent_at, is_read)
                                    VALUES (%s, %s, NOW(), 0)
                                    ''',
                                    (application_id, reschedule_message),
                                )
                            else:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, is_read)
                                    VALUES (%s, %s, 0)
                                    ''',
                                    (application_id, reschedule_message),
                                )
                            
                            # Send email notification
                            if applicant_info and applicant_info.get('email'):
                                try:
                                    applicant_name = applicant_info.get('full_name') or 'Applicant'
                                    job_title = applicant_info.get('job_title') or 'the position'
                                    email_subject = f'Interview Rescheduled - {job_title}'
                                    email_body = f"""Dear {applicant_name},

Your interview for the position "{job_title}" has been rescheduled.

New Interview Details:
- Date: {scheduled_date}
- Time: {scheduled_time}
- Mode: {interview_mode.replace('_', ' ').title()}
- Location: {location or 'To be determined'}

Please update your calendar accordingly.

Best regards,
J&T Express Recruitment Team
                                    """.strip()
                                    send_email(applicant_info.get('email'), email_subject, email_body)
                                except Exception as email_error:
                                    print(f"Email notification error: {email_error}")
                            # Do not create an admin/HR notification for reschedules here; applicant is notified directly.
                        except Exception:
                            pass
                    
                    db.commit()
                    flash('Interview updated successfully. Applicant has been notified via email.', 'success')
            
            elif action == 'cancel':
                # Block canceling for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Canceling is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                if interview_id:
                    # Verify interview belongs to HR's branch
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id, i.application_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        interview_record = cursor.fetchone()
                        if not interview_record:
                            flash('You can only cancel interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                        application_id = interview_record['application_id']
                    else:
                        cursor.execute('SELECT application_id FROM interviews WHERE interview_id = %s', (interview_id,))
                        interview_record = cursor.fetchone()
                        application_id = interview_record['application_id'] if interview_record else None
                    
                    # Update status to cancelled instead of deleting
                    cursor.execute(
                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                        ('cancelled', interview_id),
                    )
                    
                    # Get applicant email for notification
                    if application_id:
                        cursor.execute(
                            '''
                            SELECT ap.email, ap.full_name, j.job_title
                            FROM applicants ap
                            JOIN applications a ON ap.applicant_id = a.applicant_id
                            LEFT JOIN jobs j ON a.job_id = j.job_id
                            WHERE a.application_id = %s
                            LIMIT 1
                            ''',
                            (application_id,)
                        )
                        applicant_info = cursor.fetchone()
                        
                        # Create notification
                        try:
                            ensure_schema_compatibility()
                            notification_columns = set()
                            try:
                                cursor.execute('SHOW COLUMNS FROM notifications')
                                notification_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}
                            except Exception:
                                pass
                            
                            cancel_message = 'Your interview has been cancelled. Please contact HR for more information.'
                            if 'sent_at' in notification_columns:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, sent_at, is_read)
                                    VALUES (%s, %s, NOW(), 0)
                                    ''',
                                    (application_id, cancel_message),
                                )
                            else:
                                cursor.execute(
                                    '''
                                    INSERT INTO notifications (application_id, message, is_read)
                                    VALUES (%s, %s, 0)
                                    ''',
                                    (application_id, cancel_message),
                                )
                            
                            # Send email notification
                            if applicant_info and applicant_info.get('email'):
                                try:
                                    applicant_name = applicant_info.get('full_name') or 'Applicant'
                                    job_title = applicant_info.get('job_title') or 'the position'
                                    email_subject = f'Interview Cancelled - {job_title}'
                                    email_body = f"""Dear {applicant_name},

We regret to inform you that your scheduled interview for the position "{job_title}" has been cancelled.

Please contact our HR department for more information or to reschedule.

Best regards,
J&T Express Recruitment Team
                                    """.strip()
                                    send_email(applicant_info.get('email'), email_subject, email_body)
                                except Exception as email_error:
                                    print(f"Email notification error: {email_error}")
                            
                            # Do not create an admin/HR notification for cancellations here; applicant is notified directly.
                        except Exception:
                            pass
                    
                    db.commit()
                    flash('Interview cancelled successfully. Applicant has been notified via email.', 'success')
            
            elif action == 'delete':
                # Block deleting for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Deleting is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id') or request.args.get('interview_id')
                if not interview_id:
                    flash('Interview ID is required.', 'error')
                    return redirect(url_for(target_endpoint))
                
                try:
                    interview_id = int(interview_id)
                except (ValueError, TypeError):
                    flash('Invalid interview ID.', 'error')
                    return redirect(url_for(target_endpoint))
                
                # Verify interview exists and belongs to HR's branch (if scoped)
                branch_id = get_branch_scope(user)
                if branch_id:
                    cursor.execute(
                        '''
                        SELECT i.interview_id, i.application_id
                        FROM interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN jobs j ON a.job_id = j.job_id
                        WHERE i.interview_id = %s AND j.branch_id = %s
                        LIMIT 1
                        ''',
                        (interview_id, branch_id),
                    )
                    interview_record = cursor.fetchone()
                    if not interview_record:
                        flash('Interview not found or you do not have permission to delete it.', 'error')
                        return redirect(url_for(target_endpoint))
                else:
                    # HR managing all branches - verify interview exists
                    cursor.execute(
                        'SELECT interview_id, application_id FROM interviews WHERE interview_id = %s LIMIT 1',
                        (interview_id,)
                    )
                    interview_record = cursor.fetchone()
                    if not interview_record:
                        flash('Interview not found.', 'error')
                        return redirect(url_for(target_endpoint))
                
                # Delete the interview
                cursor.execute('DELETE FROM interviews WHERE interview_id = %s', (interview_id,))
                deleted_count = cursor.rowcount
                
                if deleted_count > 0:
                    db.commit()
                    print(f'✅ Interview {interview_id} deleted successfully')
                    if request.accept_mimetypes.accept_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': True, 'message': 'Interview deleted successfully.'})
                    flash('Interview deleted successfully.', 'success')
                else:
                    if request.accept_mimetypes.accept_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({'success': False, 'error': 'Interview not found or already deleted.'}), 404
                    flash('Interview not found or already deleted.', 'error')
                
                return redirect(url_for(target_endpoint))
            
            elif action == 'mark_attendance':
                # Block marking attendance for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Marking attendance is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                attendance_status = request.form.get('attendance_status', 'completed').strip()
                
                if interview_id:
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        if not cursor.fetchone():
                            flash('You can only update interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                    
                    # Get application_id for this interview
                    cursor.execute(
                        'SELECT application_id FROM interviews WHERE interview_id = %s',
                        (interview_id,)
                    )
                    interview_record = cursor.fetchone()
                    application_id = interview_record.get('application_id') if interview_record else None
                    
                    # Update interview status
                    cursor.execute(
                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                        (attendance_status, interview_id),
                    )
                    
                    # AUTOMATIC: If interview is marked as completed, update application status to 'interviewed'
                    if attendance_status == 'completed' and application_id:
                        # Only update if status is not already a final state (hired/rejected)
                        # Guard update against missing `updated_at` column
                        cursor.execute("SHOW COLUMNS FROM applications LIKE 'updated_at'")
                        _col = cursor.fetchone()
                        if _col:
                            cursor.execute(
                                'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status NOT IN (%s, %s)',
                                ('interviewed', application_id, 'hired', 'rejected'),
                            )
                        else:
                            cursor.execute(
                                'UPDATE applications SET status = %s WHERE application_id = %s AND status NOT IN (%s, %s)',
                                ('interviewed', application_id, 'hired', 'rejected'),
                            )
                        rows_updated = cursor.rowcount
                        if rows_updated > 0:
                            print(f'✅ Application status auto-updated to "interviewed" for application {application_id} (interview completed)')
                            # Intentionally do not create an admin/HR notification for interview completion triggered by applicant actions
                            # This keeps notifications focused on the applicant and avoids duplicate admin notifications.
                    
                    db.commit()
                    flash(f'Interview marked as {attendance_status.replace("_", " ")}.', 'success')
            
            elif action == 'add_evaluation':
                interview_id = request.form.get('interview_id')
                remarks = request.form.get('remarks', '').strip()
                
                # Enhanced feedback form fields
                technical_rating = request.form.get('technical_rating', '0')
                communication_rating = request.form.get('communication_rating', '0')
                cultural_rating = request.form.get('cultural_rating', '0')
                experience_rating = request.form.get('experience_rating', '0')
                strengths = request.form.get('strengths', '').strip()
                weaknesses = request.form.get('weaknesses', '').strip()
                recommendation = request.form.get('recommendation', '').strip()
                next_steps = request.form.get('next_steps', '').strip()
                
                # Legacy fields for backward compatibility
                criteria = request.form.get('criteria', '').strip()
                score = request.form.get('score', '').strip()
                
                if not interview_id:
                    flash('Interview ID is required.', 'error')
                elif not remarks and not criteria:
                    flash('Feedback comments are required.', 'error')
                else:
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        if not cursor.fetchone():
                            flash('You can only evaluate interviews for your branch.', 'error')
                            return redirect(url_for(target_endpoint))
                    
                    evaluator_id = user.get('id') or session.get('user_id')
                    
                    # Build comprehensive evaluation notes
                    evaluation_notes = []
                    if technical_rating and technical_rating != '0':
                        evaluation_notes.append(f'Technical Skills: {technical_rating}/5')
                    if communication_rating and communication_rating != '0':
                        evaluation_notes.append(f'Communication: {communication_rating}/5')
                    if cultural_rating and cultural_rating != '0':
                        evaluation_notes.append(f'Cultural Fit: {cultural_rating}/5')
                    if experience_rating and experience_rating != '0':
                        evaluation_notes.append(f'Experience: {experience_rating}/5')
                    if strengths:
                        evaluation_notes.append(f'Strengths: {strengths}')
                    if weaknesses:
                        evaluation_notes.append(f'Weaknesses: {weaknesses}')
                    if recommendation:
                        evaluation_notes.append(f'Recommendation: {recommendation.replace("_", " ").title()}')
                    if remarks:
                        evaluation_notes.append(f'Comments: {remarks}')
                    if next_steps:
                        evaluation_notes.append(f'Next Steps: {next_steps}')
                    
                    # Legacy format for backward compatibility
                    if criteria and score:
                        evaluation_notes.append(f'Criteria: {criteria}, Score: {score}')
                    
                    evaluation_text = '\n\n'.join(evaluation_notes)
                    
                    # Update interview notes with comprehensive feedback
                    cursor.execute(
                        '''
                        UPDATE interviews 
                        SET notes = CONCAT(COALESCE(notes, ''), '\n\n--- Interview Feedback ---\n', %s)
                        WHERE interview_id = %s
                        ''',
                        (evaluation_text, interview_id),
                    )
                    
                    # Try to insert into evaluation table if it exists
                    try:
                        # Store each rating as separate evaluation record
                        ratings = [
                            ('technical_skills', technical_rating),
                            ('communication', communication_rating),
                            ('cultural_fit', cultural_rating),
                            ('experience', experience_rating),
                        ]
                        for criteria_name, rating in ratings:
                            if rating and rating != '0':
                                try:
                                    cursor.execute(
                                        '''
                                        INSERT INTO evaluations (interview_id, evaluator_id, criteria, score, remarks)
                                        VALUES (%s, %s, %s, %s, %s)
                                        ON DUPLICATE KEY UPDATE score = %s, remarks = %s
                                        ''',
                                        (interview_id, evaluator_id, criteria_name, float(rating), remarks, float(rating), remarks),
                                    )
                                except Exception:
                                    # Table might not exist or have different structure, continue
                                    pass
                    except Exception:
                        # Evaluation table doesn't exist or has different structure, that's okay
                        pass
                    
                    # Update interview status to completed if not already
                    cursor.execute(
                        'UPDATE interviews SET status = %s WHERE interview_id = %s AND status = %s',
                        ('completed', interview_id, 'scheduled'),
                    )
                    
                    db.commit()
                    flash('Interview feedback recorded successfully.', 'success')
            
            elif action == 'make_decision':
                # Block making decisions for admins - they can only view
                if user.get('role') == 'admin':
                    flash('Admins can only view interviews. Making decisions is restricted to HR staff.', 'error')
                    return redirect(url_for(target_endpoint))
                
                interview_id = request.form.get('interview_id')
                decision = request.form.get('decision', '').strip()  # 'hired' or 'rejected'
                decision_notes = request.form.get('decision_notes', '').strip()
                
                if not interview_id or not decision:
                    flash('Interview ID and decision are required.', 'error')
                else:
                    branch_id = get_branch_scope(user)
                    if branch_id:
                        cursor.execute(
                            '''
                            SELECT i.interview_id, i.application_id
                            FROM interviews i
                            JOIN applications a ON i.application_id = a.application_id
                            JOIN jobs j ON a.job_id = j.job_id
                            WHERE i.interview_id = %s AND j.branch_id = %s
                            ''',
                            (interview_id, branch_id),
                        )
                        interview_record = cursor.fetchone()
                        if not interview_record:
                            flash('You can only make decisions for your branch interviews.', 'error')
                            return redirect(url_for(target_endpoint))
                        application_id = interview_record['application_id']
                    else:
                        cursor.execute('SELECT application_id FROM interviews WHERE interview_id = %s', (interview_id,))
                        interview_record = cursor.fetchone()
                        application_id = interview_record['application_id'] if interview_record else None
                    
                    if application_id:
                        # AUTOMATIC: Update application status and notify applicant
                        decision_reason = f'Interview decision: {decision_notes}' if decision_notes else ''
                        if auto_update_application_status(cursor, application_id, decision, decision_reason):
                            # Update interview status to completed (try with notes, fall back without)
                            try:
                                cursor.execute(
                                    'UPDATE interviews SET status = %s, notes = CONCAT(COALESCE(notes, ""), "\n\nDecision: ", %s, "\nNotes: ", %s) WHERE interview_id = %s',
                                    ('completed', decision, decision_notes, interview_id),
                                )
                            except Exception as notes_err:
                                if '1054' in str(notes_err) or 'Unknown column' in str(notes_err):
                                    cursor.execute(
                                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                        ('completed', interview_id),
                                    )
                                else:
                                    raise
                            db.commit()
                            flash(f'Decision recorded: {decision}. Application status automatically updated and applicant notified.', 'success')
                        else:
                            # Fallback
                            cursor.execute(
                                'UPDATE applications SET status = %s WHERE application_id = %s',
                                (decision, application_id),
                            )
                            try:
                                cursor.execute(
                                    'UPDATE interviews SET status = %s, notes = CONCAT(COALESCE(notes, ""), "\n\nDecision: ", %s, "\nNotes: ", %s) WHERE interview_id = %s',
                                    ('completed', decision, decision_notes, interview_id),
                                )
                            except Exception as notes_err:
                                if '1054' in str(notes_err) or 'Unknown column' in str(notes_err):
                                    cursor.execute(
                                        'UPDATE interviews SET status = %s WHERE interview_id = %s',
                                        ('completed', interview_id),
                                    )
                                else:
                                    raise
                            db.commit()
                            flash(f'Decision recorded: {decision}. Application status updated.', 'success')
            
            return redirect(url_for(target_endpoint))
        
        # Get application_id from query parameter for pre-filling schedule form
        prefill_application_id = request.args.get('application_id', type=int)
        
        # Apply filters - preserve all filter values even if empty for proper filtering
        filters = {
            'status': request.args.get('status', '').strip(),  # scheduled, completed, cancelled, all
            'mode': request.args.get('mode', '').strip(),  # in-person, remote, video, phone
            'date_from': request.args.get('date_from', '').strip(),
            'date_to': request.args.get('date_to', '').strip(),
            'position_id': request.args.get('position_id', type=int),
            'job_id': request.args.get('job_id', type=int),
            'applicant_id': request.args.get('applicant_id', type=int),
            'keyword': request.args.get('keyword', '').strip(),
            'view_mode': request.args.get('view_mode', 'list').strip(),  # list, calendar
        }
        # Only remove filters that are None or empty strings - keep integer 0 if it exists
        filters = {k: v for k, v in filters.items() if v is not None and v != ''}
        
        # For admin users, don't filter by branch - show all branches
        branch_id = None if user.get('role') == 'admin' else get_branch_scope(user)
        where_clauses = []
        params = []
        
        if branch_id:
            where_clauses.append('j.branch_id = %s')
            params.append(branch_id)
        
        if filters.get('status') and filters['status'].strip():
            # Normalize status filter - handle variations like 'in_progress' vs 'in-progress'
            status_value = filters['status'].strip().lower()
            # Map common status variations to database values
            status_map = {
                'scheduled': 'scheduled',
                'completed': 'completed',
                'cancelled': 'cancelled',
                'canceled': 'cancelled',  # Handle typo
                'rescheduled': 'rescheduled',
                'in_progress': 'in_progress',
                'in-progress': 'in_progress',  # Handle hyphen variation
                'inprogress': 'in_progress',  # Handle no separator
                'no_show': 'no_show',
                'no-show': 'no_show',  # Handle hyphen variation
                'noshow': 'no_show',  # Handle no separator
            }
            normalized_status = status_map.get(status_value, status_value)
            where_clauses.append('i.status = %s')
            params.append(normalized_status)
        
        if filters.get('mode'):
            where_clauses.append('i.interview_mode = %s')
            params.append(filters['mode'])
        
        if filters.get('date_from'):
            where_clauses.append('DATE(i.scheduled_date) >= %s')
            params.append(filters['date_from'])
        
        if filters.get('date_to'):
            where_clauses.append('DATE(i.scheduled_date) <= %s')
            params.append(filters['date_to'])
        
        # Position filter removed - positions table no longer exists
        # if filters.get('position_id'):
        #     where_clauses.append('j.position_id = %s')
        #     params.append(filters['position_id'])
        
        if filters.get('job_id'):
            # Ensure job_id is a valid integer
            try:
                job_id_val = int(filters['job_id']) if filters['job_id'] else None
                if job_id_val and job_id_val > 0:
                    where_clauses.append('j.job_id = %s')
                    params.append(job_id_val)
            except (ValueError, TypeError):
                print(f'⚠️ Invalid job_id filter: {filters.get("job_id")}')
        
        if filters.get('applicant_id'):
            # Ensure applicant_id is a valid integer
            try:
                applicant_id_val = int(filters['applicant_id']) if filters['applicant_id'] else None
                if applicant_id_val and applicant_id_val > 0:
                    where_clauses.append('ap.applicant_id = %s')
                    params.append(applicant_id_val)
            except (ValueError, TypeError):
                print(f'⚠️ Invalid applicant_id filter: {filters.get("applicant_id")}')
        
        if filters.get('keyword') and filters['keyword'].strip():
            keyword = f"%{filters['keyword'].strip().lower()}%"
            where_clauses.append('('
                                 'LOWER(ap.full_name) LIKE %s OR '
                                 'LOWER(ap.email) LIKE %s OR '
                                 'LOWER(COALESCE(j.job_title, "")) LIKE %s OR '
                                 'LOWER(COALESCE(b.branch_name, "")) LIKE %s OR '
                                 'LOWER(COALESCE(i.interview_mode, "")) LIKE %s OR '
                                 'LOWER(COALESCE(i.notes, "")) LIKE %s'
                                 ')')
            params.extend([keyword, keyword, keyword, keyword, keyword, keyword])
        
        where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        
        # Preload applications for scheduling dropdown
        # Updated status values to match database enum
        # Load ALL applications (not filtered by status) so any applicant can be scheduled for interview
        # For admin users, don't filter by branch - show all applications from all branches
        application_where = []
        application_params = []
        if branch_id:
            application_where.append('j.branch_id = %s')
            application_params.append(branch_id)
        # If prefill_application_id is provided, include it
        if prefill_application_id:
            application_where.append('(1=1 OR a.application_id = %s)')
            application_params.append(prefill_application_id)
        application_where_sql = ' AND '.join(application_where) if application_where else '1=1'
        # Exclude rejected applications from schedule dropdown - cannot schedule interview for rejected
        application_where_sql += ' AND a.status != %s'
        application_params_with_rejected = list(application_params) if application_params else []
        application_params_with_rejected.append('rejected')
        
        cursor.execute(
            f'''
            SELECT a.application_id,
                   a.applicant_id,
                   a.job_id,
                   ap.full_name AS applicant_name,
                   j.job_title AS job_title,
                   a.status
            FROM applications a
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE {application_where_sql}
            ORDER BY ap.full_name ASC, j.job_title ASC
            ''',
            tuple(application_params_with_rejected),
        )
        schedule_applications = cursor.fetchall() or []

        # Fetch interviews with interviewer and status info
        # Build SELECT dynamically so we include optional columns (location, notes) only when present
        try:
            try:
                cursor.execute("SHOW COLUMNS FROM interviews")
                interview_cols = {c.get('Field', '').lower() for c in (cursor.fetchall() or [])}
            except Exception:
                interview_cols = set()

            select_fields = [
                'i.interview_id',
                'i.application_id',
                'i.scheduled_date',
                "COALESCE(i.interview_mode, 'in-person') AS interview_mode",
                'i.status AS interview_status',
                'ap.applicant_id',
                'ap.full_name AS applicant_name',
                'ap.email AS applicant_email',
                'ap.phone_number AS applicant_phone',
                'j.job_id',
                "COALESCE(j.job_title, 'Untitled Job') AS job_title",
                "COALESCE(b.branch_name, 'Unassigned') AS branch_name",
                'a.status AS application_status',
            ]

            # Optionally include location and notes if the columns exist
            if 'location' in interview_cols:
                select_fields.append('i.location AS location')
            else:
                # Ensure the key exists in result dict for template consistency
                select_fields.append("'' AS location")
            if 'notes' in interview_cols:
                select_fields.append('i.notes')
            else:
                select_fields.append("NULL AS notes")

            select_sql = ',\n                   '.join(select_fields)

            cursor.execute(
                f'''
            SELECT {select_sql}
            FROM interviews i
            JOIN applications a ON i.application_id = a.application_id
            JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql}
            ORDER BY i.scheduled_date DESC
            ''',
                tuple(params) if params else None,
            )
            interviews = cursor.fetchall()
        except Exception as query_error:
            error_msg = str(query_error)
            if 'interview_type' in error_msg.lower():
                # If error is about interview_type, try to identify and fix the issue
                print(f'⚠️ Interview query error (interview_type): {error_msg}')
                # Return empty list and log the error
                flash('Error loading interviews. Please contact support if this persists.', 'error')
                interviews = []
            else:
                # Re-raise if it's a different error
                raise
        
        
        # Calculate statistics - wrapped in try-except to handle any interview_type errors
        stats_where_sql = ' AND '.join(where_clauses) if where_clauses else '1=1'
        try:
            cursor.execute(
                f'''
                SELECT 
                    COUNT(*) AS total_scheduled,
                    SUM(CASE WHEN DATE(i.scheduled_date) >= DATE_SUB(NOW(), INTERVAL 7 DAY) AND i.status = 'completed' THEN 1 ELSE 0 END) AS completed_this_week
                FROM interviews i
                JOIN applications a ON i.application_id = a.application_id
                JOIN jobs j ON a.job_id = j.job_id
                JOIN applicants ap ON a.applicant_id = ap.applicant_id
                WHERE {stats_where_sql}
                ''',
                tuple(params) if params else None,
            )
            stats_row = cursor.fetchone() or {}
        except Exception as stats_error:
            error_msg = str(stats_error)
            if 'interview_type' in error_msg.lower():
                print(f'⚠️ Interview stats query error (interview_type): {error_msg}')
                stats_row = {}
            else:
                raise
        total_scheduled = stats_row.get('total_scheduled', 0) or 0
        completed_this_week = stats_row.get('completed_this_week', 0) or 0
        
        interview_stats = {
            'total_scheduled': total_scheduled,
            'completed_this_week': completed_this_week,
        }
        
        # Separate upcoming and past interviews
        now = datetime.now()
        upcoming = []
        past = []
        
        for interview in interviews:
            scheduled = interview.get('scheduled_date')
            if scheduled and isinstance(scheduled, datetime):
                interview_status = interview.get('interview_status', 'scheduled')
                application_status = interview.get('application_status')
                
                # AUTOMATIC: If interview is completed, ensure application status shows as 'interviewed'
                # This ensures past interviews always show the correct status
                if interview_status == 'completed' and application_status not in ('hired', 'rejected'):
                    application_status = 'interviewed'
                
                interview_data = {
                    'interview_id': interview.get('interview_id'),
                    'application_id': interview.get('application_id'),
                    'applicant_id': interview.get('applicant_id'),
                    'applicant_name': interview.get('applicant_name'),
                    'applicant_email': interview.get('applicant_email'),
                    'applicant_phone': interview.get('applicant_phone'),
                    'job_id': interview.get('job_id'),
                    'job_title': interview.get('job_title'),
                    'branch_name': interview.get('branch_name'),
                    'scheduled_date': format_human_datetime(scheduled),
                    'scheduled_datetime': scheduled,  # Keep original for editing
                    'scheduled_date_raw': scheduled.strftime('%Y-%m-%d') if isinstance(scheduled, datetime) else '',
                    'scheduled_time_raw': scheduled.strftime('%H:%M') if isinstance(scheduled, datetime) else '',
                    'interview_mode': interview.get('interview_mode'),
                    'location': interview.get('location'),
                    'notes': interview.get('notes'),
                    'application_status': application_status,  # Use potentially overridden status
                    'interview_status': interview_status,
                }
                
                # Only show scheduled interviews in upcoming, completed/cancelled in past
                if interview_status == 'scheduled' and scheduled >= now:
                    upcoming.append(interview_data)
                elif interview_status in ('completed', 'cancelled', 'no_show') or (scheduled < now and interview_status != 'scheduled'):
                    past.append(interview_data)
                elif scheduled < now:
                    past.append(interview_data)
        
        # Get applicants for scheduling (all applicants who have applied to jobs)
        # Only show applicants with non-rejected and non-withdrawn applications (cannot schedule interview for rejected/withdrawn)
        # For admin users, don't filter by branch - show all applicants from all branches
        where_clauses_schedule = []
        params_schedule = []
        # Exclude rejected and withdrawn applications - cannot schedule interview for rejected/withdrawn
        where_clauses_schedule.append('a.status NOT IN (%s, %s)')
        params_schedule.extend(['rejected', 'withdrawn'])
        if branch_id:
            where_clauses_schedule.append('j.branch_id = %s')
            params_schedule.append(branch_id)
        where_sql_schedule = ' AND '.join(where_clauses_schedule) if where_clauses_schedule else '1=1'
        # Query to get applicants who have at least one eligible application (not rejected, not withdrawn)
        # Only show applicants with eligible applications (pending, reviewed, interviewed, hired)
        cursor.execute(
            f'''
            SELECT DISTINCT 
                   ap.applicant_id,
                   ap.full_name AS applicant_name,
                   ap.email AS applicant_email,
                   ap.phone_number AS applicant_phone,
                   GROUP_CONCAT(DISTINCT CONCAT(COALESCE(j.job_title, 'Untitled'), ' (', a.status, ')') SEPARATOR ' | ') AS job_titles,
                   COUNT(DISTINCT a.application_id) AS application_count,
                   GROUP_CONCAT(DISTINCT a.application_id SEPARATOR ',') AS application_ids,
                   MAX(b.branch_name) AS branch_name
            FROM applicants ap
            JOIN applications a ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            WHERE {where_sql_schedule}
            GROUP BY ap.applicant_id, ap.full_name, ap.email, ap.phone_number
            HAVING COUNT(DISTINCT a.application_id) > 0
            ORDER BY ap.full_name ASC
            LIMIT 200
            ''',
            tuple(params_schedule) if params_schedule else None,
        )
        applicants_for_schedule = cursor.fetchall()
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/interviews.html' if user.get('role') == 'hr' else 'admin/interviews.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id = session.get('branch_id')
            if branch_id:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
        # Get unique positions for filter
        # Positions table removed - return empty list
        positions = []
        
        # Get unique jobs for filter - only show jobs that have interviews
        job_where = []
        job_params = []
        job_query = '''
            SELECT DISTINCT j.job_id, COALESCE(j.job_title, 'Untitled Job') AS job_title
            FROM jobs j
            JOIN applications a ON a.job_id = j.job_id
            JOIN interviews i ON i.application_id = a.application_id
        '''
        if branch_id:
            job_where.append('j.branch_id = %s')
            job_params.append(branch_id)
        if job_where:
            job_query += ' WHERE ' + ' AND '.join(job_where)
        job_query += ' ORDER BY j.job_title ASC'
        cursor.execute(job_query, tuple(job_params) if job_params else None)
        jobs = cursor.fetchall() or []
        
        # Get unique applicants for filter - only show applicants who have interviews
        applicant_where = []
        applicant_params = []
        applicant_query = '''
            SELECT DISTINCT ap.applicant_id, ap.full_name AS applicant_name
            FROM applicants ap
            JOIN applications a ON a.applicant_id = ap.applicant_id
            JOIN interviews i ON i.application_id = a.application_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
        '''
        if branch_id:
            applicant_where.append('j.branch_id = %s')
            applicant_params.append(branch_id)
        if applicant_where:
            applicant_query += ' WHERE ' + ' AND '.join(applicant_where)
        applicant_query += ' ORDER BY ap.full_name ASC'
        cursor.execute(applicant_query, tuple(applicant_params) if applicant_params else None)
        applicants = cursor.fetchall() or []
        
        return render_template(
            template,
            interviews=upcoming + past,
            upcoming=upcoming,
            past=past,
            applications=schedule_applications,
            applicants_for_schedule=applicants_for_schedule or [],
            hr_interviewers=[],
            current_filters=filters,
            user=user,
            branch_info=branch_info,
            interview_stats=interview_stats,
            positions=positions,
            jobs=jobs,
            applicants=applicants,
            prefill_application_id=prefill_application_id,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Interviews management error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'Error: {str(exc)}. Please check the console for details.', 'error')
        template = 'hr/interviews.html' if user.get('role') == 'hr' else 'admin/interviews.html'
        return render_template(template, user=user, interviews=[], upcoming=[], past=[], applications=[], applicants_for_schedule=[], current_filters={})
    finally:
        cursor.close()


@app.route('/admin/interviews', methods=['GET', 'POST'])
@login_required('admin')
def admin_interviews():
    return _render_interviews()

@app.route('/hr/interviews', methods=['GET', 'POST'])
@login_required('hr', 'admin')
def hr_interviews():
    user = get_current_user()
    if user.get('role') != 'hr':
        return redirect(url_for('admin_interviews'))
    return _render_interviews()


@app.route('/interviews/<int:interview_id>/status', methods=['POST'])
@login_required('admin', 'hr')
def update_interview_status(interview_id: int):
    """Update interview status (scheduled → completed/cancelled/no_show). HR scoped by branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Database connection error'}), 500
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
    ensure_schema_compatibility()
    cursor = db.cursor(dictionary=True)
    try:
        # Safely get status from form or JSON
        new_status = ''
        if request.form:
            new_status = request.form.get('status', '').strip().lower()
        elif request.is_json and request.json:
            new_status = request.json.get('status', '').strip().lower()
        
        # Safely get notes from form or JSON
        notes = ''
        if request.form:
            notes = request.form.get('notes', '').strip()
        elif request.is_json and request.json:
            notes = request.json.get('notes', '').strip()
        
        # Validate status
        if not new_status or new_status not in ('completed', 'cancelled', 'no_show'):
            error_msg = 'Invalid or missing status value' if not new_status else 'Invalid status value'
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': error_msg}), 400
            flash(error_msg, 'error')
            return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
        
        # STRICT BRANCH SCOPING: HR users assigned to specific branch ONLY manage that branch
        # Admin users can manage all branches
        branch_id = None
        scope_sql = ''
        scope_params = []
        if user.get('role') == 'hr':
            branch_id = get_branch_scope(user)
            if branch_id:
                # HR with assigned branch - ENFORCE strict branch isolation
                scope_sql = ' AND j.branch_id = %s'
                scope_params = [branch_id]
        
        # First, get the application_id before updating
        cursor.execute(
            f'''
            SELECT i.application_id
            FROM interviews i
            JOIN applications a ON i.application_id = a.application_id
            JOIN jobs j ON a.job_id = j.job_id
            WHERE i.interview_id = %s{scope_sql}
            ''',
            (interview_id,) + tuple(scope_params)
        )
        interview_data = cursor.fetchone()
        application_id = interview_data.get('application_id') if interview_data else None
        
        # Verify and update interview status
        try:
            # Build update query - append to existing notes if notes provided
            # Detect if `updated_at` exists on interviews table to avoid SQL errors on older schemas
            cursor.execute("SHOW COLUMNS FROM interviews LIKE 'updated_at'")
            _icol = cursor.fetchone()
            if notes:
                if _icol:
                    cursor.execute(
                        f'''
                        UPDATE interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN jobs j ON a.job_id = j.job_id
                        SET i.status = %s, i.notes = CONCAT(COALESCE(i.notes, ''), '\n', %s), i.updated_at = NOW()
                        WHERE i.interview_id = %s{scope_sql}
                        ''',
                        (new_status, notes, interview_id) + tuple(scope_params)
                    )
                else:
                    cursor.execute(
                        f'''
                        UPDATE interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN jobs j ON a.job_id = j.job_id
                        SET i.status = %s, i.notes = CONCAT(COALESCE(i.notes, ''), '\n', %s)
                        WHERE i.interview_id = %s{scope_sql}
                        ''',
                        (new_status, notes, interview_id) + tuple(scope_params)
                    )
            else:
                if _icol:
                    cursor.execute(
                        f'''
                        UPDATE interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN jobs j ON a.job_id = j.job_id
                        SET i.status = %s, i.updated_at = NOW()
                        WHERE i.interview_id = %s{scope_sql}
                        ''',
                        (new_status, interview_id) + tuple(scope_params)
                    )
                else:
                    cursor.execute(
                        f'''
                        UPDATE interviews i
                        JOIN applications a ON i.application_id = a.application_id
                        JOIN jobs j ON a.job_id = j.job_id
                        SET i.status = %s
                        WHERE i.interview_id = %s{scope_sql}
                        ''',
                        (new_status, interview_id) + tuple(scope_params)
                    )
            if cursor.rowcount == 0:
                db.rollback()
                if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                    return jsonify({'success': False, 'error': 'Interview not found or not in scope'}), 404
                flash('Interview not found or not in your branch.', 'error')
                return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
        except Exception as update_err:
            db.rollback()
            print(f'❌ Error updating interview status in database: {update_err}')
            import traceback
            traceback.print_exc()
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
                return jsonify({'success': False, 'error': f'Database error: {str(update_err)}'}), 500
            flash(f'Database error: {str(update_err)}', 'error')
            return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))
        
        # AUTOMATIC: If interview is marked as completed, update application status from 'scheduled' to 'interviewed'
        if new_status == 'completed' and application_id:
            # Only update if status is 'scheduled' and not already a final state (hired/rejected)
            # Guard update against missing `updated_at` column
            cursor.execute("SHOW COLUMNS FROM applications LIKE 'updated_at'")
            _col = cursor.fetchone()
            if _col:
                cursor.execute(
                    'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status = %s AND status NOT IN (%s, %s)',
                    ('interviewed', application_id, 'scheduled', 'hired', 'rejected'),
                )
            else:
                cursor.execute(
                    'UPDATE applications SET status = %s WHERE application_id = %s AND status = %s AND status NOT IN (%s, %s)',
                    ('interviewed', application_id, 'scheduled', 'hired', 'rejected'),
                )
            rows_updated = cursor.rowcount
            if rows_updated > 0:
                print(f'✅ Application status auto-updated from "scheduled" to "interviewed" for application {application_id} (interview {interview_id} marked as completed)')

        # AUTOMATIC: If interview status is set to 'hired', ensure the application is marked 'hired' too.
        # This ensures actions taken on the interview page (e.g., hiring an applicant) propagate to the application record.
        if str(new_status).lower() == 'hired' and application_id:
            try:
                # Prevent overwriting a final 'hired' or 'rejected' status
                cursor.execute("SHOW COLUMNS FROM applications LIKE 'updated_at'")
                _col = cursor.fetchone()
                if _col:
                    cursor.execute(
                        'UPDATE applications SET status = %s, updated_at = NOW() WHERE application_id = %s AND status NOT IN (%s, %s)',
                        ('hired', application_id, 'hired', 'rejected'),
                    )
                else:
                    cursor.execute(
                        'UPDATE applications SET status = %s WHERE application_id = %s AND status NOT IN (%s, %s)',
                        ('hired', application_id, 'hired', 'rejected'),
                    )
                if cursor.rowcount > 0:
                    print(f'✅ Application {application_id} synced to "hired" due to interview {interview_id} action')
                    # Do NOT create an admin/HR notification for applicant hires here — applicant receives the notification/email.
                    # If you want a separate admin alert for hires, we can add a configurable flag later.
            except Exception as hire_err:
                print(f'⚠️ Failed to sync application to hired: {hire_err}')

        db.commit()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': True, 'message': 'Interview status updated', 'interview_id': interview_id, 'status': new_status})
        flash('Interview status updated.', 'success')
    except Exception as exc:
        db.rollback()
        import traceback
        print('❌ Update interview status error:', exc)
        print(traceback.format_exc())
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.accept_json:
            return jsonify({'success': False, 'error': 'Failed to update interview status'}), 500
        flash('Failed to update interview status.', 'error')
    finally:
        cursor.close()
    return redirect(url_for('hr_interviews') if user.get('role') == 'hr' else url_for('admin_interviews'))

def handle_report_export(cursor, export_format, section, export_type, period_summary, 
                        applicant_summary, applicant_summary_details, job_vacancy, job_vacancy_details,
                        hiring_outcome, hr_performance, job_title_expr, date_filter, date_params):
    """Handle report exports in PDF, Excel, or CSV format."""
    from io import StringIO, BytesIO
    import csv
    from flask import Response
    from datetime import datetime
    
    if export_format == 'csv':
        output = StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Recruitment Report'])
        if period_summary:
            writer.writerow(['Period', f"{period_summary.get('start_date', '')} to {period_summary.get('end_date', '')}"])
        writer.writerow(['Generated', period_summary.get('generated_at', datetime.now().strftime('%b %d, %Y %I:%M %p'))])
        writer.writerow([])
        
        if section == 'applicant_summary':
            writer.writerow(['Applicant Summary Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Applicants', applicant_summary.get('total_applicants', 0)])
            writer.writerow(['New Applicants', applicant_summary.get('new_applicants', 0)])
            writer.writerow(['Active Applications', applicant_summary.get('active_applications', 0)])
            writer.writerow(['With Resume', applicant_summary.get('with_resume', 0)])
            writer.writerow(['Verified Email', applicant_summary.get('verified_email', 0)])
            if export_type == 'detailed' and applicant_summary_details:
                writer.writerow([])
                writer.writerow(['Detailed Applicant Status Report'])
                writer.writerow([])
                
                # Group by status
                status_groups = {
                    'Hired': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired'],
                    'Rejected': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected'],
                    'Pending': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending'],
                    'Interview Scheduled': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled'],
                    'Interviewed': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed']
                }
                
                for status_name, apps in status_groups.items():
                    if apps:
                        writer.writerow([f'{status_name} Applicants ({len(apps)})'])
                        writer.writerow(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                        for app in apps:
                            writer.writerow([
                                app.get('applicant_name', ''),
                                app.get('job_title', ''),
                                app.get('branch_name', ''),
                                app.get('status_label', app.get('application_status', '')),
                                app.get('date_applied', ''),
                                app.get('interview_date', '—'),
                                app.get('time_to_hire', '—')
                            ])
                        writer.writerow([])
                
                # All applicants summary
                writer.writerow(['All Applicants Summary'])
                writer.writerow(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                for app in applicant_summary_details:
                    writer.writerow([
                        app.get('applicant_name', ''),
                        app.get('job_title', ''),
                        app.get('branch_name', ''),
                        app.get('status_label', app.get('application_status', '')),
                        app.get('date_applied', ''),
                        app.get('interview_date', '—'),
                        app.get('time_to_hire', '—')
                    ])
        elif section == 'job_vacancy':
            writer.writerow(['Job Vacancy Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Jobs Posted', job_vacancy.get('total_jobs', 0)])
            writer.writerow(['Active Jobs', job_vacancy.get('active_jobs', 0)])
            writer.writerow(['Closed Jobs', job_vacancy.get('closed_jobs', 0)])
            writer.writerow(['Total Applications', job_vacancy.get('total_applications', 0)])
            writer.writerow(['Avg Applications/Job', round(job_vacancy.get('avg_applications_per_job', 0), 2)])
            if export_type == 'detailed' and job_vacancy_details:
                writer.writerow([])
                writer.writerow(['Detailed Job List with Applicants'])
                writer.writerow(['Job Title', 'Branch', 'Status', 'Total Applications', 'Posted Date'])
                for job in job_vacancy_details:
                    writer.writerow([
                        job.get('job_title', ''),
                        job.get('branch_name', ''),
                        job.get('status', ''),
                        job.get('application_count', 0),
                        job.get('posted_date', '')
                    ])
                    # Get applicants for this job from applicant_summary_details
                    job_applicants = [app for app in applicant_summary_details if app.get('job_title') == job.get('job_title')]
                    if job_applicants:
                        writer.writerow([])
                        writer.writerow(['Applicants for this Job:'])
                        writer.writerow(['Applicant Name', 'Email', 'Status', 'Date Applied', 'Active Application'])
                        for app in job_applicants:
                            is_active = app.get('application_status', '').lower() in ['pending', 'scheduled', 'interviewed']
                            writer.writerow([
                                app.get('applicant_name', ''),
                                app.get('email', ''),
                                app.get('status_label', app.get('application_status', '')),
                                app.get('date_applied', ''),
                                'Yes' if is_active else 'No'
                            ])
                        writer.writerow([])
        elif section == 'hiring_outcome':
            writer.writerow(['Hiring Outcome Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Total Hired', hiring_outcome.get('total_hired', 0)])
            writer.writerow(['Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
            writer.writerow(['Avg Time to Hire', f"{hiring_outcome.get('avg_time_to_hire', 0)} days"])
            writer.writerow(['Interview to Hire Rate', f"{hiring_outcome.get('interview_to_hire_rate', 0)}%"])
            writer.writerow(['Rejection Rate', f"{hiring_outcome.get('rejection_rate', 0)}%"])
        elif section == 'hr_performance':
            writer.writerow(['HR Performance Report'])
            writer.writerow(['Metric', 'Value'])
            writer.writerow(['Interviews Scheduled', hr_performance.get('interviews_scheduled', 0)])
            writer.writerow(['Interviews Completed', hr_performance.get('interviews_completed', 0)])
            writer.writerow(['Applications Reviewed', hr_performance.get('applications_reviewed', 0)])
            writer.writerow(['Status Updates', hr_performance.get('status_updates', 0)])
            writer.writerow(['Avg Response Time', f"{hr_performance.get('avg_response_time', 0)} hours"])
        else:
            # Full report
            writer.writerow(['Applicant Summary'])
            writer.writerow(['Total Applicants', applicant_summary.get('total_applicants', 0)])
            writer.writerow(['New Applicants', applicant_summary.get('new_applicants', 0)])
            writer.writerow([])
            writer.writerow(['Job Vacancy'])
            writer.writerow(['Total Jobs', job_vacancy.get('total_jobs', 0)])
            writer.writerow(['Active Jobs', job_vacancy.get('active_jobs', 0)])
            writer.writerow([])
            writer.writerow(['Hiring Outcome'])
            writer.writerow(['Total Hired', hiring_outcome.get('total_hired', 0)])
            writer.writerow(['Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
            writer.writerow([])
            writer.writerow(['HR Performance'])
            writer.writerow(['Interviews Scheduled', hr_performance.get('interviews_scheduled', 0)])
        
        output.seek(0)
        filename = f'report_{section or "full"}_{period_summary.get("start_date", "all")}.csv'.replace('/', '-').replace(' ', '_')
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Recruitment Report"
            
            # Header
            ws['A1'] = 'Recruitment Report'
            ws['A1'].font = Font(bold=True, size=14)
            if period_summary:
                ws['A2'] = f"Period: {period_summary.get('start_date', '')} to {period_summary.get('end_date', '')}"
            ws['A3'] = f"Generated: {period_summary.get('generated_at', datetime.now().strftime('%b %d, %Y %I:%M %p'))}"
            
            row = 5
            if section == 'applicant_summary':
                ws[f'A{row}'] = 'Applicant Summary Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                for key, label in [('total_applicants', 'Total Applicants'), ('new_applicants', 'New Applicants'), 
                                  ('active_applications', 'Active Applications'), ('with_resume', 'With Resume'), 
                                  ('verified_email', 'Verified Email')]:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = applicant_summary.get(key, 0)
                    row += 1
                
                # Add detailed applicant list if requested
                if export_type == 'detailed' and applicant_summary_details:
                    row += 2
                    ws[f'A{row}'] = 'Detailed Applicant Status Report'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 2
                    
                    # Group by status
                    status_groups = {
                        'Hired': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired'],
                        'Rejected': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected'],
                        'Pending': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending'],
                        'Interview Scheduled': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled'],
                        'Interviewed': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed']
                    }
                    
                    for status_name, apps in status_groups.items():
                        if apps:
                            ws[f'A{row}'] = f'{status_name} Applicants ({len(apps)})'
                            ws[f'A{row}'].font = Font(bold=True, size=11)
                            ws[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                            for col in range(1, 9):
                                ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF')
                            row += 1
                            headers = ['Applicant Name', 'Email', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Date Updated', 'Active Application']
                            for col, header in enumerate(headers, start=1):
                                cell = ws.cell(row=row, column=col)
                                cell.value = header
                                cell.font = Font(bold=True)
                            row += 1
                            for app in apps:
                                ws[f'A{row}'] = app.get('applicant_name', '')
                                ws[f'B{row}'] = app.get('job_title', '')
                                ws[f'C{row}'] = app.get('branch_name', '')
                                ws[f'D{row}'] = app.get('status_label', app.get('application_status', ''))
                                ws[f'E{row}'] = app.get('date_applied', '')
                                ws[f'F{row}'] = app.get('interview_date', '—')
                                ws[f'G{row}'] = app.get('time_to_hire', '—')
                                row += 1
                            row += 1
                    
                    # All applicants summary sheet
                    row += 1
                    ws[f'A{row}'] = 'All Applicants Summary'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 1
                    headers = ['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire']
                    for col, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = Font(bold=True)
                    row += 1
                    for app in applicant_summary_details:
                        ws[f'A{row}'] = app.get('applicant_name', '')
                        ws[f'B{row}'] = app.get('job_title', '')
                        ws[f'C{row}'] = app.get('branch_name', '')
                        ws[f'D{row}'] = app.get('status_label', app.get('application_status', ''))
                        ws[f'E{row}'] = app.get('date_applied', '')
                        ws[f'F{row}'] = app.get('interview_date', '—')
                        ws[f'G{row}'] = app.get('time_to_hire', '—')
                        row += 1
            elif section == 'job_vacancy':
                ws[f'A{row}'] = 'Job Vacancy Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                for key, label in [('total_jobs', 'Total Jobs Posted'), ('active_jobs', 'Active Jobs'), 
                                  ('closed_jobs', 'Closed Jobs'), ('total_applications', 'Total Applications'), 
                                  ('avg_applications_per_job', 'Avg Applications/Job')]:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = job_vacancy.get(key, 0)
                    row += 1
                
                # Add detailed job list with applicants if requested
                if export_type == 'detailed' and job_vacancy_details:
                    row += 2
                    ws[f'A{row}'] = 'Detailed Job List with Applicants'
                    ws[f'A{row}'].font = Font(bold=True, size=12)
                    row += 1
                    headers = ['Job Title', 'Branch', 'Status', 'Total Applications', 'Posted Date']
                    for col, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = Font(bold=True)
                    row += 1
                    for job in job_vacancy_details:
                        ws[f'A{row}'] = job.get('job_title', '')
                        ws[f'B{row}'] = job.get('branch_name', '')
                        ws[f'C{row}'] = job.get('status', '')
                        ws[f'D{row}'] = job.get('application_count', 0)
                        ws[f'E{row}'] = job.get('posted_date', '')
                        row += 1
                        # Get applicants for this job
                        job_applicants = [app for app in applicant_summary_details if app.get('job_title') == job.get('job_title')]
                        if job_applicants:
                            ws[f'A{row}'] = f'Applicants for {job.get("job_title", "")}:'
                            ws[f'A{row}'].font = Font(bold=True, italic=True)
                            row += 1
                            app_headers = ['Applicant Name', 'Email', 'Status', 'Date Applied', 'Active Application']
                            for col, header in enumerate(app_headers, start=1):
                                cell = ws.cell(row=row, column=col)
                                cell.value = header
                                cell.font = Font(bold=True)
                            row += 1
                            for app in job_applicants:
                                is_active = app.get('application_status', '').lower() in ['pending', 'scheduled', 'interviewed']
                                ws[f'A{row}'] = app.get('applicant_name', '')
                                ws[f'B{row}'] = app.get('email', '')
                                ws[f'C{row}'] = app.get('status_label', app.get('application_status', ''))
                                ws[f'D{row}'] = app.get('date_applied', '')
                                ws[f'E{row}'] = 'Yes' if is_active else 'No'
                                row += 1
                            row += 1
            elif section == 'hiring_outcome':
                ws[f'A{row}'] = 'Hiring Outcome Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                ws[f'A{row}'] = 'Total Hired'
                ws[f'B{row}'] = hiring_outcome.get('total_hired', 0)
                row += 1
                ws[f'A{row}'] = 'Hire Rate'
                ws[f'B{row}'] = f"{hiring_outcome.get('hire_rate', 0)}%"
                row += 1
                ws[f'A{row}'] = 'Avg Time to Hire'
                ws[f'B{row}'] = f"{hiring_outcome.get('avg_time_to_hire', 0)} days"
                row += 1
                ws[f'A{row}'] = 'Interview to Hire Rate'
                ws[f'B{row}'] = f"{hiring_outcome.get('interview_to_hire_rate', 0)}%"
                row += 1
                ws[f'A{row}'] = 'Rejection Rate'
                ws[f'B{row}'] = f"{hiring_outcome.get('rejection_rate', 0)}%"
            elif section == 'hr_performance':
                ws[f'A{row}'] = 'HR Performance Report'
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1
                ws[f'A{row}'] = 'Metric'
                ws[f'B{row}'] = 'Value'
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
                row += 1
                for key, label in [('interviews_scheduled', 'Interviews Scheduled'), 
                                  ('interviews_completed', 'Interviews Completed'), 
                                  ('applications_reviewed', 'Applications Reviewed'), 
                                  ('status_updates', 'Status Updates'), 
                                  ('avg_response_time', 'Avg Response Time (hours)')]:
                    ws[f'A{row}'] = label
                    ws[f'B{row}'] = hr_performance.get(key, 0)
                    row += 1
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f'report_{section or "full"}_{period_summary.get("start_date", "all")}.xlsx'.replace('/', '-').replace(' ', '_')
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
        except ImportError:
            flash('Excel export requires openpyxl library. Please install it: pip install openpyxl', 'error')
            from flask import redirect, url_for, request
            user = get_current_user()
            redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
            return redirect(request.referrer or redirect_url)
    elif export_format == 'pdf':
        # Basic PDF export using reportlab or weasyprint
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title = Paragraph("Recruitment Report", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.2*inch))
            
            # Period info
            if period_summary:
                period_text = f"Period: {period_summary.get('start_date', '')} to {period_summary.get('end_date', '')}"
                elements.append(Paragraph(period_text, styles['Normal']))
                elements.append(Spacer(1, 0.1*inch))
            
            # Report data
            if section == 'applicant_summary':
                data = [['Metric', 'Value']]
                data.append(['Total Applicants', str(applicant_summary.get('total_applicants', 0))])
                data.append(['New Applicants', str(applicant_summary.get('new_applicants', 0))])
                data.append(['Active Applications', str(applicant_summary.get('active_applications', 0))])
                data.append(['With Resume', str(applicant_summary.get('with_resume', 0))])
                data.append(['Verified Email', str(applicant_summary.get('verified_email', 0))])
                
                # Add status breakdown summary even for summary type
                if applicant_summary_details:
                    data.append([])
                    data.append(['Applicant Status Breakdown'])
                    # Group by status to get counts
                    status_counts = {
                        'Hired': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired']),
                        'Rejected': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected']),
                        'Pending': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending']),
                        'Interview Scheduled': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled']),
                        'Interviewed': len([app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed'])
                    }
                    data.append(['Status', 'Count'])
                    for status_name, count in status_counts.items():
                        if count > 0:
                            data.append([status_name, str(count)])
                
                # Add detailed applicant list if requested
                if export_type == 'detailed' and applicant_summary_details:
                    data.append([])
                    data.append(['Detailed Applicant Status Report'])
                    data.append([])
                    
                    # Group by status
                    status_groups = {
                        'Hired': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'hired'],
                        'Rejected': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'rejected'],
                        'Pending': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'pending'],
                        'Interview Scheduled': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'scheduled'],
                        'Interviewed': [app for app in applicant_summary_details if app.get('application_status', '').lower() == 'interviewed']
                    }
                    
                    for status_name, apps in status_groups.items():
                        if apps:
                            data.append([f'{status_name} Applicants ({len(apps)})'])
                            data.append(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                            for app in apps:
                                data.append([
                                    app.get('applicant_name', ''),
                                    app.get('job_title', ''),
                                    app.get('branch_name', ''),
                                    app.get('status_label', app.get('application_status', '')),
                                    app.get('date_applied', ''),
                                    app.get('interview_date', '—'),
                                    app.get('time_to_hire', '—')
                                ])
                            data.append([])
                    
                    # All applicants summary
                    data.append(['All Applicants Summary'])
                    data.append(['Applicant Name', 'Job Applied', 'Branch', 'Status', 'Date Applied', 'Interview Date', 'Time to Hire'])
                    for app in applicant_summary_details:
                        data.append([
                            app.get('applicant_name', ''),
                            app.get('job_title', ''),
                            app.get('branch_name', ''),
                            app.get('status_label', app.get('application_status', '')),
                            app.get('date_applied', ''),
                            app.get('interview_date', '—'),
                            app.get('time_to_hire', '—')
                        ])
            elif section == 'job_vacancy':
                data = [['Metric', 'Value']]
                data.append(['Total Jobs Posted', str(job_vacancy.get('total_jobs', 0))])
                data.append(['Active Jobs', str(job_vacancy.get('active_jobs', 0))])
                data.append(['Closed Jobs', str(job_vacancy.get('closed_jobs', 0))])
                data.append(['Total Applications', str(job_vacancy.get('total_applications', 0))])
                data.append(['Avg Applications/Job', str(round(job_vacancy.get('avg_applications_per_job', 0), 2))])
                
                # Add detailed job list with applicants if requested
                if export_type == 'detailed' and job_vacancy_details:
                    data.append([])
                    data.append(['Detailed Job List with Applicants'])
                    data.append(['Job Title', 'Branch', 'Status', 'Total Applications', 'Posted Date'])
                    for job in job_vacancy_details:
                        data.append([
                            job.get('job_title', ''),
                            job.get('branch_name', ''),
                            job.get('status', ''),
                            str(job.get('application_count', 0)),
                            job.get('posted_date', '')
                        ])
                        # Get applicants for this job
                        job_applicants = [app for app in applicant_summary_details if app.get('job_title') == job.get('job_title')]
                        if job_applicants:
                            data.append([])
                            data.append([f'Applicants for {job.get("job_title", "")}:'])
                            data.append(['Applicant Name', 'Email', 'Status', 'Date Applied', 'Active Application'])
                            for app in job_applicants:
                                is_active = app.get('application_status', '').lower() in ['pending', 'scheduled', 'interviewed']
                                data.append([
                                    app.get('applicant_name', ''),
                                    app.get('email', ''),
                                    app.get('status_label', app.get('application_status', '')),
                                    app.get('date_applied', ''),
                                    'Yes' if is_active else 'No'
                                ])
                            data.append([])
            elif section == 'hiring_outcome':
                data = [['Metric', 'Value']]
                data.append(['Total Hired', str(hiring_outcome.get('total_hired', 0))])
                data.append(['Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
                data.append(['Avg Time to Hire', f"{hiring_outcome.get('avg_time_to_hire', 0)} days"])
                data.append(['Interview to Hire Rate', f"{hiring_outcome.get('interview_to_hire_rate', 0)}%"])
                data.append(['Rejection Rate', f"{hiring_outcome.get('rejection_rate', 0)}%"])
            elif section == 'hr_performance':
                data = [['Metric', 'Value']]
                data.append(['Interviews Scheduled', str(hr_performance.get('interviews_scheduled', 0))])
                data.append(['Interviews Completed', str(hr_performance.get('interviews_completed', 0))])
                data.append(['Applications Reviewed', str(hr_performance.get('applications_reviewed', 0))])
                data.append(['Status Updates', str(hr_performance.get('status_updates', 0))])
                data.append(['Avg Response Time', f"{hr_performance.get('avg_response_time', 0)} hours"])
            else:
                # Full report
                data = [['Section', 'Metric', 'Value']]
                data.append(['Applicant Summary', 'Total Applicants', str(applicant_summary.get('total_applicants', 0))])
                data.append(['Applicant Summary', 'New Applicants', str(applicant_summary.get('new_applicants', 0))])
                data.append(['Job Vacancy', 'Total Jobs', str(job_vacancy.get('total_jobs', 0))])
                data.append(['Job Vacancy', 'Active Jobs', str(job_vacancy.get('active_jobs', 0))])
                data.append(['Hiring Outcome', 'Total Hired', str(hiring_outcome.get('total_hired', 0))])
                data.append(['Hiring Outcome', 'Hire Rate', f"{hiring_outcome.get('hire_rate', 0)}%"])
                data.append(['HR Performance', 'Interviews Scheduled', str(hr_performance.get('interviews_scheduled', 0))])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            
            doc.build(elements)
            buffer.seek(0)
            filename = f'report_{section or "full"}_{period_summary.get("start_date", "all") if period_summary else "all"}.pdf'.replace('/', '-').replace(' ', '_')
            return Response(
                buffer.getvalue(),
                mimetype='application/pdf',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
        except ImportError:
            flash('PDF export requires reportlab library. Please install it: pip install reportlab', 'error')
            from flask import redirect, url_for, request
            user = get_current_user()
            redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
            return redirect(request.referrer or redirect_url)
        except Exception as pdf_error:
            print(f'⚠️ Error generating PDF: {pdf_error}')
            flash(f'Error generating PDF: {str(pdf_error)}', 'error')
            from flask import redirect, url_for, request
            user = get_current_user()
            redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
            return redirect(request.referrer or redirect_url)
    else:
        flash('Invalid export format.', 'error')
        from flask import redirect, url_for, request
        user = get_current_user()
        redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
        return redirect(request.referrer or redirect_url)

@app.route('/api/reports/data', methods=['GET'])
@login_required('admin', 'hr')
def api_get_reports_data():
    """API endpoint to fetch real report data for client-side rendering.
    
    Query parameters:
    - period: 'week', 'month', 'year', 'all'
    - branch: branch_id filter
    - position: job_title filter
    - status: application_status filter
    - dateFrom: YYYY-MM-DD
    - dateTo: YYYY-MM-DD
    """
    from datetime import datetime, timedelta
    
    user = get_current_user()
    db = get_db()
    
    if not db:
        return jsonify({'error': 'Database connection error'}), 500
    
    try:
        ensure_schema_compatibility()
        cursor = db.cursor(dictionary=True)
        
        # Get filters from query params
        period = request.args.get('period', 'month')
        branch_filter = request.args.get('branch', '').strip()
        job_filter = request.args.get('position', '').strip()
        status_filter = request.args.get('status', '').strip()
        date_from = request.args.get('dateFrom', '').strip()
        date_to = request.args.get('dateTo', '').strip()
        
        # Enforce branch scope for HR users: if an HR is assigned to a specific branch,
        # override any requested branch filter so they only see their branch data.
        branch_scope = get_branch_scope(user)
        if branch_scope is not None:
            try:
                # normalize to string to match incoming query param handling
                branch_filter = str(branch_scope)
            except Exception:
                branch_filter = branch_filter
        
        # Build WHERE clause for filtering
        where_clauses = []
        params = []
        
        # Apply period-based date filter if no custom dates provided
        if not date_from and not date_to:
            today = datetime.now().date()
            if period == 'week':
                start_date = today - timedelta(days=today.weekday())
                where_clauses.append('DATE(a.applied_at) >= %s')
                params.append(start_date)
            elif period == 'month':
                start_date = today.replace(day=1)
                where_clauses.append('DATE(a.applied_at) >= %s')
                params.append(start_date)
            elif period == 'year':
                start_date = today.replace(month=1, day=1)
                where_clauses.append('DATE(a.applied_at) >= %s')
                params.append(start_date)
            # For 'all', no date filter is applied
        
        # Branch filter
        if branch_filter and branch_filter.lower() != 'all':
            # Allow branch filter to be passed as either branch name or branch id
            try:
                branch_id_val = int(branch_filter)
            except Exception:
                branch_id_val = None

            if branch_id_val is not None:
                where_clauses.append('b.branch_id = %s')
                params.append(branch_id_val)
            else:
                where_clauses.append('b.branch_name = %s')
                params.append(branch_filter)

        # Job position filter - treat common 'all' variants as no filter
        jf_norm = job_filter.strip().lower() if job_filter else ''
        if jf_norm and jf_norm not in ('all', '', 'all positions', 'any'):
            # Accept either job id or job title
            try:
                job_id_val = int(job_filter)
            except Exception:
                job_id_val = None

            if job_id_val is not None:
                where_clauses.append('j.job_id = %s')
                params.append(job_id_val)
            else:
                # Match by title using dynamic job column detection to avoid referencing missing columns
                title_expr = job_column_expr('title', alternatives=['job_title', 'position_name'], default='j.job_title')
                where_clauses.append(f'({title_expr} = %s)')
                params.append(job_filter)
        
        # Status filter
        if status_filter and status_filter.lower() != 'all':
            where_clauses.append('a.status = %s')
            params.append(status_filter)
        
        # Date range filter (overrides period filter if provided)
        if date_from:
            try:
                date_obj = datetime.strptime(date_from, '%Y-%m-%d')
                where_clauses.append('DATE(a.applied_at) >= %s')
                params.append(date_obj.date())
            except ValueError:
                pass
        
        if date_to:
            try:
                date_obj = datetime.strptime(date_to, '%Y-%m-%d')
                where_clauses.append('DATE(a.applied_at) <= %s')
                params.append(date_obj.date())
            except ValueError:
                pass
        
        where_sql = 'WHERE ' + ' AND '.join(where_clauses) if where_clauses else ''
        
        # Fetch all applicants with their application data
        query = f'''
            SELECT 
                ap.full_name,
                COALESCE(j.job_title, 'Unknown') AS position,
                COALESCE(b.branch_name, 'Unassigned') AS branch,
                COALESCE(a.status, 'pending') AS status,
                a.applied_at,
                a.application_id,
                a.viewed_at,
                (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id) AS interview_date,
                (SELECT COUNT(*) FROM interviews i WHERE i.application_id = a.application_id AND i.status IN ('completed', 'confirmed')) AS interview_count,
                DATEDIFF(COALESCE(a.viewed_at, (SELECT MAX(i.scheduled_date) FROM interviews i WHERE i.application_id = a.application_id)), a.applied_at) AS time_to_hire
            FROM applications a
            LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
            LEFT JOIN jobs j ON a.job_id = j.job_id
            LEFT JOIN branches b ON j.branch_id = b.branch_id
            {where_sql}
            ORDER BY a.applied_at DESC
        '''
        
        # Debug: print final query and params to help diagnose empty result issues
        try:
            print('--- reports API: executing query ---')
            print(query)
            print('params:', tuple(params))
        except Exception:
            pass

        cursor.execute(query, tuple(params))
        all_records = cursor.fetchall() or []
        print(f'--- reports API: fetched {len(all_records)} records')
        
        # Calculate KPIs
        total_applicants = len(all_records)
        interviews_scheduled = len([r for r in all_records if r.get('interview_date')])
        hired = len([r for r in all_records if r.get('status') == 'hired'])
        rejected = len([r for r in all_records if r.get('status') == 'rejected'])
        pending = len([r for r in all_records if r.get('status') == 'pending'])
        
        # Calculate average time to hire
        hired_records = [r for r in all_records if r.get('status') == 'hired' and r.get('time_to_hire')]
        avg_time_to_hire = 0
        if hired_records:
            avg_time_to_hire = sum([r.get('time_to_hire') or 0 for r in hired_records]) / len(hired_records)
            avg_time_to_hire = round(avg_time_to_hire, 1)
        
        # Calculate conversion rate
        conversion_rate = round((hired / total_applicants * 100), 1) if total_applicants > 0 else 0
        
        # Get all branches from database to ensure they all display in chart
        # If a specific branch is filtered, extract actual branch names from the fetched records
        if branch_filter and branch_filter.lower() != 'all':
            # Extract distinct branch names from the filtered records (handles branch ID vs name mismatch)
            all_branches = list(set([r.get('branch', 'Unknown') for r in all_records]))
            all_branches.sort()
        else:
            cursor.execute('SELECT DISTINCT branch_name FROM branches ORDER BY branch_name')
            all_branches_result = cursor.fetchall()
            all_branches = [b.get('branch_name') for b in (all_branches_result or [])]
        
        # Group by branch performance - initialize all branches with zero values
        branch_stats = {}
        for branch_name in all_branches:
            branch_stats[branch_name] = {
                'applicants': 0,
                'hired': 0,
                'conversion_rate': 0,
                'avg_time_to_hire': None,
                '_total_time_to_hire': 0,
                '_hired_count_for_time': 0
            }

        # Now populate with actual data from filtered records
        for record in all_records:
            branch = record.get('branch', 'Unknown')
            if branch not in branch_stats:
                branch_stats[branch] = {
                    'applicants': 0,
                    'hired': 0,
                    'conversion_rate': 0,
                    'avg_time_to_hire': None,
                    '_total_time_to_hire': 0,
                    '_hired_count_for_time': 0
                }
            branch_stats[branch]['applicants'] += 1
            if record.get('status') == 'hired':
                branch_stats[branch]['hired'] += 1
                # accumulate time_to_hire for average calculation
                t = record.get('time_to_hire')
                if t is not None:
                    try:
                        branch_stats[branch]['_total_time_to_hire'] += float(t)
                        branch_stats[branch]['_hired_count_for_time'] += 1
                    except Exception:
                        pass

        # Calculate conversion rates and average time to hire per branch
        for branch in branch_stats:
            total = branch_stats[branch]['applicants']
            hired_count = branch_stats[branch]['hired']
            branch_stats[branch]['conversion_rate'] = round((hired_count / total * 100), 1) if total > 0 else 0
            if branch_stats[branch]['_hired_count_for_time'] > 0:
                avg = branch_stats[branch]['_total_time_to_hire'] / branch_stats[branch]['_hired_count_for_time']
                branch_stats[branch]['avg_time_to_hire'] = round(avg, 1)
            else:
                branch_stats[branch]['avg_time_to_hire'] = None
            # cleanup internal keys
            branch_stats[branch].pop('_total_time_to_hire', None)
            branch_stats[branch].pop('_hired_count_for_time', None)
        
        # Group by job position
        job_stats = {}
        for record in all_records:
            job = record.get('position', 'Unknown')
            if job not in job_stats:
                job_stats[job] = 0
            job_stats[job] += 1
        
        # Group by status
        status_breakdown = {
            'pending': pending,
            'scheduled': len([r for r in all_records if r.get('status') == 'scheduled']),
            'interviewed': len([r for r in all_records if r.get('status') == 'interviewed']),
            'hired': hired,
            'rejected': rejected
        }
        
        # Group by date for trend chart - adapt based on period
        daily_stats = {}
        today = datetime.now()
        
        # Determine number of days to display based on period
        if period == 'week':
            days_to_show = 7
        elif period == 'month':
            days_to_show = 30
        elif period == 'year':
            days_to_show = 365
        else:  # 'all'
            days_to_show = 30  # Default to last 30 days for 'all'
        
        # Initialize daily stats for the period and keep full date objects for correct sorting
        date_map = {}
        # Use a readable month-day label (e.g. 'Dec 02') consistently for keys so counting matches
        for i in range(days_to_show):
            date_obj = (today - timedelta(days=i))
            date_key = date_obj.strftime('%b %d')
            daily_stats[date_key] = 0
            # store the exact date object for this label to avoid ambiguous parsing later
            date_map[date_key] = date_obj
        
        # Count hires (hired status) per day — use hire date (updated_at) when available
        for record in all_records:
            # For trend of hires, prefer the updated_at date when status == 'hired'
            date_candidate = None
            try:
                if record.get('status') == 'hired' and record.get('updated_at'):
                    date_candidate = record.get('updated_at')
                else:
                    date_candidate = record.get('applied_at')
                if not date_candidate:
                    continue

                # Normalize to datetime object or parse string
                parsed = None
                if hasattr(date_candidate, 'strftime'):
                    parsed = date_candidate
                else:
                    s = str(date_candidate)
                    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                        try:
                            parsed = datetime.strptime(s, fmt)
                            break
                        except Exception:
                            parsed = None
                if not parsed:
                    continue

                date_key = parsed.strftime('%b %d')
                # Only count hires (status == 'hired') for the hiring trend
                if record.get('status') == 'hired' and date_key in daily_stats:
                    daily_stats[date_key] += 1
            except Exception:
                continue
        
        # Sort dates from oldest to newest using the stored date objects (avoids year-less parsing)
        try:
            # date_map currently maps labels like '%m/%d' to date objects; convert keys to our display format
            converted_map = { (date_map[k].strftime('%b %d') if k in date_map else k): date_map.get(k, today) for k in date_map }
            trend_labels = sorted(list(daily_stats.keys()), key=lambda k: converted_map.get(k, today))
        except Exception:
            trend_labels = list(daily_stats.keys())
        trend_data = [daily_stats[label] for label in trend_labels]
        
        cursor.close()
        
        # Return formatted JSON response
        return jsonify({
            'branchScope': branch_scope,
            'success': True,
            'kpis': {
                'applicants': total_applicants,
                'interviews': interviews_scheduled,
                'hired': hired,
                'timeToHire': avg_time_to_hire,
                'conversionRate': conversion_rate,
                'rejected': rejected,
                'pending': pending
            },
            'branchPerformance': branch_stats,
            'jobStats': job_stats,
            'statusBreakdown': status_breakdown,
            'trendChart': {
                'labels': trend_labels,
                'data': trend_data
            },
            'applicantData': [
                {
                    'name': r.get('full_name'),
                    'job': r.get('position'),
                    'branch': r.get('branch'),
                    'status': r.get('status'),
                    'date': str(r.get('applied_at')).split(' ')[0],
                    'interview': str(r.get('interview_date')).split(' ')[0] if r.get('interview_date') else '',
                    'time': int(r.get('time_to_hire')) if (r.get('status') == 'hired' and r.get('time_to_hire')) else None
                }
                for r in all_records
            ]
        })
        
    except Exception as e:
        log.exception(f'⚠️ Error fetching report data: {e}')
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            cursor.close()
        except Exception:
            # Ignore errors closing cursor
            pass


@app.route('/api/jobs', methods=['GET'])
@login_required('admin', 'hr')
def api_get_jobs():
    """Return a list of job postings visible to the current user.

    Response JSON: { success: True, jobs: [{ job_id, title, branch_name }, ...] }
    """
    try:
        user = get_current_user()
        jobs = fetch_jobs_for_user(user) or []
        out = []
        seen = set()
        for j in jobs:
            title = (j.get('job_title') or j.get('position_name') or j.get('position_title') or '').strip()
            if not title:
                continue
            # avoid duplicates by title
            if title in seen:
                continue
            seen.add(title)
            out.append({
                'job_id': j.get('job_id'),
                'title': title,
                'branch_name': j.get('branch_name')
            })

        return jsonify({'success': True, 'jobs': out})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/branches', methods=['GET'])
def api_get_branches():
    """Return a list of branches for filters.

    Response JSON: { success: True, branches: [{ branch_id, branch_name, address }, ...] }
    """
    try:
        branches = fetch_branches() or []
        out = []
        for b in branches:
            out.append({
                'branch_id': b.get('branch_id'),
                'branch_name': b.get('branch_name')
            })
        return jsonify({'success': True, 'branches': out})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/reports-analytics', endpoint='admin_reports_analytics')
@app.route('/hr/reports-analytics', endpoint='hr_reports_analytics')
@login_required('admin', 'hr')
def reports_analytics():
    """Comprehensive analytics and reporting dashboard with period filtering and export."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        template = 'hr/reports_analytics.html' if user.get('role') == 'hr' else 'admin/reports.html'
        # Get period from request
        period = request.args.get('period', 'all')
        selected_week = request.args.get('week', '')
        selected_week_month = request.args.get('week_month', '')
        selected_week_number = request.args.get('week_number', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        return render_template(
            template, 
            stats={}, 
            trends={}, 
            branch_stats=[], 
            metrics={},
            branch_comparison=[],
            system_stats={},
            job_performance=[],
            status_breakdown=[],
            funnel_data={},
            source_effectiveness=[],
            branch_info=None,
            period=period, 
            period_summary={}, 
            available_years=[], 
            selected_week=selected_week, 
            selected_week_month=selected_week_month,
            selected_week_number=selected_week_number,
            selected_month=selected_month, 
            selected_year=selected_year, 
            applicant_summary={}, 
            job_vacancy={}, 
            hiring_outcome={}, 
            hr_performance={}, 
            applicant_summary_details=[], 
            job_vacancy_details=[]
        )
    
    # Ensure schema compatibility before proceeding
    ensure_schema_compatibility()
    
    cursor = db.cursor(dictionary=True)
    
    try:
        # Get period parameters
        period = request.args.get('period', 'all')
        selected_week = request.args.get('week', '')  # Legacy support
        selected_week_month = request.args.get('week_month', '')
        selected_week_number = request.args.get('week_number', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        export_format = request.args.get('export', '')
        export_section = request.args.get('section', '')
        export_type = request.args.get('type', 'summary')
        generate_section = request.args.get('generate_section', '') or request.form.get('generate_section', '')
        
        # Calculate date range based on period
        from datetime import datetime, timedelta
        from datetime import date as date_class
        period_summary = {}
        date_filter = ''
        date_params = []
        
        if period == 'week':
            # New week selector: month + week number (1-5)
            if selected_week_month and selected_week_number:
                try:
                    year, month = map(int, selected_week_month.split('-'))
                    week_num = int(selected_week_number)
                    
                    # Calculate the first day of the month
                    month_start = datetime(year, month, 1)
                    first_day_weekday = month_start.weekday()  # 0=Monday, 6=Sunday
                    
                    # Calculate start of week (Monday)
                    # Find the Monday of the week containing the 1st
                    days_to_monday = (first_day_weekday) % 7
                    first_monday = month_start - timedelta(days=days_to_monday)
                    
                    # Week 1 starts on the first Monday on or before the 1st
                    # Week 2 starts 7 days after that, etc.
                    days_to_add = (week_num - 1) * 7
                    start_date = first_monday + timedelta(days=days_to_add)
                    
                    # Ensure start_date is within the month (at least the 1st)
                    if start_date < month_start:
                        start_date = month_start
                    
                    # End date is 6 days after start (Sunday)
                    end_date = start_date + timedelta(days=6)
                    
                    # Make sure we don't go beyond the month
                    if month == 12:
                        month_end = datetime(year + 1, 1, 1) - timedelta(days=1)
                    else:
                        month_end = datetime(year, month + 1, 1) - timedelta(days=1)
                    
                    if end_date > month_end:
                        end_date = month_end
                    
                except (ValueError, IndexError) as e:
                    print(f'Error parsing week: {e}')
                    today = datetime.now()
                    start_date = today - timedelta(days=today.weekday())
                    end_date = start_date + timedelta(days=6)
            # Legacy support for old week format
            elif selected_week:
                try:
                    year, week = map(int, selected_week.split('-W'))
                    jan4 = date_class(year, 1, 4)
                    jan4_weekday = jan4.weekday()
                    days_since_jan4 = (week - 1) * 7 - jan4_weekday
                    week_start = jan4 + timedelta(days=days_since_jan4)
                    start_date = datetime.combine(week_start, datetime.min.time())
                    end_date = start_date + timedelta(days=6)
                except (ValueError, IndexError):
                    today = datetime.now()
                    start_date = today - timedelta(days=today.weekday())
                    end_date = start_date + timedelta(days=6)
            else:
                # Default to current week (Monday to Sunday)
                today = datetime.now()
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
                # Set default selected_week_month and selected_week_number for template
                selected_week_month = f"{today.year}-{today.month:02d}"
                # Calculate which week of the month we're in
                first_day = datetime(today.year, today.month, 1)
                first_monday = first_day - timedelta(days=first_day.weekday())
                week_num = ((start_date - first_monday).days // 7) + 1
                if week_num < 1:
                    week_num = 1
                elif week_num > 5:
                    week_num = 5
                selected_week_number = str(week_num)
            
            date_filter = "AND DATE(a.applied_at) BETWEEN %s AND %s"
            date_params = [start_date.date(), end_date.date()]
            period_summary = {
                'start_date': start_date.strftime('%b %d, %Y'),
                'end_date': end_date.strftime('%b %d, %Y'),
                'total_days': (end_date.date() - start_date.date()).days + 1,
                'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p')
            }
        elif period == 'month':
            if selected_month:
                try:
                    year, month = map(int, selected_month.split('-'))
                    start_date = datetime(year, month, 1)
                    if month == 12:
                        end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
                    else:
                        end_date = datetime(year, month + 1, 1) - timedelta(days=1)
                except (ValueError, IndexError):
                    # Default to current month if parsing fails
                    today = datetime.now()
                    start_date = datetime(today.year, today.month, 1)
                    if today.month == 12:
                        end_date = datetime(today.year + 1, 1, 1) - timedelta(days=1)
                    else:
                        end_date = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
                    # Update selected_month to current month for template
                    selected_month = f"{today.year}-{today.month:02d}"
            else:
                # Default to current month
                today = datetime.now()
                start_date = datetime(today.year, today.month, 1)
                if today.month == 12:
                    end_date = datetime(today.year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(today.year, today.month + 1, 1) - timedelta(days=1)
                # Set default selected_month for template
                selected_month = f"{today.year}-{today.month:02d}"
            
            date_filter = "AND DATE(a.applied_at) BETWEEN %s AND %s"
            date_params = [start_date.date(), end_date.date()]
            period_summary = {
                'start_date': start_date.strftime('%b %d, %Y'),
                'end_date': end_date.strftime('%b %d, %Y'),
                'total_days': (end_date - start_date).days + 1,
                'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p')
            }
        elif period == 'year':
            if selected_year:
                try:
                    year = int(selected_year)
                except ValueError:
                    year = datetime.now().year
            else:
                # Default to current year
                year = datetime.now().year
                # Set default selected_year for template
                selected_year = str(year)
            
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31)
            date_filter = "AND YEAR(a.applied_at) = %s"
            date_params = [year]
            period_summary = {
                'start_date': start_date.strftime('%b %d, %Y'),
                'end_date': end_date.strftime('%b %d, %Y'),
                'total_days': 365 if year % 4 != 0 else 366,
                'generated_at': datetime.now().strftime('%b %d, %Y %I:%M %p')
            }
        
        # Get available years for dropdown
        cursor.execute("SELECT DISTINCT YEAR(applied_at) as year FROM applications WHERE applied_at IS NOT NULL ORDER BY year DESC")
        available_years = [str(row['year']) for row in cursor.fetchall() if row.get('year')]
        
        # Get user's branch scope first (needed for filtering)
        branch_id = get_branch_scope(user)
        
        # Get branches for filter dropdown
        branches = fetch_branches()
        
        # Get unique job titles for filter dropdown
        branch_filter_id = request.args.get('branch_id', type=int)
        
        # Build job filter query - filter by branch if specified, otherwise show all jobs
        job_filter_where = ""
        job_filter_params = []
        if branch_filter_id:
            job_filter_where = "WHERE j.branch_id = %s"
            job_filter_params = [branch_filter_id]
        elif branch_id:
            # If user has branch scope, limit jobs to that branch
            job_filter_where = "WHERE j.branch_id = %s"
            job_filter_params = [branch_id]
        
        cursor.execute(f'''
            SELECT DISTINCT j.job_id, COALESCE(j.job_title, 'Untitled Job') AS job_title
            FROM jobs j
            {job_filter_where}
            ORDER BY j.job_title ASC
        ''', tuple(job_filter_params) if job_filter_params else None)
        jobs = cursor.fetchall() or []
        
        # Determine available job columns for backward compatibility
        cursor.execute('SHOW COLUMNS FROM jobs')
        job_columns = {row.get('Field') for row in (cursor.fetchall() or []) if row}

        def job_expr(candidates, fallback):
            for column in candidates:
                if column in job_columns:
                    return f'j.{column}'
            return fallback

        def job_coalesce(candidates, fallback):
            expressions = [f'j.{column}' for column in candidates if column in job_columns]
            if not expressions:
                return fallback
            if len(expressions) == 1:
                return expressions[0]
            return f"COALESCE({', '.join(expressions)})"

        job_title_expr = job_expr(['job_title', 'title'], "'Untitled Job'")
        job_posted_expr = job_coalesce(['posted_at', 'created_at'], 'j.created_at')
        time_to_apply_expr = f'CASE WHEN a.applied_at IS NOT NULL AND {job_posted_expr} IS NOT NULL THEN DATEDIFF(a.applied_at, {job_posted_expr}) ELSE NULL END'
        where_templates = []
        params_reports = []
        
        # Apply branch filter from request (takes precedence over user's branch scope)
        # branch_filter_id is already retrieved above for job filtering
        if branch_filter_id:
            where_templates.append('{alias}.branch_id = %s')
            params_reports.append(branch_filter_id)
        elif branch_id:
            # Use user's branch scope only if no filter is specified
            where_templates.append('{alias}.branch_id = %s')
            params_reports.append(branch_id)
        
        # Apply job filter from request
        job_filter_id = request.args.get('job_id', type=int)
        if job_filter_id:
            where_templates.append('{alias}.job_id = %s')
            params_reports.append(job_filter_id)
        
        # Apply HR User filter from request
        hr_filter_id = request.args.get('hr_user_id', type=int)
        hr_filter_condition = None
        if hr_filter_id:
            # HR filter will be applied separately to queries that need it
            hr_filter_condition = hr_filter_id
        
        # Apply Status filter from request
        status_filter = request.args.get('status', '').strip()
        status_filter_condition = None
        if status_filter and status_filter.lower() != 'all':
            status_filter_condition = status_filter.lower()

        def build_where(alias):
            if not where_templates:
                return ''
            clause = ' AND '.join(template.format(alias=alias) for template in where_templates)
            return f'WHERE {clause}'

        base_where = build_where('j')
        
        # Apply date filter to all queries
        date_filter_clause = date_filter if date_filter else ''
        date_params_all = date_params if date_params else []
        
        # Comprehensive statistics (with period filter)
        stats = build_report_stats(user, date_filter_clause, date_params_all)
        
        # Comprehensive recruitment metrics
        metrics = {}
        try:
            if base_where:
                # For branch-scoped users, total_branches should be 1 (their branch)
                # But we'll get it from branches table to be accurate
                branch_count_query = 'SELECT COUNT(*) AS count FROM branches WHERE branch_id = %s'
                branch_count_params = (branch_id,)
                total_branches_count = fetch_count(branch_count_query, branch_count_params) or 0
                
                metrics_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT j.job_id) AS total_jobs,
                        {total_branches_count} AS total_branches,
                        COUNT(DISTINCT i.interview_id) AS total_interviews,
                        AVG({time_to_apply_expr}) AS avg_time_to_apply,
                        AVG(CASE WHEN a.applied_at IS NOT NULL AND i.scheduled_date IS NOT NULL 
                            THEN DATEDIFF(i.scheduled_date, a.applied_at) ELSE NULL END) AS avg_time_to_interview,
                        AVG(CASE WHEN a.status = 'hired' AND a.applied_at IS NOT NULL AND a.viewed_at IS NOT NULL 
                            THEN DATEDIFF(a.viewed_at, a.applied_at) ELSE NULL END) AS avg_time_to_hire,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS total_hires,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS applications_with_interviews,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) AS hired_with_interview,
                        CASE 
                            WHEN COUNT(*) > 0 THEN SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) / COUNT(*) * 100 
                            ELSE 0 
                        END AS acceptance_rate,
                        CASE 
                            WHEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) * 100
                            ELSE 0 
                        END AS interview_to_hire_rate,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS application_to_interview_rate
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                    {base_where}
                    '''
                cursor.execute(metrics_query, tuple(params_reports) if params_reports else None)
            else:
                # For admin (no branch scope), count all branches from branches table
                total_branches_count = fetch_count('SELECT COUNT(*) AS count FROM branches') or 0
                
                metrics_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT j.job_id) AS total_jobs,
                        {total_branches_count} AS total_branches,
                        COUNT(DISTINCT i.interview_id) AS total_interviews,
                        AVG({time_to_apply_expr}) AS avg_time_to_apply,
                        AVG(CASE WHEN a.applied_at IS NOT NULL AND i.scheduled_date IS NOT NULL 
                            THEN DATEDIFF(i.scheduled_date, a.applied_at) ELSE NULL END) AS avg_time_to_interview,
                        AVG(CASE WHEN a.status = 'hired' AND a.applied_at IS NOT NULL AND a.viewed_at IS NOT NULL 
                            THEN DATEDIFF(a.viewed_at, a.applied_at) ELSE NULL END) AS avg_time_to_hire,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS total_hires,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS applications_with_interviews,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) AS hired_with_interview,
                        CASE 
                            WHEN COUNT(*) > 0 THEN SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) / COUNT(*) * 100 
                            ELSE 0 
                        END AS acceptance_rate,
                        CASE 
                            WHEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) * 100
                            ELSE 0 
                        END AS interview_to_hire_rate,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS application_to_interview_rate
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                '''
                cursor.execute(metrics_query)
            metrics = cursor.fetchone() or {}
        except Exception as metrics_exc:
            print(f'⚠️ Error fetching metrics: {metrics_exc}')
            import traceback
            print(traceback.format_exc())
            metrics = {}
        
        # Recruitment funnel data
        funnel_data = {}
        try:
            if base_where:
                funnel_query = f'''
                    SELECT 
                        COUNT(DISTINCT CASE WHEN a.status = 'applied' THEN a.application_id END) AS applied,
                        COUNT(DISTINCT CASE WHEN a.status = 'under_review' THEN a.application_id END) AS under_review,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS interviewed,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                    {base_where}
                '''
                cursor.execute(funnel_query, tuple(params_reports))
            else:
                funnel_query = '''
                    SELECT 
                        COUNT(DISTINCT CASE WHEN a.status = 'applied' THEN a.application_id END) AS applied,
                        COUNT(DISTINCT CASE WHEN a.status = 'under_review' THEN a.application_id END) AS under_review,
                        COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) AS interviewed,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected
                    FROM applications a
                    LEFT JOIN interviews i ON i.application_id = a.application_id
                '''
                cursor.execute(funnel_query)
            funnel_data = cursor.fetchone() or {}
        except Exception as funnel_exc:
            print(f'⚠️ Error fetching funnel data: {funnel_exc}')
            funnel_data = {}
        
        # Source effectiveness (if source field exists in applications)
        source_effectiveness = []
        try:
            cursor.execute('SHOW COLUMNS FROM applications LIKE %s', ('source%',))
            has_source = cursor.fetchone() is not None
            if has_source and base_where:
                source_query = f'''
                    SELECT 
                        COALESCE(a.source, 'Unknown') AS source,
                        COUNT(*) AS total_applications,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hires,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS conversion_rate
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {base_where}
                    GROUP BY a.source
                    ORDER BY total_applications DESC
                '''
                cursor.execute(source_query, tuple(params_reports))
                source_effectiveness = cursor.fetchall() or []
            elif has_source:
                source_query = '''
                    SELECT 
                        COALESCE(a.source, 'Unknown') AS source,
                        COUNT(*) AS total_applications,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hires,
                        CASE 
                            WHEN COUNT(*) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(*) * 100
                            ELSE 0 
                        END AS conversion_rate
                    FROM applications a
                    GROUP BY a.source
                    ORDER BY total_applications DESC
                '''
                cursor.execute(source_query)
                source_effectiveness = cursor.fetchall() or []
        except Exception:
            source_effectiveness = []
        
        # Trends (last 30 days)
        trends = []
        try:
            trends_where_parts = ['a.applied_at >= DATE_SUB(NOW(), INTERVAL 30 DAY)']
            trends_params = []
            if base_where:
                # Extract the WHERE condition from base_where and add it
                where_condition = base_where.replace('WHERE ', '')
                trends_where_parts.append(where_condition)
                trends_params.extend(params_reports)
            
            trends_where_clause = 'WHERE ' + ' AND '.join(trends_where_parts) if trends_where_parts else ''
            trends_query = f'''
                SELECT 
                    DATE(a.applied_at) AS date,
                    COUNT(*) AS applications,
                    SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS accepted,
                    SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {trends_where_clause}
                GROUP BY DATE(a.applied_at)
                ORDER BY date DESC
            '''
            cursor.execute(trends_query, tuple(trends_params) if trends_params else None)
            trends = cursor.fetchall() or []
        except Exception as trends_exc:
            print(f'⚠️ Error fetching trends: {trends_exc}')
            trends = []
        
        # Branch statistics (admin only)
        branch_stats = []
        if user.get('role') == 'admin':
            cursor.execute(
                '''
                SELECT 
                    b.branch_id,
                    b.branch_name,
                    COUNT(DISTINCT a.application_id) AS total_applications,
                    SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS accepted,
                    SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected,
                    COUNT(DISTINCT j.job_id) AS total_jobs,
                    COUNT(DISTINCT i.interview_id) AS total_interviews
                FROM branches b
                LEFT JOIN jobs j ON j.branch_id = b.branch_id
                LEFT JOIN applications a ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                GROUP BY b.branch_id, b.branch_name
                ORDER BY total_applications DESC
                '''
            )
            branch_stats = cursor.fetchall()
        
        # Job performance
        job_performance = []
        try:
            cursor.execute(
                f'''
                SELECT 
                    j.job_id,
                    {job_title_expr} AS job_title,
                    COUNT(a.application_id) AS application_count,
                    SUM(CASE WHEN a.status = 'hired' THEN 1 ELSE 0 END) AS accepted_count,
                    SUM(CASE WHEN a.status = 'rejected' THEN 1 ELSE 0 END) AS rejected_count,
                    COUNT(DISTINCT i.interview_id) AS interview_count
                FROM jobs j
                LEFT JOIN applications a ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                {base_where if base_where else ''}
                GROUP BY j.job_id, {job_title_expr}
                ORDER BY application_count DESC
                LIMIT 10
                ''',
                tuple(params_reports) if params_reports and base_where else None,
            )
            job_performance = cursor.fetchall() or []
        except Exception as job_perf_exc:
            print(f'⚠️ Error fetching job performance: {job_perf_exc}')
            job_performance = []
        
        # Status breakdown - compute percentages in Python for compatibility
        status_breakdown = []
        try:
            status_query = f'''
                SELECT 
                    a.status,
                    COUNT(*) AS count
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {base_where if base_where else ''}
                GROUP BY a.status
                ORDER BY count DESC
            '''
            cursor.execute(status_query, tuple(params_reports) if params_reports and base_where else None)
            status_breakdown = cursor.fetchall() or []
            total_status = sum((row.get('count') or 0) for row in status_breakdown)
            for row in status_breakdown:
                count_value = row.get('count') or 0
                row['percentage'] = (count_value / total_status * 100) if total_status else 0
        except Exception as status_exc:
            print(f'⚠️ Error fetching status breakdown: {status_exc}')
            status_breakdown = []
        
        # Cross-branch comparison (only for Super Admin)
        branch_comparison = []
        if user.get('role') == 'admin' and not branch_id:
            cursor.execute(
                '''
                SELECT 
                    b.branch_id,
                    b.branch_name,
                    COUNT(DISTINCT j.job_id) AS total_jobs,
                    COUNT(DISTINCT a.application_id) AS total_applications,
                    COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                    COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.application_id END) AS pending_count,
                    COUNT(DISTINCT i.interview_id) AS total_interviews,
                    COUNT(DISTINCT a.applicant_id) AS unique_applicants
                FROM branches b
                LEFT JOIN jobs j ON j.branch_id = b.branch_id
                LEFT JOIN applications a ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                GROUP BY b.branch_id, b.branch_name
                ORDER BY total_applications DESC
                '''
            )
            branch_comparison = cursor.fetchall()
        
        # System usage statistics (only for Super Admin)
        system_stats = {}
        if user.get('role') == 'admin' and not branch_id:
            try:
                cursor.execute('SELECT COUNT(*) AS total_users FROM users')
                system_stats['total_users'] = cursor.fetchone()['total_users'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_admins FROM users WHERE user_type = "super_admin"')
                system_stats['total_admins'] = cursor.fetchone()['total_admins'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_hr FROM users WHERE user_type = "hr"')
                system_stats['total_hr'] = cursor.fetchone()['total_hr'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_applicants FROM users WHERE user_type = "applicant"')
                system_stats['total_applicants'] = cursor.fetchone()['total_applicants'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_jobs_posted FROM jobs')
                system_stats['total_jobs_posted'] = cursor.fetchone()['total_jobs_posted'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_applications_received FROM applications')
                system_stats['total_applications_received'] = cursor.fetchone()['total_applications_received'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_interviews_scheduled FROM interviews')
                system_stats['total_interviews_scheduled'] = cursor.fetchone()['total_interviews_scheduled'] or 0
                
                cursor.execute('SELECT COUNT(*) AS total_hires FROM applications WHERE status = "hired"')
                system_stats['total_hires'] = cursor.fetchone()['total_hires'] or 0
                
                # Some schemas use 'created_at' instead of 'login_time'
                cursor.execute('SELECT COUNT(*) AS total_sessions_24h FROM auth_sessions WHERE created_at >= DATE_SUB(NOW(), INTERVAL 24 HOUR)')
                system_stats['total_sessions_24h'] = cursor.fetchone()['total_sessions_24h'] or 0
            except Exception as e:
                print(f'⚠️ Error fetching system stats: {e}')
                system_stats = {
                    'total_users': 0,
                    'total_admins': 0,
                    'total_hr': 0,
                    'total_applicants': 0,
                    'total_jobs_posted': 0,
                    'total_applications_received': 0,
                    'total_interviews_scheduled': 0,
                    'total_hires': 0,
                    'total_sessions_24h': 0,
                }
        
        # Generate report sections data
        # 1. Applicant Summary Report
        applicant_summary = {}
        applicant_summary_details = []
        try:
            # Ensure date_filter is properly formatted
            if date_filter and not date_filter.strip().startswith('AND'):
                date_filter = f"AND {date_filter.strip()}"
            # Check if email_verified column exists in applicants or users table
            cursor.execute("SHOW COLUMNS FROM applicants LIKE 'email_verified'")
            has_applicant_email_verified = cursor.fetchone() is not None
            
            cursor.execute("SHOW COLUMNS FROM users LIKE 'email_verified'")
            has_user_email_verified = cursor.fetchone() is not None
            
            # Build verified_email expression
            if has_applicant_email_verified:
                verified_email_expr = "CASE WHEN ap.email_verified = 1 THEN ap.applicant_id END"
            elif has_user_email_verified:
                verified_email_expr = "CASE WHEN u.email_verified = 1 THEN ap.applicant_id END"
            else:
                verified_email_expr = "NULL"  # Return 0 if column doesn't exist
            
            # Build date filter clause for applicant query
            date_clause_applicant = ""
            if date_filter:
                # For applicant summary, filter by application submission date
                # Remove leading 'AND ' only, not all occurrences
                app_date_filter = date_filter.strip()
                if app_date_filter.startswith('AND '):
                    app_date_filter = app_date_filter[4:]  # Remove 'AND ' from start
                date_clause_applicant = f" AND {app_date_filter}"
            
            if branch_id:
                join_users = "LEFT JOIN users u ON ap.user_id = u.user_id" if has_user_email_verified and not has_applicant_email_verified else ""
                applicant_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN DATE(ap.created_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN ap.applicant_id END) AS new_applicants,
                        COUNT(DISTINCT CASE WHEN a.status IN ('pending', 'scheduled', 'interviewed') THEN a.application_id END) AS active_applications,
                        COUNT(DISTINCT CASE WHEN r.resume_id IS NOT NULL THEN ap.applicant_id END) AS with_resume,
                        COUNT(DISTINCT {verified_email_expr}) AS verified_email
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN resumes r ON ap.applicant_id = r.applicant_id
                    {join_users}
                    WHERE j.branch_id = %s{date_clause_applicant}
                '''
                applicant_params = [branch_id] + (date_params_all if date_params_all else [])
            else:
                join_users = "LEFT JOIN users u ON ap.user_id = u.user_id" if has_user_email_verified and not has_applicant_email_verified else ""
                where_clause = ""
                if date_filter:
                    # Remove leading 'AND ' only, not all occurrences
                    app_date_filter = date_filter.strip()
                    if app_date_filter.startswith('AND '):
                        app_date_filter = app_date_filter[4:]  # Remove 'AND ' from start
                    where_clause = f"WHERE {app_date_filter}"
                applicant_query = f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN DATE(ap.created_at) >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN ap.applicant_id END) AS new_applicants,
                        COUNT(DISTINCT CASE WHEN a.status IN ('pending', 'scheduled', 'interviewed') THEN a.application_id END) AS active_applications,
                        COUNT(DISTINCT CASE WHEN r.resume_id IS NOT NULL THEN ap.applicant_id END) AS with_resume,
                        COUNT(DISTINCT {verified_email_expr}) AS verified_email
                    FROM applicants ap
                    LEFT JOIN applications a ON ap.applicant_id = a.applicant_id
                    LEFT JOIN resumes r ON ap.applicant_id = r.applicant_id
                    {join_users}
                    {where_clause}
                '''
                applicant_params = date_params_all if date_params_all else []
            
            try:
                cursor.execute(applicant_query, tuple(applicant_params) if applicant_params else None)
                applicant_summary = cursor.fetchone() or {}
            except Exception as query_error:
                print(f'⚠️ Error in applicant query: {query_error}')
                applicant_summary = {}
            
            # Get detailed applicant list with job applications and status
            # Build WHERE clause using the same filters as other queries
            detail_where_parts = []
            detail_params = []
            
            # Apply branch filter (from request takes precedence, otherwise user's branch scope)
            if branch_filter_id:
                detail_where_parts.append("j.branch_id = %s")
                detail_params.append(branch_filter_id)
            elif branch_id:
                detail_where_parts.append("j.branch_id = %s")
                detail_params.append(branch_id)
            
            # Apply job filter
            if job_filter_id:
                detail_where_parts.append("j.job_id = %s")
                detail_params.append(job_filter_id)
            
            # Apply status filter
            if status_filter_condition:
                detail_where_parts.append("a.status = %s")
                detail_params.append(status_filter_condition)
            
            # Apply date filter
            if date_filter and date_params_all:
                app_date_filter = date_filter.strip()
                if app_date_filter.startswith('AND '):
                    app_date_filter = app_date_filter[4:]
                detail_where_parts.append(app_date_filter)
                detail_params.extend(date_params_all)
            
            # Build WHERE clause
            if detail_where_parts:
                detail_where_clause = "WHERE " + " AND ".join(detail_where_parts)
            else:
                detail_where_clause = ""
            
            detail_query = f'''
                SELECT 
                    ap.full_name AS applicant_name,
                    ap.email,
                    {job_title_expr} AS job_title,
                    b.branch_name,
                    a.status AS application_status,
                    a.applied_at AS date_applied,
                    a.viewed_at AS date_updated,
                    i.scheduled_date AS interview_scheduled_date,
                    CASE 
                        WHEN a.status = 'hired' AND a.applied_at IS NOT NULL AND a.viewed_at IS NOT NULL
                        THEN DATEDIFF(a.viewed_at, a.applied_at)
                        ELSE NULL
                    END AS time_to_hire_days,
                    CASE 
                        WHEN a.status = 'hired' THEN 'Hired'
                        WHEN a.status = 'rejected' THEN 'Not Hired'
                        WHEN a.status = 'pending' THEN 'Pending'
                        WHEN a.status = 'scheduled' THEN 'Interview Scheduled'
                        WHEN a.status = 'interviewed' THEN 'Interviewed'
                        ELSE 'Unknown'
                    END AS status_label,
                    CASE 
                        WHEN a.status = 'hired' THEN a.viewed_at
                        WHEN a.status = 'rejected' THEN a.viewed_at
                        WHEN a.status = 'scheduled' THEN i.scheduled_date
                        WHEN a.status = 'interviewed' THEN a.viewed_at
                        WHEN a.status = 'pending' THEN a.applied_at
                        ELSE a.viewed_at
                    END AS status_date,
                    COALESCE(
                        (SELECT ad.full_name
                         FROM activity_logs al_log
                         LEFT JOIN admins ad ON al_log.user_id = ad.admin_id
                         WHERE (al_log.action IN ('Updated application status', 'Hired applicant', 'Rejected applicant', 'Scheduled interview')
                                OR al_log.description LIKE CONCAT('%', a.application_id, '%'))
                         ORDER BY al_log.logged_at DESC
                         LIMIT 1),
                        'Unassigned'
                    ) AS hr_assigned
                FROM applicants ap
                JOIN applications a ON ap.applicant_id = a.applicant_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                LEFT JOIN (
                    SELECT i1.application_id, i1.scheduled_date
                    FROM interviews i1
                    INNER JOIN (
                        SELECT application_id, MAX(scheduled_date) AS max_date
                        FROM interviews
                        GROUP BY application_id
                    ) i2 ON i1.application_id = i2.application_id AND i1.scheduled_date = i2.max_date
                ) i ON i.application_id = a.application_id
                {detail_where_clause}
                ORDER BY a.applied_at DESC
                LIMIT 200
            '''
            
            try:
                cursor.execute(detail_query, tuple(detail_params) if detail_params else None)
                applicant_summary_details = cursor.fetchall() or []
            except Exception as detail_error:
                print(f'⚠️ Error in applicant detail query: {detail_error}')
                print(f'Query: {detail_query}')
                print(f'Params: {detail_params}')
                applicant_summary_details = []
            
            # Format dates and enhance status information
            for app in applicant_summary_details:
                # Format date_applied
                if app.get('date_applied'):
                    try:
                        if isinstance(app['date_applied'], (datetime, date_class)):
                            app['date_applied'] = app['date_applied'].strftime('%b %d, %Y %I:%M %p')
                        elif isinstance(app['date_applied'], str):
                            dt = datetime.strptime(app['date_applied'], '%Y-%m-%d %H:%M:%S')
                            app['date_applied'] = dt.strftime('%b %d, %Y %I:%M %p')
                    except Exception:
                        # Fallback to original value if formatting fails
                        app['date_applied'] = app.get('date_applied', '—')
                
                # Format interview_scheduled_date
                if app.get('interview_scheduled_date'):
                    try:
                        if isinstance(app['interview_scheduled_date'], (datetime, date_class)):
                            app['interview_date'] = app['interview_scheduled_date'].strftime('%b %d, %Y %I:%M %p')
                        elif isinstance(app['interview_scheduled_date'], str):
                            dt = datetime.strptime(app['interview_scheduled_date'], '%Y-%m-%d %H:%M:%S')
                            app['interview_date'] = dt.strftime('%b %d, %Y %I:%M %p')
                    except Exception:
                        # Fallback to original value if formatting fails
                        app['interview_date'] = app.get('interview_scheduled_date', '—')
                else:
                    app['interview_date'] = '—'
                
                # Format time_to_hire
                time_to_hire_days = app.get('time_to_hire_days')
                if time_to_hire_days is not None:
                    if time_to_hire_days == 0:
                        app['time_to_hire'] = 'Same day'
                    elif time_to_hire_days == 1:
                        app['time_to_hire'] = '1 day'
                    else:
                        app['time_to_hire'] = f'{time_to_hire_days} days'
                else:
                    app['time_to_hire'] = '—'
                
                # Format date_updated (hired date if status is hired)
                if app.get('date_updated'):
                    try:
                        if isinstance(app['date_updated'], (datetime, date_class)):
                            app['date_updated'] = app['date_updated'].strftime('%b %d, %Y %I:%M %p')
                        elif isinstance(app['date_updated'], str):
                            dt = datetime.strptime(app['date_updated'], '%Y-%m-%d %H:%M:%S')
                            app['date_updated'] = dt.strftime('%b %d, %Y %I:%M %p')
                    except Exception:
                        # Fallback to original value if formatting fails
                        app['date_updated'] = app.get('date_updated', '—')
                
                # Set last_activity for backward compatibility
                app['last_activity'] = app.get('date_applied', '—')
                app['current_status'] = app.get('status_label', app.get('application_status', '—'))
            
        except Exception as e:
            print(f'⚠️ Error fetching applicant summary: {e}')
            applicant_summary = {}
            applicant_summary_details = []
        
        # Group applicants by status after fetching details
        applicants_by_status = {
            'hired': [],
            'rejected': [],
            'pending': [],
            'scheduled': [],
            'interviewed': []
        }
        for app in applicant_summary_details:
            status = app.get('application_status', '').lower()
            if status in applicants_by_status:
                applicants_by_status[status].append(app)
        
        # 2. Job Vacancy Report
        job_vacancy = {}
        job_vacancy_details = []
        try:
            job_where = base_where
            job_params = list(params_reports) if params_reports else []
            
            # Apply date filter for job postings
            if date_filter:
                job_date_filter = date_filter.replace('a.applied_at', job_posted_expr)
                if job_where:
                    job_where += f" {job_date_filter}"
                else:
                    # Remove leading 'AND ' only, not all occurrences
                    job_date_clean = job_date_filter.strip()
                    if job_date_clean.startswith('AND '):
                        job_date_clean = job_date_clean[4:]  # Remove 'AND ' from start
                    job_where = f"WHERE {job_date_clean}"
                job_params.extend(date_params_all)
            
            try:
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT j.job_id) AS total_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'active' THEN j.job_id END) AS active_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'closed' THEN j.job_id END) AS closed_jobs,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        CASE 
                            WHEN COUNT(DISTINCT j.job_id) > 0 
                            THEN COUNT(DISTINCT a.application_id) / COUNT(DISTINCT j.job_id)
                            ELSE 0 
                        END AS avg_applications_per_job
                    FROM jobs j
                    LEFT JOIN applications a ON j.job_id = a.job_id
                    {job_where if job_where else ''}
                ''', tuple(job_params) if job_params else None)
                job_vacancy = cursor.fetchone() or {}
            except Exception as job_error:
                print(f'⚠️ Error in job vacancy query: {job_error}')
                job_vacancy = {}
            
            # Get detailed job list
            try:
                cursor.execute(f'''
                    SELECT 
                        {job_title_expr} AS job_title,
                        b.branch_name,
                        j.status,
                        COUNT(DISTINCT a.application_id) AS application_count,
                        DATE({job_posted_expr}) AS posted_date
                    FROM jobs j
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    LEFT JOIN applications a ON j.job_id = a.job_id
                    {job_where if job_where else ''}
                    GROUP BY j.job_id, {job_title_expr}, b.branch_name, j.status, posted_date
                    ORDER BY posted_date DESC
                    LIMIT 50
                ''', tuple(job_params) if job_params else None)
                job_vacancy_details = cursor.fetchall() or []
            except Exception as job_detail_error:
                print(f'⚠️ Error in job vacancy detail query: {job_detail_error}')
                job_vacancy_details = []
            
            # Format dates
            for job in job_vacancy_details:
                if job.get('posted_date'):
                    try:
                        if isinstance(job['posted_date'], (datetime, date_class)):
                            job['posted_date'] = job['posted_date'].strftime('%b %d, %Y')
                        elif isinstance(job['posted_date'], str):
                            dt = datetime.strptime(job['posted_date'], '%Y-%m-%d')
                            job['posted_date'] = dt.strftime('%b %d, %Y')
                    except Exception:
                        # Ignore formatting errors
                        pass
        except Exception as e:
            print(f'⚠️ Error fetching job vacancy: {e}')
            job_vacancy = {}
            job_vacancy_details = []
        
        # Enhanced Applicant Summary - Branch Breakdown
        applicants_by_branch = []
        try:
            # When "All Branches" is selected, start from branches table to show all branches
            # When a specific branch is selected, we can filter accordingly
            if branch_filter_id:
                # Specific branch selected - can use the original approach
                branch_applicant_where = base_where
                branch_applicant_params = list(params_reports) if params_reports else []
                
                if date_filter:
                    branch_applicant_date_filter = date_filter.replace('a.applied_at', 'a.applied_at')
                    if branch_applicant_where:
                        branch_applicant_where += f" {branch_applicant_date_filter}"
                    else:
                        branch_date_clean = branch_applicant_date_filter.strip()
                        if branch_date_clean.startswith('AND '):
                            branch_date_clean = branch_date_clean[4:]
                        branch_applicant_where = f"WHERE {branch_date_clean}"
                    branch_applicant_params.extend(date_params_all)
                
                cursor.execute(f'''
                    SELECT 
                        COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.application_id END) AS pending_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'scheduled' THEN a.application_id END) AS scheduled_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'interviewed' THEN a.application_id END) AS interviewed_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    {branch_applicant_where if branch_applicant_where else ''}
                    GROUP BY b.branch_id, b.branch_name
                    ORDER BY total_applications DESC
                ''', tuple(branch_applicant_params) if branch_applicant_params else None)
                applicants_by_branch = cursor.fetchall() or []
            else:
                # All Branches selected - start from branches table to show ALL branches
                # But respect user's branch scope if they have one
                branch_params = []
                branch_scope_clause = ""
                
                # Apply user's branch scope if they have one (HR users with branch assignment)
                if branch_id:
                    branch_scope_clause = "WHERE b.branch_id = %s"
                    branch_params.append(branch_id)
                
                # Build date filter clause for JOIN condition
                date_filter_clause = ""
                if date_filter and date_params_all:
                    # Convert WHERE-style filter to JOIN condition
                    if "BETWEEN" in date_filter:
                        date_filter_clause = "AND DATE(a.applied_at) BETWEEN %s AND %s"
                        branch_params.extend(date_params_all)
                    elif "YEAR(a.applied_at)" in date_filter:
                        date_filter_clause = "AND YEAR(a.applied_at) = %s"
                        branch_params.extend(date_params_all)
                    else:
                        date_filter_clause = date_filter.replace('AND ', 'AND ')
                        branch_params.extend(date_params_all)
                
                # Build job filter clause
                job_filter_clause = ""
                if job_filter_id:
                    job_filter_clause = "AND j.job_id = %s"
                    branch_params.append(job_filter_id)
                
                # Build the query starting from branches
                # Note: build WHERE clause after JOINs to avoid invalid SQL order
                cursor.execute(f'''
                    SELECT 
                        COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                        b.branch_id,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        COUNT(DISTINCT CASE WHEN a.status = 'pending' THEN a.application_id END) AS pending_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'scheduled' THEN a.application_id END) AS scheduled_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'interviewed' THEN a.application_id END) AS interviewed_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                    FROM branches b
                    LEFT JOIN jobs j ON j.branch_id = b.branch_id {job_filter_clause}
                    LEFT JOIN applications a ON a.job_id = j.job_id {date_filter_clause}
                    LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    {branch_scope_clause}
                    GROUP BY b.branch_id, b.branch_name
                    ORDER BY total_applications DESC, b.branch_name ASC
                ''', tuple(branch_params) if branch_params else None)
                applicants_by_branch = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching applicants by branch: {e}')
            import traceback
            print(traceback.format_exc())
            applicants_by_branch = []
        
        # Most/Least Applied Jobs
        most_applied_jobs = []
        least_applied_jobs = []
        try:
            job_applicant_where = base_where
            job_applicant_params = list(params_reports) if params_reports else []
            
            if date_filter:
                job_applicant_date_filter = date_filter.replace('a.applied_at', 'a.applied_at')
                if job_applicant_where:
                    job_applicant_where += f" {job_applicant_date_filter}"
                else:
                    job_date_clean = job_applicant_date_filter.strip()
                    if job_date_clean.startswith('AND '):
                        job_date_clean = job_date_clean[4:]
                    job_applicant_where = f"WHERE {job_date_clean}"
                job_applicant_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    {job_title_expr} AS job_title,
                    b.branch_name,
                    COUNT(DISTINCT a.application_id) AS application_count
                FROM jobs j
                LEFT JOIN applications a ON j.job_id = a.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                {job_applicant_where if job_applicant_where else ''}
                GROUP BY j.job_id, {job_title_expr}, b.branch_name
                HAVING application_count > 0
                ORDER BY application_count DESC
                LIMIT 10
            ''', tuple(job_applicant_params) if job_applicant_params else None)
            most_applied_jobs = cursor.fetchall() or []
            
            cursor.execute(f'''
                SELECT 
                    {job_title_expr} AS job_title,
                    b.branch_name,
                    COUNT(DISTINCT a.application_id) AS application_count
                FROM jobs j
                LEFT JOIN applications a ON j.job_id = a.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                {job_applicant_where if job_applicant_where else ''}
                GROUP BY j.job_id, {job_title_expr}, b.branch_name
                HAVING application_count > 0
                ORDER BY application_count ASC
                LIMIT 10
            ''', tuple(job_applicant_params) if job_applicant_params else None)
            least_applied_jobs = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching most/least applied jobs: {e}')
            most_applied_jobs = []
            least_applied_jobs = []
        
        # 3. Hiring Outcome Report
        hiring_outcome = {}
        try:
            hiring_where = base_where
            hiring_params = list(params_reports) if params_reports else []
            
            if date_filter:
                if hiring_where:
                    hiring_where += f" {date_filter}"
                else:
                    # Remove leading 'AND ' only, not all occurrences
                    date_filter_clean = date_filter.strip()
                    if date_filter_clean.startswith('AND '):
                        date_filter_clean = date_filter_clean[4:]  # Remove 'AND ' from start
                    hiring_where = f"WHERE {date_filter_clean}"
                hiring_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS total_hired,
                    CASE 
                        WHEN COUNT(*) > 0 
                        THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(*) * 100
                        ELSE 0 
                    END AS hire_rate,
                    AVG(CASE WHEN a.status = 'hired' AND a.applied_at IS NOT NULL AND a.viewed_at IS NOT NULL 
                        THEN DATEDIFF(a.viewed_at, a.applied_at) ELSE NULL END) AS avg_time_to_hire,
                    CASE 
                        WHEN COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) > 0 
                        THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' AND i.interview_id IS NOT NULL THEN a.application_id END) / COUNT(DISTINCT CASE WHEN i.interview_id IS NOT NULL THEN a.application_id END) * 100
                        ELSE 0 
                    END AS interview_to_hire_rate,
                    CASE 
                        WHEN COUNT(*) > 0 
                        THEN COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) / COUNT(*) * 100
                        ELSE 0 
                    END AS rejection_rate
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN interviews i ON i.application_id = a.application_id
                {hiring_where if hiring_where else ''}
            ''', tuple(hiring_params) if hiring_params else None)
            hiring_outcome = cursor.fetchone() or {}
            
            # Round percentages
            if hiring_outcome.get('hire_rate'):
                hiring_outcome['hire_rate'] = round(hiring_outcome['hire_rate'], 1)
            if hiring_outcome.get('interview_to_hire_rate'):
                hiring_outcome['interview_to_hire_rate'] = round(hiring_outcome['interview_to_hire_rate'], 1)
            if hiring_outcome.get('rejection_rate'):
                hiring_outcome['rejection_rate'] = round(hiring_outcome['rejection_rate'], 1)
            if hiring_outcome.get('avg_time_to_hire'):
                hiring_outcome['avg_time_to_hire'] = round(hiring_outcome['avg_time_to_hire'] or 0, 1)
        except Exception as e:
            print(f'⚠️ Error fetching hiring outcome: {e}')
            hiring_outcome = {}
        
        # Interview Schedule Report (with no-shows and cancelled)
        interview_schedule = {}
        try:
            interview_where = base_where if base_where else ""
            interview_params = list(params_reports) if params_reports else []
            
            if date_filter:
                interview_date_filter = date_filter.replace('a.applied_at', 'i.scheduled_date')
                if interview_where:
                    interview_where += f" {interview_date_filter}"
                else:
                    interview_date_clean = interview_date_filter.strip()
                    if interview_date_clean.startswith('AND '):
                        interview_date_clean = interview_date_clean[4:]
                    interview_where = f"WHERE {interview_date_clean}"
                interview_params.extend(date_params_all)
            
            # Check if interviews table has status column
            cursor.execute("SHOW COLUMNS FROM interviews LIKE 'status'")
            has_interview_status = cursor.fetchone() is not None
            
            if has_interview_status:
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                        COUNT(DISTINCT CASE WHEN i.status = 'completed' THEN i.interview_id END) AS interviews_completed,
                        COUNT(DISTINCT CASE WHEN i.status = 'cancelled' THEN i.interview_id END) AS interviews_cancelled,
                        COUNT(DISTINCT CASE WHEN i.status = 'no_show' THEN i.interview_id END) AS interviews_no_show,
                        COUNT(DISTINCT CASE WHEN i.status IN ('scheduled', 'confirmed', 'rescheduled') THEN i.interview_id END) AS interviews_upcoming
                    FROM interviews i
                    LEFT JOIN applications a ON i.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {interview_where if interview_where else ''}
                ''', tuple(interview_params) if interview_params else None)
            else:
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                        COUNT(DISTINCT CASE WHEN i.scheduled_date < NOW() THEN i.interview_id END) AS interviews_completed,
                        0 AS interviews_cancelled,
                        0 AS interviews_no_show,
                        COUNT(DISTINCT CASE WHEN i.scheduled_date >= NOW() THEN i.interview_id END) AS interviews_upcoming
                    FROM interviews i
                    LEFT JOIN applications a ON i.application_id = a.application_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {interview_where if interview_where else ''}
                ''', tuple(interview_params) if interview_params else None)
            
            interview_schedule = cursor.fetchone() or {}
            
            # Calculate completion rate
            total_scheduled = interview_schedule.get('interviews_scheduled', 0) or 0
            completed = interview_schedule.get('interviews_completed', 0) or 0
            if total_scheduled > 0:
                interview_schedule['completion_rate'] = round((completed / total_scheduled) * 100, 1)
            else:
                interview_schedule['completion_rate'] = 0
                
            # Calculate no-show rate
            no_shows = interview_schedule.get('interviews_no_show', 0) or 0
            if total_scheduled > 0:
                interview_schedule['no_show_rate'] = round((no_shows / total_scheduled) * 100, 1)
            else:
                interview_schedule['no_show_rate'] = 0
        except Exception as e:
            print(f'⚠️ Error fetching interview schedule: {e}')
            interview_schedule = {}
        
        # Enhanced Job Vacancy Activity (newly posted, updated, closed)
        job_activity = {}
        try:
            job_activity_where = base_where if base_where else ""
            job_activity_params = list(params_reports) if params_reports else []
            
            if date_filter:
                job_activity_date_filter = date_filter.replace('a.applied_at', job_posted_expr)
                if job_activity_where:
                    job_activity_where += f" {job_activity_date_filter}"
                else:
                    job_date_clean = job_activity_date_filter.strip()
                    if job_date_clean.startswith('AND '):
                        job_date_clean = job_date_clean[4:]
                    job_activity_where = f"WHERE {job_date_clean}"
                job_activity_params.extend(date_params_all)
            
            # Check if updated_at column exists in jobs table
            cursor.execute("SHOW COLUMNS FROM jobs LIKE 'updated_at'")
            has_updated_at = cursor.fetchone() is not None
            
            # Use updated_at if exists, otherwise use created_at
            job_updated_expr = 'j.updated_at' if has_updated_at else 'j.created_at'
            
            if date_params_all and len(date_params_all) >= 2:
                # Use date range for newly posted, updated, and closed
                activity_params = list(job_activity_params) + list(date_params_all[:2]) * 3
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT CASE WHEN DATE({job_posted_expr}) BETWEEN %s AND %s THEN j.job_id END) AS newly_posted,
                        COUNT(DISTINCT CASE WHEN {job_updated_expr} IS NOT NULL AND DATE({job_updated_expr}) BETWEEN %s AND %s THEN j.job_id END) AS updated_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'closed' AND DATE({job_updated_expr}) BETWEEN %s AND %s THEN j.job_id END) AS closed_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'open' THEN j.job_id END) AS active_jobs
                    FROM jobs j
                    {job_activity_where if job_activity_where else ''}
                ''', tuple(activity_params))
            else:
                # No date filter, just count active jobs
                cursor.execute(f'''
                    SELECT 
                        0 AS newly_posted,
                        0 AS updated_jobs,
                        0 AS closed_jobs,
                        COUNT(DISTINCT CASE WHEN j.status = 'open' THEN j.job_id END) AS active_jobs
                    FROM jobs j
                    {job_activity_where if job_activity_where else ''}
                ''', tuple(job_activity_params) if job_activity_params else None)
            job_activity = cursor.fetchone() or {}
        except Exception as e:
            print(f'⚠️ Error fetching job activity: {e}')
            job_activity = {}
        
        # 4. HR Performance Report (Enhanced with activity logs)
        hr_performance = {}
        hr_performance_details = []
        try:
            # Check if activity_logs table exists
            cursor.execute("SHOW TABLES LIKE 'activity_logs'")
            has_activity_logs = cursor.fetchone() is not None
            
            # Get interviews data
            interview_where = base_where if base_where else ""
            interview_params = list(params_reports) if params_reports else []
            
            if date_filter:
                interview_date_filter = date_filter.replace('a.applied_at', 'i.scheduled_date')
                if interview_where:
                    interview_where += f" {interview_date_filter}"
                else:
                    interview_date_clean = interview_date_filter.strip()
                    if interview_date_clean.startswith('AND '):
                        interview_date_clean = interview_date_clean[4:]
                    interview_where = f"WHERE {interview_date_clean}"
                interview_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    COUNT(DISTINCT i.interview_id) AS interviews_scheduled,
                    COUNT(DISTINCT CASE WHEN i.status = 'completed' THEN i.interview_id END) AS interviews_completed
                FROM interviews i
                LEFT JOIN applications a ON i.application_id = a.application_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {interview_where if interview_where else ''}
            ''', tuple(interview_params) if interview_params else None)
            interview_data = cursor.fetchone() or {}
            
            # Get status updates and reviews from applications
            status_where = base_where if base_where else ""
            status_params = list(params_reports) if params_reports else []
            
            if date_filter:
                if status_where:
                    status_where += f" {date_filter}"
                else:
                    date_filter_clean = date_filter.strip()
                    if date_filter_clean.startswith('AND '):
                        date_filter_clean = date_filter_clean[4:]
                    status_where = f"WHERE {date_filter_clean}"
                status_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    COUNT(DISTINCT CASE WHEN a.viewed_at IS NOT NULL AND a.viewed_at != a.applied_at THEN a.application_id END) AS status_updates,
                    COUNT(DISTINCT CASE WHEN a.applied_at IS NOT NULL THEN a.application_id END) AS applications_reviewed,
                    AVG(CASE WHEN a.viewed_at IS NOT NULL AND a.applied_at IS NOT NULL 
                        THEN TIMESTAMPDIFF(HOUR, a.applied_at, a.viewed_at) ELSE NULL END) AS avg_response_time
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {status_where if status_where else ''}
            ''', tuple(status_params) if status_params else None)
            status_data = cursor.fetchone() or {}
            
            # Get HR performance from activity logs if available
            hr_activity_data = []
            if has_activity_logs:
                try:
                    activity_where = ""
                    activity_params = []
                    if date_filter:
                        activity_date_clean = date_filter.strip()
                        if activity_date_clean.startswith('AND '):
                            activity_date_clean = activity_date_clean[4:]
                        activity_where = f"WHERE {activity_date_clean.replace('a.applied_at', 'al.logged_at')}"
                        activity_params = date_params_all
                    
                    cursor.execute(f'''
                        SELECT 
                            al.user_id,
                            COALESCE(ad.full_name, CONCAT('HR #', al.user_id)) AS hr_name,
                            COUNT(*) AS total_actions,
                            COUNT(DISTINCT CASE WHEN al.action LIKE '%applicant%' OR al.action LIKE '%application%' THEN al.log_id END) AS applications_processed,
                            COUNT(DISTINCT CASE WHEN al.action LIKE '%interview%' THEN al.log_id END) AS interviews_scheduled_count
                        FROM activity_logs al
                        LEFT JOIN admins ad ON al.user_id = ad.admin_id
                        {activity_where}
                        GROUP BY al.user_id, ad.full_name
                        ORDER BY total_actions DESC
                        LIMIT 10
                    ''', tuple(activity_params) if activity_params else None)
                    hr_activity_data = cursor.fetchall() or []
                except Exception as e:
                    print(f'⚠️ Error fetching HR activity: {e}')
                    hr_activity_data = []
            
            # Combine results
            hr_performance = {
                'interviews_scheduled': interview_data.get('interviews_scheduled', 0) or 0,
                'interviews_completed': interview_data.get('interviews_completed', 0) or 0,
                'status_updates': status_data.get('status_updates', 0) or 0,
                'applications_reviewed': status_data.get('applications_reviewed', 0) or 0,
                'avg_response_time': round(status_data.get('avg_response_time', 0) or 0, 1)
            }
            hr_performance_details = hr_activity_data
        except Exception as e:
            print(f'⚠️ Error fetching HR performance: {e}')
            hr_performance = {}
            hr_performance_details = []
        
        # Monthly/Yearly Trends (for monthly and yearly reports)
        monthly_trends = []
        yearly_trends = []
        branch_comparison_data = []
        fastest_hiring_branch = None
        monthly_trends = []
        
        # Fetch trend data based on period
        try:
            if period == 'week':
                # Weekly period: Get daily trends (Monday to Sunday)
                if date_filter and date_params_all:
                    trend_where = base_where if base_where else ''
                    if trend_where:
                        trend_where = trend_where + ' ' + date_filter
                    else:
                        # Remove 'AND ' prefix from date_filter if it exists
                        trend_where = date_filter.strip()
                        if trend_where.startswith('AND '):
                            trend_where = 'WHERE ' + trend_where[4:]
                        else:
                            trend_where = 'WHERE ' + trend_where
                    
                    trend_params = list(params_reports) if params_reports else []
                    trend_params.extend(date_params_all)
                    
                    # Get all applications grouped by submitted date
                    cursor.execute(f'''
                        SELECT 
                            DAYNAME(a.applied_at) AS day_name,
                            DAYOFWEEK(a.applied_at) AS day_of_week,
                            DATE(a.applied_at) AS date_key,
                            DATE_FORMAT(a.applied_at, '%a') AS day_key,
                            COUNT(DISTINCT a.application_id) AS total_applications,
                            COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                            0 AS hired_count,
                            COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {trend_where}
                        GROUP BY DAYNAME(a.applied_at), DAYOFWEEK(a.applied_at), DATE(a.applied_at), DATE_FORMAT(a.applied_at, '%a')
                    ''', tuple(trend_params))
                    daily_apps = cursor.fetchall() or []
                    
                    # Get hired applicants grouped by hire date (updated_at when status = 'hired')
                    # Build WHERE clause for hired applicants: filter by hire date (updated_at) within selected period
                    hired_where_parts = []
                    hired_params = []
                    
                    # Apply branch filter
                    if branch_filter_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_filter_id)
                    elif branch_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_id)
                    
                    # Apply job filter
                    if job_filter_id:
                        hired_where_parts.append('j.job_id = %s')
                        hired_params.append(job_filter_id)
                    
                    # Apply date filter based on hire date (updated_at) instead of submitted_at
                    if date_filter and date_params_all:
                        # Replace submitted_at with viewed_at in date filter
                        hired_date_filter = date_filter.replace('a.applied_at', 'a.viewed_at').replace('applied_at', 'viewed_at')
                        hired_date_filter = hired_date_filter.strip()
                        if hired_date_filter.startswith('AND '):
                            hired_date_filter = hired_date_filter[4:]
                        hired_where_parts.append(hired_date_filter)
                        hired_params.extend(date_params_all)
                    
                    # Add status filter
                    hired_where_parts.append("a.status = 'hired'")
                    
                    # Build final WHERE clause
                    if hired_where_parts:
                        hired_trend_where = 'WHERE ' + ' AND '.join(hired_where_parts)
                    else:
                        hired_trend_where = "WHERE a.status = 'hired'"
                    
                    cursor.execute(f'''
                        SELECT 
                            DAYNAME(a.viewed_at) AS day_name,
                            DAYOFWEEK(a.viewed_at) AS day_of_week,
                            DATE(a.viewed_at) AS date_key,
                            DATE_FORMAT(a.viewed_at, '%a') AS day_key,
                            COUNT(DISTINCT a.application_id) AS hired_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {hired_trend_where}
                        GROUP BY DAYNAME(a.viewed_at), DAYOFWEEK(a.viewed_at), DATE(a.viewed_at), DATE_FORMAT(a.viewed_at, '%a')
                        ORDER BY DATE(a.viewed_at)
                    ''', tuple(hired_params) if hired_params else None)
                    daily_hired = cursor.fetchall() or []
                    
                    # Debug: Print hired data if available
                    if daily_hired:
                        print(f'✅ Found {len(daily_hired)} hired date entries for weekly trends')
                        total_hired_count = sum(int(row.get('hired_count') or 0) for row in daily_hired)
                        print(f'   Total hired applicants: {total_hired_count}')
                        for h in daily_hired[:3]:
                            print(f'   - {h.get("date_key")}: {h.get("hired_count")} hired')
                    else:
                        log.warning('⚠️ No hired applicants found for weekly period')
                        # Check if there are any hired applicants at all (for debugging)
                        cursor.execute('''
                            SELECT COUNT(*) as count 
                            FROM applications 
                            WHERE status = 'hired'
                        ''')
                        total_hired = cursor.fetchone()
                        if total_hired and total_hired.get('count', 0) > 0:
                            print(f'   But found {total_hired.get("count")} total hired applicants in database')
                    
                    # Merge the results - use date_key for matching
                    hired_dict = {}
                    for row in daily_hired:
                        date_key = row.get('date_key')
                        if date_key:
                            # Convert to string for consistent comparison
                            date_key_str = str(date_key)
                            hired_count = int(row.get('hired_count') or 0)
                            if hired_count > 0:  # Only add if there are actually hired applicants
                                hired_dict[date_key_str] = hired_dict.get(date_key_str, 0) + hired_count
                    
                    daily_trends = []
                    # Create a set to track which dates we've added
                    added_dates = set()
                    
                    # First, add all application dates with their hired_count
                    for app_row in daily_apps:
                        date_key = app_row.get('date_key')
                        if date_key:
                            date_key_str = str(date_key)
                            # Get hired count from hired_dict
                            hired_count_from_dict = hired_dict.get(date_key_str, 0)
                            app_row['hired_count'] = hired_count_from_dict
                            daily_trends.append(app_row)
                            added_dates.add(date_key_str)
                    
                    # Then, add any hired dates that don't have applications
                    for hired_row in daily_hired:
                        date_key = hired_row.get('date_key')
                        hired_count = int(hired_row.get('hired_count') or 0)
                        if date_key and hired_count > 0:
                            date_key_str = str(date_key)
                            if date_key_str not in added_dates:
                                daily_trends.append({
                                    'day_name': hired_row.get('day_name', ''),
                                    'day_of_week': hired_row.get('day_of_week', 0),
                                    'date_key': date_key,
                                    'day_key': hired_row.get('day_key', ''),
                                    'total_applications': 0,
                                    'total_applicants': 0,
                                    'hired_count': hired_count,
                                    'rejected_count': 0
                                })
                                added_dates.add(date_key_str)
                    
                    # Sort by date_key (handle both string and date objects)
                    daily_trends.sort(key=lambda x: str(x.get('date_key', '')) if x.get('date_key') else '')
                    # Convert to monthly_trends format for consistency
                    monthly_trends = daily_trends
                    
                    # Debug output
                    if monthly_trends:
                        total_hired = sum(int(row.get('hired_count', 0) or 0) for row in monthly_trends)
                        print(f'✅ Weekly trends: {len(monthly_trends)} days, {total_hired} total hired')
                else:
                    # No date filter - still try to get hired applicants for default period
                    # Query hired applicants for current week as fallback
                    today = datetime.now()
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
                    
                    hired_where_parts = []
                    hired_params = []
                    
                    # Apply branch filter
                    if branch_filter_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_filter_id)
                    elif branch_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_id)
                    
                    # Apply job filter
                    if job_filter_id:
                        hired_where_parts.append('j.job_id = %s')
                        hired_params.append(job_filter_id)
                    
                    # Add date filter for current week
                    hired_where_parts.append("DATE(a.viewed_at) BETWEEN %s AND %s")
                    hired_params.extend([week_start.date(), week_end.date()])
                    hired_where_parts.append("a.status = 'hired'")
                    
                    hired_trend_where = 'WHERE ' + ' AND '.join(hired_where_parts)
                    
                    cursor.execute(f'''
                        SELECT 
                            DAYNAME(a.viewed_at) AS day_name,
                            DAYOFWEEK(a.viewed_at) AS day_of_week,
                            DATE(a.viewed_at) AS date_key,
                            DATE_FORMAT(a.viewed_at, '%a') AS day_key,
                            COUNT(DISTINCT a.application_id) AS hired_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {hired_trend_where}
                        GROUP BY DAYNAME(a.viewed_at), DAYOFWEEK(a.viewed_at), DATE(a.viewed_at), DATE_FORMAT(a.viewed_at, '%a')
                        ORDER BY DATE(a.viewed_at)
                    ''', tuple(hired_params) if hired_params else None)
                    daily_hired = cursor.fetchall() or []
                    
                    # Build trends from hired data
                    daily_trends = []
                    for hired_row in daily_hired:
                        daily_trends.append({
                            'day_name': hired_row.get('day_name', ''),
                            'day_of_week': hired_row.get('day_of_week', 0),
                            'date_key': hired_row.get('date_key'),
                            'day_key': hired_row.get('day_key', ''),
                            'total_applications': 0,
                            'total_applicants': 0,
                            'hired_count': hired_row.get('hired_count', 0),
                            'rejected_count': 0
                        })
                    
                    daily_trends.sort(key=lambda x: str(x.get('date_key', '')) if x.get('date_key') else '')
                    monthly_trends = daily_trends
            
            elif period == 'month':
                # Monthly period: Get weekly trends (Week 1, 2, 3, 4)
                if date_filter and date_params_all:
                    trend_where = base_where if base_where else ''
                    if trend_where:
                        trend_where = trend_where + ' ' + date_filter
                    else:
                        # Remove 'AND ' prefix from date_filter if it exists
                        trend_where = date_filter.strip()
                        if trend_where.startswith('AND '):
                            trend_where = 'WHERE ' + trend_where[4:]
                        else:
                            trend_where = 'WHERE ' + trend_where
                    
                    trend_params = list(params_reports) if params_reports else []
                    trend_params.extend(date_params_all)
                    
                    # Get all applications grouped by submitted date (weekly)
                    cursor.execute(f'''
                        SELECT 
                            WEEK(a.applied_at, 1) - WEEK(DATE_SUB(a.applied_at, INTERVAL DAY(a.applied_at)-1 DAY), 1) + 1 AS week_number,
                            CONCAT('Week ', WEEK(a.applied_at, 1) - WEEK(DATE_SUB(a.applied_at, INTERVAL DAY(a.applied_at)-1 DAY), 1) + 1) AS week_key,
                            DATE_FORMAT(a.applied_at, '%Y-%m') AS month_key,
                            COUNT(DISTINCT a.application_id) AS total_applications,
                            COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                            0 AS hired_count,
                            COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {trend_where}
                        GROUP BY WEEK(a.applied_at, 1) - WEEK(DATE_SUB(a.applied_at, INTERVAL DAY(a.applied_at)-1 DAY), 1) + 1, 
                                 DATE_FORMAT(a.applied_at, '%Y-%m')
                    ''', tuple(trend_params))
                    weekly_apps = cursor.fetchall() or []
                    
                    # Get hired applicants grouped by hire date (updated_at when status = 'hired')
                    hired_where_parts = []
                    hired_params = []
                    
                    # Apply branch filter
                    if branch_filter_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_filter_id)
                    elif branch_id:
                        hired_where_parts.append('j.branch_id = %s')
                        hired_params.append(branch_id)
                    
                    # Apply job filter
                    if job_filter_id:
                        hired_where_parts.append('j.job_id = %s')
                        hired_params.append(job_filter_id)
                    
                    # Apply date filter based on hire date (viewed_at) instead of submitted_at
                    if date_filter and date_params_all:
                        hired_date_filter = date_filter.replace('a.applied_at', 'a.viewed_at').replace('applied_at', 'viewed_at')
                        hired_date_filter = hired_date_filter.strip()
                        if hired_date_filter.startswith('AND '):
                            hired_date_filter = hired_date_filter[4:]
                        hired_where_parts.append(hired_date_filter)
                        hired_params.extend(date_params_all)
                    
                    # Add status filter
                    hired_where_parts.append("a.status = 'hired'")
                    
                    # Build final WHERE clause
                    if hired_where_parts:
                        hired_trend_where = 'WHERE ' + ' AND '.join(hired_where_parts)
                    else:
                        hired_trend_where = "WHERE a.status = 'hired'"
                    
                    cursor.execute(f'''
                        SELECT 
                            WEEK(a.viewed_at, 1) - WEEK(DATE_SUB(a.viewed_at, INTERVAL DAY(a.viewed_at)-1 DAY), 1) + 1 AS week_number,
                            CONCAT('Week ', WEEK(a.viewed_at, 1) - WEEK(DATE_SUB(a.viewed_at, INTERVAL DAY(a.viewed_at)-1 DAY), 1) + 1) AS week_key,
                            DATE_FORMAT(a.viewed_at, '%Y-%m') AS month_key,
                            COUNT(DISTINCT a.application_id) AS hired_count
                        FROM applications a
                        JOIN applicants ap ON a.applicant_id = ap.applicant_id
                        LEFT JOIN jobs j ON a.job_id = j.job_id
                        {hired_trend_where}
                        GROUP BY WEEK(a.viewed_at, 1) - WEEK(DATE_SUB(a.viewed_at, INTERVAL DAY(a.viewed_at)-1 DAY), 1) + 1, 
                                 DATE_FORMAT(a.viewed_at, '%Y-%m')
                    ''', tuple(hired_params) if hired_params else None)
                    weekly_hired = cursor.fetchall() or []
                    
                    # Merge the results
                    hired_dict = {(row['week_number'], row['month_key']): row['hired_count'] for row in weekly_hired}
                    weekly_trends = []
                    for app_row in weekly_apps:
                        key = (app_row['week_number'], app_row['month_key'])
                        app_row['hired_count'] = hired_dict.get(key, 0)
                        weekly_trends.append(app_row)
                    
                    # Add any hired weeks that don't have applications
                    for hired_row in weekly_hired:
                        key = (hired_row['week_number'], hired_row['month_key'])
                        if not any((row['week_number'], row['month_key']) == key for row in weekly_trends):
                            weekly_trends.append({
                                'week_number': hired_row['week_number'],
                                'week_key': hired_row['week_key'],
                                'month_key': hired_row['month_key'],
                                'total_applications': 0,
                                'total_applicants': 0,
                                'hired_count': hired_row['hired_count'],
                                'rejected_count': 0
                            })
                    
                    # Sort by week_number
                    weekly_trends.sort(key=lambda x: x['week_number'])
                    # Convert to monthly_trends format for consistency
                    monthly_trends = weekly_trends
                else:
                    monthly_trends = []
            
            elif period == 'year':
                # Yearly period: Get monthly trends
                # Get all applications grouped by submitted date (monthly)
                cursor.execute(f'''
                    SELECT 
                        YEAR(a.applied_at) AS year,
                        MONTH(a.applied_at) AS month,
                        DATE_FORMAT(a.applied_at, '%Y-%m') AS month_key,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT ap.applicant_id) AS total_applicants,
                        0 AS hired_count,
                        COUNT(DISTINCT CASE WHEN a.status = 'rejected' THEN a.application_id END) AS rejected_count
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {base_where if base_where else ''}
                    {date_filter if date_filter else ''}
                    GROUP BY YEAR(a.applied_at), MONTH(a.applied_at), DATE_FORMAT(a.applied_at, '%Y-%m')
                ''', tuple(list(params_reports) + date_params_all) if params_reports and date_params_all else (tuple(params_reports) if params_reports else (tuple(date_params_all) if date_params_all else None)))
                monthly_apps = cursor.fetchall() or []
                
                # Get hired applicants grouped by hire date (updated_at when status = 'hired')
                hired_where_parts = []
                hired_params = []
                
                # Apply branch filter
                if branch_filter_id:
                    hired_where_parts.append('j.branch_id = %s')
                    hired_params.append(branch_filter_id)
                elif branch_id:
                    hired_where_parts.append('j.branch_id = %s')
                    hired_params.append(branch_id)
                
                # Apply job filter
                if job_filter_id:
                    hired_where_parts.append('j.job_id = %s')
                    hired_params.append(job_filter_id)
                
                # Apply date filter based on hire date (viewed_at) instead of submitted_at
                if date_filter and date_params_all:
                    hired_date_filter = date_filter.replace('a.applied_at', 'a.viewed_at').replace('applied_at', 'viewed_at')
                    hired_date_filter = hired_date_filter.strip()
                    if hired_date_filter.startswith('AND '):
                        hired_date_filter = hired_date_filter[4:]
                    hired_where_parts.append(hired_date_filter)
                    hired_params.extend(date_params_all)
                
                # Add status filter
                hired_where_parts.append("a.status = 'hired'")
                
                # Build final WHERE clause
                if hired_where_parts:
                    hired_where_clause = 'WHERE ' + ' AND '.join(hired_where_parts)
                else:
                    hired_where_clause = "WHERE a.status = 'hired'"
                
                cursor.execute(f'''
                    SELECT 
                        YEAR(a.viewed_at) AS year,
                        MONTH(a.viewed_at) AS month,
                        DATE_FORMAT(a.viewed_at, '%Y-%m') AS month_key,
                        COUNT(DISTINCT a.application_id) AS hired_count
                    FROM applications a
                    JOIN applicants ap ON a.applicant_id = ap.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {hired_where_clause}
                    GROUP BY YEAR(a.viewed_at), MONTH(a.viewed_at), DATE_FORMAT(a.viewed_at, '%Y-%m')
                ''', tuple(hired_params) if hired_params else None)
                monthly_hired = cursor.fetchall() or []
                
                # Merge the results
                hired_dict = {row['month_key']: row['hired_count'] for row in monthly_hired}
                monthly_trends = []
                for app_row in monthly_apps:
                    month_key = app_row['month_key']
                    app_row['hired_count'] = hired_dict.get(month_key, 0)
                    monthly_trends.append(app_row)
                
                # Add any hired months that don't have applications
                for hired_row in monthly_hired:
                    month_key = hired_row['month_key']
                    if not any(row['month_key'] == month_key for row in monthly_trends):
                        monthly_trends.append({
                            'year': hired_row['year'],
                            'month': hired_row['month'],
                            'month_key': month_key,
                            'total_applications': 0,
                            'total_applicants': 0,
                            'hired_count': hired_row['hired_count'],
                            'rejected_count': 0
                        })
                
                # Sort by year, month
                monthly_trends.sort(key=lambda x: (x['year'], x['month']))
            
            if period == 'month' or period == 'year':
                
                # Branch comparison with hiring metrics
                cursor.execute(f'''
                    SELECT 
                        COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                        COUNT(DISTINCT a.application_id) AS total_applications,
                        COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) AS hired_count,
                        AVG(CASE WHEN a.status = 'hired' AND a.applied_at IS NOT NULL AND a.viewed_at IS NOT NULL 
                            THEN DATEDIFF(a.viewed_at, a.applied_at) ELSE NULL END) AS avg_time_to_hire,
                        CASE 
                            WHEN COUNT(DISTINCT a.application_id) > 0 
                            THEN COUNT(DISTINCT CASE WHEN a.status = 'hired' THEN a.application_id END) / COUNT(DISTINCT a.application_id) * 100
                            ELSE 0 
                        END AS hire_rate
                    FROM applications a
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    LEFT JOIN branches b ON j.branch_id = b.branch_id
                    {base_where if base_where else ''}
                    GROUP BY b.branch_id, b.branch_name
                    HAVING total_applications > 0
                    ORDER BY avg_time_to_hire ASC, hire_rate DESC
                ''', tuple(params_reports) if params_reports else None)
                branch_comparison_data = cursor.fetchall() or []
                
                # Find fastest hiring branch
                if branch_comparison_data:
                    fastest_hiring_branch = branch_comparison_data[0] if branch_comparison_data else None
                    if fastest_hiring_branch:
                        fastest_hiring_branch['avg_time_to_hire'] = round(fastest_hiring_branch.get('avg_time_to_hire') or 0, 1)
                        fastest_hiring_branch['hire_rate'] = round(fastest_hiring_branch.get('hire_rate') or 0, 1)
        except Exception as e:
            print(f'⚠️ Error fetching trends: {e}')
            import traceback
            print(traceback.format_exc())
            monthly_trends = []
            branch_comparison_data = []
        
        # Enhanced Applicant Summary - New applicants per job
        new_applicants_per_job = {}
        try:
            if date_filter:
                new_app_date_filter = date_filter.replace('a.applied_at', 'a.applied_at')
                new_app_where = base_where if base_where else ""
                new_app_params = list(params_reports) if params_reports else []
                
                if new_app_where:
                    new_app_where += f" {new_app_date_filter}"
                else:
                    new_date_clean = new_app_date_filter.strip()
                    if new_date_clean.startswith('AND '):
                        new_date_clean = new_date_clean[4:]
                    new_app_where = f"WHERE {new_date_clean}"
                new_app_params.extend(date_params_all)
                
                cursor.execute(f'''
                    SELECT 
                        COUNT(DISTINCT ap.applicant_id) AS new_applicants,
                        COUNT(DISTINCT j.job_id) AS active_jobs,
                        CASE 
                            WHEN COUNT(DISTINCT j.job_id) > 0 
                            THEN COUNT(DISTINCT ap.applicant_id) / COUNT(DISTINCT j.job_id)
                            ELSE 0 
                        END AS new_applicants_per_job
                    FROM applicants ap
                    JOIN applications a ON ap.applicant_id = a.applicant_id
                    LEFT JOIN jobs j ON a.job_id = j.job_id
                    {new_app_where if new_app_where else ''}
                ''', tuple(new_app_params) if new_app_params else None)
                new_applicants_per_job = cursor.fetchone() or {}
            else:
                new_applicants_per_job = {'new_applicants': 0, 'active_jobs': 0, 'new_applicants_per_job': 0}
        except Exception as e:
            print(f'⚠️ Error fetching new applicants per job: {e}')
            new_applicants_per_job = {}
        
        # 1. Time-to-Hire (Average Days) - Enhanced KPI
        avg_time_to_hire = 0
        try:
            tth_where = base_where if base_where else ""
            tth_params = list(params_reports) if params_reports else []
            
            if date_filter:
                tth_date_filter = date_filter.replace('a.applied_at', 'a.viewed_at')
                if tth_where:
                    tth_where += f" {tth_date_filter}"
                else:
                    tth_date_clean = tth_date_filter.strip()
                    if tth_date_clean.startswith('AND '):
                        tth_date_clean = tth_date_clean[4:]
                    tth_where = f"WHERE {tth_date_clean}"
                tth_params.extend(date_params_all)
            
            if tth_where:
                tth_where += " AND a.status = 'hired'"
            else:
                tth_where = "WHERE a.status = 'hired'"
            
            cursor.execute(f'''
                SELECT 
                    AVG(DATEDIFF(a.viewed_at, a.applied_at)) AS avg_time_to_hire
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {tth_where}
            ''', tuple(tth_params) if tth_params else None)
            tth_result = cursor.fetchone()
            avg_time_to_hire = round(tth_result.get('avg_time_to_hire', 0) or 0, 1) if tth_result else 0
        except Exception as e:
            print(f'⚠️ Error fetching average time-to-hire: {e}')
            avg_time_to_hire = 0
        
        # 2. Most Applied Job - Enhanced KPI
        most_applied_job_info = {}
        try:
            maj_where = base_where if base_where else ""
            maj_params = list(params_reports) if params_reports else []
            
            if date_filter:
                if maj_where:
                    maj_where += f" {date_filter}"
                else:
                    maj_date_clean = date_filter.strip()
                    if maj_date_clean.startswith('AND '):
                        maj_date_clean = maj_date_clean[4:]
                    maj_where = f"WHERE {maj_date_clean}"
                maj_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    j.job_id,
                    {job_title_expr} AS job_title,
                    COUNT(DISTINCT a.application_id) AS total_applications,
                    COUNT(DISTINCT ap.applicant_id) AS total_applicants
                FROM applications a
                JOIN applicants ap ON a.applicant_id = ap.applicant_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {maj_where if maj_where else ''}
                GROUP BY j.job_id, {job_title_expr}
                ORDER BY total_applications DESC
                LIMIT 1
            ''', tuple(maj_params) if maj_params else None)
            most_applied_job_info = cursor.fetchone() or {}
        except Exception as e:
            print(f'⚠️ Error fetching most applied job: {e}')
            most_applied_job_info = {}
        
        # 3. Application Status Distribution (for Pie/Donut Chart)
        status_distribution = []
        try:
            status_where_parts = []
            status_params = list(params_reports) if params_reports else []
            
            # Apply branch filter
            if branch_filter_id:
                status_where_parts.append('j.branch_id = %s')
                status_params.append(branch_filter_id)
            elif branch_id:
                status_where_parts.append('j.branch_id = %s')
                status_params.append(branch_id)
            
            # Apply job filter
            if job_filter_id:
                status_where_parts.append('j.job_id = %s')
                status_params.append(job_filter_id)
            
            # Apply date filter
            if date_filter and date_params_all:
                date_clean = date_filter.strip()
                if date_clean.startswith('AND '):
                    date_clean = date_clean[4:]
                status_where_parts.append(date_clean)
                status_params.extend(date_params_all)
            
            # Build WHERE clause
            if status_where_parts:
                status_where = 'WHERE ' + ' AND '.join(status_where_parts)
            else:
                status_where = ''
            
            cursor.execute(f'''
                SELECT 
                    a.status,
                    COUNT(DISTINCT a.application_id) AS count
                FROM applications a
                LEFT JOIN jobs j ON a.job_id = j.job_id
                {status_where}
                GROUP BY a.status
                ORDER BY count DESC
            ''', tuple(status_params) if status_params else None)
            status_distribution = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching status distribution: {e}')
            status_distribution = []
        
        # 4. Interview Schedule Report Details (Full List)
        interview_schedule_details = []
        try:
            interview_details_where = base_where if base_where else ""
            interview_details_params = list(params_reports) if params_reports else []
            
            if date_filter:
                interview_details_date_filter = date_filter.replace('a.applied_at', 'i.scheduled_date')
                if interview_details_where:
                    interview_details_where += f" {interview_details_date_filter}"
                else:
                    interview_details_date_clean = interview_details_date_filter.strip()
                    if interview_details_date_clean.startswith('AND '):
                        interview_details_date_clean = interview_details_date_clean[4:]
                    interview_details_where = f"WHERE {interview_details_date_clean}"
                interview_details_params.extend(date_params_all)
            
            cursor.execute(f'''
                SELECT 
                    ap.full_name AS applicant_name,
                    {job_title_expr} AS job_title,
                    COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                    i.scheduled_date AS interview_date,
                    COALESCE(i.interview_mode, 'In-person') AS interview_mode,
                    COALESCE(i.interview_mode, 'TBD') AS interview_location,
                    COALESCE(i.status, 'scheduled') AS interview_status,
                    COALESCE(ad.full_name, 'Unassigned') AS hr_assigned
                FROM interviews i
                LEFT JOIN applications a ON i.application_id = a.application_id
                LEFT JOIN applicants ap ON a.applicant_id = ap.applicant_id
                LEFT JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                LEFT JOIN activity_logs al ON al.action LIKE CONCAT('%', i.interview_id, '%') OR al.action LIKE '%interview%'
                LEFT JOIN admins ad ON al.user_id = ad.admin_id
                {interview_details_where if interview_details_where else ''}
                ORDER BY i.scheduled_date DESC
                LIMIT 500
            ''', tuple(interview_details_params) if interview_details_params else None)
            interview_schedule_details = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching interview schedule details: {e}')
            interview_schedule_details = []
        
        # 5. HR Activity Log Report Details
        hr_activity_log_details = []
        try:
            cursor.execute("SHOW TABLES LIKE 'activity_logs'")
            has_activity_logs = cursor.fetchone() is not None
            
            if has_activity_logs:
                activity_where = "WHERE u.user_type = 'hr'"
                activity_params = []
                
                if date_filter:
                    activity_date_filter = date_filter.replace('a.applied_at', 'al.logged_at')
                    activity_where += f" {activity_date_filter}"
                    activity_params.extend(date_params_all)
                
                cursor.execute(f'''
                    SELECT 
                        COALESCE(ad.full_name, CONCAT('HR #', al.user_id)) AS hr_name,
                        al.action,
                        al.description,
                        al.logged_at,
                        al.action AS affected_item
                    FROM activity_logs al
                    LEFT JOIN admins ad ON al.user_id = ad.admin_id
                    WHERE al.user_id IN (SELECT admin_id FROM admins WHERE admin_id IS NOT NULL)
                    {' AND ' + date_filter.replace('a.applied_at', 'al.logged_at').strip('WHERE AND ') if date_filter else ''}
                    ORDER BY al.logged_at DESC
                    LIMIT 500
                ''', tuple(activity_params) if activity_params else None)
                hr_activity_log_details = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching HR activity log details: {e}')
            import traceback
            traceback.print_exc()
            hr_activity_log_details = []
        
        # 6. Get HR Users list for filter
        hr_users_list = []
        try:
            cursor.execute('''
                SELECT DISTINCT ad.admin_id, ad.full_name
                FROM admins ad
                JOIN users u ON ad.user_id = u.user_id
                WHERE u.user_type = 'hr' AND u.is_active = 1 AND ad.is_active = 1
                ORDER BY ad.full_name
            ''')
            hr_users_list = cursor.fetchall() or []
        except Exception as e:
            print(f'⚠️ Error fetching HR users list: {e}')
            hr_users_list = []
        
        # Handle export requests - must be after all data is generated
        if export_format:
            try:
                return handle_report_export(cursor, export_format, export_section, export_type, period_summary, 
                                          applicant_summary, applicant_summary_details, job_vacancy, job_vacancy_details,
                                          hiring_outcome, hr_performance, job_title_expr, date_filter, date_params_all)
            except Exception as export_error:
                print(f'⚠️ Error in export: {export_error}')
                import traceback
                print(traceback.format_exc())
                flash(f'Error generating export: {str(export_error)}', 'error')
                # Redirect back to reports page without export parameter
                redirect_url = url_for('admin_reports_analytics') if user.get('role') == 'admin' else url_for('hr_reports_analytics')
                params = {'period': period}
                if selected_week_month:
                    params['week_month'] = selected_week_month
                if selected_week_number:
                    params['week_number'] = selected_week_number
                if selected_week:  # Legacy support
                    params['week'] = selected_week
                if selected_month:
                    params['month'] = selected_month
                if selected_year:
                    params['year'] = selected_year
                # Include branch and job filters in redirect
                if branch_filter_id:
                    params['branch_id'] = branch_filter_id
                if job_filter_id:
                    params['job_id'] = job_filter_id
                return redirect(url_for('admin_reports_analytics' if user.get('role') == 'admin' else 'hr_reports_analytics', **params))
        
        # Render HR template if user is HR, otherwise admin template
        template = 'hr/reports_analytics.html' if user.get('role') == 'hr' else 'admin/reports.html'
        branch_info = None
        if user.get('role') == 'hr':
            branch_id_session = session.get('branch_id')
            if branch_id_session:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id_session,))
                if branch_rows:
                    branch_info = branch_rows[0]
        # Group applicants by status for display
        applicants_by_status = {
            'hired': [],
            'rejected': [],
            'pending': [],
            'scheduled': [],
            'interviewed': []
        }
        for app in applicant_summary_details:
            status = app.get('application_status', '').lower()
            if status in applicants_by_status:
                applicants_by_status[status].append(app)
        
        # Calculate status counts
        status_counts = {
            'hired': len(applicants_by_status['hired']),
            'rejected': len(applicants_by_status['rejected']),
            'pending': len(applicants_by_status['pending']),
            'scheduled': len(applicants_by_status['scheduled']),
            'interviewed': len(applicants_by_status['interviewed'])
        }
        
        # Ensure all data is passed correctly
        return render_template(
            template,
            stats={**stats, **(metrics or {})},
            metrics=metrics or {},
            trends=trends or [],
            branch_stats=branch_stats or [],
            branch_comparison=branch_comparison or [],
            system_stats=system_stats or {},
            job_performance=job_performance or [],
            status_breakdown=status_breakdown or [],
            funnel_data=funnel_data or {},
            source_effectiveness=source_effectiveness or [],
            branch_info=branch_info,
            period=period,
            period_summary=period_summary,
            available_years=available_years,
            selected_week=selected_week,
            selected_week_month=selected_week_month,
            selected_week_number=selected_week_number,
            selected_month=selected_month,
            selected_year=selected_year,
            applicant_summary=applicant_summary,
            job_vacancy=job_vacancy,
            hiring_outcome=hiring_outcome,
            hr_performance=hr_performance,
            applicant_summary_details=applicant_summary_details,
            job_vacancy_details=job_vacancy_details,
            applicants_by_status=applicants_by_status,
            status_counts=status_counts,
            generate_section=generate_section,
            applicants_by_branch=applicants_by_branch,
            most_applied_jobs=most_applied_jobs,
            least_applied_jobs=least_applied_jobs,
            interview_schedule=interview_schedule,
            job_activity=job_activity,
            hr_performance_details=hr_performance_details,
            monthly_trends=monthly_trends,
            branch_comparison_data=branch_comparison_data,
            fastest_hiring_branch=fastest_hiring_branch,
            new_applicants_per_job=new_applicants_per_job,
            branches=branches,
            jobs=jobs,
            # New comprehensive report data
            avg_time_to_hire=avg_time_to_hire,
            most_applied_job_info=most_applied_job_info,
            status_distribution=status_distribution,
            interview_schedule_details=interview_schedule_details,
            hr_activity_log_details=hr_activity_log_details,
            hr_users_list=hr_users_list,
            hr_filter_condition=hr_filter_condition,
            status_filter_condition=status_filter_condition,
        )
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ Reports analytics error: {exc}')
        print(f'Full traceback: {error_details}')
        flash(f'An error occurred while loading reports: {str(exc)}', 'error')
        template = 'hr/reports_analytics.html' if user.get('role') == 'hr' else 'admin/reports.html'
        # Get period from request to maintain state
        period = request.args.get('period', 'all')
        selected_week = request.args.get('week', '')
        selected_week_month = request.args.get('week_month', '')
        selected_week_number = request.args.get('week_number', '')
        selected_month = request.args.get('month', '')
        selected_year = request.args.get('year', '')
        # Get available years for dropdown
        try:
            cursor.execute("SELECT DISTINCT YEAR(applied_at) as year FROM applications WHERE applied_at IS NOT NULL ORDER BY year DESC")
            available_years = [str(row['year']) for row in cursor.fetchall() if row.get('year')]
        except Exception:
            log.exception('Error fetching available years for reports')
            available_years = []
        return render_template(
            template,
            stats={},
            metrics={},
            trends=[],
            branch_stats=[],
            branch_comparison=[],
            system_stats={},
            job_performance=[],
            status_breakdown=[],
            funnel_data={},
            source_effectiveness=[],
            branch_info=None,
            period=period,
            period_summary={},
            available_years=available_years,
            selected_week=selected_week,
            selected_week_month=selected_week_month,
            selected_week_number=selected_week_number,
            selected_month=selected_month,
            selected_year=selected_year,
            applicant_summary={},
            job_vacancy={},
            hiring_outcome={},
            hr_performance={},
            applicant_summary_details=[],
            job_vacancy_details=[],
            applicants_by_status={'hired': [], 'rejected': [], 'pending': [], 'scheduled': [], 'interviewed': []},
            status_counts={'hired': 0, 'rejected': 0, 'pending': 0, 'scheduled': 0, 'interviewed': 0},
            generate_section='',
            applicants_by_branch=[],
            interview_schedule={'interviews_scheduled': 0, 'interviews_completed': 0},
            monthly_trends=[],
            branches=[],
            jobs=[],
        )
    finally:
        if cursor:
            cursor.close()


@app.route('/admin/applicants/<int:applicant_id>')
@login_required('admin', 'hr')
def view_applicant(applicant_id):
    """View detailed applicant profile."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('applicants'))
    
    cursor = db.cursor(dictionary=True)
    # Log context for debugging
    try:
        print(f"➡️ View applicant request: user={user and user.get('id')}, role={user and user.get('role')}, applicant_id={applicant_id}")
    except Exception:
        pass
    try:
        # Ensure schema compatibility before querying
        ensure_schema_compatibility()
        # Determine HR branch scope, but do not block viewing if no matching applications
        branch_id = get_branch_scope(user)
        
        # Get applicant details (isolated try/except)
        try:
            cursor.execute(
                '''
                SELECT ap.applicant_id, ap.full_name, ap.email, ap.phone_number, 
                       ap.created_at, ap.last_login
                FROM applicants ap
                WHERE ap.applicant_id = %s
                LIMIT 1
                ''',
                (applicant_id,),
            )
            applicant = cursor.fetchone()
        except Exception as e:
            applicant = None
            print(f'⚠️ Error fetching applicant details for id={applicant_id}: {e}')

        if not applicant:
            flash('Applicant not found.', 'error')
            return redirect(url_for('applicants'))
        
        # Get applicant's applications (isolated)
        try:
            where_clause = 'a.applicant_id = %s'
            params = [applicant_id]
            if branch_id:
                where_clause += ' AND j.branch_id = %s'
                params.append(branch_id)

            cursor.execute(
                f'''
                SELECT a.application_id, a.status, a.applied_at,
                       j.job_id, j.job_title AS job_title,
                       COALESCE(b.branch_name, 'Unassigned') AS branch_name,
                       COALESCE((
                           SELECT COUNT(*)
                           FROM interviews i
                           WHERE i.application_id = a.application_id
                       ), 0) AS interview_count
                FROM applications a
                JOIN jobs j ON a.job_id = j.job_id
                LEFT JOIN branches b ON j.branch_id = b.branch_id
                WHERE {where_clause}
                ORDER BY a.applied_at DESC
                ''',
                tuple(params),
            )
            applications = cursor.fetchall() or []
            # Normalize application fields defensively
            for application in applications:
                try:
                    application['interview_count'] = int(application.get('interview_count') or 0)
                    application['status'] = (application.get('status') or 'pending').strip().lower()
                except Exception:
                    application['interview_count'] = 0
                    application['status'] = 'pending'
        except Exception as e:
            applications = []
            print(f'⚠️ Error fetching applications for applicant_id={applicant_id}: {e}')
        
        # Get applicant's resumes
        try:
            cursor.execute(
                '''
                SELECT resume_id, file_name, file_path, uploaded_at
                FROM resumes
                WHERE applicant_id = %s
                ORDER BY uploaded_at DESC
                ''',
                (applicant_id,),
            )
            resumes = cursor.fetchall() or []
        except Exception as e:
            resumes = []
            print(f'⚠️ Error fetching resumes for applicant_id={applicant_id}: {e}')
        
        # Get all attachments from applications
        cursor.execute(
            '''
            SELECT DISTINCT r.resume_id, r.file_name, r.file_path, r.uploaded_at, r.file_type
            FROM application_attachments aa
            JOIN applications a ON aa.application_id = a.application_id
            JOIN resumes r ON aa.resume_id = r.resume_id
            WHERE a.applicant_id = %s
            ORDER BY r.uploaded_at DESC
            ''',
            (applicant_id,),
        )
        try:
            app_attachments = cursor.fetchall() or []
        except Exception as e:
            app_attachments = []
            print(f'⚠️ Error fetching application attachments: {e}')
        
        # Combine all resumes (both standalone and from applications)
        try:
            all_resumes = list(resumes) + list(app_attachments)
        except Exception as e:
            all_resumes = list(resumes) if resumes else []
            print(f'⚠️ Error combining resumes lists: {e}')

        # Remove duplicates robustly: prefer unique resume_id when present,
        # otherwise dedupe by file_path + file_name. Maintain recent-first order.
        seen_ids = set()
        seen_keys = set()
        unique_resumes = []
        for resume in all_resumes:
            rid = resume.get('resume_id')
            file_path = (resume.get('file_path') or '')
            file_name = (resume.get('file_name') or '').strip()
            key = None
            if rid:
                key = f'id:{rid}'
            else:
                # fallback key uses normalized path + name
                key = f'path:{file_path.replace('"', '')}|name:{file_name}'

            if key in seen_keys:
                continue
            seen_keys.add(key)
            if rid:
                seen_ids.add(rid)
            unique_resumes.append(resume)
        
        # Format resumes for template
        formatted_resumes = []
        for resume in unique_resumes:
            file_path = (resume.get('file_path') or '').replace('\\', '/')
            file_name = resume.get('file_name') or os.path.basename(file_path) or 'Resume'
            try:
                if file_path:
                    if os.path.isabs(file_path):
                        abs_path = os.path.realpath(file_path)
                    else:
                        abs_path = os.path.realpath(os.path.join(app.instance_path, file_path))
                else:
                    abs_path = None
                file_size_bytes = os.path.getsize(abs_path) if abs_path and os.path.exists(abs_path) else 0
                file_size = format_file_size(file_size_bytes) if file_size_bytes else 'Unknown'
            except Exception:
                file_size = 'Unknown'

            # Build safe view/download URLs. Prefer resume_id-based endpoints when available.
            try:
                import base64
                if resume.get('resume_id'):
                    view_u = url_for('admin_view_resume', resume_id=resume.get('resume_id'))
                    download_u = url_for('admin_download_resume', resume_id=resume.get('resume_id'))
                else:
                    encoded = base64.urlsafe_b64encode((file_path or '').encode()).decode()
                    view_u = url_for('admin_view_resume_by_path', path=encoded)
                    download_u = None
            except Exception:
                view_u = None
                download_u = None

            formatted_resumes.append({
                'resume_id': resume.get('resume_id'),
                'file_name': file_name,
                'file_size': file_size,
                'uploaded_at': format_human_datetime(resume.get('uploaded_at')) if resume.get('uploaded_at') else 'N/A',
                'view_url': view_u,
                'download_url': download_u,
                'is_pdf': file_name.lower().endswith('.pdf'),
                'file_path': file_path,
                'file_type': (resume.get('file_type') or 'resume').lower(),  # Default to 'resume' if not set
            })

        # Group documents by file_type for stable template rendering
        grouped_docs = {'resume': [], 'letter': [], 'license': []}
        for doc in formatted_resumes:
            ftype = (doc.get('file_type') or 'resume').lower()
            grouped_docs.setdefault(ftype, []).append(doc)

        # Build attachments_by_application: map application_id -> list of docs
        # Deduplicate attachments per application (by resume_id or file_path|file_name)
        attachments_by_application = {}
        attachments_seen_by_application = {}
        try:
            cursor.execute(
                '''
                SELECT aa.application_id, r.resume_id, r.file_name, r.file_path, r.uploaded_at, r.file_type
                FROM application_attachments aa
                JOIN resumes r ON aa.resume_id = r.resume_id
                JOIN applications a ON aa.application_id = a.application_id
                WHERE a.applicant_id = %s
                ORDER BY r.uploaded_at DESC
                ''',
                (applicant_id,)
            )
            rows = cursor.fetchall() or []
            for row in rows:
                aid = row.get('application_id')
                if not aid:
                    continue
                # prepare deduplication key for this application
                key = None
                if row.get('resume_id'):
                    key = f"id:{row.get('resume_id')}"
                else:
                    fp = (row.get('file_path') or '').replace('\\', '/')
                    fn = row.get('file_name') or ''
                    key = f"path:{fp}|name:{fn}"
                seen_set = attachments_seen_by_application.setdefault(aid, set())
                if key in seen_set:
                    continue
                seen_set.add(key)
                # format similar to formatted_resumes
                file_path = (row.get('file_path') or '').replace('\\', '/')
                file_name = row.get('file_name') or os.path.basename(file_path) or 'Document'
                try:
                    if file_path:
                        if os.path.isabs(file_path):
                            abs_path = os.path.realpath(file_path)
                        else:
                            abs_path = os.path.realpath(os.path.join(app.instance_path, file_path))
                    else:
                        abs_path = None
                    file_size_bytes = os.path.getsize(abs_path) if abs_path and os.path.exists(abs_path) else 0
                    file_size = format_file_size(file_size_bytes) if file_size_bytes else 'Unknown'
                except Exception:
                    file_size = 'Unknown'

                try:
                    import base64
                    if row.get('resume_id'):
                        view_u = url_for('admin_view_resume', resume_id=row.get('resume_id'))
                        download_u = url_for('admin_download_resume', resume_id=row.get('resume_id'))
                    else:
                        encoded = base64.urlsafe_b64encode((file_path or '').encode()).decode()
                        view_u = url_for('admin_view_resume_by_path', path=encoded)
                        download_u = None
                except Exception:
                    view_u = None
                    download_u = None

                doc = {
                    'resume_id': row.get('resume_id'),
                    'file_name': file_name,
                    'file_size': file_size,
                    'uploaded_at': format_human_datetime(row.get('uploaded_at')) if row.get('uploaded_at') else 'N/A',
                    'view_url': view_u,
                    'download_url': download_u,
                    'is_pdf': (file_name or '').lower().endswith('.pdf'),
                    'file_path': file_path,
                    'file_type': (row.get('file_type') or 'resume').lower(),
                }
                attachments_by_application.setdefault(aid, []).append(doc)
        except Exception as e:
            # if anything goes wrong, leave attachments_by_application empty
            attachments_by_application = {}
            print(f'⚠️ Error fetching attachments_by_application: {e}')
        
        # Mark all applications as viewed when HR/Admin views the applicant profile
        if applications:
            application_ids = [a['application_id'] for a in applications]
            # Check if viewed_at column exists, if not, add it dynamically
            try:
                cursor.execute('SHOW COLUMNS FROM applications LIKE "viewed_at"')
                has_viewed_at = cursor.fetchone() is not None
                
                if not has_viewed_at:
                    # Add viewed_at column if it doesn't exist
                    cursor.execute('ALTER TABLE applications ADD COLUMN viewed_at DATETIME NULL')
                    db.commit()
                
                # Update viewed_at for all applications (only if not already viewed)
                placeholders = ','.join(['%s'] * len(application_ids))
                cursor.execute(
                    f'''
                    UPDATE applications 
                    SET viewed_at = NOW() 
                    WHERE application_id IN ({placeholders}) 
                    AND viewed_at IS NULL
                    ''',
                    tuple(application_ids)
                )
                db.commit()
            except Exception as e:
                print(f'⚠️ Error updating viewed_at: {e}')
                # Continue even if viewed_at column doesn't exist
        
        template = 'hr/applicants.html' if (user.get('role') or '').lower() == 'hr' else 'admin/applicants.html'
        branch_info = None
        if (user.get('role') or '').lower() == 'hr':
            if branch_id:
                branch_rows = fetch_rows('SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s', (branch_id,))
                if branch_rows:
                    branch_info = branch_rows[0]
        
        # Use HR template for HR users, admin template for admin users
        # HR view uses a fixed template to avoid template parsing errors
        view_template = 'hr/view_applicant_fixed.html' if (user.get('role') or '').lower() == 'hr' else 'admin/view_applicant.html'
        
        return render_template(view_template, 
                 applicant=applicant, 
                 applications=applications,
                 resumes=formatted_resumes,
                 documents=grouped_docs,
                 attachments_by_application=attachments_by_application,
                 branch_info=branch_info)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ View applicant error: {exc} (user={user and user.get("id")}, role={user and user.get("role")}, applicant_id={applicant_id})')
        print(f'Full traceback: {error_details}')
        flash('An error occurred while loading applicant details.', 'error')
        return redirect(url_for('applicants'))
    finally:
        if cursor:
            cursor.close()


@app.route('/admin/applicants/<int:applicant_id>/verify', methods=['POST'])
@login_required('admin', 'hr')
def verify_applicant(applicant_id):
    """Verify an applicant's account. HR can only verify applicants from their branch."""
    user = get_current_user()
    db = get_db()
    if not db:
        return jsonify({'success': False, 'error': 'Database error'}), 500


    # NOTE: HR alias route removed from here and will be registered at module level
    
    cursor = db.cursor(dictionary=True)
    try:
        # If HR, verify applicant applied to their branch jobs
        branch_id = get_branch_scope(user)
        if branch_id:
            cursor.execute(
                '''
                SELECT a.applicant_id
                FROM applicants a
                JOIN applications ap ON a.applicant_id = ap.applicant_id
                JOIN jobs j ON ap.job_id = j.job_id
                WHERE a.applicant_id = %s AND j.branch_id = %s
                LIMIT 1
                ''',
                (applicant_id, branch_id),
            )
            if not cursor.fetchone():
                return jsonify({'success': False, 'error': 'You can only verify applicants from your branch.'}), 403
        
        cursor.execute(
            'UPDATE applicants SET is_verified = TRUE, email_verified_at = NOW() WHERE applicant_id = %s',
            (applicant_id,),
        )
        db.commit()
        return jsonify({'success': True})
    except Exception as exc:
        db.rollback()
        print(f'❌ Verify applicant error: {exc}')
        return jsonify({'success': False, 'error': str(exc)}), 500
    finally:
        cursor.close()


@app.route('/admin/add-job-posting')
@login_required('admin', 'hr')
def add_job_posting():
    return redirect(url_for('job_postings'))


@app.route('/admin/add-hr-account')
@login_required('admin')
def add_hr_account():
    return redirect(url_for('hr_accounts'))


@app.route('/admin/reset-all-data', methods=['GET', 'POST'])
@csrf.exempt
@login_required('admin')
def reset_all_data():
    """Reset all system data to zero while keeping admin/HR accounts and branches."""
    user = get_current_user()
    # Only allow admin users (not HR)
    if user.get('role') != 'admin':
        flash('Access denied. Only administrators can reset system data.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'GET':
        # Show confirmation page
        return render_template('admin/reset_data_confirmation.html')
    
    # POST request - validate confirmation text
    confirm_text = request.form.get('confirm_text', '').strip().upper()
    if confirm_text != 'RESET':
        flash('Invalid confirmation. Please type "RESET" to confirm the reset.', 'error')
        return redirect(url_for('reset_all_data'))
    
    # Debug: Check if CSRF token is present
    csrf_token_received = request.form.get('csrf_token')
    if not csrf_token_received:
        flash('CSRF token is missing. Please refresh the page and try again.', 'error')
        return redirect(url_for('reset_all_data'))
    
    # POST request - perform the reset
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    cursor = db.cursor(dictionary=True)
    
    try:
        # Disable foreign key checks temporarily
        cursor.execute('SET FOREIGN_KEY_CHECKS = 0')
        
        # Delete all data from tables (in order to respect foreign keys)
        # Keep: branches, admin table, users table (but only admin/hr users)
        
        tables_to_clear = [
            'activity_logs',
            'results',
            'interviews',
            'notifications',
            'saved_jobs',
            'applications',
            'jobs',
            'resumes',
            'profile_changes',
            'password_resets',
            'auth_sessions',
        ]
        
        deleted_counts = {}
        
        for table in tables_to_clear:
            try:
                cursor.execute(f'DELETE FROM {table}')
                deleted_counts[table] = cursor.rowcount
                print(f'✅ Cleared {table}: {deleted_counts[table]} rows')
            except Exception as e:
                print(f'⚠️ Error clearing {table}: {e}')
                deleted_counts[table] = 0
        
        # Delete all applicants (but keep admin/hr accounts in users table)
        cursor.execute('DELETE FROM applicants')
        applicants_deleted = cursor.rowcount
        deleted_counts['applicants'] = applicants_deleted
        print(f'✅ Cleared applicants: {applicants_deleted} rows')
        
        # Delete only applicant users from users table (keep admin and hr users)
        cursor.execute("DELETE FROM users WHERE user_type = 'applicant'")
        applicant_users_deleted = cursor.rowcount
        deleted_counts['users (applicants only)'] = applicant_users_deleted
        print(f'✅ Cleared applicant users: {applicant_users_deleted} rows')
        
        # Reset AUTO_INCREMENT for all cleared tables
        auto_increment_tables = [
            'activity_logs', 'results', 'interviews', 'notifications', 
            'saved_jobs', 'applications', 'jobs', 'resumes', 
            'profile_changes', 'password_resets', 'auth_sessions', 'applicants'
        ]
        
        for table in auto_increment_tables:
            try:
                cursor.execute(f'ALTER TABLE {table} AUTO_INCREMENT = 1')
                print(f'✅ Reset AUTO_INCREMENT for {table}')
            except Exception as e:
                print(f'⚠️ Could not reset AUTO_INCREMENT for {table}: {e}')
        
        # Re-enable foreign key checks
        cursor.execute('SET FOREIGN_KEY_CHECKS = 1')
        
        # Commit the transaction
        db.commit()
        
        total_deleted = sum(deleted_counts.values())
        flash(f'✅ System data reset successfully! Deleted {total_deleted} records. Admin/HR accounts and branches have been preserved.', 'success')
        print(f'✅ System reset complete. Total records deleted: {total_deleted}')
        
        return redirect(url_for('admin_dashboard'))
        
    except Exception as e:
        db.rollback()
        print(f'❌ Error resetting system data: {e}')
        import traceback
        traceback.print_exc()
        flash(f'Error resetting system data: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard'))
    finally:
        cursor.close()


@app.route('/admin/profile', methods=['GET', 'POST'])
@login_required('admin', 'hr')
def admin_profile():
    """Admin profile management."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('admin_dashboard'))

    cursor = db.cursor(dictionary=True)
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_profile':
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                
                if not full_name or not email:
                    flash('Full name and email are required.', 'error')
                else:
                    # Check if email is already in use in users table
                    cursor.execute(
                        '''
                        SELECT u.user_id 
                        FROM users u
                        JOIN admins a ON a.user_id = u.user_id
                        WHERE u.email = %s AND a.admin_id <> %s
                        LIMIT 1
                        ''',
                        (email, user.get('id')),
                    )
                    if cursor.fetchone():
                        flash('Email address is already in use.', 'error')
                    else:
                        # Get user_id first
                        cursor.execute(
                            'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                            (user.get('id'),),
                        )
                        admin_record = cursor.fetchone()
                        if admin_record:
                            user_id = admin_record['user_id']
                            # Update email in users table
                            cursor.execute(
                                'UPDATE users SET email = %s WHERE user_id = %s',
                                (email, user_id),
                            )
                            # Update full_name and email in admins table
                            cursor.execute(
                                'UPDATE admins SET full_name = %s, email = %s WHERE admin_id = %s',
                                (full_name, email, user.get('id')),
                            )
                            # Determine if user is admin or HR
                            user_role = user.get('role', '').lower()
                            role_label = 'Admin' if user_role == 'admin' else 'HR'
                            branch_label = user.get('branch_name') or 'Unassigned Branch'
                            
                            # Only show branch label for HR, not for admin
                            if user_role == 'admin':
                                notification_msg = f'Admin {full_name} updated their profile information.'
                            else:
                                notification_msg = f'HR {full_name} ({branch_label}) updated their profile information.'
                            
                            create_admin_notification(cursor, notification_msg)
                            db.commit()
                            session['user_name'] = full_name
                            session['user_email'] = email
                            flash('Profile updated successfully.', 'success')
                        else:
                            flash('Account not found.', 'error')
            
            elif action == 'change_password':
                current_password = request.form.get('current_password', '').strip()
                new_password = request.form.get('new_password', '').strip()
                confirm_password = request.form.get('confirm_password', '').strip()
                
                # Get user_id from admins table, then get password_hash from users table
                cursor.execute(
                    'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                    (user.get('id'),),
                )
                admin_record = cursor.fetchone()
                
                if not admin_record:
                    flash('Account not found.', 'error')
                else:
                    user_id = admin_record['user_id']
                    cursor.execute(
                        'SELECT password_hash FROM users WHERE user_id = %s LIMIT 1',
                        (user_id,),
                    )
                    record = cursor.fetchone()
                    
                    if not record or not check_password(record.get('password_hash'), current_password):
                        flash('Current password is incorrect.', 'error')
                    elif len(new_password) < 6:
                        flash('New password must be at least 6 characters.', 'error')
                    elif new_password != confirm_password:
                        flash('New passwords do not match.', 'error')
                    else:
                        try:
                            # Initiate password change verification via OTP emailed to account
                            cursor.execute('SELECT email FROM admins WHERE admin_id = %s LIMIT 1', (user.get('id'),))
                            admin_email_row = cursor.fetchone()
                            admin_email = admin_email_row.get('email') if admin_email_row else None
                            if not admin_email:
                                flash('Unable to determine account email for verification.', 'error')
                            else:
                                # Immediately apply password change for admin/hr (no OTP)
                                try:
                                    new_hash = hash_password(new_password)
                                    cursor.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', (new_hash, user_id))
                                    db.commit()
                                    # Notify admin via email about password change
                                    try:
                                        from datetime import datetime
                                        subject = 'Your account password was changed'
                                        body = f"Hi {session.get('user_name') or ''},\n\nYour account password was successfully changed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\nIf you did not perform this action, please contact support or reset your password immediately.\n"
                                        if admin_email:
                                            send_email(admin_email, subject, body)
                                    except Exception:
                                        pass
                                    flash('Your password has been changed successfully.', 'success')
                                    return redirect(url_for('admin_profile'))
                                except Exception as exc:
                                    db.rollback()
                                    print(f'❌ Error changing admin password: {exc}')
                                    flash('Unable to change password now.', 'error')
                                    return redirect(url_for('admin_profile'))
                        except Exception as exc:
                            db.rollback()
                            print(f'❌ Error initiating admin/HR password change OTP: {exc}')
                            flash('Unable to initiate password change verification. Please try again later.', 'error')
            
            return redirect(url_for('admin_profile'))
        
        # Get profile data from admins and users tables
        cursor.execute(
            '''
            SELECT 
                a.admin_id,
                a.full_name,
                a.email,
                u.user_id,
                u.user_type,
                u.is_active,
                a.last_login,
                a.created_at
            FROM admins a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.admin_id = %s
            LIMIT 1
            ''',
            (user.get('id'),),
        )
        profile = cursor.fetchone()
        if profile:
            # Map user_type to role for backward compatibility
            profile['role'] = 'admin' if profile.get('user_type') == 'super_admin' else 'hr'
        
        # Get login history using user_id from users table
        login_history = []
        if profile and profile.get('user_id'):
            try:
                cursor.execute('SHOW COLUMNS FROM auth_sessions')
                session_columns_raw = cursor.fetchall()
                session_columns = {row.get('Field') if isinstance(row, dict) else row[0] for row in session_columns_raw}
                
                if 'last_activity' in session_columns and 'logout_time' in session_columns:
                    logout_expr = 'COALESCE(last_activity, logout_time)'
                elif 'logout_time' in session_columns:
                    logout_expr = 'logout_time'
                elif 'last_activity' in session_columns:
                    logout_expr = 'last_activity'
                else:
                    logout_expr = 'NULL'

                # Determine login expression (login_time vs created_at vs last_login)
                if 'login_time' in session_columns:
                    login_expr = 'login_time'
                elif 'created_at' in session_columns:
                    login_expr = 'created_at'
                elif 'last_login' in session_columns:
                    login_expr = 'last_login'
                else:
                    login_expr = 'NULL'

                cursor.execute(
                    f'''
                    SELECT {login_expr} AS login_time, {logout_expr} AS logout_time, COALESCE(is_active, 1) AS is_active
                    FROM auth_sessions
                    WHERE user_id = %s
                    ORDER BY {login_expr} DESC
                    LIMIT 10
                    ''',
                    (profile['user_id'],),
                )
                login_history = []
                for row in cursor.fetchall() or []:
                    is_active = bool(row.get('is_active', 1))
                    logout_value = format_human_datetime(row.get('logout_time')) if row.get('logout_time') else None
                    login_history.append({
                        'login_time': format_human_datetime(row.get('login_time')),
                        'logout_time': logout_value,
                        'is_active': is_active,
                    })
            except Exception as hist_error:
                print(f'⚠️ Login history error: {hist_error}')
                login_history = []
        
        return render_template('admin/profile.html', profile=profile, login_history=login_history, branches=fetch_branches())
    except Exception as exc:
        db.rollback()
        print(f'❌ Admin profile error: {exc}')
        flash('Unable to load profile.', 'error')
        return redirect(url_for('admin_dashboard'))
    finally:
        cursor.close()


@app.route('/hr/profile', methods=['GET', 'POST'])
@login_required('hr')
def hr_profile():
    """HR profile management."""
    user = get_current_user()
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return redirect(url_for('hr_dashboard'))

    cursor = db.cursor(dictionary=True)
    try:
        if request.method == 'POST':
            action = request.form.get('action')
            
            if action == 'update_profile':
                full_name = request.form.get('full_name', '').strip()
                email = request.form.get('email', '').strip().lower()
                
                if not full_name or not email:
                    flash('Full name and email are required.', 'error')
                else:
                    # Check if email is already in use in users table
                    cursor.execute(
                        '''
                        SELECT u.user_id 
                        FROM users u
                        JOIN admins a ON a.user_id = u.user_id
                        WHERE u.email = %s AND a.admin_id <> %s
                        LIMIT 1
                        ''',
                        (email, user.get('id')),
                    )
                    if cursor.fetchone():
                        flash('Email address is already in use.', 'error')
                    else:
                        # Get user_id first
                        cursor.execute(
                            'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                            (user.get('id'),),
                        )
                        admin_record = cursor.fetchone()
                        if admin_record:
                            user_id = admin_record['user_id']
                            # Update email in users table
                            cursor.execute(
                                'UPDATE users SET email = %s WHERE user_id = %s',
                                (email, user_id),
                            )
                            # Update full_name and email in admins table
                            cursor.execute(
                                'UPDATE admins SET full_name = %s, email = %s WHERE admin_id = %s',
                                (full_name, email, user.get('id')),
                            )
                            db.commit()
                            session['user_name'] = full_name
                            session['user_email'] = email
                            flash('Profile updated successfully.', 'success')
                        else:
                            flash('Account not found.', 'error')
            
            elif action == 'change_password':
                current_password = request.form.get('current_password', '').strip()
                new_password = request.form.get('new_password', '').strip()
                confirm_password = request.form.get('confirm_password', '').strip()
                
                # Get user_id from admins table, then get password_hash from users table
                cursor.execute(
                    'SELECT user_id FROM admins WHERE admin_id = %s LIMIT 1',
                    (user.get('id'),),
                )
                admin_record = cursor.fetchone()
                
                if not admin_record:
                    flash('Account not found.', 'error')
                else:
                    user_id = admin_record['user_id']
                    cursor.execute(
                        'SELECT password_hash FROM users WHERE user_id = %s LIMIT 1',
                        (user_id,),
                    )
                    record = cursor.fetchone()
                    
                    if not record or not check_password(record.get('password_hash'), current_password):
                        flash('Current password is incorrect.', 'error')
                    elif len(new_password) < 6:
                        flash('New password must be at least 6 characters.', 'error')
                    elif new_password != confirm_password:
                        flash('New passwords do not match.', 'error')
                    else:
                        try:
                            # Initiate password change verification via OTP for HR
                            cursor.execute('SELECT email FROM admins WHERE admin_id = %s LIMIT 1', (user.get('id'),))
                            admin_email_row = cursor.fetchone()
                            hr_email = admin_email_row.get('email') if admin_email_row else None
                            if not hr_email:
                                cursor.execute('SELECT email FROM users WHERE user_id = %s LIMIT 1', (user_id,))
                                urow = cursor.fetchone()
                                hr_email = urow.get('email') if urow else None

                            if not hr_email:
                                flash('Unable to determine account email for verification.', 'error')
                            else:
                                # Immediately apply password change for HR (no OTP)
                                try:
                                    new_hash = hash_password(new_password)
                                    cursor.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', (new_hash, user_id))
                                    db.commit()
                                    # Notify HR user via email about password change
                                    try:
                                        from datetime import datetime
                                        subject = 'Your account password was changed'
                                        body = f"Hi {session.get('user_name') or ''},\n\nYour account password was successfully changed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\nIf you did not perform this action, please contact support or reset your password immediately.\n"
                                        if hr_email:
                                            send_email(hr_email, subject, body)
                                    except Exception:
                                        pass
                                    flash('Your password has been changed successfully.', 'success')
                                    return redirect(url_for('hr_profile'))
                                except Exception as exc:
                                    db.rollback()
                                    print(f'❌ Error changing HR password: {exc}')
                                    flash('Unable to change password now.', 'error')
                                    return redirect(url_for('hr_profile'))
                        except Exception as exc:
                            db.rollback()
                            print(f'❌ Error initiating HR password change OTP: {exc}')
                            flash('Unable to initiate password change verification. Please try again later.', 'error')
            
            return redirect(url_for('hr_profile'))
        
        # Get profile data from admins and users tables
        cursor.execute(
            '''
            SELECT 
                a.admin_id,
                a.full_name,
                a.email,
                u.user_id,
                u.user_type,
                u.is_active,
                a.last_login,
                a.created_at
            FROM admins a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.admin_id = %s
            LIMIT 1
            ''',
            (user.get('id'),),
        )
        profile = cursor.fetchone()
        if profile:
            profile['role'] = 'hr'
        
        # Get branch info from session (HR users manage branches via session, not direct branch_id)
        branch_info = None
        branch_id = session.get('branch_id')
        if branch_id:
            cursor.execute(
                'SELECT branch_id, branch_name, address FROM branches WHERE branch_id = %s LIMIT 1',
                (branch_id,),
            )
            branch_info = cursor.fetchone()
        
        return render_template('hr/profile.html', profile=profile, login_history=[], branch_info=branch_info)
    except Exception as exc:
        db.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f'❌ HR profile error: {exc}')
        print(f'Full traceback: {error_details}')
        flash('Unable to load profile.', 'error')
        return redirect(url_for('hr_dashboard'))
    finally:
        cursor.close()


@app.route('/admin/security', methods=['GET', 'POST'])
@login_required('admin')
def admin_security():
    """Security administration and audit logs."""
    db = get_db()
    if not db:
        flash('Database connection error.', 'error')
        return render_template('admin/security.html', audit_logs=[], activity_logs=[], security_events=[])
    
    cursor = db.cursor(dictionary=True)
    try:
        # Get activity logs from activity_logs table - show only HR users
        activity_logs = []
        try:
            cursor.execute(
                '''
                SELECT 
                    al.log_id,
                    al.action,
                    al.description,
                    al.target_table,
                    al.target_id,
                    al.logged_at,
                    COALESCE(a.full_name, 'System') AS admin_name,
                    a.email AS admin_email,
                    'HR' AS user_role
                FROM activity_logs al
                LEFT JOIN admins a ON a.admin_id = al.user_id
                WHERE a.admin_id IS NOT NULL
                ORDER BY al.logged_at DESC
                LIMIT 500
                '''
            )
            activity_logs = cursor.fetchall()
            # Format timestamps
            for log in activity_logs:
                if log.get('logged_at'):
                    log['logged_at'] = format_human_datetime(log['logged_at'])
        except Exception:
            # Table might not exist, that's okay
            pass
        
        # Get security events (failed logins, suspicious activity)
        # For now, we'll identify suspicious patterns from auth_sessions
        cursor.execute(
            '''
            SELECT 
                u.email,
                COUNT(*) AS attempt_count,
                MAX(s.created_at) AS last_attempt
            FROM auth_sessions s
            JOIN users u ON u.user_id = s.user_id
            WHERE s.created_at >= DATE_SUB(NOW(), INTERVAL 24 HOUR)
            GROUP BY u.user_id, u.email
            HAVING attempt_count > 10
            ORDER BY attempt_count DESC
            LIMIT 50
            '''
        )
        security_events = cursor.fetchall()
        # Format last_attempt timestamps for security events
        for ev in security_events:
            try:
                if ev.get('last_attempt'):
                    ev['last_attempt'] = format_human_datetime(ev.get('last_attempt'))
            except Exception:
                pass
        
        return render_template('admin/security.html', 
                             activity_logs=activity_logs,
                             security_events=security_events)
    except Exception as exc:
        db.rollback()
        print(f'❌ Security admin error: {exc}')
        import traceback
        traceback.print_exc()
        flash('Unable to load security logs.', 'error')
        return render_template('admin/security.html', audit_logs=[], activity_logs=[], security_events=[])
    finally:
        cursor.close()


@app.route('/jobs')
def jobs():
    """Public job listings with filters and smart matching.
    
    Applicants can see ALL jobs from ALL branches by default.
    Branch filtering only applies when explicitly selected via branch_id filter.
    """
    filters = {
        'keyword': request.args.get('keyword', '').strip(),
        'branch_id': request.args.get('branch_id', type=int),
        'position_id': request.args.get('position_id', type=int),
        'saved_only': request.args.get('saved_only', '').strip(),
    }
    # Keep saved_only if it's checked (value is '1'), otherwise remove it
    if filters.get('saved_only') != '1':
        filters.pop('saved_only', None)
    # Only keep filters with actual values (remove empty/None)
    filters = {k: v for k, v in filters.items() if v}
    
    applicant_id = session.get('user_id') if is_logged_in() and session.get('user_role') == 'applicant' else None
    
    # IMPORTANT: Pass filters dict (even if empty) or None - applicants see ALL branches by default
    # Only filter by branch if branch_id is explicitly provided in request
    job_listings = fetch_open_jobs(filters if filters else None, applicant_id)
    
    branches = fetch_branches()
    positions = fetch_positions()
    
    return render_template(
        'applicant/jobs.html',
        jobs=job_listings,
        branches=branches,
        positions=positions,
        current_filters=filters,
    )


@app.route('/about')
def about():
    return render_template('about.html')


@app.route('/logout')
def logout():
    """Logout user and redirect to login page."""
    # Logout user first (updates database)
    if is_logged_in():
        logout_user()
    
    # Force clear ALL session data to prevent any redirect loops
    session.clear()
    
    # Ensure session is completely empty
    for key in list(session.keys()):
        session.pop(key, None)
    
    # Flash logout message
    flash('You have been logged out successfully.', 'success')
    
    # Always redirect to login page using immediate redirect
    return immediate_redirect(url_for('login', _external=True))

# Handle favicon requests to prevent 404 errors
@app.route('/favicon.ico')
def favicon():
    """Handle favicon requests."""
    try:
        return send_from_directory('static/images', 'whitehat_logo.jpg', mimetype='image/jpeg')
    except Exception:
        # Return empty response if favicon not found
        return ('', 204)

# Handle Chrome DevTools requests to prevent 404 errors in logs
@app.route('/.well-known/appspecific/com.chrome.devtools.json')
def chrome_devtools():
    """Handle Chrome DevTools requests."""
    return ('', 204)


# Handle Vite dev client requests to avoid noisy 404s in logs
@app.route('/@vite/client')
def vite_client_placeholder():
    """Return an empty response for Vite dev client placeholder requests."""
    # Some browsers/tools may request this during development;
    # returning 204 prevents unnecessary 404 logs without affecting functionality.
    return ('', 204)


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    """Handle CSRF token errors with proper JSON for AJAX and 302 for normal requests."""
    log.exception(f'❌ CSRF Error: {e}')
    flash('Security error: Your session has expired. Please refresh the page and try again.', 'error')
    wants_json = (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or request.accept_mimetypes.accept_json
        or request.is_json
        or ('application/json' in (request.headers.get('Accept') or '').lower())
    )
    if request.endpoint == 'login' or '/login' in (request.path or ''):
        if wants_json:
            return jsonify({'success': False, 'error': 'CSRF failed', 'detail': 'Session expired'}), 401
        return render_template('login.html'), 400
    if wants_json:
        return jsonify({'success': False, 'error': 'CSRF failed', 'detail': 'Session expired'}), 401
    referrer = request.referrer or url_for('index', _external=True)
    return immediate_redirect(referrer)


# Security headers
@app.after_request
def set_security_headers(response):
    """Add security headers to all responses."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    
    # Use Content-Security-Policy instead of X-Frame-Options
    response.headers['Content-Security-Policy'] = "frame-ancestors 'none'"
    
    # Only add HSTS in production with HTTPS
    if not app.debug and request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains; preload'
    
    # Set cache control for static resources
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
    
    # Set UTF-8 charset for HTML responses
    if response.mimetype and 'text/html' in response.mimetype:
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
    elif response.mimetype == 'application/json':
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
    
    # Remove redirect pages - ensure all redirects are immediate HTTP 302 redirects
    if response.status_code in (301, 302, 303, 307, 308):
        if 'Location' in response.headers:
            # Add cache control headers to prevent redirect page caching
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            
            # Remove redirect page HTML if present
            if hasattr(response, 'data') and response.data:
                try:
                    body_text = response.data.decode('utf-8', errors='ignore')
                    # Check if it's Flask's default redirect page
                    if 'Redirecting' in body_text or 'should be redirected' in body_text.lower() or 'redirect automatically' in body_text.lower():
                        # Replace redirect page with empty body for immediate redirect
                        response.data = b''
                except Exception:
                    # Ignore errors decoding response body
                    pass
    
    return response


if __name__ == '__main__':
    import os
    import sys
    from threading import Thread
    import time
    
    print('[*] Starting J&T Express Recruitment System...')
    
    # Flag to track initialization status
    init_complete = False
    init_error = None
    
    def initialize_in_background():
        """Initialize database and default accounts in background thread."""
        global init_complete, init_error
        try:
            with app.app_context():
                print('[*] Checking database connection...')
                try:
                    # Try to connect (will fail fast if MySQL not running due to 3s timeout)
                    db = get_db()
                    if not db:
                        print('⚠️ Warning: Database connection failed.')
                        print('   Please ensure MySQL is running in XAMPP Control Panel.')
                        print('   Server will start, but database features may not work.')
                        init_complete = True
                        return
                    
                    # Quick connection test
                    try:
                        cursor = db.cursor()
                        cursor.execute('SELECT 1 AS test')
                        cursor.fetchone()
                        cursor.close()
                        print('[*] Database connected successfully.')
                    except Exception as test_err:
                        print(f'⚠️ Database connection test failed: {test_err}')
                        init_complete = True
                        return
                    
                    # Initialize default accounts (skip schema checks to speed up)
                    print('[*] Initializing default accounts...')
                    try:
                        # Temporarily skip schema compatibility check for faster startup
                        ensure_default_accounts()
                        print('[*] Database initialization complete.')
                    except Exception as init_err:
                        print(f'⚠️ Warning during account initialization: {init_err}')
                        print('   This is usually not critical - server will continue.')
                    
                    init_complete = True
                except TimeoutError:
                    print('⚠️ Database connection timed out.')
                    print('   Please check if MySQL is running in XAMPP.')
                    init_complete = True
                except Exception as db_err:
                    print(f'⚠️ Warning: Database error: {db_err}')
                    print('   Server will start, but database features may not work.')
                    print('   Please ensure MySQL is running in XAMPP Control Panel.')
                    init_complete = True
        except Exception as e:
            init_error = str(e)
            print(f'⚠️ Warning: Error during initialization: {e}')
            print('[*] Server will continue to start, but some features may not work.')
            traceback.print_exc()
            init_complete = True
    
    # Start initialization in background thread with timeout
    init_thread = Thread(target=initialize_in_background, daemon=True)
    init_thread.start()
    
    # Wait for initialization with shorter timeout (max 3 seconds)
    # If MySQL is running, it should connect quickly
    timeout = 3
    start_time = time.time()
    while not init_complete and (time.time() - start_time) < timeout:
        time.sleep(0.1)
    
    if not init_complete:
        print('[*] Starting server immediately - database initialization continues in background.')
        print('   If you see database errors, ensure MySQL is running in XAMPP Control Panel.')
    
    try:
        debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
        port = int(os.environ.get('PORT', 5000))
        host = os.environ.get('HOST', '0.0.0.0')
        
        print(f'[*] Starting Flask server on {host}:{port} (debug={debug_mode})...')
        print('[*] Server is ready! Press Ctrl+C to stop.')
        if init_error:
            print(f'[*] Note: {init_error}')
        
        # Disable reloader to prevent issues
        app.run(debug=debug_mode, host=host, port=port, use_reloader=False, threaded=True)
    except KeyboardInterrupt:
        print('\n[*] Server stopped by user.')
        sys.exit(0)
    except OSError as e:
        if 'Address already in use' in str(e) or 'address is already in use' in str(e).lower():
            print(f'❌ Error: Port {port} is already in use.')
            log.warning('   Please stop the other process or use a different port.')
            log.warning('   Set PORT environment variable to use a different port.')
        else:
            log.exception('❌ Critical error starting server: %s', e)
        sys.exit(1)
    except Exception as run_error:
        print(f'❌ Critical error starting server: {run_error}')
        traceback.print_exc()
        sys.exit(1)
